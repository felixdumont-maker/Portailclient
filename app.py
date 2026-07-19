# app.py
import os
import re
import math
import json
import sqlite3
import pathlib
import uuid
import secrets
from functools import wraps
from typing import Tuple
from datetime import datetime, timedelta
import unicodedata
from rapidfuzz import fuzz

from flask import (
    Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify, make_response, Response
)
from flask_bcrypt import Bcrypt
from flask_mail import Mail, Message
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import CSRFProtect
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadTimeSignature, BadSignature
from authlib.integrations.flask_client import OAuth
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename

# ── Chargement .env (PythonAnywhere : .env à la racine du projet) ──
from dotenv import load_dotenv
from pathlib import Path
ENV_PATH = (Path(__file__).resolve().parent / ".." / ".env").resolve()
load_dotenv(dotenv_path=ENV_PATH)
from drive_service import create_folder, upload_file, get_folder_link, list_files_in_folder, list_subfolders, make_folder_public, make_file_public, make_file_public, share_folder_with_user
from calendar_service import (
    create_production_event, delete_production_event,
    get_available_meeting_slots, format_slot_fr, create_meeting_event,
    load_watch_state, save_watch_state, register_calendar_watch,
    stop_calendar_watch, list_changed_events,
)

# Tâches automatiques par type de service
TACHES_PAR_SERVICE = {
    'web': [
        'Confirmer le cahier des charges',
        'Wireframe des pages',
        'Intégrer le contenu client',
        'Tests mobile & tablette',
        'Soumettre pour révision',
        'Mise en ligne',
        'Envoyer le lien final',
    ],
    'graphisme': [
        'Brief créatif',
        'Propositions & moodboard',
        'Déclinaisons visuelles',
        'Export fichiers finaux',
        'Envoyer le kit graphique',
    ],
    'photo': [
        'Confirmer la séance',
        'Liste de shots',
        'Transfert & tri',
        'Retouche & sélection',
        'Livraison Drive',
        'Courriel de livraison',
    ],
    'video': [
        'Confirmer le script',
        'Vérifier le matériel',
        'Montage & étalonnage',
        'Soumettre la révision',
        'Export & livraison',
    ],
    'immobilier': [
        'Confirmer l\'accès',
        'Préparer le matériel',
        'Retouche des photos',
        'Livraison dans les 48h',
    ],
    'info': [
        'Analyser les besoins',
        'Préparer l\'environnement',
        'Tester la solution',
        'Former le client',
        'Documenter & livrer',
    ],
    'default': [
        'Confirmer les détails',
        'Préparer les livrables',
        'Soumettre pour révision',
        'Livraison finale',
    ],
}

# Pistes de révision génériques par service — insérées automatiquement quand un projet
# passe en révision sans items personnalisés saisis à la main (voir start_revision /
# api_admin_start_revision). Rédigées pour un client qui reçoit son tout premier site web.
REVISION_ITEMS_PAR_SERVICE = {
    'Site Web Vitrine': [
        'Vérifier que votre nom, votre adresse et vos coordonnées sont exacts partout sur le site',
        'Relire les textes de chaque page (accueil, services, à propos, contact) pour toute correction souhaitée',
        'Confirmer que les photos et le logo vous représentent bien',
        'Vérifier les prix et la description de vos services',
        "Tester le formulaire de contact (envoyer un message pour vous assurer qu'il fonctionne)",
        "Parcourir le site sur votre téléphone pour vérifier que tout s'affiche bien",
    ],
}

from invoice_service import creer_facture_projet, generer_numero_facture
from email_templates import (
    email_bienvenue, email_projet_cree, email_documents_requis,
    email_travaux_en_cours, email_en_revision, email_livraison,
    email_archive, email_nouvelle_facture, _base_confirm, email_documents_recus,
    email_travaux_en_cours_avec_date, email_annulation, _invitation_client,
    email_nouveau_client, email_identite_visuelle_prete, email_reset_password,
    email_revision_site_web
)
# ───────────────────────────────────────────────────────────
# App init
# ───────────────────────────────────────────────────────────
app = Flask(__name__)
app.config['PREFERRED_URL_SCHEME'] = 'https'  # à mettre APRÈS la création de app
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)  # derrière proxy (PA)

# ───────────────────────────────────────────────────────────
# Config (tirée du .env)
# ───────────────────────────────────────────────────────────
class Config:
    SECRET_KEY = os.getenv("SECRET_KEY", "change-me")
    SESSION_COOKIE_HTTPONLY = True
    SESSION_COOKIE_SAMESITE = "Lax"
    SESSION_COOKIE_SECURE = os.getenv("FLASK_ENV", "development") == "production"
    SESSION_COOKIE_NAME = "cocktailmedia_session"
    PERMANENT_SESSION_LIFETIME = timedelta(days=30)

    MAIL_SERVER = os.getenv("MAIL_SERVER", "smtp.gmail.com")
    MAIL_PORT = int(os.getenv("MAIL_PORT", "587"))
    MAIL_USE_TLS = True
    MAIL_USERNAME = os.getenv("MAIL_USERNAME", "")
    MAIL_PASSWORD = os.getenv("MAIL_PASSWORD", "")
    MAIL_DEFAULT_SENDER = os.getenv("MAIL_DEFAULT_SENDER", os.getenv("MAIL_USERNAME", ""))

    GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "")
    GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "")

    BASE_DIR = os.path.abspath(os.path.dirname(__file__))
    INSTANCE_DIR = os.path.join(BASE_DIR, "instance")
    os.makedirs(INSTANCE_DIR, exist_ok=True)
    DB_PATH = os.getenv("DB_PATH", os.path.join(INSTANCE_DIR, "portail.db"))

    UPLOAD_ROOT = os.getenv("UPLOAD_ROOT", os.path.join(os.getcwd(), "uploads"))
    MAX_CONTENT_LENGTH = int(os.getenv("MAX_UPLOAD_MB", 100)) * 1024 * 1024
    ALLOWED_EXTENSIONS = set(
        (os.getenv("ALLOWED_EXTENSIONS", "pdf,png,jpg,jpeg,webp,doc,docx,xls,xlsx,zip"))
        .replace(" ", "").split(",")
    )

app.config.from_object(Config)

if app.config['SECRET_KEY'] in ('change-me', '', None):
    raise RuntimeError("SECRET_KEY n'est pas défini dans .env — les tokens HMAC ne sont pas sécurisés.")

# Ensure upload root exists
pathlib.Path(app.config["UPLOAD_ROOT"]).mkdir(parents=True, exist_ok=True)

bcrypt = Bcrypt(app)
mail = Mail(app)
s = URLSafeTimedSerializer(app.config['SECRET_KEY'])
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri="redis://redis:6379",
)

@app.errorhandler(429)
def ratelimit_handler(e):
    return jsonify({"error": "Trop de tentatives. Attendez une minute avant de réessayer."}), 429

csrf = CSRFProtect(app)

# ───────────────────────────────────────────────────────────
# Security headers
# ───────────────────────────────────────────────────────────
@app.after_request
def set_security_headers(response):
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self'; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com data:; "
        "img-src 'self' data: blob:; "
        "connect-src 'self' https://fonts.googleapis.com https://fonts.gstatic.com; "
        "frame-ancestors 'none';"
    )
    return response

# OAuth
oauth = OAuth(app)
oauth.register(
    name="google",
    client_id=app.config["GOOGLE_CLIENT_ID"],
    client_secret=app.config["GOOGLE_CLIENT_SECRET"],
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={"scope": "openid email profile"},
)


# ───────────────────────────────────────────────────────────
# Google OAuth Routes
# ───────────────────────────────────────────────────────────

@app.route("/login/google", endpoint="google_login")
def oauth_login_google():
    redirect_uri = url_for("google_callback", _external=True)
    return oauth.google.authorize_redirect(redirect_uri)

@app.route("/google-callback", endpoint="google_callback")
def oauth_google_callback():
    token = oauth.google.authorize_access_token()
    user_info = token.get("userinfo") or oauth.google.parse_id_token(token)

    email = (user_info.get("email") or "").strip().lower()
    nom   = user_info.get("name") or ""
    if not email:
        flash("Impossible de récupérer l'adresse email Google.", "error")
        return redirect(url_for("accueil"))

    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    if not user:
        conn.execute("""
            INSERT INTO clients (nom_complet, email, auth_provider, is_email_confirmed, is_admin)
            VALUES (?, ?, 'google', 1, 0)
        """, (nom or email.split("@")[0], email))
        conn.commit()
        user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    elif not int(user['is_email_confirmed'] or 0):
        # Google a vérifié l'email — on peut confirmer le compte
        conn.execute("UPDATE clients SET is_email_confirmed = 1 WHERE id = ?", (user['id'],))
        conn.commit()
        user = conn.execute("SELECT * FROM clients WHERE id = ?", (user['id'],)).fetchone()
    conn.close()

    session.permanent = True
    session['user_id'] = user['id']
    session['user_name'] = user['nom_complet']
    session['is_admin'] = bool(user['is_admin'])
    session['has_outils'] = bool(user['has_outils']) if 'has_outils' in user.keys() else False
    session['has_entrainement'] = bool(user['has_entrainement']) if 'has_entrainement' in user.keys() else False

    flash("Connexion Google réussie!", "success")
    return redirect(url_for('admin_dashboard' if session['is_admin'] else 'dashboard'))


# ───────────────────────────────────────────────────────────
# DB helpers
# ───────────────────────────────────────────────────────────
def get_db_connection():
    # isolation_level=None = autocommit — conn.commit() est no-op, chaque stmt s'auto-commit
    conn = sqlite3.connect(app.config["DB_PATH"], timeout=10, isolation_level=None)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn

# ───────────────────────────────────────────────────────────
# Comptabilité — Catégories de dépenses (travailleur autonome QC)
# Chaque catégorie est rattachée à sa ligne fiscale sur les deux formulaires :
#   • t2125 = formulaire fédéral T2125 (ARC), partie 4/5 « Dépenses »
#   • tp80  = formulaire québécois TP-80 (Revenu Québec), section 3 « Dépenses »
# Numéros vérifiés sur T2125 et TP-80 (2025-10). Certains regroupements
# (abonnements/logiciels, frais bancaires) sont des choix prudents à confirmer
# avec la comptable selon la nature exacte de la dépense.
# ───────────────────────────────────────────────────────────
LIGNES_FISCALES = {
    'Fournitures et matériel':            {'t2125': '8811', 'tp80': '226'},
    'Marchandise et matières premières':  {'t2125': '8320', 'tp80': '134'},
    'Frais de bureau':                    {'t2125': '8810', 'tp80': '222'},
    'Publicité et marketing':             {'t2125': '8521', 'tp80': '200'},
    'Transport et déplacements':          {'t2125': '9200', 'tp80': '236'},
    'Repas et représentation':            {'t2125': '8523', 'tp80': '218'},
    'Frais de véhicule':                  {'t2125': '9281', 'tp80': '220'},
    'Télécommunications':                 {'t2125': '9220', 'tp80': '238'},
    'Honoraires professionnels':          {'t2125': '8860', 'tp80': '228'},
    'Abonnements et logiciels':           {'t2125': '8760', 'tp80': '204'},
    'Loyer et local':                     {'t2125': '8910', 'tp80': '232'},
    'Assurances':                         {'t2125': '8690', 'tp80': '210'},
    'Frais bancaires':                    {'t2125': '8710', 'tp80': '212'},
    'Autre':                              {'t2125': '9270', 'tp80': '246'},
}
CATEGORIES_DEPENSES = list(LIGNES_FISCALES.keys())

# ───────────────────────────────────────────────────────────
# Comptabilité — Catégories de revenus (bien plus simple que les dépenses)
#   • Ventes/honoraires  → T2125 8000 · TP-80 110 (défaut, factures)
#   • Autres revenus      → T2125 8230 · TP-80 128 (subventions, intérêts, comptant…)
# Le grand livre des revenus vit dans `transactions` (type='revenu'), alimenté par
# plusieurs sources (source = 'facture' | 'manuel' | 'square' | 'shopify').
# ───────────────────────────────────────────────────────────
LIGNES_FISCALES_REVENUS = {
    'Ventes et honoraires professionnels': {'t2125': '8000', 'tp80': '110'},
    'Autres revenus':                      {'t2125': '8230', 'tp80': '128'},
}
CATEGORIES_REVENUS = list(LIGNES_FISCALES_REVENUS.keys())

def materialiser_revenu_facture(conn, facture_id, date_paiement=None):
    """Crée ou met à jour la ligne de revenu du grand livre liée à une facture payée.
    Idempotent : une facture = au plus une ligne revenu (source='facture')."""
    f = conn.execute("SELECT * FROM factures WHERE id = ?", (facture_id,)).fetchone()
    if not f:
        return
    cli = conn.execute(
        "SELECT nom_entreprise, nom_complet, is_test_client FROM clients WHERE id = ?", (f['id_client'],)
    ).fetchone()
    if cli and int(cli['is_test_client'] or 0):
        # Client de test : jamais comptabilisé dans le grand livre (revenus/bilan/taxes).
        conn.execute(
            "DELETE FROM transactions WHERE type='revenu' AND source='facture' AND id_facture = ?",
            (facture_id,)
        )
        return
    nom_client = ((cli['nom_entreprise'] or cli['nom_complet'] or '').strip()) if cli else ''
    date_rev = (date_paiement or f['date_paiement'] or f['date_emission']
                or datetime.now().strftime('%Y-%m-%d'))
    description = f"Facture {f['numero']}" + (f" — {nom_client}" if nom_client else '')
    cat = 'Ventes et honoraires professionnels'
    lg = LIGNES_FISCALES_REVENUS[cat]
    existing = conn.execute(
        "SELECT id FROM transactions WHERE type='revenu' AND source='facture' AND id_facture = ?",
        (facture_id,)
    ).fetchone()
    if existing:
        conn.execute("""
            UPDATE transactions SET date_transaction=?, description=?, categorie=?,
                montant_avant_taxes=?, montant_tps=?, montant_tvq=?, montant_total=?,
                ligne_t2125=?, ligne_tp80=?
            WHERE id=?
        """, (date_rev, description, cat, f['sous_total'] or 0, f['tps'] or 0,
              f['tvq'] or 0, f['total'] or 0, lg['t2125'], lg['tp80'], existing['id']))
    else:
        conn.execute("""
            INSERT INTO transactions
                (type, date_transaction, description, categorie,
                 montant_avant_taxes, montant_tps, montant_tvq, montant_total,
                 source, id_facture, ligne_t2125, ligne_tp80)
            VALUES ('revenu', ?, ?, ?, ?, ?, ?, ?, 'facture', ?, ?, ?)
        """, (date_rev, description, cat, f['sous_total'] or 0, f['tps'] or 0,
              f['tvq'] or 0, f['total'] or 0, facture_id, lg['t2125'], lg['tp80']))

def supprimer_revenu_facture(conn, facture_id):
    """Retire la ligne revenu liée à une facture (décochée payée, annulée ou supprimée)."""
    conn.execute(
        "DELETE FROM transactions WHERE type='revenu' AND source='facture' AND id_facture = ?",
        (facture_id,)
    )

# ───────────────────────────────────────────────────────────
# Intégrations externes (Phase 2) : Square, puis Shopify.
# UNE app CocktailOS (créée dans le Square Developer Dashboard) ; en mode SaaS
# multi-abonnés plus tard, chaque abonné autorisera cette app (OAuth par marchand).
# Les identifiants viennent de l'environnement — inertes tant qu'ils sont vides.
# ───────────────────────────────────────────────────────────
SQUARE_APP_ID     = os.getenv('SQUARE_APP_ID', '')
SQUARE_APP_SECRET = os.getenv('SQUARE_APP_SECRET', '')
SQUARE_ENV        = os.getenv('SQUARE_ENV', 'sandbox')  # 'sandbox' ou 'production'
SQUARE_API_BASE   = 'https://connect.squareup.com' if SQUARE_ENV == 'production' else 'https://connect.squareupsandbox.com'
SQUARE_SCOPES     = 'MERCHANT_PROFILE_READ PAYMENTS_READ ORDERS_READ'

def enregistrer_revenu_square(conn, payment, order=None, organisation_id=None):
    """Mappe un paiement Square encaissé vers une ligne du grand livre (source='square').
    Idempotent via source_ref = id du paiement Square (rejouer ne crée pas de doublon).
    `order` (facultatif) fournit la ventilation des taxes TPS/TVQ."""
    pid = payment.get('id')
    if not pid:
        return None
    if (payment.get('status') or '').upper() != 'COMPLETED':
        return None  # seulement les paiements réellement encaissés
    total = round(((payment.get('total_money') or {}).get('amount') or 0) / 100.0, 2)
    tps = tvq = 0.0
    total_tax = 0.0
    if order:
        total_tax = round(((order.get('total_tax_money') or {}).get('amount') or 0) / 100.0, 2)
        for t in (order.get('taxes') or []):
            nom = (t.get('name') or '').upper()
            montant = round(((t.get('applied_money') or {}).get('amount') or 0) / 100.0, 2)
            if 'TPS' in nom or 'GST' in nom:
                tps += montant
            elif 'TVQ' in nom or 'QST' in nom:
                tvq += montant
        # Si les 2 taxes ne sont pas nommées séparément dans Square, on ne devine pas
        if round(tps + tvq, 2) != total_tax:
            tps = tvq = 0.0  # ventilation à affiner selon la config Square réelle
    avant = round(total - total_tax, 2)
    date_rev = (payment.get('created_at') or '')[:10] or datetime.now().strftime('%Y-%m-%d')
    description = f"Vente Square {pid[-8:]}"
    lg = LIGNES_FISCALES_REVENUS['Ventes et honoraires professionnels']
    existing = conn.execute(
        "SELECT id FROM transactions WHERE type='revenu' AND source='square' AND source_ref = ?",
        (pid,)
    ).fetchone()
    if existing:
        conn.execute("""
            UPDATE transactions SET date_transaction=?, description=?,
                montant_avant_taxes=?, montant_tps=?, montant_tvq=?, montant_total=?,
                organisation_id=?, ligne_t2125=?, ligne_tp80=?
            WHERE id=?
        """, (date_rev, description, avant, tps, tvq, total, organisation_id,
              lg['t2125'], lg['tp80'], existing['id']))
        return existing['id']
    cur = conn.execute("""
        INSERT INTO transactions
            (organisation_id, type, date_transaction, description, categorie,
             montant_avant_taxes, montant_tps, montant_tvq, montant_total,
             source, source_ref, ligne_t2125, ligne_tp80)
        VALUES (?, 'revenu', ?, ?, 'Ventes et honoraires professionnels',
                ?, ?, ?, ?, 'square', ?, ?, ?)
    """, (organisation_id, date_rev, description, avant, tps, tvq, total,
          pid, lg['t2125'], lg['tp80']))
    return cur.lastrowid

# ───────────────────────────────────────────────────────────
# Analyse de reçus par photo (LLM vision) — extraction + classement.
# Provider-agnostique via env : SCAN_PROVIDER (défaut 'gemini').
# Gemini accepte JPG/PNG/WebP/HEIC/PDF en base64. Inerte si la clé est absente.
# ───────────────────────────────────────────────────────────
SCAN_PROVIDER  = os.getenv('SCAN_PROVIDER', 'gemini')
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY', '')
GEMINI_MODEL   = os.getenv('GEMINI_MODEL', 'gemini-2.5-flash')

def analyser_recu(image_b64, mime_type, categories, type_libelle, sens_contexte=None):
    """Envoie une image/PDF de reçu au LLM vision et extrait des champs structurés.
    `categories` = liste des catégories permises (dépenses ou revenus).
    `sens_contexte` (optionnel) = nom/courriel du propriétaire du compte ; si fourni,
    l'IA détermine si la facture est REÇUE (facturée à nous) ou ÉMISE (par nous) —
    garde-fou anti-erreur pour l'ingestion Gmail.
    Retourne un dict de champs pré-remplis, ou {'error': ...}."""
    import json as _json, requests
    if SCAN_PROVIDER != 'gemini' or not GEMINI_API_KEY:
        return {'error': "Analyse par photo non configurée (clé API manquante)."}
    prompt = (
        f"Tu analyses un reçu ou une facture pour une {type_libelle} d'un travailleur "
        f"autonome au Québec. Extrais les informations et réponds uniquement en JSON. "
        f"La catégorie DOIT être exactement l'une de ces valeurs : {', '.join(categories)}. "
        f"Choisis la plus pertinente selon la nature de l'achat. Les montants sont en "
        f"dollars canadiens. Si la TPS/GST (5 %) ou la TVQ/QST (9,975 %) est visible, "
        f"indique chacune; sinon 0. montant_total = total payé taxes incluses. "
        f"date_transaction au format AAAA-MM-JJ. description = un libellé court "
        f"(marchand + nature). Si une info est absente, mets une chaîne vide ou 0."
    )
    props = {
        "date_transaction": {"type": "string"},
        "fournisseur":      {"type": "string"},
        "description":      {"type": "string"},
        "categorie":        {"type": "string", "enum": categories},
        "montant_avant_taxes": {"type": "number"},
        "montant_tps":      {"type": "number"},
        "montant_tvq":      {"type": "number"},
        "montant_total":    {"type": "number"},
        "confiance":        {"type": "string", "enum": ["haute", "moyenne", "basse"]},
    }
    if sens_contexte:
        prompt += (
            f" IMPORTANT — analyse QUI émet la facture. Sur une facture il y a un ÉMETTEUR "
            f"(le vendeur/fournisseur, souvent en haut avec son logo, ses coordonnées et son "
            f"numéro de TPS/TVQ — c'est lui qui réclame le paiement) et un DESTINATAIRE "
            f"(le client, souvent après « Facturé à / Facturer à / Bill to »). "
            f"Le titulaire du compte est : « {sens_contexte} ». "
            f"Mets 'emetteur' = le nom de la société qui ÉMET la facture. "
            f"Puis mets « sens » = 'emis' si l'ÉMETTEUR est le titulaire du compte "
            f"(c.-à-d. c'est SA facture à un client — À ÉCARTER), ou 'recu' si l'émetteur "
            f"est quelqu'un d'AUTRE qui facture le titulaire (= une dépense). "
            f"En cas de doute, mets 'emis'."
        )
        props["emetteur"] = {"type": "string"}
        props["sens"] = {"type": "string", "enum": ["recu", "emis"]}
    schema = {
        "type": "object",
        "properties": props,
        "required": ["date_transaction", "description", "categorie", "montant_total"],
    }
    body = {
        "contents": [{"parts": [
            {"inline_data": {"mime_type": mime_type, "data": image_b64}},
            {"text": prompt},
        ]}],
        "generationConfig": {"responseMimeType": "application/json", "responseSchema": schema},
    }
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"
    try:
        r = requests.post(url, headers={"x-goog-api-key": GEMINI_API_KEY,
                                        "Content-Type": "application/json"},
                          json=body, timeout=45)
        if r.status_code != 200:
            return {'error': f"Erreur d'analyse ({r.status_code})."}
        text = r.json()["candidates"][0]["content"]["parts"][0]["text"]
        fields = _json.loads(text)
    except Exception as e:
        return {'error': f"Analyse impossible : {e}"}
    cat = (fields.get('categorie') or '').strip()
    if cat not in categories:
        cat = categories[-1]  # catch-all (« Autre » / « Autres revenus »)
    fields['categorie'] = cat
    return fields

_DRIVE_RECUS_FOLDER = {'id': os.getenv('DRIVE_RECUS_FOLDER_ID', '') or None}
DRIVE_RECUS_NOM = os.getenv('DRIVE_RECUS_NOM', 'Reçus (comptabilité)')

def _dossier_recus_drive():
    """Trouve (ou crée une seule fois) le dossier Drive des pièces justificatives,
    sous GOOGLE_DRIVE_ROOT_FOLDER_ID. Retourne son ID, ou None si Drive indispo."""
    if _DRIVE_RECUS_FOLDER['id']:
        return _DRIVE_RECUS_FOLDER['id']
    racine = os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
    if not racine:
        return None
    try:
        from drive_service import list_subfolders, create_folder
        for f in list_subfolders(racine):
            if f['name'] == DRIVE_RECUS_NOM:
                _DRIVE_RECUS_FOLDER['id'] = f['id']
                return f['id']
        _DRIVE_RECUS_FOLDER['id'] = create_folder(DRIVE_RECUS_NOM, parent_id=racine)
        return _DRIVE_RECUS_FOLDER['id']
    except Exception as e:
        print(f"[RECU] dossier Drive indisponible: {e}")
        return None

def sauver_piece_jointe(raw, filename, mime):
    """Sauvegarde le reçu comme pièce justificative — sur Google Drive de préférence
    (retourne le lien Drive), avec repli local si Drive est indisponible."""
    import uuid, tempfile
    ext = os.path.splitext(filename or '')[1].lower() or {
        'image/jpeg': '.jpg', 'image/png': '.png', 'image/heic': '.heic',
        'image/heif': '.heic', 'image/webp': '.webp', 'application/pdf': '.pdf',
    }.get(mime, '.bin')
    nom = f"recu_{datetime.now().strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}{ext}"

    # 1) Tentative Drive
    folder = _dossier_recus_drive()
    if folder:
        tmp = os.path.join(tempfile.gettempdir(), nom)
        try:
            with open(tmp, 'wb') as fh:
                fh.write(raw)
            from drive_service import upload_file
            _fid, lien = upload_file(tmp, nom, folder)
            return lien  # lien Drive cliquable, stocké dans piece_jointe
        except Exception as e:
            print(f"[RECU] upload Drive échoué, repli local: {e}")
        finally:
            try: os.remove(tmp)
            except Exception: pass

    # 2) Repli local (si Drive non configuré ou en échec)
    dossier = os.path.join(app.config['UPLOAD_ROOT'], 'pieces_justificatives')
    os.makedirs(dossier, exist_ok=True)
    with open(os.path.join(dossier, nom), 'wb') as fh:
        fh.write(raw)
    return f"pieces_justificatives/{nom}"

def init_db():
    """Crée les tables manquantes et sème les valeurs par défaut essentielles."""
    # WAL mode est persistant — à activer une seule fois au démarrage, pas à chaque connexion
    _wal_conn = sqlite3.connect(app.config["DB_PATH"], timeout=10, isolation_level=None)
    _wal_conn.execute("PRAGMA journal_mode = WAL;")
    _wal_conn.close()

    schema_sql = """
    CREATE TABLE IF NOT EXISTS clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom_complet TEXT NOT NULL,
        email TEXT NOT NULL UNIQUE,
        nom_entreprise TEXT,
        telephone TEXT,
        mot_de_passe_hash TEXT,
        auth_provider TEXT NOT NULL DEFAULT 'password',
        is_email_confirmed INTEGER NOT NULL DEFAULT 0,
        is_admin INTEGER NOT NULL DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS projets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom_projet TEXT NOT NULL,
        statut TEXT DEFAULT 'Nouveau',
        lien_gdrive TEXT,
        id_client INTEGER NOT NULL,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_client) REFERENCES clients(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS services (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom_service TEXT NOT NULL UNIQUE,
        description TEXT
    );

    CREATE TABLE IF NOT EXISTS checklistes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_projet INTEGER NOT NULL,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_projet) REFERENCES projets(id) ON DELETE CASCADE
    );

    /* Modèle (gabarit) d'items par service */
    CREATE TABLE IF NOT EXISTS checklist_model_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_service INTEGER NOT NULL,
        nom_item TEXT NOT NULL,
        requires_file INTEGER NOT NULL DEFAULT 0,
        is_required INTEGER NOT NULL DEFAULT 1,
        position INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_service) REFERENCES services(id) ON DELETE CASCADE
    );

    /* Items réels d'une checklist de projet */
    CREATE TABLE IF NOT EXISTS checklist_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_checklist INTEGER NOT NULL,
        nom_item TEXT NOT NULL,
        requires_file INTEGER NOT NULL DEFAULT 0,
        is_required INTEGER NOT NULL DEFAULT 1,
        est_coche INTEGER NOT NULL DEFAULT 0,
        file_path TEXT,
        important INTEGER NOT NULL DEFAULT 0,
        position INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_checklist) REFERENCES checklistes(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS uploads (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_item INTEGER NOT NULL,
        filename TEXT NOT NULL,
        filepath TEXT NOT NULL,
        uploaded_by TEXT NOT NULL CHECK (uploaded_by IN ('client','admin')),
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_item) REFERENCES checklist_items(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS notification_settings (
        id INTEGER PRIMARY KEY CHECK (id = 1),
        admin_emails TEXT DEFAULT '',
        client_updates INTEGER NOT NULL DEFAULT 1,
        admin_updates INTEGER NOT NULL DEFAULT 1
    );

    CREATE TABLE IF NOT EXISTS identite_visuelle (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_projet INTEGER NOT NULL UNIQUE,
        is_complete INTEGER NOT NULL DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_projet) REFERENCES projets(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS iv_logos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_iv INTEGER NOT NULL,
        variante TEXT NOT NULL CHECK (variante IN ('principal','icone','variante')),
        drive_file_id TEXT,
        filename TEXT,
        FOREIGN KEY (id_iv) REFERENCES identite_visuelle(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS iv_fonts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_iv INTEGER NOT NULL,
        nom_font TEXT NOT NULL,
        google_font_url TEXT,
        usage TEXT,
        FOREIGN KEY (id_iv) REFERENCES identite_visuelle(id) ON DELETE CASCADE
    );

    /* Index utiles */
    CREATE INDEX IF NOT EXISTS idx_projets_client ON projets(id_client);
    CREATE INDEX IF NOT EXISTS idx_checklist_projet ON checklistes(id_projet);
    CREATE INDEX IF NOT EXISTS idx_items_checklist ON checklist_items(id_checklist);
    CREATE INDEX IF NOT EXISTS idx_model_items_service ON checklist_model_items(id_service);

    /* ── Module comptabilité ── */
    CREATE TABLE IF NOT EXISTS parametres_facturation (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        organisation_id INTEGER,
        charge_taxes INTEGER NOT NULL DEFAULT 0,
        neq TEXT,
        numero_tps TEXT,
        numero_tvq TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        organisation_id INTEGER,
        type TEXT NOT NULL CHECK (type IN ('revenu','depense')),
        date_transaction TEXT NOT NULL,
        description TEXT NOT NULL,
        categorie TEXT,
        montant_avant_taxes REAL NOT NULL DEFAULT 0,
        montant_tps REAL NOT NULL DEFAULT 0,
        montant_tvq REAL NOT NULL DEFAULT 0,
        montant_total REAL NOT NULL DEFAULT 0,
        source TEXT NOT NULL DEFAULT 'manuel',
        id_facture INTEGER,
        piece_jointe TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    CREATE INDEX IF NOT EXISTS idx_transactions_type ON transactions(type);
    CREATE INDEX IF NOT EXISTS idx_transactions_date ON transactions(date_transaction);
    CREATE TABLE IF NOT EXISTS decision_boards (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_projet INTEGER NOT NULL UNIQUE,
        accent_color TEXT DEFAULT '#E83B14',
        accent_color_rgb TEXT DEFAULT '232,59,20',
        show_directions INTEGER DEFAULT 1,
        directions_json TEXT,
        show_names INTEGER DEFAULT 1,
        names_json TEXT,
        show_icons INTEGER DEFAULT 1,
        icons_json TEXT,
        show_typos INTEGER DEFAULT 1,
        typos_json TEXT,
        show_palettes INTEGER DEFAULT 1,
        palettes_json TEXT,
        show_logos INTEGER DEFAULT 1,
        logos_json TEXT,
        is_active INTEGER DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_projet) REFERENCES projets(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS decision_board_choices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_projet INTEGER NOT NULL UNIQUE,
        choix_directions TEXT,
        choix_noms TEXT,
        choix_icones TEXT,
        choix_typos TEXT,
        choix_palettes TEXT,
        choix_logos TEXT,
        nom_suggestion TEXT,
        commentaires TEXT,
        submitted_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_projet) REFERENCES projets(id) ON DELETE CASCADE
    );

    CREATE INDEX IF NOT EXISTS idx_decision_board_projet ON decision_boards(id_projet);
    CREATE INDEX IF NOT EXISTS idx_decision_choices_projet ON decision_board_choices(id_projet);

    CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_client INTEGER NOT NULL,
        id_projet INTEGER NOT NULL,
        message TEXT NOT NULL,
        is_read INTEGER NOT NULL DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_client) REFERENCES clients(id) ON DELETE CASCADE,
        FOREIGN KEY (id_projet) REFERENCES projets(id) ON DELETE CASCADE
    );
    CREATE INDEX IF NOT EXISTS idx_notifications_client ON notifications(id_client, is_read);

    /* Notifications internes (admin) — to-do, Todoist, assignations, etc.
       destinataire = email de l'admin ciblé, ou NULL/'' = visible par tous les admins */
    CREATE TABLE IF NOT EXISTS admin_notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        destinataire TEXT,
        type TEXT NOT NULL DEFAULT 'info',
        titre TEXT NOT NULL,
        message TEXT,
        lien TEXT,
        is_read INTEGER NOT NULL DEFAULT 0,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE INDEX IF NOT EXISTS idx_admin_notifs ON admin_notifications(destinataire, is_read, created_at);

    /* Abonnements Web Push (notifs téléphone/desktop portail fermé) */
    CREATE TABLE IF NOT EXISTS push_subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT,
        endpoint TEXT NOT NULL UNIQUE,
        p256dh TEXT NOT NULL,
        auth TEXT NOT NULL,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    /* ── Module Pigistes ── */
    CREATE TABLE IF NOT EXISTS pigistes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom_complet TEXT NOT NULL,
        email TEXT NOT NULL UNIQUE,
        mot_de_passe_hash TEXT,
        telephone TEXT,
        adresse TEXT,
        ville TEXT,
        province TEXT DEFAULT 'Québec',
        code_postal TEXT,
        numero_tps TEXT,
        numero_tvq TEXT,
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS mandats (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_pigiste INTEGER NOT NULL REFERENCES pigistes(id) ON DELETE CASCADE,
        id_projet INTEGER REFERENCES projets(id) ON DELETE SET NULL,
        titre TEXT NOT NULL,
        description TEXT,
        date_debut TEXT,
        date_echeance TEXT,
        montant_convenu REAL DEFAULT 0,
        statut TEXT NOT NULL DEFAULT 'en_attente',
        notes_admin TEXT,
        drive_folder_id TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS mandats_livrables (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_mandat INTEGER NOT NULL REFERENCES mandats(id) ON DELETE CASCADE,
        filename TEXT NOT NULL,
        drive_file_id TEXT,
        public_url TEXT,
        uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS factures_pigiste (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_pigiste INTEGER NOT NULL REFERENCES pigistes(id) ON DELETE CASCADE,
        numero TEXT NOT NULL UNIQUE,
        date_emission TEXT NOT NULL,
        date_echeance TEXT,
        statut TEXT NOT NULL DEFAULT 'brouillon',
        montant_ht REAL DEFAULT 0,
        tps REAL DEFAULT 0,
        tvq REAL DEFAULT 0,
        montant_total REAL DEFAULT 0,
        notes TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS factures_pigiste_lignes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        id_facture INTEGER NOT NULL REFERENCES factures_pigiste(id) ON DELETE CASCADE,
        id_mandat INTEGER REFERENCES mandats(id) ON DELETE SET NULL,
        description TEXT NOT NULL,
        quantite REAL DEFAULT 1,
        taux REAL DEFAULT 0,
        montant REAL DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS mediatech_gabarits (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL,
        description TEXT,
        drive_folder_id TEXT,
        preview_drive_id TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS short_links (
        code TEXT PRIMARY KEY,
        url TEXT NOT NULL,
        expires_at TEXT NOT NULL,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS rendez_vous (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        id_client      INTEGER NOT NULL,
        calendar_event_id TEXT,
        start_utc      TEXT NOT NULL,
        end_utc        TEXT NOT NULL,
        meet_link      TEXT,
        label_fr       TEXT,
        created_at     TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (id_client) REFERENCES clients(id)
    );

    CREATE TABLE IF NOT EXISTS client_ressources (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        id_client      INTEGER,
        titre          TEXT NOT NULL,
        description    TEXT,
        categorie      TEXT DEFAULT 'guide',
        type_source    TEXT NOT NULL DEFAULT 'lien',
        drive_file_id  TEXT,
        url            TEXT,
        created_at     TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (id_client) REFERENCES clients(id)
    );

    CREATE TABLE IF NOT EXISTS ressource_images (
        id             INTEGER PRIMARY KEY AUTOINCREMENT,
        id_ressource   INTEGER NOT NULL,
        drive_file_id  TEXT NOT NULL,
        nom_fichier    TEXT,
        legende        TEXT,
        ordre          INTEGER DEFAULT 0,
        created_at     TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (id_ressource) REFERENCES client_ressources(id)
    );

    CREATE TABLE IF NOT EXISTS guide_sections (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        id_ressource  INTEGER NOT NULL,
        ordre         INTEGER DEFAULT 0,
        titre         TEXT NOT NULL,
        intro         TEXT,
        astuce        TEXT,
        etapes_json   TEXT DEFAULT '[]',
        created_at    TEXT DEFAULT (datetime('now')),
        FOREIGN KEY (id_ressource) REFERENCES client_ressources(id)
    );

    CREATE TABLE IF NOT EXISTS todos_perso (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        texte TEXT NOT NULL,
        est_coche INTEGER DEFAULT 0,
        priorite TEXT DEFAULT 'normale',
        date_echeance TEXT,
        calendar_event_id TEXT,
        projet_id INTEGER,
        projet_nom TEXT,
        created_at TEXT DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS sites (
        id                   INTEGER PRIMARY KEY AUTOINCREMENT,
        template             TEXT NOT NULL,
        slug                 TEXT NOT NULL,
        business_name        TEXT NOT NULL,
        owner_name           TEXT,
        owner_title          TEXT,
        tagline              TEXT,
        description          TEXT,
        address              TEXT,
        city                 TEXT,
        province             TEXT DEFAULT 'QC',
        postal_code          TEXT,
        phone                TEXT,
        email                TEXT,
        acuity_url           TEXT,
        instagram            TEXT,
        facebook             TEXT,
        linkedin             TEXT,
        hero_style           TEXT,
        style_variant        TEXT,
        direction            TEXT,
        seo_meta_title       TEXT,
        seo_meta_description TEXT,
        seo_keywords         TEXT,
        seo_og_image         TEXT,
        seo_twitter_handle   TEXT,
        seo_logo_url         TEXT,
        seo_business_type    TEXT,
        seo_price_range      TEXT DEFAULT '$$',
        client_email         TEXT,
        resend_api_key       TEXT,
        site_url             TEXT,
        github_repo          TEXT,
        sanity_project_id    TEXT,
        vercel_project_id    TEXT,
        vercel_url           TEXT,
        status               TEXT DEFAULT 'draft',
        error_message        TEXT,
        created_at           TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at           TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS soumissions (
        id                 INTEGER PRIMARY KEY AUTOINCREMENT,
        id_client          INTEGER NOT NULL,
        titre              TEXT NOT NULL,
        message_intro      TEXT,
        statut             TEXT DEFAULT 'envoyee',
        option_acceptee_id INTEGER,
        date_expiration    TEXT,
        created_at         TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at         TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_client) REFERENCES clients(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS soumission_options (
        id                          INTEGER PRIMARY KEY AUTOINCREMENT,
        id_soumission               INTEGER NOT NULL,
        nom                         TEXT NOT NULL,
        description                 TEXT,
        prix_setup                  REAL DEFAULT 0,
        prix_mensuel                REAL DEFAULT 0,
        prix_horaire                REAL DEFAULT 0,
        delai_livraison             TEXT,
        conditions_paiement         TEXT,
        inclus_json                 TEXT,
        couts_tiers_json            TEXT,
        couts_supplementaires_json  TEXT,
        scenarios_json              TEXT,
        est_recommande              INTEGER DEFAULT 0,
        badge_texte                 TEXT,
        ordre                       INTEGER DEFAULT 0,
        FOREIGN KEY (id_soumission) REFERENCES soumissions(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS soumission_templates (
        id                    INTEGER PRIMARY KEY AUTOINCREMENT,
        nom                   TEXT NOT NULL,
        description           TEXT,
        message_intro_template TEXT,
        titre_template        TEXT NOT NULL DEFAULT 'Soumission - {nom_entreprise}',
        est_actif             INTEGER NOT NULL DEFAULT 1,
        created_at            TEXT DEFAULT CURRENT_TIMESTAMP,
        updated_at            TEXT DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS soumission_template_options (
        id                         INTEGER PRIMARY KEY AUTOINCREMENT,
        id_template                INTEGER NOT NULL,
        nom                        TEXT NOT NULL,
        description                TEXT,
        prix_setup                 REAL DEFAULT 0,
        prix_mensuel               REAL DEFAULT 0,
        prix_horaire               REAL DEFAULT 0,
        delai_livraison            TEXT,
        conditions_paiement        TEXT,
        badge_texte                TEXT,
        est_recommande             INTEGER DEFAULT 0,
        ordre                      INTEGER DEFAULT 0,
        features_json              TEXT DEFAULT '{}',
        inclus_json                TEXT DEFAULT '[]',
        couts_tiers_json           TEXT DEFAULT '[]',
        couts_supplementaires_json TEXT DEFAULT '[]',
        scenarios_json             TEXT DEFAULT '[]',
        rachat_disponible          INTEGER DEFAULT 0,
        prix_rachat                REAL DEFAULT 0,
        inclus_rachat_json         TEXT DEFAULT '[]',
        created_at                 TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (id_template) REFERENCES soumission_templates(id) ON DELETE CASCADE
    );
    """
    conn = get_db_connection()
    try:
        conn.executescript(schema_sql)
        # Seed défaut notification_settings (id=1) si absent
        row = conn.execute("SELECT 1 FROM notification_settings WHERE id = 1").fetchone()
        if not row:
            conn.execute("""
                INSERT INTO notification_settings (id, admin_emails, client_updates, admin_updates)
                VALUES (1, '', 1, 1)
            """)
        # Seed défaut parametres_facturation (id=1) si absent
        row_param = conn.execute("SELECT 1 FROM parametres_facturation WHERE id = 1").fetchone()
        if not row_param:
            conn.execute("""
                INSERT INTO parametres_facturation (id, charge_taxes, neq, numero_tps, numero_tvq)
                VALUES (1, 0, '', '', '')
            """)
        # Migration additive : branding (nom entreprise + couleur) sur parametres_facturation
        for col, ddl in [
            ('nom_entreprise', "ALTER TABLE parametres_facturation ADD COLUMN nom_entreprise TEXT DEFAULT ''"),
            ('couleur_marque', "ALTER TABLE parametres_facturation ADD COLUMN couleur_marque TEXT DEFAULT '#c0321a'"),
        ]:
            try:
                conn.execute(ddl)
            except sqlite3.OperationalError:
                pass  # colonne déjà présente

        # Migration additive : type de notification client (pour icône Lucide côté UI)
        try:
            conn.execute("ALTER TABLE notifications ADD COLUMN type TEXT NOT NULL DEFAULT 'info'")
        except sqlite3.OperationalError:
            pass  # colonne déjà présente

        # Migration additive : marquer un client comme compte de test — exclu de tous
        # les agrégats/statistiques admin (revenus, compteurs) mais visible normalement
        # dans sa propre fiche et son propre portail.
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN is_test_client INTEGER NOT NULL DEFAULT 0")
        except sqlite3.OperationalError:
            pass  # colonne déjà présente

        # Migration additive : documents de marque du client (logo, favicon, couleurs)
        for col, ddl in [
            ('logo_url', "ALTER TABLE clients ADD COLUMN logo_url TEXT"),
            ('favicon_url', "ALTER TABLE clients ADD COLUMN favicon_url TEXT"),
            ('couleur_primaire', "ALTER TABLE clients ADD COLUMN couleur_primaire TEXT"),
            ('couleur_secondaire', "ALTER TABLE clients ADD COLUMN couleur_secondaire TEXT"),
        ]:
            try:
                conn.execute(ddl)
            except sqlite3.OperationalError:
                pass  # colonne déjà présente

        # Migration additive : catalogue de services — actif/inactif + durée affichée (texte libre)
        try:
            conn.execute("ALTER TABLE services ADD COLUMN actif INTEGER NOT NULL DEFAULT 1")
        except sqlite3.OperationalError:
            pass  # colonne déjà présente
        try:
            conn.execute("ALTER TABLE services ADD COLUMN duree_affichee TEXT")
        except sqlite3.OperationalError:
            pass  # colonne déjà présente

        # Migration additive : extras par service (catalogue consultable par la création de projet)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS service_extras (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                id_service INTEGER NOT NULL REFERENCES services(id) ON DELETE CASCADE,
                nom TEXT NOT NULL,
                prix REAL NOT NULL DEFAULT 0,
                position INTEGER DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Seed one-shot : migre les services de l'ancienne "bibliothèque" (DEFAULT_SERVICES du
        # frontend) qui ne sont pas encore de vraies lignes en base. Idempotent (vérifie le nom,
        # insensible à la casse, avant chaque insertion) — sans effet sur les runs suivants.
        _SERVICES_SEED = [
            {
                'slug': 'site-web-vitrine', 'nom_service': 'Site Web Vitrine',
                'description': "Création de site vitrine client (Next.js + Sanity + Vercel)",
                'icon': 'web', 'categorie': 'Sites Web', 'prix': 500,
                'localisation_requise': 0, 'documents_requis': 1, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 0, 'duree_production_minutes': 180, 'duree_finalisation_minutes': 120,
                'duree_affichee': '3h production + 2h finalisation',
                'items': [
                    {'nom_item': 'Nom du business', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Logo (SVG)', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'vecteur', 'field_type': 'file'},
                    {'nom_item': 'Logo 2', 'requires_file': 1, 'is_required': 0, 'item_type': 'document', 'file_category': 'vecteur', 'field_type': 'file'},
                    {'nom_item': 'Logo 2-2', 'requires_file': 1, 'is_required': 0, 'item_type': 'document', 'file_category': 'vecteur', 'field_type': 'file'},
                    {'nom_item': "Description de l'entreprise", 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Adresse', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Téléphone', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Courriel professionnel', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Équipe', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'members'},
                    {'nom_item': "Mission de l'entreprise", 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': "Vision de l'entreprise", 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': "Valeurs de l'entreprise", 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Texte section À propos', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Photos du salon / environnement', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'photo', 'field_type': 'file'},
                ],
            },
            {
                'slug': 'site-web-shopify', 'nom_service': 'Design Shopify',
                'description': "Personnalisation de thème Shopify existant — client a déjà ses produits, on s'occupe du visuel",
                'icon': 'web', 'categorie': 'Sites Web', 'prix': 500,
                'localisation_requise': 0, 'documents_requis': 1, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 0, 'duree_production_minutes': 240, 'duree_finalisation_minutes': 120,
                'duree_affichee': '4h production + 2h finalisation',
                'items': [
                    {'nom_item': 'Accès collaborateur Shopify (URL boutique)', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Domaine personnalisé (ex: monsite.com)', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Nom du business', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Logo (SVG)', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'vecteur', 'field_type': 'file'},
                    {'nom_item': 'Logo secondaire', 'requires_file': 1, 'is_required': 0, 'item_type': 'document', 'file_category': 'vecteur', 'field_type': 'file'},
                    {'nom_item': 'Palette de couleurs', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'photo', 'field_type': 'color-palette'},
                    {'nom_item': 'Typographie principale', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Typographie secondaire', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': "Exemples de design / sites d'inspiration", 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Pages à personnaliser (ex: Accueil, À propos, Contact…)', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': "Texte hero / page d'accueil", 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Texte section À propos', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Photos hero / bannière principale', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'photo', 'field_type': 'file'},
                    {'nom_item': 'Photos ambiance / lifestyle', 'requires_file': 1, 'is_required': 0, 'item_type': 'document', 'file_category': 'photo', 'field_type': 'file'},
                    {'nom_item': 'Liens réseaux sociaux', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Courriel de contact affiché sur le site', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Politique de livraison', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'document', 'field_type': 'file-or-textarea'},
                    {'nom_item': 'Politique de retour / remboursement', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'document', 'field_type': 'file-or-textarea'},
                ],
            },
            {
                'slug': 'site-web-transactionnel', 'nom_service': 'Site Web Transactionnel',
                'description': 'Boutique en ligne — vente de produits ou services (Next.js + Stripe + Sanity)',
                'icon': 'web', 'categorie': 'Sites Web', 'prix': 1000,
                'localisation_requise': 0, 'documents_requis': 1, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 0, 'duree_production_minutes': 180, 'duree_finalisation_minutes': 120,
                'duree_affichee': '3h production + 2h finalisation',
                'items': [
                    {'nom_item': 'Nom du business', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Logo (SVG)', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'vecteur', 'field_type': 'file'},
                    {'nom_item': "Description de l'entreprise", 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Adresse', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Téléphone', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Courriel professionnel', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Équipe', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'members'},
                    {'nom_item': 'Liste des produits / services', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Descriptions détaillées des produits / services', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Prix des produits / services', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Photos des produits / services', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'photo', 'field_type': 'file'},
                    {'nom_item': 'Plateforme de paiement (ex: Stripe)', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Méthodes de paiement acceptées', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Zones de livraison', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Tarifs de livraison', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Politique de confidentialité', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'document', 'field_type': 'file-or-textarea'},
                    {'nom_item': 'Conditions générales de vente', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'document', 'field_type': 'file-or-textarea'},
                    {'nom_item': 'Politique de retour / remboursement', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'document', 'field_type': 'file-or-textarea'},
                    {'nom_item': 'Photos ambiance / boutique', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'photo', 'field_type': 'file'},
                ],
            },
            {
                'slug': 'video-corporatif', 'nom_service': 'Vidéo corporatif',
                'description': "Vidéo de présentation d'entreprise — tournage sur place",
                'icon': 'video', 'categorie': "Vidéo d'entreprise", 'prix': 300,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 180, 'duree_production_minutes': 180, 'duree_finalisation_minutes': 0,
                'duree_affichee': '3h tournage + 3h post-prod', 'items': [],
            },
            {
                'slug': 'couverture-evenements', 'nom_service': "Couverture d'évènements / 3H",
                'description': "Couverture vidéo d'un événement — 3 heures sur place",
                'icon': 'video', 'categorie': "Vidéo d'entreprise", 'prix': 300,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 240, 'duree_production_minutes': 240, 'duree_finalisation_minutes': 0,
                'duree_affichee': '4h tournage + 4h post-prod', 'items': [],
            },
            {
                'slug': 'video-immobilier', 'nom_service': 'Vidéos immobiliers',
                'description': "Vidéo de présentation d'une propriété — tournage sur place",
                'icon': 'video', 'categorie': "Vidéo d'entreprise", 'prix': 200,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 60, 'duree_production_minutes': 60, 'duree_finalisation_minutes': 0,
                'duree_affichee': '1h tournage + 1h post-prod', 'items': [],
            },
            {
                'slug': 'video-aerien', 'nom_service': 'Vidéos aériens',
                'description': 'Prise de vue aérienne par drone — tournage sur place',
                'icon': 'video', 'categorie': "Vidéo d'entreprise", 'prix': 200,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 30, 'duree_production_minutes': 45, 'duree_finalisation_minutes': 0,
                'duree_affichee': '30 min tournage + 45 min post-prod', 'items': [],
            },
            {
                'slug': 'forfait-short-reel', 'nom_service': 'Forfait Short/Reel',
                'description': 'Création de contenu court format pour réseaux sociaux — tournage sur place',
                'icon': 'video', 'categorie': "Vidéo d'entreprise", 'prix': 330,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 120, 'duree_production_minutes': 120, 'duree_finalisation_minutes': 0,
                'duree_affichee': '2h tournage + 2h post-prod', 'items': [],
            },
            {
                'slug': 'video-unite-short-reel', 'nom_service': 'Vidéo unité Short/Reel',
                'description': 'Courte vidéo unitaire pour réseaux sociaux — tournage sur place',
                'icon': 'video', 'categorie': "Vidéo d'entreprise", 'prix': 0,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 30, 'duree_production_minutes': 20, 'duree_finalisation_minutes': 0,
                'duree_affichee': '30 min tournage + 20 min post-prod', 'items': [],
            },
            {
                'slug': 'photos-produits', 'nom_service': 'Photos de produits',
                'description': 'Photographie de produits au bureau — 1001 rang St-Malo, Trois-Rivières',
                'icon': 'photo', 'categorie': 'Photographie', 'prix': 175,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 60, 'duree_production_minutes': 60, 'duree_finalisation_minutes': 0,
                'duree_affichee': '1h prise de vue + 1h retouche', 'items': [],
            },
            {
                'slug': 'photos-en-action', 'nom_service': 'Photos en actions',
                'description': "Photographie en action — tournage à l'adresse du client",
                'icon': 'photo', 'categorie': 'Photographie', 'prix': 250,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 60, 'duree_production_minutes': 60, 'duree_finalisation_minutes': 0,
                'duree_affichee': '1h prise de vue + 1h retouche', 'items': [],
            },
            {
                'slug': 'couverture-evenement-photo', 'nom_service': "Couverture d'évènement / 3H",
                'description': "Couverture photo d'un événement — adresse du client ou de l'événement",
                'icon': 'photo', 'categorie': 'Photographie', 'prix': 250,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 240, 'duree_production_minutes': 120, 'duree_finalisation_minutes': 0,
                'duree_affichee': '4h prise de vue + 2h retouche', 'items': [],
            },
            {
                'slug': 'portraits-pro', 'nom_service': 'Portraits professionnels / 3 personnes',
                'description': 'Portraits pro au bureau — 1001 rang St-Malo, Trois-Rivières',
                'icon': 'photo', 'categorie': 'Photographie', 'prix': 100,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 15, 'duree_production_minutes': 30, 'duree_finalisation_minutes': 0,
                'duree_affichee': '15 min prise de vue + 30 min retouche', 'items': [],
            },
            {
                'slug': 'retouches-photos', 'nom_service': 'Retouches de photos existantes',
                'description': 'Retouche et post-production de photos fournies par le client',
                'icon': 'photo', 'categorie': 'Photographie', 'prix': 50,
                'localisation_requise': 0, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 0, 'duree_production_minutes': 30, 'duree_finalisation_minutes': 0,
                'duree_affichee': '30 min retouche', 'items': [],
            },
            {
                'slug': 'photo-immobiliere', 'nom_service': 'Photographies immobilières (drone incl.)',
                'description': "Photographie immobilière incluant prises de vue par drone — adresse du client",
                'icon': 'photo', 'categorie': 'Photographie', 'prix': 150,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 45, 'duree_production_minutes': 25, 'duree_finalisation_minutes': 0,
                'duree_affichee': '45 min prise de vue + 25 min retouche', 'items': [],
            },
            {
                'slug': 'photo-drone', 'nom_service': 'Photographies par drone',
                'description': 'Prise de vue aérienne par drone — adresse du client',
                'icon': 'photo', 'categorie': 'Photographie', 'prix': 200,
                'localisation_requise': 1, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 45, 'duree_production_minutes': 25, 'duree_finalisation_minutes': 0,
                'duree_affichee': '45 min prise de vue + 25 min retouche', 'items': [],
            },
            {
                'slug': 'creation-logo', 'nom_service': "Création d'un logo personnalisé",
                'description': 'Création de logo sur mesure — rencontre créative Google Meet incluse',
                'icon': 'graphisme', 'categorie': 'Graphisme', 'prix': 200,
                'localisation_requise': 0, 'documents_requis': 1, 'appel_exploratoire_requis': 1, 'decision_board_requis': 1,
                'duree_seance_minutes': 15, 'duree_tournage_minutes': 0, 'duree_production_minutes': 120, 'duree_finalisation_minutes': 60,
                'duree_affichee': '15 min rencontre + 2h travaux + révision + 1h finalisation',
                'items': [
                    {'nom_item': "Nom de l'entreprise", 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Couleurs souhaitées', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Inspirations / références visuelles', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                ],
            },
            {
                'slug': 'refonte-identite-visuelle', 'nom_service': "Refonte d'identité visuelle",
                'description': "Refonte complète de l'identité visuelle — rencontre créative Google Meet incluse",
                'icon': 'graphisme', 'categorie': 'Graphisme', 'prix': 150,
                'localisation_requise': 0, 'documents_requis': 1, 'appel_exploratoire_requis': 1, 'decision_board_requis': 1,
                'duree_seance_minutes': 15, 'duree_tournage_minutes': 0, 'duree_production_minutes': 120, 'duree_finalisation_minutes': 60,
                'duree_affichee': '15 min rencontre + 2h travaux + révision + 1h finalisation',
                'items': [
                    {'nom_item': "Nom de l'entreprise", 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Logo existant', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'vecteur', 'field_type': 'file'},
                    {'nom_item': 'Couleurs existantes', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Polices utilisées', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Couleurs souhaitées', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Inspirations / références visuelles', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                ],
            },
            {
                'slug': 'support-imprimable-1', 'nom_service': 'Support imprimable — 1 visuel',
                'description': "Création d'un support imprimable (carte, affiche, flyer…)",
                'icon': 'graphisme', 'categorie': 'Graphisme', 'prix': 75,
                'localisation_requise': 0, 'documents_requis': 1, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 0, 'duree_production_minutes': 30, 'duree_finalisation_minutes': 0,
                'duree_affichee': '30 min / visuel',
                'items': [
                    {'nom_item': 'Texte / contenu à intégrer', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Couleurs / charte graphique', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Logo (SVG)', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'vecteur', 'field_type': 'file'},
                    {'nom_item': 'Format ou dimensions souhaités', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Exemples de style / références', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                ],
            },
            {
                'slug': 'support-imprimable-4', 'nom_service': 'Support imprimable — 4 visuels',
                'description': 'Création de 4 supports imprimables (cartes, affiches, flyers…)',
                'icon': 'graphisme', 'categorie': 'Graphisme', 'prix': 200,
                'localisation_requise': 0, 'documents_requis': 1, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 0, 'duree_production_minutes': 120, 'duree_finalisation_minutes': 0,
                'duree_affichee': '30 min × 4 visuels',
                'items': [
                    {'nom_item': 'Texte / contenu à intégrer', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Couleurs / charte graphique', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Logo (SVG)', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'vecteur', 'field_type': 'file'},
                    {'nom_item': 'Format ou dimensions souhaités', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Exemples de style / références', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                ],
            },
            {
                'slug': 'support-numerique-1', 'nom_service': 'Support numérique — 1 visuel',
                'description': "Création d'un support numérique (bannière, publication, story…)",
                'icon': 'graphisme', 'categorie': 'Graphisme', 'prix': 75,
                'localisation_requise': 0, 'documents_requis': 1, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 0, 'duree_production_minutes': 30, 'duree_finalisation_minutes': 0,
                'duree_affichee': '30 min / visuel',
                'items': [
                    {'nom_item': 'Texte / contenu à intégrer', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Couleurs / charte graphique', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Logo (SVG)', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'vecteur', 'field_type': 'file'},
                    {'nom_item': 'Format ou dimensions souhaités', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Exemples de style / références', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                ],
            },
            {
                'slug': 'presentation-powerpoint', 'nom_service': 'Présentation PowerPoint',
                'description': "Création d'une présentation PowerPoint professionnelle",
                'icon': 'graphisme', 'categorie': "Infographie d'entreprise", 'prix': 100,
                'localisation_requise': 0, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 0, 'duree_production_minutes': 60, 'duree_finalisation_minutes': 0,
                'duree_affichee': '1h production', 'items': [],
            },
            {
                'slug': 'plan-affaires', 'nom_service': "Création de plan d'affaires",
                'description': "Rédaction et mise en page d'un plan d'affaires complet",
                'icon': 'graphisme', 'categorie': "Infographie d'entreprise", 'prix': 200,
                'localisation_requise': 0, 'documents_requis': 0, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 0, 'duree_production_minutes': 60, 'duree_finalisation_minutes': 0,
                'duree_affichee': '1h production', 'items': [],
            },
            {
                'slug': 'support-numerique-4', 'nom_service': 'Support numérique — 4 visuels',
                'description': 'Création de 4 supports numériques (bannières, publications, stories…)',
                'icon': 'graphisme', 'categorie': 'Graphisme', 'prix': 200,
                'localisation_requise': 0, 'documents_requis': 1, 'appel_exploratoire_requis': 0, 'decision_board_requis': 0,
                'duree_seance_minutes': 0, 'duree_tournage_minutes': 0, 'duree_production_minutes': 120, 'duree_finalisation_minutes': 0,
                'duree_affichee': '30 min × 4 visuels',
                'items': [
                    {'nom_item': 'Texte / contenu à intégrer', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                    {'nom_item': 'Couleurs / charte graphique', 'requires_file': 0, 'is_required': 1, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Logo (SVG)', 'requires_file': 1, 'is_required': 1, 'item_type': 'document', 'file_category': 'vecteur', 'field_type': 'file'},
                    {'nom_item': 'Format ou dimensions souhaités', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'text'},
                    {'nom_item': 'Exemples de style / références', 'requires_file': 0, 'is_required': 0, 'item_type': 'document', 'file_category': 'donnees', 'field_type': 'textarea'},
                ],
            },
        ]
        for _svc in _SERVICES_SEED:
            _existing = conn.execute(
                "SELECT 1 FROM services WHERE lower(nom_service) = lower(?)", (_svc['nom_service'],)
            ).fetchone()
            if _existing:
                continue
            _cur = conn.execute(
                """INSERT INTO services
                   (nom_service, description, icon, prix, localisation_requise, documents_requis,
                    appel_exploratoire_requis, decision_board_requis, duree_seance_minutes,
                    duree_tournage_minutes, duree_production_minutes, duree_finalisation_minutes,
                    slug, categorie, actif, duree_affichee)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)""",
                (_svc['nom_service'], _svc['description'], _svc['icon'], _svc['prix'],
                 _svc['localisation_requise'], _svc['documents_requis'], _svc['appel_exploratoire_requis'],
                 _svc['decision_board_requis'], _svc['duree_seance_minutes'], _svc['duree_tournage_minutes'],
                 _svc['duree_production_minutes'], _svc['duree_finalisation_minutes'], _svc['slug'],
                 _svc['categorie'], _svc['duree_affichee']),
            )
            _service_id = _cur.lastrowid
            for _idx, _item in enumerate(_svc['items']):
                conn.execute(
                    """INSERT INTO checklist_model_items
                       (id_service, nom_item, requires_file, is_required, item_type, file_category, field_type, position)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (_service_id, _item['nom_item'], _item['requires_file'], _item['is_required'],
                     _item['item_type'], _item['file_category'], _item['field_type'], _idx),
                )

        # Migration: colonne statut_updated_at sur projets
        try:
            conn.execute("ALTER TABLE projets ADD COLUMN statut_updated_at TEXT")
            conn.execute("UPDATE projets SET statut_updated_at = created_at WHERE statut_updated_at IS NULL")
        except Exception:
            pass  # colonne déjà présente
        # Migration: colonne titre_affiche sur projets
        try:
            conn.execute("ALTER TABLE projets ADD COLUMN titre_affiche TEXT")
        except Exception:
            pass  # colonne déjà présente
        # Migration: commentaire client sur les items de révision (distinct de text_value,
        # qui a la sémantique "renseigner = coché" pour les champs de saisie classiques —
        # un commentaire de révision ne doit jamais cocher l'item automatiquement)
        try:
            conn.execute("ALTER TABLE checklist_items ADD COLUMN commentaire TEXT")
        except Exception:
            pass  # colonne déjà présente
        # Migration: admin_resolu — coche admin distincte de est_coche (qui reflète la
        # réponse du CLIENT). Sert à suivre, item par item, quelles corrections demandées
        # ont été effectivement faites par l'équipe.
        try:
            conn.execute("ALTER TABLE checklist_items ADD COLUMN admin_resolu INTEGER DEFAULT 0")
        except Exception:
            pass  # colonne déjà présente
        # Migration: role admin (gestion/production) — distingue qui gère/assigne (accès
        # complet à l'admin) de qui produit (seul destinataire valide des mandats/tâches).
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN role TEXT")
        except Exception:
            pass  # colonne déjà présente
        # Migration: is_producteur_principal — seul(s) pigiste(s) marqué(s) ainsi peu(ven)t
        # être ciblé(s) par une création de mandat (voir api_admin_mandats_create /
        # api_admin_mandats_pigistes_create).
        try:
            conn.execute("ALTER TABLE pigistes ADD COLUMN is_producteur_principal INTEGER DEFAULT 0")
        except Exception:
            pass  # colonne déjà présente
        # Migration: assigne_admin_id — NULL = tâche partagée/visible par toute l'équipe
        # admin (comportement historique, préservé pour les tâches existantes), sinon
        # id du compte admin (clients.id) auquel la tâche appartient.
        try:
            conn.execute("ALTER TABLE todos_perso ADD COLUMN assigne_admin_id INTEGER")
        except Exception:
            pass  # colonne déjà présente
        # Migration: liaison roadmap (planification de campagne) <-> tâche perso
        # assignée <-> post marketing programmé, pour que cocher à n'importe lequel
        # des 3 endroits se propage aux autres (voir _sync_roadmap_todo_completion).
        try:
            conn.execute("ALTER TABLE roadmap_todos ADD COLUMN assigne_admin_id INTEGER")
        except Exception:
            pass  # colonne déjà présente
        try:
            conn.execute("ALTER TABLE roadmap_todos ADD COLUMN linked_todo_perso_id INTEGER")
        except Exception:
            pass  # colonne déjà présente
        try:
            conn.execute("ALTER TABLE roadmap_todos ADD COLUMN linked_marketing_post_id INTEGER")
        except Exception:
            pass  # colonne déjà présente
        try:
            conn.execute("ALTER TABLE todos_perso ADD COLUMN linked_roadmap_todo_id INTEGER")
        except Exception:
            pass  # colonne déjà présente
        try:
            conn.execute("ALTER TABLE marketing_posts ADD COLUMN linked_roadmap_todo_id INTEGER")
        except Exception:
            pass  # colonne déjà présente
        # Migration: assignation multi-personnes (une tâche/item peut être assigné à
        # plusieurs membres de l'équipe, pas juste un·e) — remplace l'usage de
        # assigne_admin_id (colonne conservée mais plus lue/écrite après cette migration).
        try:
            conn.execute('''CREATE TABLE IF NOT EXISTS todo_assignees (
                todo_id INTEGER NOT NULL REFERENCES todos_perso(id) ON DELETE CASCADE,
                admin_id INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
                PRIMARY KEY (todo_id, admin_id)
            )''')
        except Exception:
            pass
        try:
            conn.execute('''CREATE TABLE IF NOT EXISTS roadmap_todo_assignees (
                roadmap_todo_id INTEGER NOT NULL REFERENCES roadmap_todos(id) ON DELETE CASCADE,
                admin_id INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
                PRIMARY KEY (roadmap_todo_id, admin_id)
            )''')
        except Exception:
            pass
        # Migration: 5e silo de tâches identifié lors de l'audit du module Tâches
        # (2026-07-16) — les todos des phases de la roadmap produit CocktailOS
        # (page /admin/roadmaps, onglet CocktailOS) vivaient uniquement en
        # localStorage côté navigateur (clé cocktailos_todos_phase_<id>) : perdus au
        # vidage du cache, invisibles d'un autre appareil ou pour un autre admin.
        # Convergence Phase 5 : stockage serveur, pas d'injection dans todos_perso/PWA
        # Tâches — ce sont des items de planification produit interne sans
        # assigné·e ni échéance, pas des tâches d'équipe individuelles.
        try:
            conn.execute('''CREATE TABLE IF NOT EXISTS cocktailos_vision_todos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                phase_id INTEGER NOT NULL,
                texte TEXT NOT NULL,
                est_coche INTEGER NOT NULL DEFAULT 0,
                position INTEGER NOT NULL DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )''')
        except Exception:
            pass
        # Backfill unique : copie assigne_admin_id existant vers les tables de jonction
        # (idempotent — INSERT OR IGNORE sur la PK composite)
        try:
            conn.execute("""
                INSERT OR IGNORE INTO todo_assignees (todo_id, admin_id)
                SELECT id, assigne_admin_id FROM todos_perso WHERE assigne_admin_id IS NOT NULL
            """)
            conn.execute("""
                INSERT OR IGNORE INTO roadmap_todo_assignees (roadmap_todo_id, admin_id)
                SELECT id, assigne_admin_id FROM roadmap_todos WHERE assigne_admin_id IS NOT NULL
            """)
            conn.commit()
        except Exception as e:
            print(f"[MIGRATION] backfill todo_assignees: {e}")
        # Migration: visuels déposés sur un post marketing (upload direct sur le
        # portail, comme les fichiers de checklist projet).
        try:
            conn.execute('''CREATE TABLE IF NOT EXISTS marketing_post_fichiers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                id_post INTEGER NOT NULL REFERENCES marketing_posts(id) ON DELETE CASCADE,
                filename TEXT NOT NULL,
                filepath TEXT NOT NULL,
                uploaded_by INTEGER,
                created_at TEXT DEFAULT (datetime('now'))
            )''')
        except Exception:
            pass
        # Migration: content_key — déclare explicitement quel champ Sanity (ou sentinel
        # "membre[]") un item de checklist doit alimenter, pour remplacer la correspondance
        # de libellés devinée par le prefill de création de site.
        try:
            conn.execute("ALTER TABLE checklist_model_items ADD COLUMN content_key TEXT")
        except Exception:
            pass  # colonne déjà présente
        try:
            conn.execute("ALTER TABLE checklist_items ADD COLUMN content_key TEXT")
        except Exception:
            pass  # colonne déjà présente
        # Migration: id_projet sur sites — pour savoir quelle checklist pousser vers Sanity
        # à la création (voir _push_checklist_content_to_sanity)
        try:
            conn.execute("ALTER TABLE sites ADD COLUMN id_projet INTEGER")
        except Exception:
            pass  # colonne déjà présente
        # Migration: id_projet sur rendez_vous — permet de lier une réservation confirmée au
        # projet qui l'attendait (statut "En attente de rendez-vous") pour faire avancer son
        # statut automatiquement, sans deviner (voir _lier_rendez_vous_au_projet)
        try:
            conn.execute("ALTER TABLE rendez_vous ADD COLUMN id_projet INTEGER")
        except Exception:
            pass  # colonne déjà présente
        # Migration : coordonnées de contact directement sur une tâche (PWA Tâches, phase 2) —
        # ex. transférer un contact iPhone (.vcf) sur une tâche sans créer de fiche client.
        for _col in ('contact_nom', 'contact_telephone', 'contact_courriel'):
            try:
                conn.execute(f"ALTER TABLE todos_perso ADD COLUMN {_col} TEXT")
            except Exception:
                pass  # colonne déjà présente

        # Réconciliation one-shot de la checklist "Site Web Vitrine" : le seed _SERVICES_SEED
        # ci-dessus ne retouche jamais un service déjà en base, donc les items déjà créés
        # (avant l'introduction de content_key) ne l'auraient jamais reçu sans ce bloc dédié.
        # Idempotent : chaque UPDATE/INSERT/DELETE est sans effet s'il a déjà été appliqué.
        try:
            _vitrine_svc = conn.execute(
                "SELECT id FROM services WHERE lower(nom_service) = lower('Site Web Vitrine')"
            ).fetchone()
            if _vitrine_svc:
                _vid = _vitrine_svc['id']
                # Items morts : aucun slot d'asset ne les consomme (confirmé sur des projets réels)
                conn.execute(
                    "DELETE FROM checklist_model_items WHERE id_service = ? AND nom_item IN ('Logo 2', 'Logo 2-2')",
                    (_vid,)
                )
                # content_key sur les items existants — gabarit (checklist_model_items) ET
                # rétroactivement sur les checklist_items déjà instanciés sur de vrais projets
                # (backfill non destructif : ne touche que content_key, jamais text_value).
                # Certains items existants ont été créés avant l'ajout des accents au texte
                # source (ex. "Telephone"/"Texte section A propos" sans accent en DB) — les
                # deux graphies sont couvertes ici ; celle qui ne correspond à aucune ligne
                # ne fait simplement rien.
                _CONTENT_KEY_MAP = {
                    'Mission de l\'entreprise':  'pageEquipe.missionTexte',
                    'Vision de l\'entreprise':   'pageEquipe.visionTexte',
                    'Valeurs de l\'entreprise':  'pageEquipe.valeursTexte',
                    'Texte section À propos':    'pageEquipe.heroSousTitre',
                    'Texte section A propos':    'pageEquipe.heroSousTitre',
                    'Adresse':                   'siteSettings.adresse',
                    'Téléphone':                 'siteSettings.telephone',
                    'Telephone':                 'siteSettings.telephone',
                    'Courriel professionnel':    'siteSettings.courriel',
                    'Équipe':                    'membre[]',
                }
                for _nom, _key in _CONTENT_KEY_MAP.items():
                    conn.execute(
                        "UPDATE checklist_model_items SET content_key = ? WHERE id_service = ? AND nom_item = ?",
                        (_key, _vid, _nom)
                    )
                    conn.execute(
                        """UPDATE checklist_items SET content_key = ?
                           WHERE nom_item = ? AND id_checklist IN (
                               SELECT ck.id FROM checklistes ck JOIN projets p ON p.id = ck.id_projet
                               WHERE p.id_service = ?
                           )""",
                        (_key, _nom, _vid)
                    )
                # Nouveaux items réseaux sociaux — absents jusqu'ici, alors que le prefill de
                # création de site essaie déjà de les lire depuis la checklist.
                _existing_names = {r['nom_item'] for r in conn.execute(
                    "SELECT nom_item FROM checklist_model_items WHERE id_service = ?", (_vid,)
                ).fetchall()}
                _pos = conn.execute(
                    "SELECT COALESCE(MAX(position), 0) AS m FROM checklist_model_items WHERE id_service = ?", (_vid,)
                ).fetchone()['m']
                for _label, _key in (
                    ('Instagram', 'siteSettings.instagram'),
                    ('Facebook',  'siteSettings.facebook'),
                    ('LinkedIn',  'siteSettings.linkedin'),
                ):
                    if _label not in _existing_names:
                        _pos += 1
                        conn.execute(
                            """INSERT INTO checklist_model_items
                               (id_service, nom_item, requires_file, is_required, item_type, file_category, field_type, position, content_key)
                               VALUES (?, ?, 0, 0, 'document', 'donnees', 'text', ?, ?)""",
                            (_vid, _label, _pos, _key)
                        )
        except Exception:
            pass  # réconciliation déjà appliquée ou service absent

        # Migration: liaison Todoist sur todos_perso
        try:
            conn.execute("ALTER TABLE todos_perso ADD COLUMN todoist_task_id TEXT")
        except Exception:
            pass  # colonne déjà présente
        try:
            conn.execute("ALTER TABLE todos_perso ADD COLUMN source TEXT DEFAULT 'portail'")
        except Exception:
            pass  # colonne déjà présente
        # Migration: assignation directe d'un todo à un client (en plus du projet)
        try:
            conn.execute("ALTER TABLE todos_perso ADD COLUMN client_id INTEGER")
        except Exception:
            pass  # colonne déjà présente
        # Anti-doublon Todoist (poll multi-workers) : une tâche Todoist = un seul todo
        try:
            conn.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_todos_todoist "
                "ON todos_perso(todoist_task_id) WHERE todoist_task_id IS NOT NULL"
            )
        except Exception:
            pass
        # Migration: livrables logo(s) vectorisé(s) — déposés par l'admin, fournis au client à la livraison
        conn.execute('''CREATE TABLE IF NOT EXISTS projet_logo_fichiers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            id_projet INTEGER NOT NULL,
            filename TEXT NOT NULL,
            filepath TEXT NOT NULL,
            drive_file_id TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (id_projet) REFERENCES projets(id) ON DELETE CASCADE
        )''')
        # Migration: colonne sections_json sur client_ressources (sections d'un guide, pour classer les captures)
        try:
            conn.execute("ALTER TABLE client_ressources ADD COLUMN sections_json TEXT")
        except Exception:
            pass  # colonne déjà présente
        # Migration: table ressource_bundles + colonne bundle_id
        conn.execute('''CREATE TABLE IF NOT EXISTS ressource_bundles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            description TEXT,
            icone TEXT DEFAULT "folder",
            ordre INTEGER DEFAULT 0,
            created_at TEXT
        )''')
        try:
            conn.execute("ALTER TABLE client_ressources ADD COLUMN bundle_id INTEGER")
        except Exception:
            pass  # colonne déjà présente
        # Migration: table ressource_assignations (envoyer une ressource du catalogue à un client précis, depuis un projet, sans la rendre exclusive)
        conn.execute('''CREATE TABLE IF NOT EXISTS ressource_assignations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            id_ressource INTEGER NOT NULL,
            id_client INTEGER NOT NULL,
            id_projet INTEGER,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (id_ressource) REFERENCES client_ressources(id) ON DELETE CASCADE,
            FOREIGN KEY (id_client) REFERENCES clients(id) ON DELETE CASCADE,
            UNIQUE(id_ressource, id_client)
        )''')
        # Migration: colonne section_id sur ressource_images (rattacher une capture à une section du guide)
        try:
            conn.execute("ALTER TABLE ressource_images ADD COLUMN section_id TEXT")
        except Exception:
            pass  # colonne déjà présente
        # Migration: supprimer le CHECK constraint sur iv_logos.variante
        try:
            row = conn.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='iv_logos'").fetchone()
            if row and "CHECK" in row['sql']:
                conn.executescript("""
                    PRAGMA foreign_keys = OFF;
                    CREATE TABLE iv_logos_new (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        id_iv INTEGER NOT NULL,
                        variante TEXT NOT NULL,
                        drive_file_id TEXT,
                        public_url TEXT,
                        filename TEXT,
                        preview_file_id TEXT,
                        preview_url TEXT,
                        FOREIGN KEY (id_iv) REFERENCES identite_visuelle(id) ON DELETE CASCADE
                    );
                    INSERT INTO iv_logos_new (id, id_iv, variante, drive_file_id, public_url, filename, preview_file_id, preview_url)
                        SELECT id, id_iv, variante, drive_file_id,
                               CASE WHEN typeof(public_url) = 'null' THEN NULL ELSE public_url END,
                               filename,
                               CASE WHEN typeof(preview_file_id) = 'null' THEN NULL ELSE preview_file_id END,
                               CASE WHEN typeof(preview_url) = 'null' THEN NULL ELSE preview_url END
                        FROM iv_logos;
                    DROP TABLE iv_logos;
                    ALTER TABLE iv_logos_new RENAME TO iv_logos;
                    PRAGMA foreign_keys = ON;
                """)
        except Exception:
            pass
        # Trigger SQLite — maintient statut_updated_at automatiquement à chaque changement de statut
        conn.execute("""
            CREATE TRIGGER IF NOT EXISTS trg_statut_updated_at
            AFTER UPDATE OF statut ON projets
            BEGIN
                UPDATE projets SET statut_updated_at = CURRENT_TIMESTAMP WHERE id = NEW.id;
            END
        """)
        # Migration: type_prestation et quantite sur mandats
        try:
            conn.execute("ALTER TABLE mandats ADD COLUMN type_prestation TEXT")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE mandats ADD COLUMN quantite INTEGER DEFAULT 1")
        except Exception:
            pass
        # Migration: confirm_token sur clients (token à usage unique — protection scanners)
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN confirm_token TEXT")
        except Exception:
            pass
        # Migration: must_change_password (mot de passe temporaire admin)
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN must_change_password INTEGER DEFAULT 0")
        except Exception:
            pass
        # Migration: thème (police + couleurs) sur sites
        try:
            conn.execute("ALTER TABLE sites ADD COLUMN theme_font_pair TEXT")
            conn.execute("ALTER TABLE sites ADD COLUMN theme_accent_color TEXT")
            conn.execute("ALTER TABLE sites ADD COLUMN theme_bg_color TEXT")
        except Exception:
            pass
        # Migration: gabarit visuel nommé (refonte design, ex. "1a") sur sites
        # — colonne conservée mais inerte, système abandonné le 2026-07-07, voir `direction` ci-dessous
        try:
            conn.execute("ALTER TABLE sites ADD COLUMN style_variant TEXT")
        except Exception:
            pass
        # Migration: direction artistique Baseline (ex. "editorial") sur sites — template "vitrine"
        try:
            conn.execute("ALTER TABLE sites ADD COLUMN direction TEXT")
        except Exception:
            pass
        # Migration: colonnes facturation/Drive sur clients — déjà utilisées dans le code
        # mais jamais créées via migration (ajoutées hors-bande sur la DB de prod à l'origine)
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN mode_facturation TEXT DEFAULT 'projet'")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN adresse_facturation TEXT")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN ville_facturation TEXT")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN province_facturation TEXT DEFAULT 'Québec'")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN code_postal_facturation TEXT")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN pays_facturation TEXT DEFAULT 'Canada'")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN drive_folder_id TEXT")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN factures_folder_id TEXT")
        except Exception:
            pass
        # Migration: pipeline CRM — statut de relation prospect→client sur clients
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN statut_relation TEXT NOT NULL DEFAULT 'actif'")
        except Exception:
            pass
        try:
            conn.execute("ALTER TABLE clients ADD COLUMN prochain_suivi TEXT")
        except Exception:
            pass
        # Migration: champs fiche client CRM (onglet Aperçu — infos étendues)
        for _col in (
            "source_acquisition TEXT",
            "site_web TEXT",
            "adresse TEXT",
            "ville TEXT",
            "contact_secondaire_nom TEXT",
            "contact_secondaire_role TEXT",
            "contact_secondaire_email TEXT",
        ):
            try:
                conn.execute(f"ALTER TABLE clients ADD COLUMN {_col}")
            except Exception:
                pass
        conn.execute("""
            CREATE TABLE IF NOT EXISTS client_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                id_client INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
                contenu TEXT NOT NULL,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Journal d'activité client — uniquement pour les événements NON dérivables
        # des autres tables (ex. changement de statut). Le reste de la timeline
        # (création, notes, factures, RDV) est reconstruit à la lecture.
        conn.execute("""
            CREATE TABLE IF NOT EXISTS client_activite (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                id_client INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
                type TEXT NOT NULL,
                titre TEXT NOT NULL,
                meta TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        # Migration: n° de ligne fiscale (T2125 fédéral + TP-80 Québec) sur les dépenses
        for _col in ("ligne_t2125 TEXT", "ligne_tp80 TEXT"):
            try:
                conn.execute(f"ALTER TABLE transactions ADD COLUMN {_col}")
            except Exception:
                pass  # colonne déjà présente
        # Backfill des dépenses existantes selon leur catégorie
        for _cat, _lignes in LIGNES_FISCALES.items():
            conn.execute(
                "UPDATE transactions SET ligne_t2125 = ?, ligne_tp80 = ? "
                "WHERE type = 'depense' AND categorie = ? "
                "AND (ligne_t2125 IS NULL OR ligne_tp80 IS NULL)",
                (_lignes['t2125'], _lignes['tp80'], _cat)
            )
        # Migration: revenus multi-source — réf. externe + date de paiement des factures
        try:
            conn.execute("ALTER TABLE transactions ADD COLUMN source_ref TEXT")
        except Exception:
            pass  # colonne déjà présente
        try:
            conn.execute("ALTER TABLE factures ADD COLUMN date_paiement TEXT")
        except Exception:
            pass  # colonne déjà présente
        # Anti-doublon : une facture = au plus une ligne revenu matérialisée
        try:
            conn.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_revenu_facture "
                "ON transactions(id_facture) "
                "WHERE type='revenu' AND source='facture' AND id_facture IS NOT NULL"
            )
        except Exception:
            pass
        # Backfill : matérialiser le revenu des factures déjà payées
        try:
            for _f in conn.execute("SELECT id FROM factures WHERE statut='payee'").fetchall():
                materialiser_revenu_facture(conn, _f['id'])
        except Exception:
            pass
        # Phase 2 — Intégrations externes (Square, Shopify) : jetons OAuth par organisation
        conn.execute("""
            CREATE TABLE IF NOT EXISTS integrations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                organisation_id INTEGER,
                provider TEXT NOT NULL,
                merchant_id TEXT,
                access_token TEXT,
                refresh_token TEXT,
                token_expires_at TEXT,
                scopes TEXT,
                statut TEXT NOT NULL DEFAULT 'actif',
                connected_at TEXT DEFAULT CURRENT_TIMESTAMP,
                last_sync_at TEXT,
                meta TEXT
            )
        """)
        # Une connexion par fournisseur et par organisation (org NULL = instance actuelle)
        try:
            conn.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_integrations_provider_org "
                "ON integrations(provider, COALESCE(organisation_id, 0))"
            )
        except Exception:
            pass
        # Idempotence des revenus externes : un paiement source = une seule ligne
        try:
            conn.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_transactions_source_ref "
                "ON transactions(source, source_ref) WHERE source_ref IS NOT NULL"
            )
        except Exception:
            pass
        # Appareils jumelés pour l'app de capture (/capture) : jeton par appareil,
        # rattaché au compte de l'abonné (user_id) — indépendant de la session CRM.
        conn.execute("""
            CREATE TABLE IF NOT EXISTS capture_devices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                organisation_id INTEGER,
                token_hash TEXT NOT NULL UNIQUE,
                label TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                last_used_at TEXT,
                revoked INTEGER NOT NULL DEFAULT 0
            )
        """)
        # Appareils jumelés pour la PWA Tâches (/taches) : même principe que capture_devices,
        # mais un jeton par PERSONNE de l'équipe (chacun jumelle son propre téléphone avec
        # son propre compte) — indépendant de la session CRM.
        conn.execute("""
            CREATE TABLE IF NOT EXISTS task_devices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                token_hash TEXT NOT NULL UNIQUE,
                label TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                last_used_at TEXT,
                revoked INTEGER NOT NULL DEFAULT 0
            )
        """)
        # Boîte de réception comptable : factures détectées (Gmail, photo…) EN ATTENTE
        # de validation. Rien n'entre au grand livre sans approbation (filet anti-erreur).
        conn.execute("""
            CREATE TABLE IF NOT EXISTS factures_a_valider (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                organisation_id INTEGER,
                source TEXT NOT NULL DEFAULT 'gmail',
                source_ref TEXT,
                sens TEXT NOT NULL DEFAULT 'recu',
                expediteur TEXT,
                date_transaction TEXT,
                fournisseur TEXT,
                description TEXT,
                categorie TEXT,
                montant_avant_taxes REAL DEFAULT 0,
                montant_tps REAL DEFAULT 0,
                montant_tvq REAL DEFAULT 0,
                montant_total REAL DEFAULT 0,
                piece_jointe TEXT,
                confiance TEXT,
                note TEXT,
                statut TEXT NOT NULL DEFAULT 'en_attente',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                traite_at TEXT
            )
        """)
        # Anti-doublon : un même courriel/pièce source ne crée qu'une entrée
        try:
            conn.execute(
                "CREATE UNIQUE INDEX IF NOT EXISTS idx_avalider_source_ref "
                "ON factures_a_valider(source, source_ref) WHERE source_ref IS NOT NULL"
            )
        except Exception:
            pass
        conn.commit()
    finally:
        conn.close()

# Initialisation DB au démarrage du processus (Flask 3.x n'a plus before_first_request)
init_db()


def log_activite(conn, client_id, type_, titre, meta=None):
    """Journalise un événement client non dérivable (ex. changement de statut).
    N'échoue jamais le flux appelant : encapsulé en try/except."""
    try:
        conn.execute(
            "INSERT INTO client_activite (id_client, type, titre, meta) VALUES (?,?,?,?)",
            (client_id, type_, titre, meta),
        )
    except Exception as e:
        print(f"[ACTIVITE] log échoué (client {client_id}): {e}")

# ───────────────────────────────────────────────────────────
# Liens courts (raccourcisseur interne)
# ───────────────────────────────────────────────────────────
PORTAIL_URL = os.getenv('PORTAIL_URL', 'https://portail.cocktailmedia.ca')

# ── Tarifs pigistes (source de vérité interne) ──────────────────────────────
TARIFS_PIGISTE = [
    {'id': 'montage_court',     'categorie': 'Montage vidéo',  'label': 'Montage vidéo court simple',                  'prix': 15.0,  'unite': 'capsule'},
    {'id': 'montage_format',    'categorie': 'Montage vidéo',  'label': 'Adaptation d\'un montage → autre format',     'prix': 15.0,  'unite': 'format'},
    {'id': 'canva_declinaison', 'categorie': 'Graphisme Canva','label': 'Déclinaison / adaptation d\'un visuel existant','prix': 15.0, 'unite': 'visuel'},
    {'id': 'canva_simple',      'categorie': 'Graphisme Canva','label': 'Visuel simple à partir d\'un brief clair',    'prix': 20.0,  'unite': 'visuel'},
    {'id': 'canva_affiche',     'categorie': 'Graphisme Canva','label': 'Affiche ou flyer',                            'prix': 25.0,  'unite': 'création'},
    {'id': 'canva_lot4',        'categorie': 'Graphisme Canva','label': 'Petit lot de 4 visuels cohérents',            'prix': 100.0, 'unite': 'lot'},
]
_TARIFS_MAP = {t['id']: t for t in TARIFS_PIGISTE}

def create_short_link(url, expires_in=3600):
    """Stocke un lien et retourne son URL courte (ex: portail.cocktailmedia.ca/r/abc123)."""
    code = secrets.token_urlsafe(6)
    expires_at = (datetime.utcnow() + timedelta(seconds=expires_in)).isoformat()
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO short_links (code, url, expires_at) VALUES (?, ?, ?)",
        (code, url, expires_at)
    )
    conn.commit()
    conn.close()
    return f"{PORTAIL_URL}/r/{code}"

@app.route('/r/<code>')
def short_redirect(code):
    conn = get_db_connection()
    row = conn.execute(
        "SELECT url, expires_at FROM short_links WHERE code = ?", (code,)
    ).fetchone()
    conn.close()
    if not row or datetime.utcnow().isoformat() > row['expires_at']:
        return redirect(PORTAIL_URL)
    return redirect(row['url'])

# Scheduler facturation mensuelle
from billing_scheduler import init_scheduler, ajouter_ligne_facture_mensuelle
_scheduler = init_scheduler(app, mail)

# ───────────────────────────────────────────────────────────
# Utils / Security
# ───────────────────────────────────────────────────────────
def is_password_strong(password: str) -> bool:
    if len(password) < 8: return False
    if not re.search(r"[a-z]", password): return False
    if not re.search(r"[A-Z]", password): return False
    if not re.search(r"\d", password): return False
    if not re.search(r"[!@#$%^&*]", password): return False
    return True
def normalize_company_name(name: str) -> str:
    if not name:
        return ""
    nfkd = unicodedata.normalize("NFKD", name.lower())
    ascii_str = "".join(c for c in nfkd if not unicodedata.combining(c))
    cleaned = re.sub(r"[^\w\s]", "", ascii_str)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    cleaned = re.sub(r"s$", "", cleaned)
    return cleaned

def find_matching_company(nom_entreprise: str, conn: sqlite3.Connection, seuil: int = 85):
    needle = normalize_company_name(nom_entreprise)
    if not needle:
        return None
    rows = conn.execute(
        "SELECT id, nom_entreprise, drive_folder_id FROM clients WHERE nom_entreprise IS NOT NULL AND nom_entreprise != ''"
    ).fetchall()
    for row in rows:
        score = fuzz.ratio(needle, normalize_company_name(row["nom_entreprise"]))
        if score >= seuil:
            return row
    return None

def send_email(to_list, subject, body, html=None, cc=None, attachments=None):
    if not to_list:
        return
    if isinstance(to_list, str):
        to_list = [to_list]
    subject = subject.replace('\r', '').replace('\n', '')
    try:
        msg = Message(
            subject,
            sender=app.config['MAIL_DEFAULT_SENDER'] or app.config['MAIL_USERNAME'],
            recipients=to_list,
            cc=cc or []
        )
        msg.body = body
        if html:
            msg.html = html
        if attachments:
            for fname, fdata in attachments:
                msg.attach(fname, 'application/pdf', fdata)
        mail.send(msg)
    except Exception as e:
        print(f"[MAIL] Erreur d'envoi: {e}")
def send_email_client(client_id_or_row, subject, body, html=None, attachments=None):
    """Envoie un email à un client SEULEMENT si son compte est confirmé."""
    conn = get_db_connection()
    if isinstance(client_id_or_row, int):
        client = conn.execute("SELECT email, is_email_confirmed FROM clients WHERE id=?", (client_id_or_row,)).fetchone()
    else:
        client = client_id_or_row
    conn.close()
    if not client:
        return
    if not int(client['is_email_confirmed'] or 0):
        print(f"[MAIL] Email bloqué (compte non confirmé) : {client['email']} — {subject}")
        return
    send_email(client['email'], subject, body, html=html, attachments=attachments)

def push_notification(conn, id_client, id_projet, message, type='info'):
    """Insère une notification in-app pour un client."""
    try:
        conn.execute(
            "INSERT INTO notifications (id_client, id_projet, message, type) VALUES (?, ?, ?, ?)",
            (id_client, id_projet, message, type)
        )
    except Exception as e:
        print(f"[NOTIF] Erreur insertion notification: {e}")

def push_admin_notif(conn, titre, message='', type='info', lien=None, destinataire=None):
    """Insère une notification interne pour un admin (to-do, Todoist, assignation…).
       destinataire = email admin ciblé, ou None = visible par tous les admins."""
    try:
        conn.execute(
            "INSERT INTO admin_notifications (destinataire, type, titre, message, lien) VALUES (?, ?, ?, ?, ?)",
            (destinataire or None, type, titre, message or '', lien or None)
        )
    except Exception as e:
        print(f"[NOTIF-ADMIN] Erreur insertion: {e}")
    # Web Push (notif téléphone/desktop même portail fermé)
    try:
        send_web_push_to_admins(conn, titre, message or '', lien or '/admin', destinataire)
    except Exception as e:
        print(f"[WEBPUSH] envoi depuis notif: {e}")

VAPID_PUBLIC_KEY  = os.getenv('VAPID_PUBLIC_KEY', '')
VAPID_PRIVATE_KEY = os.getenv('VAPID_PRIVATE_KEY', '')
VAPID_SUBJECT     = os.getenv('VAPID_SUBJECT', 'mailto:felix.dumont@cocktailmedia.ca')

def send_web_push_to_admins(conn, titre, message, lien, destinataire=None):
    """Envoie une notif Web Push aux abonnements admin. Nettoie les abonnements expirés (404/410)."""
    if not VAPID_PRIVATE_KEY:
        return
    try:
        from pywebpush import webpush, WebPushException
    except ImportError:
        return
    import json as _json
    # destinataire None/'' = tous les abonnements ; sinon seulement celui ciblé
    if destinataire:
        subs = conn.execute("SELECT * FROM push_subscriptions WHERE email=? OR email IS NULL", (destinataire,)).fetchall()
    else:
        subs = conn.execute("SELECT * FROM push_subscriptions").fetchall()
    payload = _json.dumps({'title': titre, 'body': message, 'url': lien})
    for s in subs:
        try:
            webpush(
                subscription_info={'endpoint': s['endpoint'], 'keys': {'p256dh': s['p256dh'], 'auth': s['auth']}},
                data=payload,
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims={'sub': VAPID_SUBJECT},
            )
        except WebPushException as e:
            code = getattr(e.response, 'status_code', None)
            if code in (404, 410):
                conn.execute("DELETE FROM push_subscriptions WHERE id=?", (s['id'],))
                print(f"[WEBPUSH] abonnement expiré supprimé ({code})")
            else:
                print(f"[WEBPUSH] échec envoi: {e}")
        except Exception as e:
            print(f"[WEBPUSH] erreur: {e}")

FILE_CATEGORIES = {
    'photo':     {'label': '📸 Photo',            'extensions': {'jpg','jpeg','png','heic','webp'}},
    'vecteur':   {'label': '🎨 Vecteur/Graphique', 'extensions': {'svg','ai','psd','png','eps'}},
    'video':     {'label': '🎬 Vidéo',             'extensions': {'mp4','mov','avi','mkv'}},
    'document':  {'label': '📄 Document',          'extensions': {'pdf','doc','docx','odt'}},
    'donnees':   {'label': '📊 Données',           'extensions': {'xls','xlsx','csv'}},
    'archive':   {'label': '📦 Archive',           'extensions': {'zip','rar'}},
    'autre':     {'label': '🔗 Autre',             'extensions': set()},
}

def allowed(filename: str) -> bool:
    return "." in filename and filename.rsplit(".",1)[1].lower() in app.config["ALLOWED_EXTENSIONS"]

def allowed_for_category(filename: str, category: str) -> bool:
    """Vérifie si l'extension du fichier correspond à la catégorie attendue."""
    if not filename or category not in FILE_CATEGORIES:
        return True  # Pas de restriction si catégorie inconnue
    cat = FILE_CATEGORIES[category]
    if not cat['extensions']:
        return True  # 'autre' = tout accepté
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return ext in cat['extensions']

def login_required(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if 'user_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({"error": "Non authentifié"}), 401
            return redirect(url_for('accueil'))
        return f(*args, **kwargs)
    return wrap

def admin_required(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if not bool(session.get('is_admin', False)):
            if request.path.startswith('/api/'):
                return jsonify({"error": "Non autorisé"}), 401
            return redirect(url_for('accueil'))
        return f(*args, **kwargs)
    return wrap

def pigiste_required(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if not session.get('pigiste_id'):
            return jsonify({'error': 'Non authentifié'}), 401
        return f(*args, **kwargs)
    return wrap

import redis as _redis_mod
_redis_client = _redis_mod.from_url("redis://redis:6379", decode_responses=True)

def _redis_rate_limit(key: str, ttl_seconds: int) -> bool:
    """Retourne True si l'action est autorisée, False si elle est limitée."""
    return bool(_redis_client.set(key, 1, ex=ttl_seconds, nx=True))

def _redis_login_locked(email: str) -> bool:
    """Retourne True si le compte est verrouillé (≥10 échecs en 15 min)."""
    return int(_redis_client.get(f"login_fail:{email}") or 0) >= 10

def _redis_login_fail_inc(email: str) -> None:
    """Incrémente le compteur d'échecs de login pour cet email."""
    key = f"login_fail:{email}"
    c = _redis_client.incr(key)
    if c == 1:
        _redis_client.expire(key, 900)

def _redis_login_reset(email: str) -> None:
    """Efface le compteur d'échecs après un login réussi."""
    _redis_client.delete(f"login_fail:{email}")

# ───────────────────────────────────────────────────────────
# Notifications settings
# ───────────────────────────────────────────────────────────
def get_notification_settings() -> dict:
    conn = get_db_connection()
    row = conn.execute("""
        SELECT admin_emails, client_updates, admin_updates
        FROM notification_settings WHERE id = 1
    """).fetchone()
    conn.close()
    if not row:
        return {"admin_emails": [], "client_updates": 1, "admin_updates": 1}
    raw = (row['admin_emails'] or '').strip()
    try:
        import json as _j
        parsed = _j.loads(raw)
        admin_emails = [e.strip() for e in (parsed if isinstance(parsed, list) else [parsed]) if e.strip()]
    except Exception:
        admin_emails = [e.strip() for e in raw.split(',') if e.strip()]
    return {
        "admin_emails": admin_emails,
        "client_updates": int(row['client_updates'] or 1),
        "admin_updates": int(row['admin_updates'] or 1),
    }

# ───────────────────────────────────────────────────────────
# Readiness / Pastille
# ───────────────────────────────────────────────────────────
def table_has_column(conn: sqlite3.Connection, table: str, column: str) -> bool:
    try:
        info = conn.execute(f"PRAGMA table_info({table})").fetchall()
    except sqlite3.Error:
        return False
    return any(
        (row["name"] if hasattr(row, "keys") else row[1]) == column
        for row in info
    )

def compute_checklist_readiness(project_id: int) -> Tuple[bool, int, int]:
    """
    Retourne (is_ready, done_count, total_required)
    - is_ready == True si TOUS les items obligatoires sont cochés
      et, s'ils nécessitent un fichier, qu'un file_path est présent.
    """
    conn = get_db_connection()
    row = conn.execute("""
        SELECT ch.id AS id_checklist
        FROM checklistes ch WHERE ch.id_projet = ?
    """, (project_id,)).fetchone()

    if not row:
        conn.close()
        return (False, 0, 0)

    id_checklist = row['id_checklist'] if 'id_checklist' in row.keys() else row['id']
    has_required_column = table_has_column(conn, "checklist_items", "is_required")

    select_columns = "id, est_coche, requires_file, file_path"
    if has_required_column:
        select_columns += ", is_required"

    items = conn.execute(
        f"""
        SELECT {select_columns}
        FROM checklist_items
        WHERE id_checklist = ?
    """,
        (id_checklist,),
    ).fetchall()
    conn.close()

    total_required = 0
    done_required = 0
    for it in items:
        is_required = 1
        if has_required_column:
            is_required = int(it["is_required"] or 0)
        if is_required != 1:
            continue
        total_required += 1
        ok = int(it['est_coche'] or 0) == 1
        if int(it['requires_file'] or 0) == 1:
            ok = ok and bool(it['file_path'])
        if ok:
            done_required += 1

    return (done_required == total_required and total_required > 0, done_required, total_required)

def _notify_checklist_complete(conn, projet):
    """Courriel interne à Félix quand la checklist de documents d'un projet passe au
    complet (statut 'Documents à donner' -> 'Documents reçus')."""
    try:
        client = conn.execute("SELECT nom_complet FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
        corps = (
            f"La checklist de documents est complète pour « {projet['nom_projet']} »"
            f" ({client['nom_complet'] if client else 'client inconnu'}).\n\n"
            f"Voir le projet : {PORTAIL_URL}/admin/projet/{projet['id']}"
        )
        send_email('felix.dumont@cocktailmedia.ca', f"Checklist complétée — {projet['nom_projet']}", corps)
    except Exception as e:
        print(f"[MAIL] _notify_checklist_complete: {e}")

def compute_revision_readiness(project_id: int) -> Tuple[bool, bool]:
    """Retourne (all_resolved, has_comments) pour les items de révision (is_revision=1) d'un
       projet — un item est résolu s'il est coché (approuvé) OU porte un commentaire (une
       demande de changement du client compte comme traité de son côté, pas comme en attente).
       Le commentaire du client est stocké dans text_value (saisi via /api/v1/item/text/,
       le même champ que le reste du checklist) — pas dans la colonne commentaire, qui n'est
       jamais alimentée par l'UI actuelle."""
    conn = get_db_connection()
    row = conn.execute("SELECT id FROM checklistes WHERE id_projet = ?", (project_id,)).fetchone()
    if not row:
        conn.close()
        return (False, False)
    items = conn.execute(
        "SELECT est_coche, text_value FROM checklist_items WHERE id_checklist = ? AND is_revision = 1",
        (row['id'],)
    ).fetchall()
    conn.close()
    if not items:
        return (False, False)
    all_resolved = all(int(it['est_coche'] or 0) == 1 or (it['text_value'] or '').strip() for it in items)
    has_comments = any((it['text_value'] or '').strip() for it in items)
    return (all_resolved, has_comments)

def pastille_color(is_ready: bool) -> str:
    # Rouge si pas prêt, Orange si prêt
    return "orange" if is_ready else "red"

    # ───────────────────────────────────────────────────────────
# Phases Projet (4 statuts + couleurs) + helpers
# ───────────────────────────────────────────────────────────
PROJECT_PHASES = [
    ("Documents à donner", "red"),
    ("Documents reçus", "blue"),
    ("Travaux en cours", "orange"),
    ("En révision", "purple"),
    ("Finalisation", "teal"),
    ("Travaux terminés", "green"),
]

# Variantes acceptées → label normalisé
_PHASE_ALIAS = {
    "documents a donner": "Documents à donner",
    "document a donner": "Documents à donner",
    "doc a donner": "Documents à donner",
    "documents à donner": "Documents à donner",

    "documents reçus": "Documents reçus",
    "documents recus": "Documents reçus",
    "doc recu": "Documents reçus",
    "doc reçu": "Documents reçus",

    "travaux en cours": "Travaux en cours",
    "en cours": "Travaux en cours",

    "en révision": "En révision",
    "en revision": "En révision",
    "révision": "En révision",
    "revision": "En révision",

    "finalisation": "Finalisation",
    "finalisation vf": "Finalisation",
    "vf": "Finalisation",

    "travaux terminés": "Travaux terminés",
    "travaux termines": "Travaux terminés",
    "terminé": "Travaux terminés",
    "termine": "Travaux terminés",
    "terminés": "Travaux terminés",
    "fini": "Travaux terminés",
}

def normalize_status(raw: str) -> str:
    if not raw:
        return "Documents à donner"
    key = raw.strip().lower()
    return _PHASE_ALIAS.get(key, raw.strip())

def status_color(status_label: str) -> str:
    """Couleur (red/blue/orange/green/purple) selon la phase."""
    lab = normalize_status(status_label)
    color_map = {
        "Documents à donner": "red",
        "Documents reçus":    "blue",
        "Travaux en cours":   "orange",
        "En révision":        "purple",
        "Finalisation":       "teal",
        "Travaux terminés":   "green",
        "Complété":           "green",
        "Annulé":             "grey",
        "En attente de rendez-vous": "grey",
    }
    return color_map.get(lab, "red")
def status_badge_class(status_label: str) -> str:
    """Classe CSS badge en fonction de la couleur de phase."""
    color = status_color(status_label)
    return f"badge badge-{color}"

# ───────────────────────────────────────────────────────────
# Pipeline de phase — API moderne (/api/v1, portail-next)
# Distinct de _PHASE_ALIAS/normalize_status ci-dessus, qui reste utilisé tel quel par
# l'ancienne UI Jinja (legacy, hors scope de ce système). Source unique de vérité pour
# la progression % et pour la liste d'étapes affichée par service (pipeline_for_service) —
# le frontend n'a plus à deviner quelles étapes s'appliquent à un service donné.
PHASE_CONFIG = {
    'En attente de rendez-vous': {'progress': 0,   'pipeline': False},
    'Documents à donner':        {'progress': 15,  'pipeline': True},
    'Documents reçus':           {'progress': 30,  'pipeline': True},
    'Travaux en cours':          {'progress': 45,  'pipeline': True},
    'En révision':               {'progress': 80,  'pipeline': True},
    'Corrections en cours':      {'progress': 85,  'pipeline': False},
    'Complété':                  {'progress': 100, 'pipeline': True},
    'Annulé':                    {'progress': 0,   'pipeline': False},
}

def phase_progress(statut: str) -> int:
    return PHASE_CONFIG.get(statut, {}).get('progress', 0)

def pipeline_for_service(service_row) -> list:
    """Liste ordonnée des statuts pertinents pour CE service. appel_exploratoire_requis et
       documents_requis sont deux exigences INDÉPENDANTES (un service peut demander les
       deux, une seule, ou aucune) — pas un choix exclusif. Même logique que
       initial_statut_for_service ci-dessous, qui doit toujours rester synchronisée avec
       cette fonction : le statut initial d'un projet doit toujours être une étape présente
       dans son propre pipeline."""
    appel_requis = bool(service_row['appel_exploratoire_requis']) if service_row is not None else False
    documents_requis = bool(service_row['documents_requis']) if service_row is not None else True
    steps = []
    if appel_requis:
        steps.append('En attente de rendez-vous')
    if documents_requis:
        steps += ['Documents à donner', 'Documents reçus']
    steps += ['Travaux en cours', 'En révision', 'Complété']
    return steps

def initial_statut_for_service(service_row) -> str:
    """Statut de départ d'un projet selon les exigences réelles de son service — rendez-vous
       d'abord si requis, sinon documents si requis, sinon directement au travail (voir
       _do_start_travaux appelé synchroneement dans ce cas par les points de création)."""
    appel_requis = bool(service_row['appel_exploratoire_requis']) if service_row is not None else False
    documents_requis = bool(service_row['documents_requis']) if service_row is not None else True
    if appel_requis:
        return 'En attente de rendez-vous'
    if documents_requis:
        return 'Documents à donner'
    return 'Travaux en cours'

def pipeline_with_current_statut(steps: list, statut: str) -> list:
    """'Corrections en cours' n'est pas une étape systématique du pipeline (seuls les
       projets où le client a demandé des changements pendant la révision y passent) —
       on l'insère seulement dans la liste affichée quand le projet y est actuellement,
       entre 'En révision' et 'Complété', pour que le fil d'étapes reste cohérent."""
    if statut == 'Corrections en cours' and 'Corrections en cours' not in steps:
        steps = list(steps)
        if 'Complété' in steps:
            steps.insert(steps.index('Complété'), 'Corrections en cours')
        else:
            steps.append('Corrections en cours')
    return steps

import json as _json
app.jinja_env.filters['fromjson'] = lambda s: _json.loads(s) if s else []
# Exposer les helpers aux templates
app.jinja_env.globals.update(
    normalize_status=normalize_status,
    status_badge_class=status_badge_class
)

import json as _json
from datetime import datetime as _dt, timedelta as _td

@app.template_filter('from_json')
def from_json_filter(value):
    try:
        return _json.loads(value)
    except:
        return []

@app.template_filter('mois_precedent')
def mois_precedent_filter(mois):
    try:
        d = _dt.strptime(mois, '%Y-%m')
        d2 = d.replace(day=1) - _td(days=1)
        return d2.strftime('%Y-%m')
    except:
        return mois

@app.template_filter('mois_suivant')
def mois_suivant_filter(mois):
    try:
        d = _dt.strptime(mois, '%Y-%m')
        d2 = d.replace(day=28) + _td(days=4)
        return d2.strftime('%Y-%m')
    except:
        return mois

@app.template_filter('jours_du_mois')
def jours_du_mois_filter(mois):
    try:
        import calendar
        d = _dt.strptime(mois, '%Y-%m')
        premier_jour = d.weekday()  # 0=lundi
        nb_jours = calendar.monthrange(d.year, d.month)[1]
        cells = [None] * premier_jour + list(range(1, nb_jours + 1))
        # Compléter jusqu'à multiple de 7
        while len(cells) % 7 != 0:
            cells.append(None)
        return cells
    except:
        return []

# ───────────────────────────────────────────────────────────
# API v1 — JSON endpoints pour Next.js
# ───────────────────────────────────────────────────────────
from flask import jsonify
from flask_cors import CORS

_cors_origins = ["https://portail.cocktailmedia.ca", os.getenv("PORTAIL_URL", "")]
if os.getenv("FLASK_ENV", "development") != "production":
    _cors_origins.append("http://localhost:3001")
CORS(app, origins=_cors_origins, supports_credentials=True)

@app.route('/api/v1/auth/login', methods=['POST'])
@limiter.limit("5 per minute")
def api_login():
    data = request.get_json() or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')

    if email and _redis_login_locked(email):
        return jsonify({"error": "Trop de tentatives. Compte temporairement verrouillé (15 min)."}), 429

    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()

    if user and user['auth_provider'] == 'password':
        if not user['mot_de_passe_hash']:
            conn.close()
            _redis_login_fail_inc(email)
            return jsonify({"error": "Votre compte n'est pas encore activé. Vérifiez votre courriel pour le lien d'invitation."}), 403
        try:
            hash_ok = bcrypt.check_password_hash(user['mot_de_passe_hash'], password)
        except Exception:
            conn.close()
            _redis_login_fail_inc(email)
            return jsonify({"error": "Votre compte n'est pas encore activé. Vérifiez votre courriel pour le lien d'invitation."}), 403
        if hash_ok:
            conn.close()
            if not int(user['is_email_confirmed'] or 0):
                return jsonify({"error": "Email non confirmé. Vérifiez votre boîte courriel."}), 403
            _redis_login_reset(email)
            session.permanent = True
            session['user_id'] = user['id']
            session['user_name'] = user['nom_complet']
            session['is_admin'] = bool(user['is_admin'])
            session['has_outils'] = bool(user['has_outils'])
            session['has_entrainement'] = bool(user['has_entrainement']) if 'has_entrainement' in user.keys() else False
            force_pw = bool(user['must_change_password']) if 'must_change_password' in user.keys() else False
            return jsonify({
                "success": True,
                "force_password_change": force_pw,
                "user": {
                    "id": user['id'],
                    "nom": user['nom_complet'],
                    "email": user['email'],
                    "is_admin": bool(user['is_admin']),
                    "role": "admin" if user['is_admin'] else "client",
                    "entrainement_only": bool(user['entrainement_only']) if 'entrainement_only' in user.keys() else False
                }
            })

    # Fallback: check pigistes table
    pigiste = conn.execute("SELECT * FROM pigistes WHERE email = ? AND is_active = 1", (email,)).fetchone()
    conn.close()
    if pigiste and pigiste['mot_de_passe_hash'] and bcrypt.check_password_hash(pigiste['mot_de_passe_hash'], password):
        _redis_login_reset(email)
        session.clear()
        session.permanent = True
        session['pigiste_id'] = pigiste['id']
        session['pigiste_nom'] = pigiste['nom_complet']
        session['role'] = 'pigiste'
        return jsonify({
            "success": True,
            "user": {
                "id": pigiste['id'],
                "nom": pigiste['nom_complet'],
                "email": pigiste['email'],
                "role": "pigiste"
            }
        })

    _redis_login_fail_inc(email)
    return jsonify({"error": "Email ou mot de passe incorrect"}), 401

@app.route('/api/v1/auth/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({"success": True})

@app.route('/api/v1/auth/me', methods=['GET'])
def api_me():
    if session.get('pigiste_id'):
        return jsonify({
            "id": session['pigiste_id'],
            "nom": session['pigiste_nom'],
            "role": "pigiste",
            "is_admin": False
        })
    if 'user_id' not in session:
        return jsonify({"error": "Non authentifié"}), 401
    # Repli DB pour les sessions ouvertes avant l'ajout des flags (évite une reconnexion forcée)
    if 'has_entrainement' not in session or 'entrainement_only' not in session:
        conn = get_db_connection()
        row = conn.execute("SELECT has_entrainement, entrainement_only FROM clients WHERE id = ?", (session['user_id'],)).fetchone()
        conn.close()
        session['has_entrainement'] = bool(row['has_entrainement']) if row and row['has_entrainement'] is not None else False
        session['entrainement_only'] = bool(row['entrainement_only']) if row and row['entrainement_only'] is not None else False
    return jsonify({
        "id": session['user_id'],
        "nom": session['user_name'],
        "is_admin": session.get('is_admin', False),
        "role": "admin" if session.get('is_admin') else "client",
        "has_outils": session.get('has_outils', False),
        "has_entrainement": session.get('has_entrainement', False),
        "entrainement_only": session.get('entrainement_only', False),
    })


# ──────────────────────────────────────────────────────────────
#  ENTRAÎNEMENT — plan personnel par client (jour-par-jour, suivi)
# ──────────────────────────────────────────────────────────────

@app.route('/api/v1/entrainement/me', methods=['GET'])
@login_required
def api_entrainement_me():
    """Renvoie le plan actif du client connecté (isolation par session)."""
    client_id = session['user_id']
    conn = get_db_connection()
    plan = conn.execute(
        """SELECT id, titre, note, contenu_json, created_at
             FROM entrainement_plans
            WHERE client_id = ? AND actif = 1
            ORDER BY created_at DESC, id DESC
            LIMIT 1""",
        (client_id,)
    ).fetchone()
    conn.close()
    if not plan:
        return jsonify({"error": "Aucun plan"}), 404
    try:
        contenu = json.loads(plan['contenu_json'] or '{}')
    except (ValueError, TypeError):
        contenu = {}
    return jsonify({
        "id": plan['id'],
        "titre": plan['titre'],
        "note": plan['note'],
        "contenu": contenu,
        "created_at": plan['created_at'],
    })


@app.route('/api/v1/entrainement/progress', methods=['GET'])
@login_required
def api_entrainement_progress_get():
    """Renvoie les cases cochées du client pour un plan donné."""
    client_id = session['user_id']
    plan_id = request.args.get('plan_id', type=int)
    if not plan_id:
        return jsonify({"error": "plan_id requis"}), 400
    conn = get_db_connection()
    # Vérifie que le plan appartient bien au client
    owns = conn.execute(
        "SELECT 1 FROM entrainement_plans WHERE id = ? AND client_id = ?", (plan_id, client_id)
    ).fetchone()
    if not owns:
        conn.close()
        return jsonify({"error": "Introuvable"}), 404
    rows = conn.execute(
        """SELECT exercice_key, date FROM entrainement_progress
            WHERE client_id = ? AND plan_id = ? AND done = 1""",
        (client_id, plan_id)
    ).fetchall()
    conn.close()
    return jsonify({"entries": [{"exercice_key": r['exercice_key'], "date": r['date']} for r in rows]})


@app.route('/api/v1/entrainement/progress', methods=['POST'])
@login_required
def api_entrainement_progress_post():
    """Coche / décoche un exercice pour une date donnée."""
    client_id = session['user_id']
    data = request.get_json() or {}
    plan_id = data.get('plan_id')
    exercice_key = (data.get('exercice_key') or '').strip()
    date = (data.get('date') or '').strip()
    done = bool(data.get('done', True))
    if not plan_id or not exercice_key or not date:
        return jsonify({"error": "plan_id, exercice_key et date requis"}), 400
    conn = get_db_connection()
    owns = conn.execute(
        "SELECT 1 FROM entrainement_plans WHERE id = ? AND client_id = ?", (plan_id, client_id)
    ).fetchone()
    if not owns:
        conn.close()
        return jsonify({"error": "Introuvable"}), 404
    if done:
        conn.execute(
            """INSERT INTO entrainement_progress (client_id, plan_id, exercice_key, date, done)
                    VALUES (?, ?, ?, ?, 1)
               ON CONFLICT(client_id, plan_id, exercice_key, date)
                    DO UPDATE SET done = 1""",
            (client_id, plan_id, exercice_key, date)
        )
    else:
        conn.execute(
            "DELETE FROM entrainement_progress WHERE client_id = ? AND plan_id = ? AND exercice_key = ? AND date = ?",
            (client_id, plan_id, exercice_key, date)
        )
    conn.close()
    return jsonify({"success": True, "done": done})


@app.route('/api/v1/auth/register', methods=['POST'])
@limiter.limit("5 per minute")
def api_register():
    data = request.get_json() or {}
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')
    password2 = data.get('password2', '')
    nom = data.get('nom_complet', '').strip()
    nom_entreprise = data.get('nom_entreprise', '').strip()
    telephone = data.get('telephone', '').strip()

    if not nom_entreprise:
        return jsonify({"error": "Le nom d'entreprise est obligatoire"}), 400
    if password != password2:
        return jsonify({"error": "Les mots de passe ne correspondent pas"}), 400
    if not is_password_strong(password):
        return jsonify({"error": "Mot de passe trop faible — min. 8 caractères, majuscule, minuscule, chiffre et caractère spécial"}), 400

    conn = get_db_connection()
    exists = conn.execute("SELECT is_email_confirmed FROM clients WHERE email = ?", (email,)).fetchone()
    if exists:
        conn.close()
        if not int(exists['is_email_confirmed'] or 0):
            # Renvoi silencieux — ne révèle pas l'état du compte
            try:
                new_token = s.dumps(email, salt='email-confirm-salt')
                confirm_url = f"{PORTAIL_URL}/confirm-email?token={new_token}"
                conn2 = get_db_connection()
                u = conn2.execute("SELECT nom_complet FROM clients WHERE email = ?", (email,)).fetchone()
                conn2.execute("UPDATE clients SET confirm_token = ? WHERE email = ?", (new_token, email))
                conn2.commit()
                conn2.close()
                html_confirm = _base_confirm(u['nom_complet'] if u else '', confirm_url)
                send_email(email, "Confirmez votre compte — Cocktail Média",
                           f"Confirmez votre compte : {confirm_url}", html=html_confirm)
            except Exception:
                pass
            return jsonify({"error": "Si cette adresse n'est pas encore utilisée, un courriel de confirmation vous sera envoyé.", "unconfirmed": True}), 400
        return jsonify({"error": "Si cette adresse n'est pas encore utilisée, un courriel de confirmation vous sera envoyé."}), 400

    hashed = bcrypt.generate_password_hash(password).decode('utf-8')

    confirm_token = s.dumps(email, salt='email-confirm-salt')
    conn.execute("""
        INSERT INTO clients (nom_complet, email, nom_entreprise, telephone, mot_de_passe_hash, auth_provider, is_email_confirmed, is_admin, confirm_token)
        VALUES (?, ?, ?, ?, ?, 'password', 0, 0, ?)
    """, (nom, email, nom_entreprise, telephone, hashed, confirm_token))
    conn.commit()
    conn.close()

    # Drive + email : opérations non-bloquantes (fire-and-forget)
    import threading

    def _post_register_tasks():
        # send_email (Flask-Mail) a besoin d'un contexte applicatif — sans ce
        # `with`, l'envoi échoue silencieusement ("Working outside of application
        # context"), avalé par le try/except : le client ne reçoit jamais son
        # courriel de confirmation.
        with app.app_context():
            try:
                new_folder_id = create_folder(nom_entreprise, parent_id=os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID'))
                factures_folder_id = create_folder("Factures", parent_id=new_folder_id)
                c = get_db_connection()
                c.execute("UPDATE clients SET drive_folder_id = ?, factures_folder_id = ? WHERE email = ?",
                          (new_folder_id, factures_folder_id, email))
                c.close()
            except Exception as e:
                print(f"[DRIVE] Création dossier client échouée: {e}")
            try:
                confirm_url = f"{PORTAIL_URL}/confirm-email?token={confirm_token}"
                html_confirm = _base_confirm(nom, confirm_url)
                send_email(email, "Confirmez votre compte — Cocktail Média",
                           f"Bonjour {nom}, confirmez votre compte : {confirm_url}", html=html_confirm)
            except Exception as e:
                print(f"[MAIL] Confirmation échouée: {e}")

    threading.Thread(target=_post_register_tasks, daemon=True).start()
    return jsonify({"success": True})

@app.route('/api/v1/dashboard', methods=['GET'])
@login_required
def api_dashboard():
    user_id = session['user_id']
    conn = get_db_connection()
    client_row = conn.execute(
        "SELECT email, nom_entreprise, nom_complet, drive_folder_id FROM clients WHERE id = ?", (user_id,)
    ).fetchone()
    drive_folder_id = client_row['drive_folder_id'] if client_row else None

    if client_row and client_row['email']:
        try:
            from drive_service import get_drive_service as _gds
            _svc = _gds()
            folder_ok = False
            if drive_folder_id:
                try:
                    _svc.files().get(fileId=drive_folder_id, fields='id', supportsAllDrives=True).execute()
                    folder_ok = True
                except Exception:
                    drive_folder_id = None  # dossier introuvable — on va le recréer

            if not folder_ok:
                nom_dossier = client_row['nom_entreprise'] or client_row['nom_complet']
                drive_folder_id = create_folder(nom_dossier, parent_id=os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID'))
                create_folder("Factures", parent_id=drive_folder_id)
                conn.execute(
                    "UPDATE clients SET drive_folder_id = ? WHERE id = ?", (drive_folder_id, user_id)
                )
                conn.commit()

            share_folder_with_user(drive_folder_id, client_row['email'])

            # Partager tous les dossiers IV des projets du client
            iv_folders = conn.execute("""
                SELECT iv.iv_folder_id FROM identite_visuelle iv
                JOIN projets p ON p.id = iv.id_projet
                WHERE p.id_client = ? AND iv.iv_folder_id IS NOT NULL
            """, (user_id,)).fetchall()
            for row in iv_folders:
                try:
                    share_folder_with_user(row['iv_folder_id'], client_row['email'])
                except Exception:
                    pass
        except Exception as _e:
            print(f"[DASHBOARD] Drive folder check/share échoué: {_e}")

    projets_actifs = conn.execute("""
        SELECT p.*, s.icon as service_icon, s.nom_service as service_nom
        FROM projets p
        LEFT JOIN services s ON s.id = p.id_service
        WHERE p.id_client = ? AND (p.is_archived = 0 OR p.is_archived IS NULL)
        ORDER BY p.created_at DESC
    """, (user_id,)).fetchall()
    projets_archives = conn.execute("""
        SELECT p.*, s.icon as service_icon, s.nom_service as service_nom
        FROM projets p
        LEFT JOIN services s ON s.id = p.id_service
        WHERE p.id_client = ? AND p.is_archived = 1
        ORDER BY p.created_at DESC
    """, (user_id,)).fetchall()
    conn.close()

    def format_projet(p):
        ready, done, total = compute_checklist_readiness(p['id'])
        return {
            "id": p['id'],
            "nom_projet": p['nom_projet'],
            "statut": p['statut'],
            "lien_gdrive": p['lien_gdrive'],
            "service_icon": p['service_icon'],
            "service_nom": p['service_nom'],
            "date_livraison_estimee": p['date_livraison_estimee'],
            "is_archived": p['is_archived'],
            "created_at": p['created_at'],
            "checklist": {"ready": ready, "done": done, "total": total}
        }

    return jsonify({
        "projets_actifs": [format_projet(p) for p in projets_actifs],
        "projets_archives": [format_projet(p) for p in projets_archives],
        "user": {
            "id": session['user_id'],
            "nom": session['user_name'],
            "is_admin": session.get('is_admin', False),
            "drive_folder_id": drive_folder_id,
        }
    })


# ───────────────────────────────────────────────────────────
# API v1 — Admin Clients
# ───────────────────────────────────────────────────────────

def _provision_client_drive_folder(conn, client_id, dossier_nom):
    """Crée le dossier Drive racine + sous-dossier Factures pour un client qui n'en a pas encore."""
    try:
        drive_folder_id = create_folder(dossier_nom, parent_id=os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID'))
        factures_folder_id = create_folder("Factures", parent_id=drive_folder_id)
        conn.execute("UPDATE clients SET drive_folder_id=?, factures_folder_id=? WHERE id=?",
                     (drive_folder_id, factures_folder_id, client_id))
        conn.commit()
    except Exception as e:
        print(f"[DRIVE] Création dossier client échouée: {e}")


@app.route('/api/v1/admin/client/<int:client_id>', methods=['GET'])
@admin_required
def api_admin_get_client(client_id):
    conn = get_db_connection()
    client = conn.execute(
        'SELECT * FROM clients WHERE id = ?', (client_id,)
    ).fetchone()
    conn.close()
    if not client:
        return jsonify({'error': 'Client introuvable'}), 404
    return jsonify({
        'id': client['id'],
        'nom_complet': client['nom_complet'],
        'email': client['email'],
        'nom_entreprise': client['nom_entreprise'],
        'telephone': client['telephone'],
        'mode_facturation': client['mode_facturation'] or 'projet',
        'adresse_facturation': client['adresse_facturation'],
        'ville_facturation': client['ville_facturation'],
        'province_facturation': client['province_facturation'] or 'Québec',
        'code_postal_facturation': client['code_postal_facturation'],
        'pays_facturation': client['pays_facturation'] or 'Canada',
        'is_admin': client['is_admin'],
        'is_email_confirmed': bool(client['is_email_confirmed']),
        'created_at': client['created_at'],
        'statut_relation': client['statut_relation'] or 'actif',
        'prochain_suivi': client['prochain_suivi'],
        'source_acquisition': client['source_acquisition'],
        'site_web': client['site_web'],
        'adresse': client['adresse'],
        'ville': client['ville'],
        'contact_secondaire_nom': client['contact_secondaire_nom'],
        'contact_secondaire_role': client['contact_secondaire_role'],
        'contact_secondaire_email': client['contact_secondaire_email'],
        'lien_gdrive': get_folder_link(client['drive_folder_id']) if client['drive_folder_id'] else None,
    })

@app.route('/api/v1/admin/client/<int:client_id>/factures', methods=['GET'])
@admin_required
def api_admin_client_factures(client_id):
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT id, numero, statut, total, date_emission, date_echeance, stripe_payment_url "
        "FROM factures WHERE id_client = ? AND statut != 'annulee' ORDER BY date_emission DESC",
        (client_id,)
    ).fetchall()
    conn.close()
    from datetime import date
    today = date.today().isoformat()
    factures, total_paye, total_attente = [], 0.0, 0.0
    for f in rows:
        montant = f['total'] or 0
        statut = f['statut']
        en_retard = statut != 'payee' and bool(f['date_echeance']) and f['date_echeance'] < today
        if statut == 'payee':
            total_paye += montant
        else:
            total_attente += montant
        factures.append({
            'id': f['id'],
            'numero': f['numero'],
            'statut': statut,
            'en_retard': bool(en_retard),
            'total': montant,
            'date_emission': f['date_emission'],
            'date_echeance': f['date_echeance'],
            'stripe_payment_url': f['stripe_payment_url'],
        })
    return jsonify({
        'factures': factures,
        'total_paye': round(total_paye, 2),
        'total_attente': round(total_attente, 2),
        'nb_factures': len(factures),
    })

@app.route('/api/v1/admin/client/<int:client_id>/activite', methods=['GET'])
@admin_required
def api_admin_client_activite(client_id):
    """Timeline d'activité : événements dérivés des tables existantes
    (création, notes, factures, RDV) + événements loggés (changements de statut)."""
    conn = get_db_connection()
    client = conn.execute("SELECT created_at FROM clients WHERE id = ?", (client_id,)).fetchone()
    if not client:
        conn.close()
        return jsonify({'error': 'Client introuvable'}), 404

    events = []

    def q(sql, params=()):
        try:
            return conn.execute(sql, params).fetchall()
        except Exception:
            return []

    # Création
    if client['created_at']:
        events.append({'type': 'creation', 'titre': 'Client créé', 'date': client['created_at']})

    # Notes
    for n in q("SELECT contenu, created_at FROM client_notes WHERE id_client = ?", (client_id,)):
        extrait = (n['contenu'] or '').strip().replace('\n', ' ')
        if len(extrait) > 80:
            extrait = extrait[:80] + '…'
        events.append({'type': 'note', 'titre': 'Note ajoutée', 'detail': extrait, 'date': n['created_at']})

    # Factures
    _lblf = {'payee': 'payée', 'envoyee': 'envoyée', 'ouverte': 'en attente', 'brouillon': 'brouillon'}
    for f in q("SELECT numero, statut, date_emission FROM factures WHERE id_client = ? AND statut != 'annulee'", (client_id,)):
        events.append({'type': 'facture',
                       'titre': f"Facture {f['numero']} — {_lblf.get(f['statut'], f['statut'])}",
                       'date': f['date_emission']})

    # Rendez-vous
    for r in q("SELECT label_fr, start_utc FROM rendez_vous WHERE id_client = ?", (client_id,)):
        events.append({'type': 'rdv', 'titre': r['label_fr'] or 'Rendez-vous', 'date': r['start_utc']})

    # Événements loggés (changements de statut, etc.)
    for a in q("SELECT type, titre, created_at FROM client_activite WHERE id_client = ?", (client_id,)):
        events.append({'type': a['type'], 'titre': a['titre'], 'date': a['created_at']})

    conn.close()
    # Tri par date décroissante (les dates ISO se comparent lexicographiquement)
    events.sort(key=lambda e: (e.get('date') or ''), reverse=True)
    return jsonify(events)

@app.route('/api/v1/admin/client/<int:client_id>', methods=['PUT'])
@admin_required
def api_admin_update_client(client_id):
    data = request.get_json() or {}
    statut_relation = data.get('statut_relation') or 'actif'
    conn = get_db_connection()

    if statut_relation == 'actif':
        existing = conn.execute(
            "SELECT drive_folder_id, nom_entreprise, nom_complet FROM clients WHERE id = ?", (client_id,)
        ).fetchone()
        if existing and not existing['drive_folder_id']:
            dossier_nom = existing['nom_entreprise'] or data.get('nom_entreprise') or existing['nom_complet']
            _provision_client_drive_folder(conn, client_id, dossier_nom)

    conn.execute("""
        UPDATE clients SET
            nom_complet = ?,
            email = ?,
            nom_entreprise = ?,
            telephone = ?,
            mode_facturation = ?,
            adresse_facturation = ?,
            ville_facturation = ?,
            province_facturation = ?,
            code_postal_facturation = ?,
            pays_facturation = ?,
            statut_relation = ?,
            prochain_suivi = ?,
            source_acquisition = ?,
            site_web = ?,
            adresse = ?,
            ville = ?,
            contact_secondaire_nom = ?,
            contact_secondaire_role = ?,
            contact_secondaire_email = ?
        WHERE id = ?
    """, (
        data.get('nom_complet'),
        (data.get('email') or '').strip().lower() or None,
        (data.get('nom_entreprise') or '').strip() or None,
        (data.get('telephone') or '').strip() or None,
        data.get('mode_facturation') or 'projet',
        (data.get('adresse_facturation') or '').strip() or None,
        (data.get('ville_facturation') or '').strip() or None,
        (data.get('province_facturation') or 'Québec').strip(),
        (data.get('code_postal_facturation') or '').strip() or None,
        (data.get('pays_facturation') or 'Canada').strip(),
        statut_relation,
        (data.get('prochain_suivi') or '').strip() or None,
        (data.get('source_acquisition') or '').strip() or None,
        (data.get('site_web') or '').strip() or None,
        (data.get('adresse') or '').strip() or None,
        (data.get('ville') or '').strip() or None,
        (data.get('contact_secondaire_nom') or '').strip() or None,
        (data.get('contact_secondaire_role') or '').strip() or None,
        (data.get('contact_secondaire_email') or '').strip().lower() or None,
        client_id
    ))
    conn.commit()
    # Même raison que sur /adresse-facturation : sans ça, les PDF déjà générés gardent
    # l'ancien nom/courriel/adresse jusqu'à la prochaine modif de ligne sur chaque facture.
    factures_existantes = conn.execute(
        "SELECT id FROM factures WHERE id_client = ? AND pdf_path IS NOT NULL", (client_id,)
    ).fetchall()
    for f in factures_existantes:
        try:
            regenerer_pdf_facture(f['id'], conn)
        except Exception as e:
            print(f"[INVOICE] Régénération PDF échouée pour facture {f['id']}: {e}")
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/client/<int:client_id>', methods=['DELETE'])
@admin_required
def api_admin_delete_client(client_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM clients WHERE id = ?', (client_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ───────────────────────────────────────────────────────────
# API v1 — Admin Projets
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/projet/<int:projet_id>', methods=['GET'])
@admin_required
def api_admin_get_projet(projet_id):
    conn = get_db_connection()
    p = conn.execute("""
        SELECT p.*, c.nom_complet as client_nom, c.email as client_email,
               c.telephone as client_telephone, s.nom_service, s.categorie as service_categorie,
               s.documents_requis as svc_documents_requis,
               s.appel_exploratoire_requis as svc_appel_exploratoire_requis
        FROM projets p
        LEFT JOIN clients c ON c.id = p.id_client
        LEFT JOIN services s ON s.id = p.id_service
        WHERE p.id = ?
    """, (projet_id,)).fetchone()
    if not p:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    logo_fichiers = conn.execute(
        "SELECT id, filename FROM projet_logo_fichiers WHERE id_projet=? ORDER BY created_at", (projet_id,)
    ).fetchall()
    conn.close()
    pipeline_steps = pipeline_for_service({
        'documents_requis': p['svc_documents_requis'] if p['svc_documents_requis'] is not None else 1,
        'appel_exploratoire_requis': p['svc_appel_exploratoire_requis'] or 0,
    })
    pipeline_steps = pipeline_with_current_statut(pipeline_steps, p['statut'])
    is_site_web = (p['service_categorie'] == 'Sites Web') or bool(
        re.search(r'(site|web|shopify|vercel|transactionnel|vitrine)', (p['nom_service'] or ''), re.IGNORECASE)
    )
    return jsonify({
        'id': p['id'],
        'nom_projet': p['nom_projet'],
        'titre_affiche': p['titre_affiche'],
        'statut': p['statut'],
        'pipeline_steps': pipeline_steps,
        'progress_pct': phase_progress(p['statut']),
        'is_site_web': is_site_web,
        'lien_gdrive': p['lien_gdrive'],
        'drive_folder_id': p['drive_folder_id'],
        'date_livraison_estimee': p['date_livraison_estimee'],
        'localisation': p['localisation'],
        'is_archived': p['is_archived'],
        'created_at': p['created_at'],
        'client_id': p['id_client'],
        'client_nom': p['client_nom'],
        'client_email': p['client_email'],
        'client_telephone': p['client_telephone'],
        'nom_service': p['nom_service'],
        'facturation_mode': p['facturation_mode'],
        'lien_site_test': p['lien_site_test'],
        'logo_fichiers': [{'id': f['id'], 'filename': f['filename']} for f in logo_fichiers],
    })

@app.route('/api/v1/admin/projet/<int:projet_id>', methods=['PUT'])
@admin_required
def api_admin_update_projet(projet_id):
    data = request.get_json() or {}
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (projet_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404

    # facturation_mode : null/'' => facturer le client ; sinon raison de non-facturation
    ancien_mode = projet['facturation_mode']          # None = on facture
    nouveau_mode = (data.get('facturation_mode') or '').strip() or None

    conn.execute("""
        UPDATE projets SET
            nom_projet = ?,
            titre_affiche = ?,
            statut = ?,
            localisation = ?,
            date_livraison_estimee = ?,
            lien_gdrive = ?,
            lien_site_test = ?,
            id_client = ?,
            facturation_mode = ?
        WHERE id = ?
    """, (
        data.get('nom_projet'),
        (data.get('titre_affiche') or '').strip() or None,
        data.get('statut'),
        (data.get('localisation') or '').strip() or None,
        data.get('date_livraison_estimee') or None,
        (data.get('lien_gdrive') or '').strip() or None,
        (data.get('lien_site_test') or '').strip() or None,
        data.get('id_client'),
        nouveau_mode,
        projet_id
    ))
    conn.commit()

    # ── Réconcilier la facture selon le changement de mode de facturation ──
    # On facture quand facturation_mode est NULL ; on ne facture pas sinon.
    facture_action = None
    ancien_facturer = ancien_mode is None
    nouveau_facturer = nouveau_mode is None
    if ancien_facturer != nouveau_facturer:
        try:
            fact = conn.execute("""
                SELECT f.id, f.numero, f.statut
                FROM factures f
                JOIN facture_lignes fl ON fl.id_facture = f.id
                WHERE fl.id_projet = ?
                ORDER BY f.id DESC LIMIT 1
            """, (projet_id,)).fetchone()

            if ancien_facturer and not nouveau_facturer:
                # facturer → ne pas facturer : annuler la facture existante (sauf payée / déjà annulée)
                if fact and fact['statut'] not in ('annulee', 'payee'):
                    conn.execute("UPDATE factures SET statut='annulee' WHERE id=?", (fact['id'],))
                    conn.commit()
                    facture_action = {'type': 'annulee', 'numero': fact['numero']}
            else:
                # ne pas facturer → facturer : réactiver une facture annulée, sinon en créer une (silencieux)
                if fact and fact['statut'] == 'annulee':
                    from datetime import date as _d, timedelta as _td
                    today = _d.today()
                    conn.execute(
                        "UPDATE factures SET statut='envoyee', date_emission=?, date_echeance=? WHERE id=?",
                        (today.strftime('%Y-%m-%d'), (today + _td(days=15)).strftime('%Y-%m-%d'), fact['id'])
                    )
                    conn.commit()
                    facture_action = {'type': 'reactivee', 'numero': fact['numero']}
                elif not fact:
                    nouvelle = creer_facture_projet(projet_id, conn)  # création silencieuse : pas de courriel ni Drive
                    if nouvelle:
                        facture_action = {'type': 'creee', 'numero': nouvelle['numero']}
        except Exception as e:
            print(f"[INVOICE] update_projet facturation: {e}")

    conn.close()
    return jsonify({'success': True, 'facture_action': facture_action})

@app.route('/api/v1/admin/projet/<int:project_id>/logo', methods=['POST'])
@admin_required
def api_admin_upload_projet_logo(project_id):
    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'Aucun fichier reçu'}), 400
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404

    logo_folder_id = None
    try:
        parent = projet['drive_folder_id'] or os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
        if parent:
            logo_folder_id = create_folder("Logo", parent_id=parent)
    except Exception as e:
        print(f"[DRIVE] Création dossier Logo échouée: {e}")

    uploaded = []
    for file in files:
        if not file or not file.filename:
            continue
        ext = os.path.splitext(file.filename)[1].lower().lstrip('.')
        if ext not in {'svg', 'ai', 'eps', 'pdf', 'png'}:
            continue
        original_name = secure_filename(file.filename)
        safe_name = f"{uuid.uuid4().hex[:8]}_{original_name}"
        base_dir = os.path.join(app.config["UPLOAD_ROOT"], f"projet_{project_id}", "logo")
        pathlib.Path(base_dir).mkdir(parents=True, exist_ok=True)
        save_path = os.path.join(base_dir, safe_name)
        file.save(save_path)
        drive_file_id = None
        try:
            if logo_folder_id:
                drive_file_id, _ = upload_file(save_path, original_name, logo_folder_id)
        except Exception as e:
            print(f"[DRIVE] Upload logo échoué: {e}")
        cur = conn.execute(
            "INSERT INTO projet_logo_fichiers (id_projet, filename, filepath, drive_file_id) VALUES (?, ?, ?, ?)",
            (project_id, original_name, save_path, drive_file_id)
        )
        uploaded.append({'id': cur.lastrowid, 'filename': original_name})
    conn.commit()
    conn.close()
    if not uploaded:
        return jsonify({'error': 'Aucun fichier valide (svg, ai, eps, pdf, png)'}), 400
    return jsonify({'success': True, 'fichiers': uploaded})

@app.route('/api/v1/admin/projet/<int:project_id>/logo/<int:file_id>', methods=['DELETE'])
@admin_required
def api_admin_delete_projet_logo(project_id, file_id):
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM projet_logo_fichiers WHERE id=? AND id_projet=?", (file_id, project_id)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Fichier introuvable'}), 404
    conn.execute("DELETE FROM projet_logo_fichiers WHERE id=?", (file_id,))
    conn.commit()
    conn.close()
    try:
        if row['filepath'] and os.path.exists(row['filepath']):
            os.remove(row['filepath'])
    except Exception:
        pass
    return jsonify({'success': True})

@app.route('/api/v1/projet/<int:project_id>/logo/<int:file_id>', methods=['GET'])
@login_required
def api_projet_download_logo(project_id, file_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT id_client FROM projets WHERE id=?", (project_id,)).fetchone()
    fichier = conn.execute("SELECT * FROM projet_logo_fichiers WHERE id=? AND id_projet=?", (file_id, project_id)).fetchone()
    conn.close()
    if not projet or not fichier:
        return jsonify({'error': 'Logo introuvable'}), 404
    is_owner = (projet['id_client'] == session.get('user_id'))
    is_admin = bool(session.get('is_admin', False))
    if not (is_owner or is_admin):
        return jsonify({'error': 'Non autorisé'}), 403
    return send_file(fichier['filepath'], as_attachment=True, download_name=fichier['filename'])

@app.route('/api/v1/admin/projet/<int:projet_id>', methods=['DELETE'])
@admin_required
def api_admin_delete_projet(projet_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM projets WHERE id = ?', (projet_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ───────────────────────────────────────────────────────────
# API v1 — Admin Clients liste + création
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/client/add', methods=['POST'])
@admin_required
def api_admin_create_client():
    data = request.get_json(force=True) or {}
    nom = (data.get('nom_complet') or '').strip()
    email = (data.get('email') or '').strip().lower()
    entreprise = (data.get('nom_entreprise') or '').strip() or None
    telephone = (data.get('telephone') or '').strip() or None
    statut_relation = data.get('statut_relation') or 'actif'
    is_prospect = statut_relation == 'prospect'
    adresse_facturation    = (data.get('adresse_facturation') or '').strip() or None
    ville_facturation      = (data.get('ville_facturation') or '').strip() or None
    province_facturation   = (data.get('province_facturation') or 'Québec').strip()
    code_postal_facturation = (data.get('code_postal_facturation') or '').strip() or None

    if not nom or not email:
        return jsonify({'error': 'Nom et email obligatoires.'}), 400

    conn = get_db_connection()
    try:
        token = s.dumps(email, salt='invitation-client-salt')
        conn.execute("""
            INSERT INTO clients (nom_complet, email, nom_entreprise, telephone, mot_de_passe_hash, auth_provider, is_email_confirmed, is_admin, confirm_token, statut_relation,
                adresse_facturation, ville_facturation, province_facturation, code_postal_facturation)
            VALUES (?, ?, ?, ?, NULL, 'password', 0, 0, ?, ?, ?, ?, ?, ?)
        """, (nom, email, entreprise, telephone, token, statut_relation,
              adresse_facturation, ville_facturation, province_facturation, code_postal_facturation))
        client_id = conn.execute("SELECT id FROM clients WHERE email=?", (email,)).fetchone()['id']
        conn.commit()

        # Un prospect n'a pas encore besoin d'un dossier Drive ni d'un accès portail —
        # provisionnés plus tard, à la promotion en client actif.
        if not is_prospect:
            _provision_client_drive_folder(conn, client_id, entreprise if entreprise else nom)

            try:
                long_invite_url = f"{PORTAIL_URL}/invitation/{token}"
                invite_url = create_short_link(long_invite_url, expires_in=604800)
                html_invite = _invitation_client(nom, invite_url)
                send_email(email,
                    "Bienvenue chez Cocktail Média — Créez votre accès",
                    f"Bonjour {nom}, créez votre accès au portail : {invite_url}",
                    html=html_invite)
            except Exception as e:
                print(f"[MAIL] Email invitation échoué: {e}")

        return jsonify({'id': client_id, 'nom_complet': nom, 'email': email})

    except Exception as e:
        conn.rollback()
        if 'UNIQUE' in str(e):
            return jsonify({'error': f"L'email '{email}' existe déjà."}), 409
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/v1/admin/client/<int:client_id>/adresse-facturation', methods=['PUT'])
@admin_required
def api_admin_update_client_adresse_facturation(client_id):
    # Route dédiée — ne touche QUE nom + courriel + adresse de facturation, contrairement au
    # PUT complet /admin/client/<id> qui exige tous les champs (utilisé depuis la fiche
    # facture, où on n'a pas — et ne veut pas devoir renvoyer — tout le profil du client).
    # Modifiable même sur une facture fermée/payée : correction possible en tout temps.
    data = request.get_json(force=True) or {}
    nom_complet = (data.get('nom_complet') or '').strip()
    email = (data.get('email') or '').strip().lower()
    if not nom_complet:
        return jsonify({'error': 'Le nom est obligatoire.'}), 400
    if not email:
        return jsonify({'error': 'Le courriel est obligatoire.'}), 400
    conn = get_db_connection()
    try:
        client = conn.execute("SELECT id FROM clients WHERE id=?", (client_id,)).fetchone()
        if not client:
            return jsonify({'error': 'Client introuvable.'}), 404
        conn.execute("""
            UPDATE clients SET
                nom_complet = ?,
                email = ?,
                adresse_facturation = ?,
                ville_facturation = ?,
                province_facturation = ?,
                code_postal_facturation = ?
            WHERE id = ?
        """, (
            nom_complet,
            email,
            (data.get('adresse_facturation') or '').strip() or None,
            (data.get('ville_facturation') or '').strip() or None,
            (data.get('province_facturation') or 'Québec').strip(),
            (data.get('code_postal_facturation') or '').strip() or None,
            client_id,
        ))
        conn.commit()
        # Régénère le PDF de toutes les factures déjà émises pour ce client — sinon le PDF
        # déjà sur disque garde l'ancien nom/courriel/adresse jusqu'à la prochaine modif de ligne.
        factures_existantes = conn.execute(
            "SELECT id FROM factures WHERE id_client = ? AND pdf_path IS NOT NULL", (client_id,)
        ).fetchall()
        for f in factures_existantes:
            try:
                regenerer_pdf_facture(f['id'], conn)
            except Exception as e:
                print(f"[INVOICE] Régénération PDF échouée pour facture {f['id']}: {e}")
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        if 'UNIQUE' in str(e):
            return jsonify({'error': f"Le courriel '{email}' est déjà utilisé par un autre client."}), 409
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


STAGES_PIPELINE_CLIENT = ('prospect', 'contacte', 'devis_envoye', 'actif', 'inactif')


@app.route('/api/v1/admin/client/<int:client_id>/statut-relation', methods=['PUT'])
@admin_required
def api_admin_update_client_statut_relation(client_id):
    data = request.get_json(force=True) or {}
    statut_relation = data.get('statut_relation')
    if statut_relation not in STAGES_PIPELINE_CLIENT:
        return jsonify({'error': 'Étape de pipeline invalide.'}), 400

    conn = get_db_connection()
    client = conn.execute(
        "SELECT drive_folder_id, nom_entreprise, nom_complet, statut_relation FROM clients WHERE id = ?", (client_id,)
    ).fetchone()
    if not client:
        conn.close()
        return jsonify({'error': 'Client introuvable'}), 404

    if statut_relation == 'actif' and not client['drive_folder_id']:
        _provision_client_drive_folder(conn, client_id, client['nom_entreprise'] or client['nom_complet'])

    ancien = client['statut_relation'] or 'actif'
    conn.execute("UPDATE clients SET statut_relation = ? WHERE id = ?", (statut_relation, client_id))
    if ancien != statut_relation:
        _lbl = {'prospect': 'Prospect', 'contacte': 'Contacté', 'devis_envoye': 'Devis envoyé', 'actif': 'Actif', 'inactif': 'Inactif'}
        log_activite(conn, client_id, 'statut',
                     f"Statut : {_lbl.get(ancien, ancien)} → {_lbl.get(statut_relation, statut_relation)}")
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'statut_relation': statut_relation})


@app.route('/api/v1/admin/client/<int:client_id>/mode-facturation', methods=['PUT'])
@admin_required
def api_admin_update_client_mode_facturation(client_id):
    data = request.get_json(force=True) or {}
    mode_facturation = data.get('mode_facturation')
    if mode_facturation not in ('projet', 'mensuel'):
        return jsonify({'error': 'mode_facturation invalide'}), 400

    conn = get_db_connection()
    client = conn.execute("SELECT mode_facturation FROM clients WHERE id = ?", (client_id,)).fetchone()
    if not client:
        conn.close()
        return jsonify({'error': 'Client introuvable'}), 404

    ancien = client['mode_facturation'] or 'projet'
    conn.execute("UPDATE clients SET mode_facturation = ? WHERE id = ?", (mode_facturation, client_id))
    if ancien != mode_facturation:
        _lbl = {'projet': 'Par projet', 'mensuel': 'Facture ouverte (mensuel)'}
        log_activite(conn, client_id, 'statut',
                     f"Facturation : {_lbl.get(ancien, ancien)} → {_lbl.get(mode_facturation, mode_facturation)}")
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'mode_facturation': mode_facturation})


@app.route('/api/v1/admin/client/<int:client_id>/resend-invitation', methods=['POST'])
@admin_required
def api_admin_resend_invitation(client_id):
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    conn.close()
    if not user:
        return jsonify({'error': 'Client introuvable'}), 404
    has_password = bool(user['mot_de_passe_hash'])
    if int(user['is_email_confirmed'] or 0) and has_password:
        return jsonify({'error': 'Ce client a déjà activé son compte'}), 409
    conn2 = get_db_connection()
    token = s.dumps(user['email'], salt='invitation-client-salt')
    conn2.execute("UPDATE clients SET is_email_confirmed = 0, confirm_token = ? WHERE id = ?", (token, client_id))
    conn2.close()
    try:
        long_invite_url = f"{PORTAIL_URL}/invitation/{token}"
        invite_url = create_short_link(long_invite_url, expires_in=604800)
        html_invite = _invitation_client(user['nom_complet'], invite_url)
        send_email(user['email'],
            "Bienvenue chez Cocktail Média — Créez votre accès",
            f"Bonjour {user['nom_complet']}, créez votre accès au portail : {invite_url}",
            html=html_invite)
    except Exception as e:
        print(f"[MAIL] Resend invitation échoué: {e}")
        return jsonify({'error': "Erreur lors de l'envoi de l'invitation"}), 500
    return jsonify({'success': True})


BUFFER_SEANCE = timedelta(hours=1)    # séances en personne / production
BUFFER_RDV    = timedelta(minutes=15) # rendez-vous en ligne


def _get_busy_combined(debut_fenetre: datetime, fin_fenetre: datetime) -> list:
    """Retourne les plages occupées avec buffer différencié selon le type.
    - Rendez-vous en ligne (table rendez_vous) → buffer 15 min
    - Séances / autres événements Google Calendar   → buffer 1h
    Chaque élément : {'start', 'end', 'buffer': timedelta}
    """
    from calendar_service import get_busy_slots
    try:
        gcal_busy = get_busy_slots(debut_fenetre, fin_fenetre)
    except Exception as e:
        print(f"[CALENDAR] freebusy échoué, fallback DB seul: {e}")
        gcal_busy = []

    conn = get_db_connection()
    db_rows = conn.execute(
        "SELECT start_utc, end_utc FROM rendez_vous WHERE start_utc < ? AND end_utc > ?",
        (fin_fenetre.isoformat(), debut_fenetre.isoformat())
    ).fetchall()
    conn.close()

    # Index des rendez-vous en ligne pour identification rapide
    rdv_index = {(r[0], r[1]) for r in db_rows}

    result = []
    gcal_index = set()
    for b in gcal_busy:
        s_norm = b['start'].replace('Z', '').replace('+00:00', '')
        e_norm = b['end'].replace('Z', '').replace('+00:00', '')
        gcal_index.add((s_norm, e_norm))
        is_online = (s_norm, e_norm) in rdv_index
        result.append({
            'start':  b['start'],
            'end':    b['end'],
            'buffer': BUFFER_RDV if is_online else BUFFER_SEANCE,
        })

    # Rendez-vous DB absents du calendrier (fallback)
    for r in db_rows:
        if (r[0], r[1]) not in gcal_index:
            result.append({'start': r[0], 'end': r[1], 'buffer': BUFFER_RDV})

    return result


def _slot_libre(slot_debut: datetime, slot_fin: datetime, busy: list) -> bool:
    """Retourne True si le créneau ne chevauche aucune plage occupée.
    Utilise le buffer propre à chaque événement (15 min online, 1h séance)."""
    for b in busy:
        buf     = b.get('buffer', BUFFER_SEANCE)
        b_start = datetime.fromisoformat(b['start'].replace('Z', '').replace('+00:00', '')) - buf
        b_end   = datetime.fromisoformat(b['end'].replace('Z', '').replace('+00:00', ''))   + buf
        if slot_debut < b_end and slot_fin > b_start:
            return False
    return True


def _toronto_offset_now() -> timedelta:
    """Retourne l'offset UTC→Toronto en tenant compte de l'heure avancée (EDT/EST)."""
    from zoneinfo import ZoneInfo
    from datetime import timezone as _tz
    tz = ZoneInfo("America/Toronto")
    now_utc = datetime.now(_tz.utc)
    offset = now_utc.astimezone(tz).utcoffset()
    return timedelta(seconds=-offset.total_seconds())


@app.route('/api/v1/admin/client/<int:client_id>/envoyer-agenda', methods=['POST'])
@admin_required
def api_admin_envoyer_agenda(client_id):
    from email_templates import email_agenda_rendez_vous
    conn = get_db_connection()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    conn.close()
    if not client:
        return jsonify({'error': 'Client introuvable'}), 404
    try:
        from calendar_service import HEURES_REUNION
        tz_offset = _toronto_offset_now()
        duree = timedelta(hours=1)
        aujourd_hui = datetime.utcnow().date()
        raw_slots = []
        for delta in range(1, 30):
            if len(raw_slots) >= 4:
                break
            jour = aujourd_hui + timedelta(days=delta)
            if jour.weekday() >= 5:
                continue
            debut_fenetre = datetime(jour.year, jour.month, jour.day, 8, 0) + tz_offset
            fin_fenetre   = datetime(jour.year, jour.month, jour.day, 18, 0) + tz_offset
            busy = _get_busy_combined(debut_fenetre, fin_fenetre)
            for heure in HEURES_REUNION:
                h, m = int(heure), 30 if heure % 1 else 0
                slot_debut = datetime(jour.year, jour.month, jour.day, h, m) + tz_offset
                slot_fin   = slot_debut + duree
                if _slot_libre(slot_debut, slot_fin, busy):
                    raw_slots.append((slot_debut, slot_fin))
                    break
        if not raw_slots:
            return jsonify({'error': 'Aucun créneau disponible dans les prochains jours'}), 503
        from urllib.parse import quote
        slots = []
        for start_utc, end_utc in raw_slots:
            token = s.dumps(
                {'c': client_id, 's': start_utc.isoformat(), 'e': end_utc.isoformat()},
                salt='booking-confirm-salt'
            )
            slots.append({
                'label': format_slot_fr(start_utc),
                'url': f"{PORTAIL_URL}/api/v1/booking/confirm?token={quote(token, safe='')}",
            })
        # Texte fallback plaintext
        lignes = [f"Bonjour {client['nom_complet']},\n",
                  "Voici les prochains créneaux disponibles :\n"]
        for sl in slots:
            lignes.append(f"  • {sl['label']} → {sl['url']}")
        lignes.append("\n— L'équipe Cocktail Média")
        send_email(
            client['email'],
            "Planifiez votre rendez-vous — Cocktail Média",
            "\n".join(lignes),
            html=email_agenda_rendez_vous(client['nom_complet'], slots, booking_url=os.getenv('BOOKING_URL', ''))
        )
    except Exception as e:
        print(f"[MAIL] envoyer_agenda: {e}")
        return jsonify({'error': "Erreur lors de l'envoi de l'email"}), 500
    return jsonify({'success': True})


@app.route('/api/v1/admin/client/<int:client_id>/envoyer-agenda-jour', methods=['POST'])
@admin_required
def api_admin_envoyer_agenda_jour(client_id):
    """Envoie au client tous les créneaux disponibles d'une journée précise."""
    from email_templates import email_agenda_rendez_vous
    from calendar_service import format_slot_fr, HEURES_REUNION
    from datetime import date as date_cls, timedelta

    data = request.get_json(force=True) or {}
    date_str = (data.get('date') or '').strip()
    if not date_str:
        return jsonify({'error': 'Date requise (YYYY-MM-DD)'}), 400

    try:
        target = date_cls.fromisoformat(date_str)
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400

    conn = get_db_connection()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    conn.close()
    if not client:
        return jsonify({'error': 'Client introuvable'}), 404

    try:
        tz_offset = _toronto_offset_now()
        duree = timedelta(hours=1)
        debut_fenetre = datetime(target.year, target.month, target.day, 8, 0) + tz_offset
        fin_fenetre   = datetime(target.year, target.month, target.day, 18, 0) + tz_offset
        busy = _get_busy_combined(debut_fenetre, fin_fenetre)

        slots_raw = []
        for heure in HEURES_REUNION:
            h, m = int(heure), 30 if heure % 1 else 0
            slot_debut = datetime(target.year, target.month, target.day, h, m) + tz_offset
            slot_fin   = slot_debut + duree
            if _slot_libre(slot_debut, slot_fin, busy):
                slots_raw.append((slot_debut, slot_fin))

        if not slots_raw:
            return jsonify({'error': f'Aucun créneau disponible le {date_str}'}), 503

        from urllib.parse import quote
        slots = []
        for start_utc, end_utc in slots_raw:
            token = s.dumps(
                {'c': client_id, 's': start_utc.isoformat(), 'e': end_utc.isoformat()},
                salt='booking-confirm-salt'
            )
            slots.append({
                'label': format_slot_fr(start_utc),
                'url': f"{PORTAIL_URL}/api/v1/booking/confirm?token={quote(token, safe='')}",
            })

        lignes = [f"Bonjour {client['nom_complet']},\n",
                  f"Voici les créneaux disponibles pour le {date_str} :\n"]
        for sl in slots:
            lignes.append(f"  • {sl['label']} → {sl['url']}")
        lignes.append("\n— L'équipe Cocktail Média")

        send_email(
            client['email'],
            f"Disponibilités du {date_str} — Cocktail Média",
            "\n".join(lignes),
            html=email_agenda_rendez_vous(client['nom_complet'], slots, booking_url=os.getenv('BOOKING_URL', ''))
        )
    except Exception as e:
        print(f"[MAIL] envoyer_agenda_jour: {e}")
        return jsonify({'error': "Erreur lors de l'envoi de l'email"}), 500

    return jsonify({'success': True, 'nb_slots': len(slots)})


@app.route('/api/v1/admin/client/<int:client_id>/envoyer-agenda-plage', methods=['POST'])
@admin_required
def api_admin_envoyer_agenda_plage(client_id):
    """Envoie tous les créneaux disponibles sur une plage de dates (du_date → au_date inclus)."""
    from email_templates import email_agenda_rendez_vous
    from calendar_service import format_slot_fr, HEURES_REUNION
    from datetime import date as date_cls, timedelta

    data = request.get_json(force=True) or {}
    du_str  = (data.get('du')  or '').strip()
    au_str  = (data.get('au')  or '').strip()
    if not du_str or not au_str:
        return jsonify({'error': 'Dates requises (du et au en YYYY-MM-DD)'}), 400

    try:
        du = date_cls.fromisoformat(du_str)
        au = date_cls.fromisoformat(au_str)
    except ValueError:
        return jsonify({'error': 'Format de date invalide'}), 400

    if au < du:
        return jsonify({'error': 'La date de fin doit être après la date de début'}), 400
    if (au - du).days > 13:
        return jsonify({'error': 'Maximum 14 jours à la fois'}), 400

    conn = get_db_connection()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    conn.close()
    if not client:
        return jsonify({'error': 'Client introuvable'}), 404

    try:
        tz_offset = _toronto_offset_now()
        duree = timedelta(hours=1)
        from urllib.parse import quote

        slots = []
        current = du
        while current <= au:
            if current.weekday() < 5:  # lun-ven seulement
                debut_f = datetime(current.year, current.month, current.day, 8, 0) + tz_offset
                fin_f   = datetime(current.year, current.month, current.day, 18, 0) + tz_offset
                busy = _get_busy_combined(debut_f, fin_f)
                for heure in HEURES_REUNION:
                    h, m = int(heure), 30 if heure % 1 else 0
                    sd = datetime(current.year, current.month, current.day, h, m) + tz_offset
                    sf = sd + duree
                    if _slot_libre(sd, sf, busy):
                        token = s.dumps(
                            {'c': client_id, 's': sd.isoformat(), 'e': sf.isoformat()},
                            salt='booking-confirm-salt'
                        )
                        slots.append({
                            'label': format_slot_fr(sd),
                            'url': f"{PORTAIL_URL}/api/v1/booking/confirm?token={quote(token, safe='')}",
                        })
            current += timedelta(days=1)

        if not slots:
            return jsonify({'error': f'Aucun créneau disponible du {du_str} au {au_str}'}), 503

        lignes = [f"Bonjour {client['nom_complet']},\n",
                  f"Voici les créneaux disponibles du {du_str} au {au_str} :\n"]
        for sl in slots:
            lignes.append(f"  • {sl['label']} → {sl['url']}")
        lignes.append("\n— L'équipe Cocktail Média")

        send_email(
            client['email'],
            f"Disponibilités du {du_str} au {au_str} — Cocktail Média",
            "\n".join(lignes),
            html=email_agenda_rendez_vous(client['nom_complet'], slots, booking_url=os.getenv('BOOKING_URL', ''))
        )
    except Exception as e:
        print(f"[MAIL] envoyer_agenda_plage: {e}")
        return jsonify({'error': "Erreur lors de l'envoi de l'email"}), 500

    return jsonify({'success': True, 'nb_slots': len(slots)})


@app.route('/api/v1/booking/confirm', methods=['GET'])
def api_booking_confirm():
    """Route publique — le client clique depuis le courriel pour confirmer un créneau."""
    from email_templates import _base, _p
    token = request.args.get('token', '')
    if not token:
        return jsonify({'error': 'Token manquant'}), 400
    try:
        data = s.loads(token, salt='booking-confirm-salt', max_age=604800)  # 7 jours
    except SignatureExpired:
        return jsonify({'error': 'Ce lien a expiré (7 jours). Demandez un nouvel email.'}), 410
    except Exception:
        return jsonify({'error': 'Lien invalide'}), 400

    client_id = data.get('c')
    start_utc = data.get('s')
    end_utc   = data.get('e')
    if not all([client_id, start_utc, end_utc]):
        return jsonify({'error': 'Données de réservation incomplètes'}), 400

    conn = get_db_connection()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    conn.close()
    if not client:
        return jsonify({'error': 'Client introuvable'}), 404

    from datetime import datetime as dt
    start_dt = dt.fromisoformat(start_utc)
    end_dt   = dt.fromisoformat(end_utc)

    # Idempotence — évite les doublons si le client clique le lien plusieurs fois
    conn = get_db_connection()
    existing = conn.execute(
        "SELECT id, meet_link, label_fr FROM rendez_vous WHERE id_client=? AND start_utc=?",
        (client_id, start_utc)
    ).fetchone()
    conn.close()
    if existing:
        from urllib.parse import quote
        label = existing['label_fr'] or format_slot_fr(start_dt)
        return redirect(
            f"{PORTAIL_URL}/booking/confirme"
            f"?nom={quote(client['nom_complet'])}"
            f"&creneau={quote(label)}"
            f"&meet={quote(existing['meet_link'] or '')}"
            f"&rdv={existing['id']}"
            f"&start={quote(start_utc)}"
            f"&end={quote(end_utc)}"
        )

    try:
        event_id, meet_link = create_meeting_event(start_dt, end_dt, client['nom_complet'], client['email'])
        label = format_slot_fr(start_dt)
        conn = get_db_connection()
        cur = conn.execute(
            "INSERT INTO rendez_vous (id_client, calendar_event_id, start_utc, end_utc, meet_link, label_fr) VALUES (?,?,?,?,?,?)",
            (client_id, event_id, start_utc, end_utc, meet_link, label)
        )
        rdv_id = cur.lastrowid
        conn.commit()
        conn.close()
        _lier_rendez_vous_au_projet(client_id, rdv_id)
        from email_templates import email_rendez_vous_confirme
        google_url, outlook_url, ics_url = _build_calendar_urls(start_dt, end_dt, meet_link, rdv_id)
        send_email(
            client['email'],
            "Rendez-vous confirmé — Cocktail Média",
            f"Bonjour {client['nom_complet']}, votre rendez-vous est confirmé : {label}. Meet : {meet_link}",
            html=email_rendez_vous_confirme(client['nom_complet'], label, meet_link, google_url, outlook_url, ics_url)
        )
        send_email(
            'felix.dumont@cocktailmedia.ca',
            f"Nouveau rendez-vous — {client['nom_complet']}",
            f"{client['nom_complet']} ({client['email']}) a confirmé : {label}.\nMeet : {meet_link}"
        )
        _creer_tache_depuis_booking(client, f"Rendez-vous confirmé — {client['nom_complet']} — {label}", start_dt.date().isoformat())
    except Exception as e:
        print(f"[BOOKING] Création événement échouée: {e}")
        return jsonify({'error': "Erreur lors de la création du rendez-vous"}), 500

    from urllib.parse import quote
    return redirect(
        f"{PORTAIL_URL}/booking/confirme"
        f"?nom={quote(client['nom_complet'])}"
        f"&creneau={quote(label)}"
        f"&meet={quote(meet_link or '')}"
        f"&rdv={rdv_id}"
        f"&start={quote(start_utc)}"
        f"&end={quote(end_utc)}"
    )


def _find_service_slots(duree_seance_minutes: int, n: int = 4, jours_max: int = 45) -> list:
    """Trouve les prochains blocs libres pour réserver une séance de service en présentiel.
    Le bloc réservé dans l'agenda inclut 1h avant et 1h après la séance (préparation/transport),
    mais seule la fenêtre centrale (la séance elle-même) est proposée au client.
    Retourne une liste de tuples (bloc_debut, bloc_fin, seance_debut, seance_fin) en UTC."""
    from calendar_service import HEURES_REUNION
    tz_offset = _toronto_offset_now()
    duree_seance = timedelta(minutes=duree_seance_minutes)
    duree_bloc = BUFFER_SEANCE + duree_seance + BUFFER_SEANCE
    aujourd_hui = datetime.utcnow().date()
    resultats = []
    for delta in range(1, jours_max):
        if len(resultats) >= n:
            break
        jour = aujourd_hui + timedelta(days=delta)
        if jour.weekday() >= 5:
            continue
        debut_fenetre = datetime(jour.year, jour.month, jour.day, 8, 0) + tz_offset
        fin_fenetre   = datetime(jour.year, jour.month, jour.day, 18, 0) + tz_offset
        busy = _get_busy_combined(debut_fenetre, fin_fenetre)
        for heure in HEURES_REUNION:
            h, m = int(heure), 30 if heure % 1 else 0
            bloc_debut = datetime(jour.year, jour.month, jour.day, h, m) + tz_offset
            bloc_fin = bloc_debut + duree_bloc
            if bloc_fin > fin_fenetre:
                continue
            if _slot_libre(bloc_debut, bloc_fin, busy):
                seance_debut = bloc_debut + BUFFER_SEANCE
                seance_fin = seance_debut + duree_seance
                resultats.append((bloc_debut, bloc_fin, seance_debut, seance_fin))
                break
    return resultats


def _creer_tache_depuis_booking(client_row, texte, date_echeance, id_projet=None, projet_nom=None, lien='/admin'):
    """Crée une tâche à la confirmation d'une réservation cliente, assignée par défaut
    au rôle gestion. Comble l'angle mort identifié par l'audit du module de tâches
    (2026-07-16) : jusqu'ici, un booking confirmé ne générait ni tâche ni notification
    in-app, seulement un courriel à felix.dumont@cocktailmedia.ca.
    Assignée (pas partagée/broadcast) — une tâche non assignée est invisible dans « Mes
    tâches » (PWA et desktop filtrent strictement sur l'assignation), donc restait sans
    propriétaire et sans être vue (2026-07-18)."""
    conn = get_db_connection()
    try:
        gestion_id = _admin_id_for_role(conn, 'gestion')
        cur = conn.execute(
            "INSERT INTO todos_perso (texte, priorite, date_echeance, is_titre, projet_id, projet_nom, client_id) "
            "VALUES (?, 'normale', ?, 0, ?, ?, ?)",
            (texte, date_echeance, id_projet, projet_nom, client_row['id'])
        )
        todo_id = cur.lastrowid
        destinataire = None
        if gestion_id:
            conn.execute("INSERT OR IGNORE INTO todo_assignees (todo_id, admin_id) VALUES (?, ?)", (todo_id, gestion_id))
            row = conn.execute("SELECT email FROM clients WHERE id=?", (gestion_id,)).fetchone()
            destinataire = row['email'] if row else None
        conn.commit()
        push_admin_notif(conn, "Nouvelle réservation", texte, type='todo', lien=lien, destinataire=destinataire)
    except Exception as e:
        print(f"[BOOKING] Création tâche de suivi échouée: {e}")
    finally:
        conn.close()


def _creer_projet_depuis_booking(id_client: int, id_service: int, date_seance: str, heure_seance: str, localisation: str):
    """Crée automatiquement un projet suite à la réservation d'une séance par un client
    (dossier Drive, checklist, événement de séance, courriel) — retourne (id_projet, nom_projet)."""
    conn = get_db_connection()
    try:
        service_row = conn.execute("SELECT * FROM services WHERE id=?", (id_service,)).fetchone()
        client_row = conn.execute("SELECT * FROM clients WHERE id=?", (id_client,)).fetchone()
        if not service_row or not client_row:
            return None, None

        nom_service = service_row['nom_service']
        documents_requis = bool(service_row['documents_requis'])
        statut = initial_statut_for_service(service_row)
        duree_seance_minutes = int(service_row['duree_seance_minutes'] or 60)
        nom_projet = f"{date_seance} — {nom_service} — {localisation}" if localisation else f"{date_seance} — {nom_service}"

        cur = conn.cursor()
        cur.execute("""
            INSERT INTO projets (nom_projet, heure_seance, duree_seance_minutes, statut, id_client, localisation, id_service)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (nom_projet, heure_seance, duree_seance_minutes, statut, id_client, localisation, id_service))
        id_projet = cur.lastrowid

        cur.execute("INSERT INTO checklistes (id_projet) VALUES (?)", (id_projet,))
        id_checklist = cur.lastrowid
        items_modele = conn.execute("""
            SELECT nom_item, requires_file, is_required, item_type, video_url, is_revision_item, file_category, field_type
            FROM checklist_model_items WHERE id_service = ?
        """, (id_service,)).fetchall()
        for m in items_modele:
            if not int(m['is_revision_item'] or 0):
                cur.execute("""
                    INSERT INTO checklist_items (id_checklist, nom_item, requires_file, is_required, item_type, video_url, is_revision, file_category, field_type)
                    VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?)
                """, (id_checklist, m['nom_item'], int(m['requires_file'] or 0), int(m['is_required'] or 1), m['item_type'] or 'document', m['video_url'], m['file_category'] or 'autre', m['field_type'] or 'check'))

        conn.commit()

        try:
            parent = client_row['drive_folder_id'] if client_row['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            projet_folder_id = create_folder(nom_projet, parent_id=parent)
            make_folder_public(projet_folder_id)
            depot_folder_id = None
            if documents_requis:
                depot_folder_id = create_folder("Dépôt de fichiers", parent_id=projet_folder_id)
                make_folder_public(depot_folder_id)
            lien_gdrive = get_folder_link(projet_folder_id)
            conn.execute("UPDATE projets SET lien_gdrive=?, drive_folder_id=?, depot_folder_id=? WHERE id=?",
                         (lien_gdrive, projet_folder_id, depot_folder_id, id_projet))
            if service_row['drive_subfolders']:
                for nom_sf in service_row['drive_subfolders'].split('|'):
                    nom_sf = nom_sf.strip()
                    if nom_sf:
                        sf_id = create_folder(nom_sf, parent_id=projet_folder_id)
                        make_folder_public(sf_id)
            conn.commit()
        except Exception as drive_e:
            print(f"[DRIVE] Création dossier projet (auto-booking) échouée: {drive_e}")

        try:
            from calendar_service import create_seance_event
            create_seance_event(nom_projet, date_seance, heure_seance, duree_seance_minutes, localisation, client_row['email'])
        except Exception as e:
            print(f"[CALENDAR] Invitation séance (auto-booking) échouée: {e}")

        try:
            if client_row['email']:
                lien_projet = url_for('project_detail', project_id=id_projet, _external=True)
                send_email_client(client_row,
                    f"Nouveau projet — {nom_projet}",
                    f"Bonjour {client_row['nom_complet']}, un nouveau projet vous a été assigné : {nom_projet}",
                    html=email_projet_cree(client_row['nom_complet'], nom_projet, lien_projet))
        except Exception as e:
            print(f"[MAIL] Email projet créé (auto-booking) échoué: {e}")

        return id_projet, nom_projet
    except Exception as e:
        conn.rollback()
        print(f"[BOOKING] Création projet automatique échouée: {e}")
        return None, None
    finally:
        conn.close()


@app.route('/api/v1/admin/service/<int:service_id>/envoyer-dispo', methods=['POST'])
@admin_required
def api_admin_service_envoyer_dispo(service_id):
    """Envoie au client les prochains créneaux disponibles pour réserver une séance d'un service précis
    (en présentiel) — le bloc réservé dans l'agenda inclut 1h de battement avant/après la séance."""
    from email_templates import email_agenda_rendez_vous
    from calendar_service import format_slot_fr
    data = request.get_json(force=True) or {}
    client_id = data.get('id_client')
    if not client_id:
        return jsonify({'error': 'Client requis'}), 400

    conn = get_db_connection()
    service_row = conn.execute("SELECT * FROM services WHERE id=?", (service_id,)).fetchone()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    conn.close()
    if not service_row:
        return jsonify({'error': 'Service introuvable'}), 404
    if not client:
        return jsonify({'error': 'Client introuvable'}), 404

    duree_seance_minutes = int(service_row['duree_seance_minutes'] or 60)
    try:
        blocs = _find_service_slots(duree_seance_minutes, n=4)
        if not blocs:
            return jsonify({'error': 'Aucun créneau disponible dans les prochaines semaines'}), 503

        from urllib.parse import quote
        slots = []
        for bloc_debut, bloc_fin, seance_debut, seance_fin in blocs:
            token = s.dumps({
                'c': client_id,
                'sv': service_id,
                'bs': bloc_debut.isoformat(),
                'be': bloc_fin.isoformat(),
                'ss': seance_debut.isoformat(),
                'se': seance_fin.isoformat(),
            }, salt='booking-service-salt')
            slots.append({
                'label': format_slot_fr(seance_debut),
                'url': f"{PORTAIL_URL}/api/v1/booking/confirm-service?token={quote(token, safe='')}",
            })

        lignes = [f"Bonjour {client['nom_complet']},\n",
                  f"Voici les prochains créneaux disponibles pour réserver votre séance « {service_row['nom_service']} » :\n"]
        for sl in slots:
            lignes.append(f"  • {sl['label']} → {sl['url']}")
        lignes.append("\n— L'équipe Cocktail Média")

        send_email(
            client['email'],
            f"Réservez votre séance « {service_row['nom_service']} » — Cocktail Média",
            "\n".join(lignes),
            html=email_agenda_rendez_vous(client['nom_complet'], slots, booking_url=os.getenv('BOOKING_URL', ''), nom_service=service_row['nom_service'])
        )
    except Exception as e:
        print(f"[MAIL] envoyer_dispo_service: {e}")
        return jsonify({'error': "Erreur lors de l'envoi de l'email"}), 500

    return jsonify({'success': True, 'nb_slots': len(slots)})


@app.route('/api/v1/booking/confirm-service', methods=['GET'])
def api_booking_confirm_service():
    """Route publique — le client clique depuis le courriel pour réserver une séance de service.
    Redirige vers une page où il fournit l'adresse de la séance avant la confirmation finale."""
    from calendar_service import format_slot_fr
    from datetime import datetime as dt
    token = request.args.get('token', '')
    if not token:
        return jsonify({'error': 'Token manquant'}), 400
    try:
        data = s.loads(token, salt='booking-service-salt', max_age=604800)  # 7 jours
    except SignatureExpired:
        return jsonify({'error': 'Ce lien a expiré (7 jours). Demandez un nouvel email.'}), 410
    except Exception:
        return jsonify({'error': 'Lien invalide'}), 400

    client_id = data.get('c')
    service_id = data.get('sv')
    bloc_debut_str = data.get('bs')
    seance_debut_str = data.get('ss')
    if not all([client_id, service_id, bloc_debut_str, seance_debut_str]):
        return jsonify({'error': 'Données de réservation incomplètes'}), 400

    conn = get_db_connection()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    service_row = conn.execute("SELECT nom_service FROM services WHERE id=?", (service_id,)).fetchone()
    existing = conn.execute(
        "SELECT id, label_fr FROM rendez_vous WHERE id_client=? AND start_utc=?",
        (client_id, bloc_debut_str)
    ).fetchone()
    conn.close()
    if not client or not service_row:
        return jsonify({'error': 'Client ou service introuvable'}), 404

    label = format_slot_fr(dt.fromisoformat(seance_debut_str))
    from urllib.parse import quote

    if existing:
        return redirect(
            f"{PORTAIL_URL}/booking/confirme"
            f"?nom={quote(client['nom_complet'])}"
            f"&creneau={quote(existing['label_fr'] or label)}"
            f"&rdv={existing['id']}"
        )

    return redirect(
        f"{PORTAIL_URL}/booking/localisation"
        f"?token={quote(token, safe='')}"
        f"&nom={quote(client['nom_complet'])}"
        f"&service={quote(service_row['nom_service'])}"
        f"&creneau={quote(label)}"
    )


@app.route('/api/v1/booking/confirm-service/finaliser', methods=['POST'])
def api_booking_confirm_service_finaliser():
    """Route publique — finalise la réservation d'une séance de service une fois l'adresse fournie :
    crée l'événement (bloc complet incluant le battement), le rendez-vous, ET le projet automatiquement."""
    from email_templates import email_rendez_vous_confirme
    from calendar_service import format_slot_fr
    from datetime import datetime as dt, timezone as _tz
    from zoneinfo import ZoneInfo

    data_in = request.get_json(force=True, silent=True) or {}
    token = (data_in.get('token') or '').strip()
    localisation = (data_in.get('localisation') or '').strip()
    if not token:
        return jsonify({'error': 'Token manquant'}), 400
    if not localisation:
        return jsonify({'error': "L'adresse de la séance est requise"}), 400
    try:
        data = s.loads(token, salt='booking-service-salt', max_age=604800)
    except SignatureExpired:
        return jsonify({'error': 'Ce lien a expiré (7 jours). Demandez un nouvel email.'}), 410
    except Exception:
        return jsonify({'error': 'Lien invalide'}), 400

    client_id = data.get('c')
    id_service = data.get('sv')
    bloc_debut_str   = data.get('bs')
    bloc_fin_str     = data.get('be')
    seance_debut_str = data.get('ss')
    if not all([client_id, id_service, bloc_debut_str, bloc_fin_str, seance_debut_str]):
        return jsonify({'error': 'Données de réservation incomplètes'}), 400

    bloc_debut   = dt.fromisoformat(bloc_debut_str)
    bloc_fin     = dt.fromisoformat(bloc_fin_str)
    seance_debut = dt.fromisoformat(seance_debut_str)
    label = format_slot_fr(seance_debut)

    conn = get_db_connection()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    existing = conn.execute(
        "SELECT id, meet_link, label_fr FROM rendez_vous WHERE id_client=? AND start_utc=?",
        (client_id, bloc_debut_str)
    ).fetchone()
    conn.close()
    if not client:
        return jsonify({'error': 'Client introuvable'}), 404

    # Idempotence — évite les doublons si le client soumet plusieurs fois
    if existing:
        return jsonify({
            'success': True, 'deja': True,
            'nom': client['nom_complet'], 'creneau': existing['label_fr'] or label,
            'meet': existing['meet_link'] or '', 'rdv': existing['id'],
            'start': bloc_debut_str, 'end': bloc_fin_str,
        })

    try:
        event_id, meet_link = create_meeting_event(bloc_debut, bloc_fin, client['nom_complet'], client['email'])
        conn = get_db_connection()
        cur = conn.execute(
            "INSERT INTO rendez_vous (id_client, calendar_event_id, start_utc, end_utc, meet_link, label_fr) VALUES (?,?,?,?,?,?)",
            (client_id, event_id, bloc_debut_str, bloc_fin_str, meet_link, label)
        )
        rdv_id = cur.lastrowid
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[BOOKING] Création événement (service) échouée: {e}")
        return jsonify({'error': "Erreur lors de la création du rendez-vous"}), 500

    # Conversion UTC → heure locale Toronto pour la date/heure du projet (create_seance_event interprète en heure locale)
    tz = ZoneInfo("America/Toronto")
    local_seance = seance_debut.replace(tzinfo=_tz.utc).astimezone(tz)
    date_seance = local_seance.date().isoformat()
    heure_seance = f"{local_seance.hour:02d}:{local_seance.minute:02d}"

    id_projet, nom_projet = _creer_projet_depuis_booking(client_id, id_service, date_seance, heure_seance, localisation)

    try:
        google_url, outlook_url, ics_url = _build_calendar_urls(bloc_debut, bloc_fin, meet_link, rdv_id)
        send_email(
            client['email'],
            "Rendez-vous confirmé — Cocktail Média",
            f"Bonjour {client['nom_complet']}, votre séance est confirmée : {label} à {localisation}.",
            html=email_rendez_vous_confirme(client['nom_complet'], label, meet_link, google_url, outlook_url, ics_url)
        )
        notif_felix = f"{client['nom_complet']} ({client['email']}) a réservé une séance « {label} » — adresse : {localisation}."
        if id_projet:
            notif_felix += f"\nProjet créé automatiquement : {nom_projet} (#{id_projet})"
        send_email('felix.dumont@cocktailmedia.ca', f"Nouvelle réservation de service — {client['nom_complet']}", notif_felix)
    except Exception as e:
        print(f"[MAIL] Confirmation réservation service échouée: {e}")

    if id_projet:
        _creer_tache_depuis_booking(client, f"Préparer la séance — {nom_projet}", date_seance, id_projet=id_projet, projet_nom=nom_projet, lien=f'/admin/projet/{id_projet}')
    else:
        _creer_tache_depuis_booking(client, f"Séance réservée — {client['nom_complet']} — {label}", date_seance)

    return jsonify({
        'success': True,
        'nom': client['nom_complet'], 'creneau': label,
        'meet': meet_link or '', 'rdv': rdv_id,
        'start': bloc_debut_str, 'end': bloc_fin_str,
        'projet_id': id_projet, 'nom_projet': nom_projet,
    })


@app.route('/api/v1/webhook/calendar', methods=['POST'])
def api_webhook_calendar():
    """Reçoit les push notifications Google Calendar (Appointment Scheduling)."""
    resource_state = request.headers.get('X-Goog-Resource-State', '')

    # Handshake initial — Google envoie 'sync' à l'enregistrement
    if resource_state != 'exists':
        return '', 200

    try:
        state = load_watch_state()
        events, new_token = list_changed_events(state.get('sync_token'))
        state['sync_token'] = new_token
        save_watch_state(state)

        conn = get_db_connection()
        for event in events:
            if event.get('eventType') != 'appointmentBooking':
                continue
            if event.get('status') == 'cancelled':
                continue

            event_id = event.get('id', '')
            if conn.execute("SELECT id FROM rendez_vous WHERE calendar_event_id=?", (event_id,)).fetchone():
                continue

            # Email du client (attendee != organisateur)
            client_email = next(
                (a['email'] for a in event.get('attendees', [])
                 if a.get('email') and a['email'] != CALENDAR_ID),
                None
            )
            if not client_email:
                continue

            client = conn.execute(
                "SELECT * FROM clients WHERE LOWER(email)=LOWER(?)", (client_email,)
            ).fetchone()
            if not client:
                print(f"[WEBHOOK] Booking ignoré — email inconnu: {client_email}")
                continue

            # Parse dates
            from datetime import datetime as _dt, timezone as _tz
            start_raw = event['start'].get('dateTime', event['start'].get('date', ''))
            end_raw   = event['end'].get('dateTime', event['end'].get('date', ''))
            start_dt  = _dt.fromisoformat(start_raw).astimezone(_tz.utc).replace(tzinfo=None)
            end_dt    = _dt.fromisoformat(end_raw).astimezone(_tz.utc).replace(tzinfo=None)

            entry_points = event.get('conferenceData', {}).get('entryPoints', [{}])
            meet_link = event.get('hangoutLink') or entry_points[0].get('uri', '') if entry_points else ''
            label = format_slot_fr(start_dt)

            cur = conn.execute(
                "INSERT INTO rendez_vous (id_client, calendar_event_id, start_utc, end_utc, meet_link, label_fr)"
                " VALUES (?,?,?,?,?,?)",
                (client['id'], event_id, start_dt.isoformat(), end_dt.isoformat(), meet_link, label)
            )
            rdv_id = cur.lastrowid
            conn.commit()
            _lier_rendez_vous_au_projet(client['id'], rdv_id)

            from email_templates import email_rendez_vous_confirme
            google_url, outlook_url, ics_url = _build_calendar_urls(start_dt, end_dt, meet_link, rdv_id)
            send_email(
                client['email'],
                "Rendez-vous confirmé — Cocktail Média",
                f"Bonjour {client['nom_complet']}, votre rendez-vous est confirmé : {label}.",
                html=email_rendez_vous_confirme(client['nom_complet'], label, meet_link, google_url, outlook_url, ics_url)
            )
            send_email(
                'felix.dumont@cocktailmedia.ca',
                f"Nouveau rendez-vous (Google Agenda) — {client['nom_complet']}",
                f"{client['nom_complet']} ({client['email']}) a réservé via Google Agenda : {label}.\nMeet : {meet_link}"
            )
            print(f"[WEBHOOK] RDV créé — {client['nom_complet']} ({client_email}) — {label}")

        conn.close()
    except Exception as e:
        print(f"[WEBHOOK] calendar error: {e}")

    return '', 200


@app.route('/api/v1/admin/webhook/calendar/register', methods=['POST'])
@admin_required
def api_admin_register_calendar_watch():
    """Enregistre ou renouvelle le push notification watch Google Calendar."""
    state = load_watch_state()
    if state.get('channel_id') and state.get('resource_id'):
        stop_calendar_watch(state['channel_id'], state['resource_id'])

    try:
        new_state = register_calendar_watch(f"{PORTAIL_URL}/api/v1/webhook/calendar")
        exp_ms = new_state.get('expiration_ms', 0)
        from datetime import datetime as _dt
        exp_str = _dt.utcfromtimestamp(exp_ms / 1000).strftime('%Y-%m-%d %H:%M UTC') if exp_ms else 'inconnu'
        return jsonify({'success': True, 'expiration': exp_str, 'channel_id': new_state['channel_id']})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/v1/admin/webhook/calendar/status', methods=['GET'])
@admin_required
def api_admin_webhook_calendar_status():
    """Retourne l'état du push notification watch actuel."""
    state = load_watch_state()
    if not state:
        return jsonify({'active': False})
    from datetime import datetime as _dt
    exp_ms = state.get('expiration_ms', 0)
    exp_str = _dt.utcfromtimestamp(exp_ms / 1000).strftime('%Y-%m-%d %H:%M UTC') if exp_ms else None
    now_ms = _dt.utcnow().timestamp() * 1000
    return jsonify({
        'active': bool(state.get('channel_id')),
        'channel_id': state.get('channel_id', ''),
        'expiration': exp_str,
        'expires_soon': exp_ms > 0 and (exp_ms - now_ms) < 7 * 24 * 3600 * 1000,
    })


@app.route('/api/v1/admin/client/<int:client_id>/creer-rendez-vous', methods=['POST'])
@admin_required
def api_admin_creer_rendez_vous(client_id):
    data = request.get_json() or {}
    date_str  = data.get('date', '')   # YYYY-MM-DD
    heure_str = data.get('heure', '')  # HH:MM
    if not date_str or not heure_str:
        return jsonify({'error': 'Date et heure requises'}), 400

    conn = get_db_connection()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    conn.close()
    if not client:
        return jsonify({'error': 'Client introuvable'}), 404

    try:
        from datetime import datetime as dt
        from zoneinfo import ZoneInfo
        from datetime import timezone as _tz, timedelta as _td
        tz = ZoneInfo("America/Toronto")
        local_start = dt.strptime(f"{date_str} {heure_str}", "%Y-%m-%d %H:%M").replace(tzinfo=tz)
        local_end   = local_start + _td(hours=1)
        start_utc   = local_start.astimezone(_tz.utc).replace(tzinfo=None)
        end_utc     = local_end.astimezone(_tz.utc).replace(tzinfo=None)

        event_id, meet_link = create_meeting_event(start_utc, end_utc, client['nom_complet'], client['email'])
        label = format_slot_fr(start_utc)

        conn = get_db_connection()
        cur = conn.execute(
            "INSERT INTO rendez_vous (id_client, calendar_event_id, start_utc, end_utc, meet_link, label_fr) VALUES (?,?,?,?,?,?)",
            (client_id, event_id, start_utc.isoformat(), end_utc.isoformat(), meet_link, label)
        )
        rdv_id = cur.lastrowid
        conn.commit()
        conn.close()
        _lier_rendez_vous_au_projet(client_id, rdv_id)

        from email_templates import email_rendez_vous_confirme
        google_url, outlook_url, ics_url = _build_calendar_urls(start_utc, end_utc, meet_link, rdv_id)
        send_email(
            client['email'],
            "Rendez-vous confirmé — Cocktail Média",
            f"Bonjour {client['nom_complet']}, un rendez-vous a été planifié : {label}. Meet : {meet_link}",
            html=email_rendez_vous_confirme(client['nom_complet'], label, meet_link, google_url, outlook_url, ics_url)
        )
    except Exception as e:
        print(f"[BOOKING] creer_rendez_vous: {e}")
        return jsonify({'error': str(e)}), 500

    return jsonify({'success': True, 'label': label, 'meet_link': meet_link})


def _build_calendar_urls(start_utc, end_utc, meet_link: str, rdv_id: int) -> tuple:
    """Retourne (google_url, outlook_url, ics_url) pour un rendez-vous."""
    from urllib.parse import urlencode
    fmt = lambda dt: dt.strftime('%Y%m%dT%H%M%S') + 'Z'
    details = f"Rejoignez la réunion Google Meet : {meet_link}" if meet_link else "Réunion Cocktail Média"
    google = 'https://calendar.google.com/calendar/render?' + urlencode({
        'action': 'TEMPLATE',
        'text': 'Réunion — Cocktail Média',
        'dates': f"{fmt(start_utc)}/{fmt(end_utc)}",
        'details': details,
        'location': meet_link or '',
    })
    outlook = 'https://outlook.live.com/owa/?' + urlencode({
        'path': '/calendar/action/compose', 'rru': 'addevent',
        'startdt': start_utc.strftime('%Y-%m-%dT%H:%M:%SZ'),
        'enddt':   end_utc.strftime('%Y-%m-%dT%H:%M:%SZ'),
        'subject': 'Réunion — Cocktail Média',
        'body': details,
        'location': meet_link or '',
    })
    ics = f"{PORTAIL_URL}/api/v1/rendez-vous/{rdv_id}/ics"
    return google, outlook, ics


@app.route('/api/v1/rendez-vous/<int:rdv_id>/ics', methods=['GET'])
def api_rendez_vous_ics(rdv_id):
    conn = get_db_connection()
    rdv = conn.execute("SELECT * FROM rendez_vous WHERE id=?", (rdv_id,)).fetchone()
    conn.close()
    if not rdv:
        return jsonify({'error': 'Introuvable'}), 404
    from datetime import datetime as dt
    start = dt.fromisoformat(rdv['start_utc'])
    end   = dt.fromisoformat(rdv['end_utc'])
    fmt   = lambda d: d.strftime('%Y%m%dT%H%M%S') + 'Z'
    meet  = rdv['meet_link'] or ''
    ics = '\r\n'.join([
        'BEGIN:VCALENDAR', 'VERSION:2.0',
        'PRODID:-//Cocktail Média//Portail//FR', 'METHOD:PUBLISH',
        'BEGIN:VEVENT',
        f'DTSTART:{fmt(start)}', f'DTEND:{fmt(end)}',
        'SUMMARY:Réunion — Cocktail Média',
        f'DESCRIPTION:Rejoignez la réunion Google Meet : {meet}',
        f'URL:{meet}', f'LOCATION:{meet}',
        'ORGANIZER;CN=Cocktail Média:mailto:felix.dumont@cocktailmedia.ca',
        'END:VEVENT', 'END:VCALENDAR',
    ])
    from flask import Response
    return Response(ics, mimetype='text/calendar',
                    headers={'Content-Disposition': 'attachment; filename="rendez-vous-cocktail.ics"'})


@app.route('/api/v1/client/rendez-vous/<int:rdv_id>', methods=['DELETE'])
@login_required
def api_client_annuler_rdv(rdv_id):
    client_id = session['user_id']
    conn = get_db_connection()
    rdv = conn.execute("SELECT * FROM rendez_vous WHERE id=? AND id_client=?", (rdv_id, client_id)).fetchone()
    if not rdv:
        conn.close()
        return jsonify({'error': 'Rendez-vous introuvable'}), 404
    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    conn.close()
    try:
        from calendar_service import delete_calendar_event
        if rdv['calendar_event_id']:
            delete_calendar_event(rdv['calendar_event_id'])
        conn = get_db_connection()
        conn.execute("DELETE FROM rendez_vous WHERE id=?", (rdv_id,))
        conn.commit()
        conn.close()
        send_email(
            'felix.dumont@cocktailmedia.ca',
            f"Rendez-vous annulé — {client['nom_complet']}",
            f"{client['nom_complet']} a annulé son rendez-vous : {rdv['label_fr']}."
        )
    except Exception as e:
        print(f"[BOOKING] annuler_rdv: {e}")
        return jsonify({'error': str(e)}), 500
    return jsonify({'success': True})


@app.route('/api/v1/client/rendez-vous/slots', methods=['GET'])
@login_required
def api_client_rdv_slots():
    """Retourne les prochains créneaux disponibles pour replanifier."""
    try:
        raw = get_available_meeting_slots(n=4, duree_minutes=60)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    return jsonify([{
        'label':     format_slot_fr(s),
        'start_utc': s.isoformat(),
        'end_utc':   e.isoformat(),
    } for s, e in raw])


@app.route('/api/v1/client/rendez-vous/<int:rdv_id>', methods=['PUT'])
@login_required
def api_client_modifier_rdv(rdv_id):
    """Annule l'ancien créneau et en crée un nouveau."""
    client_id = session['user_id']
    data = request.get_json() or {}
    start_utc_str = data.get('start_utc', '')
    end_utc_str   = data.get('end_utc', '')
    if not start_utc_str or not end_utc_str:
        return jsonify({'error': 'Créneau requis'}), 400

    conn = get_db_connection()
    rdv = conn.execute("SELECT * FROM rendez_vous WHERE id=? AND id_client=?", (rdv_id, client_id)).fetchone()
    if not rdv:
        conn.close()
        return jsonify({'error': 'Rendez-vous introuvable'}), 404

    client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
    conn.close()

    try:
        from datetime import datetime as dt
        from calendar_service import delete_calendar_event
        if rdv['calendar_event_id']:
            delete_calendar_event(rdv['calendar_event_id'])

        start_dt = dt.fromisoformat(start_utc_str)
        end_dt   = dt.fromisoformat(end_utc_str)
        event_id, meet_link = create_meeting_event(start_dt, end_dt, client['nom_complet'], client['email'])
        label = format_slot_fr(start_dt)

        conn = get_db_connection()
        conn.execute(
            "UPDATE rendez_vous SET calendar_event_id=?, start_utc=?, end_utc=?, meet_link=?, label_fr=? WHERE id=?",
            (event_id, start_utc_str, end_utc_str, meet_link, label, rdv_id)
        )
        conn.commit()
        rdv_id_val = rdv_id
        conn.close()

        from email_templates import email_rendez_vous_confirme
        google_url, outlook_url, ics_url = _build_calendar_urls(start_dt, end_dt, meet_link, rdv_id_val)
        send_email(
            client['email'],
            "Rendez-vous modifié — Cocktail Média",
            f"Votre rendez-vous a été déplacé : {label}. Meet : {meet_link}",
            html=email_rendez_vous_confirme(client['nom_complet'], label, meet_link, google_url, outlook_url, ics_url)
        )
        send_email(
            'felix.dumont@cocktailmedia.ca',
            f"Rendez-vous modifié — {client['nom_complet']}",
            f"{client['nom_complet']} a déplacé son rendez-vous : {label}."
        )
    except Exception as e:
        print(f"[BOOKING] modifier_rdv: {e}")
        return jsonify({'error': str(e)}), 500

    return jsonify({'success': True, 'label': label, 'meet_link': meet_link,
                    'start_utc': start_utc_str, 'end_utc': end_utc_str})


@app.route('/api/v1/client/rendez-vous', methods=['GET'])
@login_required
def api_client_rendez_vous():
    client_id = session['user_id']
    conn = get_db_connection()
    rdvs = conn.execute(
        "SELECT * FROM rendez_vous WHERE id_client=? AND start_utc >= datetime('now','-1 hour') ORDER BY start_utc",
        (client_id,)
    ).fetchall()
    conn.close()
    return jsonify([{
        'id':               r['id'],
        'start_utc':        r['start_utc'],
        'end_utc':          r['end_utc'],
        'meet_link':        r['meet_link'] or '',
        'label_fr':         r['label_fr'] or '',
    } for r in rdvs])


@app.route('/api/v1/admin/clients', methods=['GET'])
@admin_required
def api_admin_get_clients():
    conn = get_db_connection()
    clients = conn.execute("""
        SELECT c.*, COUNT(p.id) as nb_projets
        FROM clients c
        LEFT JOIN projets p ON p.id_client = c.id
        WHERE c.is_admin = 0
        GROUP BY c.id
        ORDER BY c.created_at DESC
    """).fetchall()
    conn.close()
    return jsonify([{
        'id': c['id'],
        'nom_complet': c['nom_complet'],
        'email': c['email'],
        'nom_entreprise': c['nom_entreprise'],
        'telephone': c['telephone'],
        'created_at': c['created_at'],
        'nb_projets': c['nb_projets'],
        'is_email_confirmed': bool(c['is_email_confirmed']),
        'statut_relation': c['statut_relation'] or 'actif',
        'is_test_client': bool(c['is_test_client']) if 'is_test_client' in c.keys() else False,
    } for c in clients])


@app.route('/api/v1/admin/client/<int:client_id>/notes', methods=['GET'])
@admin_required
def api_admin_get_client_notes(client_id):
    conn = get_db_connection()
    notes = conn.execute(
        "SELECT * FROM client_notes WHERE id_client = ? ORDER BY created_at DESC", (client_id,)
    ).fetchall()
    conn.close()
    return jsonify([{'id': n['id'], 'contenu': n['contenu'], 'created_at': n['created_at']} for n in notes])


@app.route('/api/v1/admin/client/<int:client_id>/notes', methods=['POST'])
@admin_required
def api_admin_add_client_note(client_id):
    data = request.get_json(force=True) or {}
    contenu = (data.get('contenu') or '').strip()
    if not contenu:
        return jsonify({'error': 'Le contenu de la note est obligatoire.'}), 400
    conn = get_db_connection()
    cur = conn.execute(
        "INSERT INTO client_notes (id_client, contenu) VALUES (?, ?)", (client_id, contenu)
    )
    note = conn.execute(
        "SELECT * FROM client_notes WHERE id = ?", (cur.lastrowid,)
    ).fetchone()
    conn.commit()
    conn.close()
    return jsonify({'id': note['id'], 'contenu': note['contenu'], 'created_at': note['created_at']})


@app.route('/api/v1/admin/client/<int:client_id>/notes/<int:note_id>', methods=['DELETE'])
@admin_required
def api_admin_delete_client_note(client_id, note_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM client_notes WHERE id = ? AND id_client = ?", (note_id, client_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/client/<int:client_id>/projets', methods=['GET'])
@admin_required
def api_admin_get_client_projets(client_id):
    conn = get_db_connection()
    projets = conn.execute("""
        SELECT p.*, s.nom_service
        FROM projets p
        LEFT JOIN services s ON s.id = p.id_service
        WHERE p.id_client = ?
        ORDER BY p.created_at DESC
    """, (client_id,)).fetchall()
    conn.close()
    return jsonify([{
        'id': p['id'],
        'nom_projet': p['nom_projet'],
        'statut': p['statut'],
        'date_livraison_estimee': p['date_livraison_estimee'],
        'is_archived': p['is_archived'],
        'created_at': p['created_at'],
        'nom_service': p['nom_service'],
    } for p in projets])

# ───────────────────────────────────────────────────────────
# API v1 — Admin Projets liste + création
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/projets/new', methods=['POST'])
@admin_required
def api_admin_create_projet():
    from datetime import date as date_cls
    data = request.get_json(force=True) or {}
    id_client = data.get('id_client')
    id_service = data.get('id_service')
    date_seance = data.get('date_seance') or str(date_cls.today())
    heure_seance = (data.get('heure_seance') or '').strip() or None
    titre_projet = (data.get('titre_projet') or '').strip() or None
    localisation = (data.get('localisation') or '').strip() or None
    lien_reunion = (data.get('lien_reunion') or '').strip() or None
    generer_meet = bool(data.get('generer_meet'))
    date_reunion = (data.get('date_reunion') or '').strip() or None
    heure_reunion = (data.get('heure_reunion') or '').strip() or None
    extras = data.get('extras', [])
    facturation_mode = (data.get('facturation_mode') or '').strip() or None

    if not id_client or not id_service:
        return jsonify({'error': 'Client et service obligatoires.'}), 400

    conn = get_db_connection()
    try:
        service_row = conn.execute("SELECT * FROM services WHERE id=?", (id_service,)).fetchone()
        if not service_row:
            return jsonify({'error': 'Service introuvable.'}), 400

        nom_service = service_row['nom_service']
        documents_requis = bool(service_row['documents_requis'])
        statut = initial_statut_for_service(service_row)

        if not heure_seance and service_row['heure_seance_defaut']:
            heure_seance = service_row['heure_seance_defaut']

        duree_seance_minutes = int(service_row['duree_seance_minutes'] or 60)

        appel_requis = bool(service_row['appel_exploratoire_requis'])
        if appel_requis and date_reunion and heure_reunion and not lien_reunion:
            generer_meet = True

        lien_gdrive = None
        if generer_meet and date_reunion and heure_reunion:
            try:
                from calendar_service import create_meet_event
                lien_reunion = create_meet_event(titre_projet or nom_service, date_reunion, heure_reunion)
            except Exception as e:
                print(f"[MEET] Génération lien échouée: {e}")

        nom_projet = f"{date_seance} — {nom_service} — {localisation}" if localisation else f"{date_seance} — {nom_service}"

        cur = conn.cursor()
        cur.execute("""
            INSERT INTO projets (nom_projet, titre_projet, heure_seance, lien_reunion, duree_seance_minutes, statut, lien_gdrive, id_client, localisation, id_service, facturation_mode)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (nom_projet, titre_projet, heure_seance, lien_reunion, duree_seance_minutes, statut, lien_gdrive, id_client, localisation, id_service, facturation_mode))
        id_projet = cur.lastrowid

        for extra in extras:
            nom_extra = (extra.get('nom') or '').strip()
            prix_extra = round(float(extra.get('prix', 0) or 0), 2)
            km_extra = int(extra.get('km', 0) or 0)
            if nom_extra:
                cur.execute(
                    "INSERT INTO projet_extras (id_projet, nom, prix, km) VALUES (?, ?, ?, ?)",
                    (id_projet, nom_extra, prix_extra, km_extra)
                )

        cur.execute("INSERT INTO checklistes (id_projet) VALUES (?)", (id_projet,))
        id_checklist = cur.lastrowid
        items_modele = conn.execute("""
            SELECT nom_item, requires_file, is_required, item_type, video_url, is_revision_item, file_category, field_type
            FROM checklist_model_items WHERE id_service = ?
        """, (id_service,)).fetchall()
        for m in items_modele:
            if not int(m['is_revision_item'] or 0):
                cur.execute("""
                    INSERT INTO checklist_items (id_checklist, nom_item, requires_file, is_required, item_type, video_url, is_revision, file_category, field_type)
                    VALUES (?, ?, ?, ?, ?, ?, 0, ?, ?)
                """, (id_checklist, m['nom_item'], int(m['requires_file'] or 0), int(m['is_required'] or 1), m['item_type'] or 'document', m['video_url'], m['file_category'] or 'autre', m['field_type'] or 'check'))

        conn.commit()

        try:
            client_row = conn.execute("SELECT * FROM clients WHERE id=?", (id_client,)).fetchone()
            parent = client_row['drive_folder_id'] if client_row and client_row['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            projet_folder_id = create_folder(nom_projet, parent_id=parent)
            make_folder_public(projet_folder_id)
            depot_folder_id = None
            if documents_requis:
                depot_folder_id = create_folder("Dépôt de fichiers", parent_id=projet_folder_id)
                make_folder_public(depot_folder_id)
            lien_gdrive_new = get_folder_link(projet_folder_id)
            conn.execute("UPDATE projets SET lien_gdrive=?, drive_folder_id=?, depot_folder_id=? WHERE id=?",
                         (lien_gdrive_new, projet_folder_id, depot_folder_id, id_projet))
            if service_row['drive_subfolders']:
                for nom_sf in service_row['drive_subfolders'].split('|'):
                    nom_sf = nom_sf.strip()
                    if nom_sf:
                        sf_id = create_folder(nom_sf, parent_id=projet_folder_id)
                        make_folder_public(sf_id)
            conn.commit()
        except Exception as drive_e:
            print(f"[DRIVE] Création dossier projet échouée: {drive_e}")

        if heure_seance and localisation:
            try:
                from calendar_service import create_seance_event
                client_row2 = conn.execute("SELECT * FROM clients WHERE id=?", (id_client,)).fetchone()
                create_seance_event(nom_projet, date_seance, heure_seance, duree_seance_minutes, localisation, client_row2['email'])
            except Exception as e:
                print(f"[CALENDAR] Invitation séance échouée: {e}")

        try:
            client_notif = conn.execute("SELECT * FROM clients WHERE id=?", (id_client,)).fetchone()
            if client_notif and client_notif['email']:
                lien_projet = url_for('project_detail', project_id=id_projet, _external=True)
                send_email_client(client_notif,
                    f"Nouveau projet — {nom_projet}",
                    f"Bonjour {client_notif['nom_complet']}, un nouveau projet vous a été assigné : {nom_projet}",
                    html=email_projet_cree(client_notif['nom_complet'], nom_projet, lien_projet)

                )
        except Exception as e:
            print(f"[MAIL] Email projet créé échoué: {e}")

        # Ni rendez-vous ni documents requis pour ce service : rien à attendre, on démarre
        # tout de suite (facture, événement Calendar, todos, courriel) — le bouton
        # « Démarrer les travaux » reste disponible ensuite pour relancer/confirmer au besoin.
        if statut == 'Travaux en cours':
            try:
                projet_frais = conn.execute("SELECT * FROM projets WHERE id=?", (id_projet,)).fetchone()
                client_frais = conn.execute("SELECT * FROM clients WHERE id=?", (id_client,)).fetchone()
                _do_start_travaux(conn, projet_frais, client_frais)
            except Exception as e:
                print(f"[START] Auto-start à la création échoué: {e}")

        return jsonify({'id': id_projet, 'nom_projet': nom_projet})

    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/v1/admin/projets', methods=['GET'])
@admin_required
def api_admin_get_projets():
    conn = get_db_connection()
    projets = conn.execute("""
        SELECT p.*, c.nom_complet as client_nom, c.is_test_client as client_is_test,
               s.nom_service, s.categorie as service_categorie,
               pg.nom_complet as responsable_nom
        FROM projets p
        LEFT JOIN clients c ON c.id = p.id_client
        LEFT JOIN services s ON s.id = p.id_service
        LEFT JOIN pigistes pg ON pg.id = p.id_pigiste
        ORDER BY p.created_at DESC
    """).fetchall()
    # Ratios de checklist par projet (pour la vraie progression quand elle existe)
    chk = {}
    try:
        for row in conn.execute("""
            SELECT ck.id_projet AS pid, COUNT(ci.id) AS total,
                   SUM(CASE WHEN ci.est_coche THEN 1 ELSE 0 END) AS done
            FROM checklistes ck JOIN checklist_items ci ON ci.id_checklist = ck.id
            GROUP BY ck.id_projet
        """).fetchall():
            chk[row['pid']] = (row['done'] or 0, row['total'] or 0)
    except Exception:
        pass
    conn.close()

    # Progression : ratio de checklist si dispo, sinon repli sur le statut (PHASE_CONFIG)
    def progress_for(p):
        done, total = chk.get(p['id'], (0, 0))
        if total > 0:
            return round(100 * done / total)
        if p['statut'] in ('Travaux terminés',):  # alias legacy hors PHASE_CONFIG
            return 100
        return phase_progress(p['statut'])

    return jsonify([{
        'id': p['id'],
        'nom_projet': p['nom_projet'],
        'statut': p['statut'],
        'client_nom': p['client_nom'],
        'nom_service': p['nom_service'],
        'service_categorie': p['service_categorie'],
        'responsable_nom': p['responsable_nom'],
        'progress': progress_for(p),
        'date_livraison_estimee': p['date_livraison_estimee'],
        'is_archived': p['is_archived'],
        'created_at': p['created_at'],
        'is_test_client': bool(p['client_is_test']) if 'client_is_test' in p.keys() else False,
    } for p in projets])

# ───────────────────────────────────────────────────────────
# API v1 — Admin Services
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/services', methods=['GET'])
@admin_required
def api_admin_get_services():
    conn = get_db_connection()
    services = conn.execute('SELECT * FROM services ORDER BY nom_service').fetchall()
    result = []
    for s in services:
        items = conn.execute(
            'SELECT * FROM checklist_model_items WHERE id_service = ? ORDER BY position',
            (s['id'],)
        ).fetchall()
        extras = conn.execute(
            'SELECT * FROM service_extras WHERE id_service = ? ORDER BY position, id',
            (s['id'],)
        ).fetchall()
        result.append({
            'id': s['id'],
            'nom_service': s['nom_service'],
            'description': s['description'],
            'icon': s['icon'],
            'slug': s['slug'],
            'categorie': s['categorie'],
            'localisation_requise': bool(s['localisation_requise']),
            'appel_exploratoire_requis': bool(s['appel_exploratoire_requis']),
            'decision_board_requis': bool(s['decision_board_requis']),
            'prix': s['prix'],
            'duree_affichee': s['duree_affichee'],
            'actif': bool(s['actif']) if s['actif'] is not None else True,
            'nb_items': len(items),
            'items': [{'id': i['id'], 'nom_item': i['nom_item'], 'is_required': bool(i['is_required'])} for i in items],
            'extras': [{'id': e['id'], 'nom': e['nom'], 'prix': e['prix']} for e in extras],
        })
    conn.close()
    return jsonify(result)

# ───────────────────────────────────────────────────────────
# API v1 — Admin Dashboard stats
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/dashboard', methods=['GET'])
@admin_required
def api_admin_dashboard():
    conn = get_db_connection()
    total_clients = conn.execute(
        'SELECT COUNT(*) FROM clients WHERE is_admin = 0 AND is_test_client = 0'
    ).fetchone()[0]
    projets_actifs = conn.execute("""
        SELECT COUNT(*) FROM projets p LEFT JOIN clients c ON c.id = p.id_client
        WHERE (p.is_archived = 0 OR p.is_archived IS NULL)
          AND (c.is_test_client = 0 OR c.is_test_client IS NULL)
    """).fetchone()[0]
    en_revision = conn.execute("""
        SELECT COUNT(*) FROM projets p LEFT JOIN clients c ON c.id = p.id_client
        WHERE p.statut = 'En révision'
          AND (c.is_test_client = 0 OR c.is_test_client IS NULL)
    """).fetchone()[0]
    archives = conn.execute("""
        SELECT COUNT(*) FROM projets p LEFT JOIN clients c ON c.id = p.id_client
        WHERE p.is_archived = 1
          AND (c.is_test_client = 0 OR c.is_test_client IS NULL)
    """).fetchone()[0]
    projets_recents = conn.execute("""
        SELECT p.*, c.nom_complet as client_nom
        FROM projets p
        LEFT JOIN clients c ON c.id = p.id_client
        WHERE (p.is_archived = 0 OR p.is_archived IS NULL)
          AND (c.is_test_client = 0 OR c.is_test_client IS NULL)
        ORDER BY p.created_at DESC LIMIT 5
    """).fetchall()
    top_clients = conn.execute("""
        SELECT c.id, c.nom_complet, c.nom_entreprise
        FROM clients c
        WHERE c.is_admin = 0 AND c.is_test_client = 0
        ORDER BY c.created_at DESC LIMIT 5
    """).fetchall()
    mois_courant = datetime.now().strftime('%Y-%m')
    visuels_a_creer = conn.execute("""
        SELECT id, titre, date_publication FROM marketing_posts
        WHERE strftime('%Y-%m', date_publication) = ? AND todo_felix_done = 0
          AND linked_roadmap_todo_id IS NULL
        ORDER BY date_publication ASC LIMIT 5
    """, (mois_courant,)).fetchall()
    a_publier = conn.execute("""
        SELECT id, titre, date_publication FROM marketing_posts
        WHERE strftime('%Y-%m', date_publication) = ? AND statut = 'visuels prêts'
        ORDER BY date_publication ASC LIMIT 5
    """, (mois_courant,)).fetchall()
    conn.close()
    return jsonify({
        'total_clients': total_clients,
        'projets_actifs': projets_actifs,
        'en_revision': en_revision,
        'archives': archives,
        'projets_recents': [{
            'id': p['id'],
            'nom_projet': p['nom_projet'],
            'statut': p['statut'],
            'client_nom': p['client_nom'],
            'date_livraison_estimee': p['date_livraison_estimee'],
        } for p in projets_recents],
        'top_clients': [{
            'id': c['id'],
            'nom_complet': c['nom_complet'],
            'nom_entreprise': c['nom_entreprise'],
            'initiales': ''.join([n[0] for n in c['nom_complet'].split()[:2]]).upper(),
            'couleur': 'bg-[#e83b14]',
        } for c in top_clients],
        'visuels_a_creer': [{
            'id': p['id'],
            'titre': p['titre'],
            'date_publication': p['date_publication'],
        } for p in visuels_a_creer],
        'a_publier': [{
            'id': p['id'],
            'titre': p['titre'],
            'date_publication': p['date_publication'],
        } for p in a_publier],
    })

# ───────────────────────────────────────────────────────────
# API v1 — Changelog de développement CocktailOS
# ───────────────────────────────────────────────────────────

CHANGELOG_PATH = os.getenv('CHANGELOG_PATH', '/data/changelog.json')

def _lire_changelog():
    try:
        with open(CHANGELOG_PATH, 'r', encoding='utf-8') as f:
            data = _json.load(f)
            return data.get('entries', [])
    except (FileNotFoundError, ValueError):
        return []

def _ecrire_changelog(entries):
    with open(CHANGELOG_PATH, 'w', encoding='utf-8') as f:
        _json.dump({'entries': entries}, f, ensure_ascii=False, indent=2)

@app.route('/api/v1/admin/changelog', methods=['GET'])
@admin_required
def api_admin_changelog_get():
    entries = _lire_changelog()
    entries = sorted(entries, key=lambda e: e.get('date', ''), reverse=True)
    return jsonify({'entries': entries})

@app.route('/api/v1/admin/changelog', methods=['POST'])
@admin_required
def api_admin_changelog_post():
    data = request.get_json(force=True, silent=True) or {}
    aujourdhui = datetime.now().strftime('%Y-%m-%d')
    entries = _lire_changelog()
    nb_du_jour = sum(1 for e in entries if e.get('id', '').startswith(aujourdhui))
    nouvel_id = f"{aujourdhui}-{nb_du_jour + 1:03d}"
    entree = {
        'id': nouvel_id,
        'date': aujourdhui,
        'session': data.get('session', ''),
        'category': data.get('category', ''),
        'module': data.get('module', ''),
        'description': data.get('description', ''),
        'files_modified': data.get('files_modified', []),
        'author': 'Claude Code',
    }
    entries.insert(0, entree)
    _ecrire_changelog(entries)
    return jsonify({'success': True, 'entry': entree})

# ───────────────────────────────────────────────────────────
# API v1 — Checklist items projet
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/projet/<int:projet_id>/checklist', methods=['GET'])
@admin_required
def api_admin_get_checklist(projet_id):
    conn = get_db_connection()
    checklist = conn.execute(
        'SELECT * FROM checklistes WHERE id_projet = ?', (projet_id,)
    ).fetchone()
    if not checklist:
        conn.close()
        return jsonify([])
    items = conn.execute(
        'SELECT * FROM checklist_items WHERE id_checklist = ? ORDER BY position',
        (checklist['id'],)
    ).fetchall()
    conn.close()
    return jsonify([{
        'id': i['id'],
        'nom_item': i['nom_item'],
        'est_coche': bool(i['est_coche']),
        'requires_file': i['requires_file'],
        'is_required': i['is_required'],
        'is_revision': bool(i['is_revision']),
        'admin_resolu': bool(i['admin_resolu']),
        'item_type': i['item_type'],
        'file_category': i['file_category'],
        'field_type': i['field_type'] if i['field_type'] else 'check',
        'text_value': i['text_value'] if i['text_value'] else None,
        'has_file': bool(i['file_path']),
        'file_name': os.path.basename(i['file_path']) if i['file_path'] else None,
    } for i in items])


@app.route('/api/v1/admin/item/<int:item_id>/file')
@admin_required
def api_admin_serve_item_file(item_id):
    conn = get_db_connection()
    item = conn.execute("SELECT * FROM checklist_items WHERE id = ?", (item_id,)).fetchone()
    conn.close()
    if not item or not item['file_path']:
        return jsonify({'error': 'Fichier introuvable'}), 404
    upload_root = os.path.realpath(app.config["UPLOAD_ROOT"])
    filepath = os.path.realpath(item['file_path'])
    if not filepath.startswith(upload_root + os.sep):
        return jsonify({'error': 'Accès non autorisé'}), 403
    if not os.path.exists(filepath):
        return jsonify({'error': 'Fichier manquant sur le disque'}), 404
    return send_file(filepath, as_attachment=False, download_name=os.path.basename(filepath))


# ───────────────────────────────────────────────────────────
# API v1 — Admin Projet Actions
# ───────────────────────────────────────────────────────────

def _lier_rendez_vous_au_projet(id_client, rdv_id):
    """Si ce client a EXACTEMENT un projet en attente de rendez-vous ('En attente de
       rendez-vous'), lie la réservation qui vient d'être confirmée à ce projet (id_projet
       sur rendez_vous) et fait avancer son statut automatiquement — vers 'Documents à
       donner' si le service en requiert, sinon directement 'Travaux en cours'. Si 0 ou
       plusieurs projets correspondent (ambigu), ne touche à rien : le bouton manuel sur la
       fiche projet admin reste le filet de sécurité pour ces cas."""
    conn = get_db_connection()
    try:
        candidats = conn.execute(
            "SELECT * FROM projets WHERE id_client=? AND statut='En attente de rendez-vous'",
            (id_client,)
        ).fetchall()
        if len(candidats) != 1:
            return
        projet = candidats[0]
        conn.execute("UPDATE rendez_vous SET id_projet=? WHERE id=?", (projet['id'], rdv_id))
        conn.commit()
        service_row = conn.execute("SELECT * FROM services WHERE id=?", (projet['id_service'],)).fetchone() if projet['id_service'] else None
        documents_requis = bool(service_row['documents_requis']) if service_row else True
        client = conn.execute("SELECT * FROM clients WHERE id=?", (id_client,)).fetchone()
        if documents_requis:
            conn.execute("UPDATE projets SET statut='Documents à donner' WHERE id=?", (projet['id'],))
            conn.commit()
            try:
                if client and client['email']:
                    lien = url_for('project_detail', project_id=projet['id'], _external=True)
                    send_email_client(client, f"Documents requis — {projet['nom_projet']}",
                        f"Bonjour {client['nom_complet']}, nous avons besoin de vos documents.",
                        html=email_documents_requis(client['nom_complet'], projet['nom_projet'], lien))
                    push_notification(conn, id_client, projet['id'], f"Documents requis pour votre projet « {projet['nom_projet']} ».", type='documents_requis')
                    conn.commit()
            except Exception as e:
                print(f"[RDV] courriel documents_requis: {e}")
        else:
            _do_start_travaux(conn, projet, client)
    except Exception as e:
        print(f"[RDV] _lier_rendez_vous_au_projet: {e}")
    finally:
        conn.close()


def _do_start_travaux(conn, projet, client, date_livraison_override=None):
    """Démarre les travaux : événement Calendar (si pas déjà fait), facture (si applicable),
       todos par défaut (si pas déjà créés), courriel + notification au client. Fonction
       partagée entre le bouton dédié « Démarrer travaux » ET /force-status — un seul
       chemin de logique, plus de divergence de comportement selon l'entrée.

       date_livraison_override : date de livraison choisie manuellement par l'admin (ex:
       "2026-08-01"), sinon le calcul automatique habituel (find_next_available_slot via
       create_production_event) reste utilisé. Ne change QUE la date affichée/promise au
       client (todos + courriel) — l'événement Calendar interne de l'équipe garde son propre
       créneau auto-calculé, pas déplacé par cet override."""
    project_id = projet['id']
    service_row = conn.execute("SELECT * FROM services WHERE id=?", (projet['id_service'],)).fetchone()
    date_livraison = projet['date_livraison_estimee']
    if isinstance(date_livraison, str) and date_livraison:
        try:
            from datetime import datetime as _dt
            date_livraison = _dt.strptime(date_livraison[:10], "%Y-%m-%d").date()
        except ValueError:
            date_livraison = None

    override_date = None
    if date_livraison_override:
        try:
            from datetime import datetime as _dt2
            override_date = _dt2.strptime(str(date_livraison_override)[:10], "%Y-%m-%d").date()
        except ValueError:
            override_date = None

    if not projet['calendar_event_id']:
        duree_minutes = service_row['duree_production_minutes'] if service_row and service_row['duree_production_minutes'] else 60
        delai_fixe = service_row['delai_fixe_heures'] if service_row and service_row['delai_fixe_heures'] else 0
        icon_service = service_row['icon'] if service_row and service_row['icon'] else 'default'
        taches_cal = TACHES_PAR_SERVICE.get(icon_service, TACHES_PAR_SERVICE['default'])
        event_id, date_livraison = create_production_event(projet['nom_projet'], icon_service, duree_minutes, delai_fixe, taches=taches_cal)
        if override_date:
            date_livraison = override_date
        conn.execute(
            "UPDATE projets SET statut='Travaux en cours', calendar_event_id=?, date_livraison_estimee=? WHERE id=?",
            (event_id, str(date_livraison) if date_livraison else None, project_id)
        )
    elif override_date:
        date_livraison = override_date
        conn.execute(
            "UPDATE projets SET statut='Travaux en cours', date_livraison_estimee=? WHERE id=?",
            (str(override_date), project_id)
        )
    else:
        conn.execute("UPDATE projets SET statut='Travaux en cours' WHERE id=?", (project_id,))

    existing_todos = conn.execute("SELECT COUNT(*) FROM todos_perso WHERE projet_id=?", (project_id,)).fetchone()[0]
    if existing_todos == 0:
        icon_service = service_row['icon'] if service_row and service_row['icon'] else 'default'
        taches = TACHES_PAR_SERVICE.get(icon_service, TACHES_PAR_SERVICE['default'])
        date_echeance = str(date_livraison) if date_livraison else None
        # Assignées par défaut au rôle gestion — sinon invisibles dans « Mes tâches »
        # (filtrée sur l'assignation) et perdues dans « Toutes ». Le regroupement « Par
        # projet » du module Tâches (projet_id) les affiche déjà comme des sous-tâches
        # sous une barre de progression par projet, une fois assignées (2026-07-18).
        gestion_id = _admin_id_for_role(conn, 'gestion')
        for texte in taches:
            cur = conn.execute(
                """INSERT INTO todos_perso (texte, priorite, date_echeance, projet_id, projet_nom)
                   VALUES (?, 'normale', ?, ?, ?)""",
                (texte, date_echeance, project_id, projet['nom_projet'])
            )
            if gestion_id:
                conn.execute("INSERT OR IGNORE INTO todo_assignees (todo_id, admin_id) VALUES (?, ?)", (cur.lastrowid, gestion_id))
    conn.commit()

    facture = None
    pdf_path = None
    try:
        mode = client['mode_facturation'] if client and 'mode_facturation' in client.keys() else 'projet'
        # facturation_mode non NULL = projet marqué « ne pas facturer » → aucune facture générée
        facturer = not projet['facturation_mode']
        if client and mode == 'mensuel' and facturer:
            from datetime import date as _date
            prix = float(service_row['prix'] or 0) if service_row else 0
            if prix > 0:
                ajouter_ligne_facture_mensuelle(client['id'], project_id, service_row['nom_service'] if service_row else projet['nom_projet'], _date.today().strftime("%Y-%m-%d"), projet['localisation'] or None, prix, conn)
        if client and mode == 'projet' and facturer:
            facture = creer_facture_projet(project_id, conn)
            if facture:
                pdf_path = facture['pdf_path']
                try:
                    cr = conn.execute("SELECT factures_folder_id FROM clients WHERE id=?", (client['id'],)).fetchone()
                    if cr and cr['factures_folder_id']:
                        fid, _ = upload_file(pdf_path, f"{facture['numero']}.pdf", cr['factures_folder_id'])
                        conn.execute("UPDATE factures SET drive_file_id=? WHERE id=?", (fid, facture['id']))
                        conn.commit()
                except Exception as e:
                    print(f"[DRIVE] Upload facture _do_start_travaux: {e}")
    except Exception as e:
        print(f"[INVOICE] _do_start_travaux: {e}")

    try:
        if client and client['email']:
            lien = url_for('project_detail', project_id=project_id, _external=True)
            from calendar_service import format_date_fr
            date_str = format_date_fr(date_livraison) if date_livraison else 'à déterminer'
            msg = Message(f"Les travaux sont en cours — {projet['nom_projet']}", sender=app.config['MAIL_DEFAULT_SENDER'], recipients=[client['email']])
            msg.body = f"Bonjour {client['nom_complet']}, les travaux sont en cours. Livraison estimée : {date_str}"
            msg.html = email_travaux_en_cours_avec_date(client['nom_complet'], projet['nom_projet'], lien, date_str,
                facture_numero=facture['numero'] if facture else None,
                facture_total=facture['total'] if facture else None,
                facture_echeance=facture['date_echeance'] if facture else None)
            if pdf_path and os.path.exists(pdf_path):
                with open(pdf_path, 'rb') as f:
                    msg.attach(f"{facture['numero']}.pdf", 'application/pdf', f.read())
            mail.send(msg)
            push_notification(conn, client['id'], project_id, f"Les travaux sont en cours sur votre projet « {projet['nom_projet']} ».", type='travaux_en_cours')
            conn.commit()
    except Exception as e:
        print(f"[MAIL] _do_start_travaux: {e}")
    return date_livraison


def _check_revision_auto_transition(conn, project_id):
    """Appelée après qu'un item de révision soit coché/commenté. Si le client a répondu à
    TOUS les items (coché ou commenté) et qu'au moins un commentaire (une correction
    demandée) existe, la balle revient dans le camp de l'équipe — passe le projet en
    'Corrections en cours' et prévient l'admin. Si tout est simplement approuvé sans aucun
    commentaire, on ne bouge rien : l'admin reste libre de compléter manuellement quand il
    le juge à propos (pas d'auto-complétion, qui demande un choix de ressources à assigner)."""
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet or projet['statut'] != 'En révision':
        return
    all_resolved, has_comments = compute_revision_readiness(project_id)
    if all_resolved and has_comments:
        conn.execute("UPDATE projets SET statut='Corrections en cours' WHERE id=?", (project_id,))
        conn.commit()
        push_admin_notif(
            conn,
            titre=f"Corrections demandées — {projet['nom_projet']}",
            message="Le client a terminé sa révision avec au moins une demande de changement.",
            type='info',
            lien=f"/admin/projet/{project_id}",
        )
        conn.commit()
        try:
            checklist = conn.execute("SELECT id FROM checklistes WHERE id_projet=?", (project_id,)).fetchone()
            a_corriger = conn.execute(
                "SELECT nom_item, text_value FROM checklist_items "
                "WHERE id_checklist=? AND is_revision=1 AND text_value IS NOT NULL AND TRIM(text_value) != ''",
                (checklist['id'],)
            ).fetchall() if checklist else []
            lignes = "\n".join(f"- {it['nom_item']} : {it['text_value']}" for it in a_corriger)
            corps = (
                f"Le client a terminé sa révision pour « {projet['nom_projet']} » avec des demandes de changement.\n\n"
                f"À corriger :\n{lignes}\n\n"
                f"Voir le projet : {PORTAIL_URL}/admin/projet/{project_id}"
            )
            send_email('felix.dumont@cocktailmedia.ca', f"Corrections demandées — {projet['nom_projet']}", corps)
        except Exception as e:
            print(f"[MAIL] _check_revision_auto_transition (notif corrections): {e}")

def _check_corrections_completed_auto_transition(conn, project_id):
    """Appelée après que l'admin ait coché un item « à corriger » (is_revision=1,
    coché + commenté par le client) comme réglé (admin_resolu=1). Si TOUS les items à
    corriger du projet sont maintenant réglés, les remet à neuf (le client n'a besoin
    de re-vérifier que ceux-là — les items simplement approuvés restent tels quels),
    repasse le projet en 'En révision' et renvoie automatiquement le courriel de
    révision pour ce dernier passage."""
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet or projet['statut'] != 'Corrections en cours':
        return
    checklist = conn.execute("SELECT id FROM checklistes WHERE id_projet=?", (project_id,)).fetchone()
    if not checklist:
        return
    a_corriger = conn.execute(
        "SELECT * FROM checklist_items "
        "WHERE id_checklist=? AND is_revision=1 AND est_coche=1 AND text_value IS NOT NULL AND TRIM(text_value) != ''",
        (checklist['id'],)
    ).fetchall()
    if not a_corriger or not all(int(it['admin_resolu'] or 0) == 1 for it in a_corriger):
        return
    for it in a_corriger:
        conn.execute(
            "UPDATE checklist_items SET est_coche=0, text_value=NULL, file_path=NULL, admin_resolu=0 WHERE id=?",
            (it['id'],)
        )
    conn.commit()
    _do_start_revision(conn, projet, [])

def _do_start_revision(conn, projet, items_revision=None):
    """Passe le projet en révision : seed la checklist de révision (items fournis, ou
       défauts par service si aucun), courriel + notification. Partagée entre le bouton
       dédié et /force-status."""
    project_id = projet['id']
    items_revision = [str(n).strip() for n in (items_revision or []) if str(n).strip()]
    checklist = conn.execute("SELECT id FROM checklistes WHERE id_projet=?", (project_id,)).fetchone()
    if not checklist:
        conn.execute("INSERT INTO checklistes (id_projet) VALUES (?)", (project_id,))
        conn.commit()
        checklist = conn.execute("SELECT id FROM checklistes WHERE id_projet=?", (project_id,)).fetchone()
    if not items_revision:
        deja_existants = conn.execute(
            "SELECT COUNT(*) AS n FROM checklist_items WHERE id_checklist=? AND is_revision=1",
            (checklist['id'],)
        ).fetchone()['n']
        if not deja_existants:
            service_row = conn.execute("SELECT nom_service FROM services WHERE id=?", (projet['id_service'],)).fetchone() if projet['id_service'] else None
            nom_service = service_row['nom_service'] if service_row else None
            items_revision = REVISION_ITEMS_PAR_SERVICE.get(nom_service, [])
    for nom in items_revision:
        conn.execute("""
            INSERT INTO checklist_items (id_checklist, nom_item, requires_file, is_required, is_revision, field_type)
            VALUES (?, ?, 0, 1, 1, 'review')
        """, (checklist['id'], nom))
    conn.execute("UPDATE projets SET statut='En révision' WHERE id=?", (project_id,))
    conn.commit()
    try:
        client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
        if client and client['email']:
            projet_fresh = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
            _envoyer_courriel_revision(conn, projet_fresh, client)
            push_notification(conn, client['id'], project_id,
                f"Votre projet « {projet_fresh['nom_projet']} » est en révision.", type='revision')
            conn.commit()
    except Exception as e:
        print(f"[MAIL] _do_start_revision: {e}")


def _do_complete(conn, projet, ressource_ids=None):
    """Complète le projet : assigne les ressources choisies, coche les todos, courriel de
       livraison + notification. Partagée entre le bouton dédié et /force-status."""
    project_id = projet['id']
    id_client = projet['id_client']
    ressources_assignees = []
    for rid in (ressource_ids or []):
        try:
            rid = int(rid)
        except (TypeError, ValueError):
            continue
        row = conn.execute("SELECT * FROM client_ressources WHERE id=?", (rid,)).fetchone()
        if not row or (row['id_client'] is not None and row['id_client'] != id_client):
            continue
        conn.execute("""
            INSERT INTO ressource_assignations (id_ressource, id_client, id_projet)
            VALUES (?, ?, ?)
            ON CONFLICT(id_ressource, id_client) DO NOTHING
        """, (rid, id_client, project_id))
        ressources_assignees.append(dict(_ressource_to_dict(row)))

    conn.execute("UPDATE projets SET statut='Complété' WHERE id=?", (project_id,))
    conn.execute("UPDATE todos_perso SET est_coche=1 WHERE projet_id=? AND est_coche=0", (project_id,))
    conn.commit()
    try:
        client = conn.execute("SELECT * FROM clients WHERE id=?", (id_client,)).fetchone()
        if client and client['email']:
            lien = url_for('project_detail', project_id=project_id, _external=True)
            for r in ressources_assignees:
                if r['url'] and not r['url'].startswith('http'):
                    r['url'] = PORTAIL_URL + r['url']
            logo_fichiers = conn.execute(
                "SELECT id, filename FROM projet_logo_fichiers WHERE id_projet=? ORDER BY created_at", (project_id,)
            ).fetchall()
            logos = [{
                'filename': f['filename'],
                'url': url_for('api_projet_download_logo', project_id=project_id, file_id=f['id'], _external=True),
            } for f in logo_fichiers]
            send_email_client(client,
                f"Votre projet est terminé — {projet['nom_projet']}",
                f"Bonjour {client['nom_complet']}, votre projet est terminé.",
                html=email_livraison(client['nom_complet'], projet['nom_projet'], lien, projet['lien_gdrive'] or None, ressources_assignees, projet['lien_site_test'] or None, logos))
            push_notification(conn, client['id'], project_id,
                f"Votre projet « {projet['nom_projet']} » est terminé !", type='termine')
            conn.commit()
    except Exception as e:
        print(f"[MAIL] _do_complete: {e}")
    return ressources_assignees


@app.route('/api/v1/admin/projet/<int:project_id>/start', methods=['POST'])
@admin_required
def api_admin_start_work(project_id):
    data = request.get_json(silent=True) or {}
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
    date_livraison = _do_start_travaux(conn, projet, client, date_livraison_override=data.get('date_livraison'))
    conn.close()
    return jsonify({'success': True, 'date_livraison': str(date_livraison) if date_livraison else None})


@app.route('/api/v1/admin/projet/<int:project_id>/revision', methods=['POST'])
@admin_required
def api_admin_start_revision(project_id):
    data = request.get_json() or {}
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    _do_start_revision(conn, projet, data.get('items', []))
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/item/<int:item_id>/resoudre', methods=['POST'])
@admin_required
def api_admin_toggle_item_resolu(item_id):
    """Coche/décoche « corrigé » côté admin sur un item de révision signalé par le
    client. Quand tous les items signalés d'un projet sont corrigés, ils sont remis à
    neuf et le projet retourne automatiquement en révision (voir
    _check_corrections_completed_auto_transition)."""
    conn = get_db_connection()
    item = conn.execute("SELECT * FROM checklist_items WHERE id = ?", (item_id,)).fetchone()
    if not item:
        conn.close()
        return jsonify({'error': 'Item introuvable'}), 404
    checklist = conn.execute("SELECT id_projet FROM checklistes WHERE id = ?", (item['id_checklist'],)).fetchone()
    new_val = 1 - int(item['admin_resolu'] or 0)
    conn.execute("UPDATE checklist_items SET admin_resolu = ? WHERE id = ?", (new_val, item_id))
    conn.commit()
    if new_val == 1 and int(item['is_revision'] or 0) == 1:
        _check_corrections_completed_auto_transition(conn, checklist['id_projet'])
    conn.close()
    return jsonify({'success': True, 'admin_resolu': bool(new_val)})


@app.route('/api/v1/admin/projet/<int:project_id>/complete', methods=['POST'])
@admin_required
def api_admin_complete_project(project_id):
    data = request.get_json(silent=True) or {}
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    ressources_assignees = _do_complete(conn, projet, data.get('ressource_ids', []))
    conn.close()
    return jsonify({'success': True, 'ressources_assignees': len(ressources_assignees)})


@app.route('/api/v1/admin/projet/<int:project_id>/force-status', methods=['POST'])
@admin_required
def api_admin_force_status(project_id):
    data = request.get_json() or {}
    statut = data.get('statut', '').strip()
    if not statut:
        return jsonify({'error': 'Statut requis'}), 400
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
    statut_norm = normalize_status(statut)
    nom_projet = projet['nom_projet']

    # Transitions connues du pipeline : déléguer aux mêmes fonctions que les boutons
    # dédiés — plus de divergence de comportement (facture/calendrier/courriel) selon
    # que l'admin utilise le bouton ou force le statut manuellement.
    if statut_norm == "Travaux en cours":
        _do_start_travaux(conn, projet, client, date_livraison_override=data.get('date_livraison'))
        conn.close()
        return jsonify({'success': True})
    if statut_norm == "En révision":
        _do_start_revision(conn, projet, [])
        conn.close()
        return jsonify({'success': True})
    if statut_norm in ("Complété", "Travaux terminés"):
        _do_complete(conn, projet, [])
        conn.close()
        return jsonify({'success': True})

    # Cas hors pipeline normal (reset, annulation) : mise à jour simple + courriel ponctuel.
    conn.execute("UPDATE projets SET statut=? WHERE id=?", (statut, project_id))
    conn.commit()
    try:
        if client and client['email']:
            lien = url_for('project_detail', project_id=project_id, _external=True)
            subject = body_txt = body_html = None
            if statut_norm == "Documents à donner":
                subject = f"Documents requis — {nom_projet}"
                body_txt = f"Bonjour {client['nom_complet']}, nous avons besoin de vos documents."
                body_html = email_documents_requis(client['nom_complet'], nom_projet, lien)
                push_notification(conn, client['id'], project_id, f"Documents requis pour votre projet « {nom_projet} ».", type='documents_requis')
            elif statut_norm == "Annulé":
                subject = f"Projet annulé — {nom_projet}"
                body_txt = f"Bonjour {client['nom_complet']}, votre projet a été annulé."
                body_html = email_annulation(client['nom_complet'], nom_projet)
            if subject and body_txt:
                send_email_client(client, subject, body_txt, html=body_html)
            conn.commit()
    except Exception as e:
        print(f"[MAIL] api_force_status: {e}")
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/projet/<int:project_id>/notifier-revision', methods=['POST'])
@admin_required
def api_admin_notifier_revision(project_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
    iv = conn.execute("SELECT is_complete FROM identite_visuelle WHERE id_projet=?", (project_id,)).fetchone()
    conn.close()
    try:
        if iv and int(iv['is_complete'] or 0) == 1:
            lien = url_for('projet_identite', project_id=project_id, _external=True)
            send_email_client(client, f"Votre identité visuelle est prête — {projet['nom_projet']}", f"Bonjour {client['nom_complet']}, votre identité visuelle est prête pour révision.", html=email_identite_visuelle_prete(client['nom_complet'], projet['nom_projet'], lien))
        else:
            lien = url_for('project_detail', project_id=project_id, _external=True)
            send_email_client(client, f"Votre projet est en révision — {projet['nom_projet']}", f"Bonjour {client['nom_complet']}, votre projet est en révision.", html=email_en_revision(client['nom_complet'], projet['nom_projet'], lien))
    except Exception as e:
        print(f"[MAIL] api_notifier_revision: {e}")
        return jsonify({'error': 'Erreur envoi email'}), 500
    return jsonify({'success': True})


@app.route('/api/v1/admin/projet/<int:project_id>/rappel-documents', methods=['POST'])
@admin_required
def api_admin_rappel_documents(project_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
    conn.close()
    if not client:
        return jsonify({'error': 'Client introuvable'}), 404
    try:
        lien = url_for('project_detail', project_id=project_id, _external=True)
        send_email_client(
            client,
            f"Rappel — Documents requis pour votre projet",
            f"Bonjour {client['nom_complet']}, nous attendons toujours vos documents pour le projet : {projet['nom_projet']}.",
            html=email_documents_requis(client['nom_complet'], projet['titre_affiche'] or projet['nom_projet'], lien)
        )
    except Exception as e:
        print(f"[MAIL] rappel_documents: {e}")
        return jsonify({'error': 'Erreur envoi email'}), 500
    return jsonify({'success': True})


@app.route('/api/v1/admin/projet/<int:project_id>/archive', methods=['POST'])
@admin_required
def api_admin_archive_project(project_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    conn.execute("UPDATE projets SET is_archived=1 WHERE id=?", (project_id,))
    conn.commit()
    push_notification(conn, projet['id_client'], project_id, f"Votre projet « {projet['nom_projet']} » a été archivé.", type='archive')
    conn.commit()
    try:
        client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
        if client and client['email'] and int(client['is_email_confirmed'] or 0):
            send_email(client['email'], f"Projet archivé — {projet['nom_projet']}", f"Bonjour {client['nom_complet']}, votre projet a été archivé.", html=email_archive(client['nom_complet'], projet['nom_projet']))
    except Exception as e:
        print(f"[MAIL] api_archive: {e}")
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/projet/<int:project_id>/unarchive', methods=['POST'])
@admin_required
def api_admin_unarchive_project(project_id):
    conn = get_db_connection()
    conn.execute("UPDATE projets SET is_archived=0 WHERE id=?", (project_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/projet/<int:project_id>/recreate-drive', methods=['POST'])
@admin_required
def api_admin_recreate_drive(project_id):
    conn = get_db_connection()
    try:
        p = conn.execute(
            'SELECT p.nom_projet, p.id_client, s.documents_requis as svc_docs, s.drive_subfolders FROM projets p LEFT JOIN services s ON s.id=p.id_service WHERE p.id=?',
            (project_id,)
        ).fetchone()
        if not p:
            return jsonify({'error': 'Projet introuvable'}), 404

        client = conn.execute('SELECT drive_folder_id FROM clients WHERE id=?', (p['id_client'],)).fetchone()
        parent = (client['drive_folder_id'] if client and client['drive_folder_id'] else None) or os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')

        folder_id = create_folder(p['nom_projet'], parent_id=parent)
        make_folder_public(folder_id)
        lien = get_folder_link(folder_id)

        docs_requis = bool(p['svc_docs'])
        depot_id = None
        if docs_requis:
            depot_id = create_folder('Dépôt de fichiers', parent_id=folder_id)
            make_folder_public(depot_id)

        if p['drive_subfolders']:
            for nom_sf in p['drive_subfolders'].split('|'):
                nom_sf = nom_sf.strip()
                if nom_sf:
                    sf_id = create_folder(nom_sf, parent_id=folder_id)
                    make_folder_public(sf_id)

        conn.execute(
            'UPDATE projets SET lien_gdrive=?, drive_folder_id=?, depot_folder_id=? WHERE id=?',
            (lien, folder_id, depot_id, project_id)
        )
        conn.commit()
        return jsonify({'success': True, 'lien_gdrive': lien, 'drive_folder_id': folder_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:project_id>/checklist', methods=['PUT'])
@admin_required
def api_admin_edit_checklist_items(project_id):
    data = request.get_json() or {}
    items = data.get('items', [])
    conn = get_db_connection()
    checklist = conn.execute("SELECT id FROM checklistes WHERE id_projet=?", (project_id,)).fetchone()
    if not checklist:
        conn.close()
        return jsonify({'error': 'Checklist introuvable'}), 404
    TYPE_MAP = {
        'photo':      ('document', 'photo',    1),
        'vecteur':    ('document', 'vecteur',  1),
        'video_file': ('document', 'video',    1),
        'document':   ('document', 'document', 1),
        'donnees':    ('document', 'donnees',  1),
        'archive':    ('document', 'archive',  1),
        'video_url':  ('video',    'autre',    0),
        'autre':      ('document', 'autre',    0),
    }
    for item in items:
        iid = item.get('id')
        nom = str(item.get('nom', '')).strip()
        type_unified = item.get('type', 'document')
        is_required = 1 if item.get('is_required') else 0
        item_type, file_category, requires_file = TYPE_MAP.get(type_unified, ('document', 'autre', 0))
        if nom and iid:
            conn.execute("""
                UPDATE checklist_items
                SET nom_item=?, item_type=?, file_category=?, requires_file=?, is_required=?
                WHERE id=? AND id_checklist=?
            """, (nom, item_type, file_category, requires_file, is_required, iid, checklist['id']))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ───────────────────────────────────────────────────────────
# API v1 — Auth confirm email + reset password
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/auth/confirm-email/<token>', methods=['GET'])
def api_confirm_email(token):
    try:
        email = s.loads(token, salt='email-confirm-salt', max_age=604800)  # 7 jours
    except (SignatureExpired, BadSignature):
        # Signature expirée ou invalide — vérifier quand même si le compte est déjà actif
        # (cas: scanner de sécurité qui a déjà consommé le lien, ou token expiré après confirmation manuelle)
        try:
            is_valid, maybe_email = s.loads_unsafe(token, salt='email-confirm-salt')
            if is_valid and maybe_email:
                conn = get_db_connection()
                u = conn.execute("SELECT is_email_confirmed FROM clients WHERE email = ?", (maybe_email,)).fetchone()
                conn.close()
                if u and int(u['is_email_confirmed'] or 0):
                    return jsonify({'error': 'already_confirmed'}), 400
        except Exception:
            pass
        return jsonify({'error': 'expired'}), 400

    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    if not user:
        conn.close()
        return jsonify({'error': 'Compte introuvable'}), 404

    if int(user['is_email_confirmed'] or 0):
        # Déjà confirmé (y compris si un scanner a déjà consommé ce lien)
        conn.execute("UPDATE clients SET confirm_token = NULL WHERE email = ?", (email,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'already_confirmed': True})

    # Vérification usage unique : le token doit correspondre au dernier envoyé
    stored = user['confirm_token'] if 'confirm_token' in user.keys() else None
    if stored is not None and stored != token:
        conn.close()
        return jsonify({'error': 'expired'}), 400

    # Compte créé par l'admin (pas de hash) → invitation flow uniquement
    if not user['mot_de_passe_hash']:
        conn.close()
        return jsonify({'error': 'Votre compte a été créé par un administrateur. Utilisez le lien d\'invitation reçu par courriel pour activer votre accès.'}), 400

    conn.execute("UPDATE clients SET is_email_confirmed = 1, confirm_token = NULL WHERE email = ?", (email,))
    conn.commit()
    conn.close()

    try:
        send_email(email, "Bienvenue chez Cocktail Média !",
                   f"Bonjour {user['nom_complet']}, votre compte est maintenant actif.",
                   html=email_bienvenue(user['nom_complet'], email))
    except Exception as e:
        print(f"[MAIL] Bienvenue échoué: {e}")
    try:
        lien_admin = f"{PORTAIL_URL}/admin/client/{user['id']}"
        send_email(
            'felix.dumont@cocktailmedia.ca',
            f"Nouveau client inscrit — {user['nom_complet']}",
            f"Nouveau client : {user['nom_complet']} ({email}) — {user['nom_entreprise'] or 'sans entreprise'}",
            html=email_nouveau_client(user['nom_complet'], email, user['nom_entreprise'], lien_admin)
        )
        conn2 = get_db_connection()
        push_admin_notif(
            conn2,
            titre=f"Nouveau client inscrit — {user['nom_complet']}",
            message=f"{user['nom_complet']} ({email}) — {user['nom_entreprise'] or 'sans entreprise'}",
            type='info',
            lien=f"/admin/client/{user['id']}",
        )
        conn2.commit()
        conn2.close()
    except Exception as e:
        print(f"[MAIL] Notif admin échouée: {e}")
    return jsonify({'success': True})

@app.route('/api/v1/auth/forgot-password', methods=['POST'])
@limiter.limit("10 per minute")
def api_forgot_password():
    data = request.get_json() or {}
    email = data.get('email', '').strip().lower()
    if not email:
        return jsonify({'error': 'Email requis'}), 400
    if not _redis_rate_limit(f"reset_rate:{email}", 900):
        return jsonify({'success': True})
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    conn.close()
    if user:
        try:
            token = s.dumps([email, user['mot_de_passe_hash']], salt='password-reset-salt')
            long_reset_url = f"{PORTAIL_URL}/reset-password?token={token}"
            short_url = create_short_link(long_reset_url, expires_in=3600)
            nom = user['nom_complet']
            html = email_reset_password(nom, short_url)
            send_email(
                email,
                "Réinitialisation de votre mot de passe — Portail Client",
                f"Bonjour {nom}, réinitialisez votre mot de passe : {short_url}",
                html=html
            )
        except Exception as e:
            print(f"[MAIL] Reset API échoué: {e}")
    return jsonify({'success': True})


@app.route('/api/v1/auth/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.get_json() or {}
    new_password = data.get('password', '')
    current_password = data.get('current_password', '')
    if not new_password:
        return jsonify({'error': 'Mot de passe requis'}), 400
    if not is_password_strong(new_password):
        return jsonify({'error': 'Mot de passe trop faible — min. 8 caractères, majuscule, minuscule, chiffre et caractère spécial'}), 400
    user_id = session['user_id']
    conn = get_db_connection()
    user = conn.execute("SELECT mot_de_passe_hash, must_change_password FROM clients WHERE id = ?", (user_id,)).fetchone()
    if not user:
        conn.close()
        return jsonify({'error': 'Compte introuvable'}), 404
    force_change = bool(user['must_change_password']) if 'must_change_password' in user.keys() else False
    if not force_change:
        if not current_password:
            conn.close()
            return jsonify({'error': 'Mot de passe actuel requis'}), 400
        if not user['mot_de_passe_hash'] or not bcrypt.check_password_hash(user['mot_de_passe_hash'], current_password):
            conn.close()
            return jsonify({'error': 'Mot de passe actuel incorrect'}), 403
    hashed = bcrypt.generate_password_hash(new_password).decode('utf-8')
    conn.execute("UPDATE clients SET mot_de_passe_hash = ?, must_change_password = 0 WHERE id = ?", (hashed, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/client/<int:client_id>/set-temp-password', methods=['POST'])
@admin_required
def api_admin_set_temp_password(client_id):
    data = request.get_json() or {}
    temp_password = data.get('password', '').strip()
    if not temp_password:
        return jsonify({'error': 'Mot de passe requis'}), 400
    if not is_password_strong(temp_password):
        return jsonify({'error': 'Mot de passe trop faible — min. 8 caractères, majuscule, minuscule, chiffre et caractère spécial (!@#$%^&*)'}), 400
    conn = get_db_connection()
    client = conn.execute("SELECT id, email, is_email_confirmed FROM clients WHERE id = ?", (client_id,)).fetchone()
    if not client:
        conn.close()
        return jsonify({'error': 'Client introuvable'}), 404
    hashed = bcrypt.generate_password_hash(temp_password).decode('utf-8')
    conn.execute("""
        UPDATE clients SET mot_de_passe_hash = ?, must_change_password = 1, is_email_confirmed = 1
        WHERE id = ?
    """, (hashed, client_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/auth/resend-confirmation', methods=['POST'])
@limiter.limit("10 per minute")
def api_resend_confirmation():
    data = request.get_json() or {}
    email = data.get('email', '').strip().lower()
    if not email:
        return jsonify({'error': 'Email requis'}), 400
    if not _redis_rate_limit(f"resend_confirm:{email}", 300):
        return jsonify({'success': True})
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    conn.close()
    if user and not int(user['is_email_confirmed'] or 0) and user['mot_de_passe_hash']:
        try:
            new_token = s.dumps(email, salt='email-confirm-salt')
            conn2 = get_db_connection()
            conn2.execute("UPDATE clients SET confirm_token = ? WHERE email = ?", (new_token, email))
            conn2.commit()
            conn2.close()
            confirm_url = f"{PORTAIL_URL}/confirm-email?token={new_token}"
            html_confirm = _base_confirm(user['nom_complet'], confirm_url)
            send_email(email, "Confirmez votre compte — Cocktail Média",
                       f"Bonjour {user['nom_complet']}, confirmez votre compte : {confirm_url}",
                       html=html_confirm)
        except Exception as e:
            print(f"[MAIL] Resend confirmation échoué: {e}")
    return jsonify({'success': True})

@app.route('/api/v1/auth/reset-password', methods=['POST'])
def api_reset_password():
    data = request.get_json() or {}
    token = data.get('token')
    password = data.get('password')
    if not token or not password:
        return jsonify({'error': 'Données manquantes'}), 400
    if not is_password_strong(password):
        return jsonify({'error': 'Mot de passe trop faible'}), 400
    try:
        payload = s.loads(token, salt='password-reset-salt', max_age=3600)
        email, token_hash = payload
    except (SignatureExpired, BadTimeSignature, ValueError, TypeError):
        return jsonify({'error': 'Lien invalide ou expiré'}), 400
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    if not user or user['mot_de_passe_hash'] != token_hash:
        conn.close()
        return jsonify({'error': 'Ce lien a déjà été utilisé ou est invalide'}), 400
    hashed = bcrypt.generate_password_hash(password).decode('utf-8')
    conn.execute(
        'UPDATE clients SET mot_de_passe_hash = ?, is_email_confirmed = 1 WHERE email = ?', (hashed, email)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/auth/invitation-info/<token>', methods=['GET'])
def api_invitation_info(token):
    try:
        email = s.loads(token, salt='invitation-client-salt', max_age=604800)
    except (SignatureExpired, BadSignature):
        return jsonify({'error': 'Lien invalide ou expiré'}), 400
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    conn.close()
    if not user:
        return jsonify({'error': 'Compte introuvable'}), 404
    if int(user['is_email_confirmed'] or 0) and bool(user['mot_de_passe_hash']):
        return jsonify({'error': 'Ce lien a déjà été utilisé'}), 409
    stored_token = user['confirm_token'] if 'confirm_token' in user.keys() else None
    if stored_token and stored_token != token:
        return jsonify({'error': 'Lien invalide ou expiré'}), 400
    return jsonify({
        'email': email,
        'nom': user['nom_complet'],
        'nom_entreprise': user['nom_entreprise'] or '',
        'telephone': user['telephone'] or '',
        'adresse_facturation': user['adresse_facturation'] or '',
        'ville_facturation': user['ville_facturation'] or '',
        'province_facturation': user['province_facturation'] or 'Québec',
        'code_postal_facturation': user['code_postal_facturation'] or '',
    })


@app.route('/api/v1/auth/accept-invitation', methods=['POST'])
def api_accept_invitation():
    data = request.get_json() or {}
    token = data.get('token', '')
    password = data.get('password', '')
    if not token or not password:
        return jsonify({'error': 'Données manquantes'}), 400
    if not is_password_strong(password):
        return jsonify({'error': 'Mot de passe trop faible'}), 400
    try:
        email = s.loads(token, salt='invitation-client-salt', max_age=604800)
    except (SignatureExpired, BadSignature):
        return jsonify({'error': 'Lien invalide ou expiré'}), 400
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    if not user:
        conn.close()
        return jsonify({'error': 'Compte introuvable'}), 404
    if int(user['is_email_confirmed'] or 0) and bool(user['mot_de_passe_hash']):
        conn.close()
        return jsonify({'error': 'Ce lien a déjà été utilisé'}), 409
    stored_token = user['confirm_token'] if 'confirm_token' in user.keys() else None
    if stored_token and stored_token != token:
        conn.close()
        return jsonify({'error': 'Lien invalide ou expiré'}), 400
    nom_entreprise = data.get('nom_entreprise', '').strip() or None
    telephone = data.get('telephone', '').strip() or None
    adresse = data.get('adresse_facturation', '').strip() or None
    ville = data.get('ville_facturation', '').strip() or None
    province = data.get('province_facturation', 'Québec').strip() or 'Québec'
    code_postal = data.get('code_postal_facturation', '').strip() or None
    hashed = bcrypt.generate_password_hash(password).decode('utf-8')
    conn.execute("""
        UPDATE clients SET mot_de_passe_hash=?, is_email_confirmed=1, must_change_password=0, confirm_token=NULL,
        nom_entreprise=COALESCE(?, nom_entreprise),
        telephone=COALESCE(?, telephone),
        adresse_facturation=?, ville_facturation=?,
        province_facturation=?, code_postal_facturation=?
        WHERE email=?
    """, (hashed, nom_entreprise, telephone, adresse, ville, province, code_postal, email))
    conn.commit()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    conn.close()
    session.clear()
    session.permanent = True
    session['user_id'] = user['id']
    session['user_name'] = user['nom_complet']
    session['is_admin'] = bool(user['is_admin'])
    session['has_outils'] = bool(user['has_outils']) if 'has_outils' in user.keys() else False
    session['has_entrainement'] = bool(user['has_entrainement']) if 'has_entrainement' in user.keys() else False

    # Emails bienvenue + notif admin : non-bloquants
    _nom = user['nom_complet']
    _entreprise = user['nom_entreprise']
    _uid = user['id']

    def _post_invitation_emails():
        # send_email (Flask-Mail) a besoin d'un contexte applicatif — sans ce `with`,
        # l'envoi échoue silencieusement ("Working outside of application context").
        with app.app_context():
            try:
                send_email(email, "Bienvenue chez Cocktail Média !",
                           f"Bonjour {_nom}, votre compte est maintenant actif.",
                           html=email_bienvenue(_nom, email))
            except Exception as e:
                print(f"[MAIL] Bienvenue invitation échoué: {e}")
            try:
                lien_admin = f"{PORTAIL_URL}/admin/client/{_uid}"
                send_email(
                    'felix.dumont@cocktailmedia.ca',
                    f"Nouveau client activé — {_nom}",
                    f"Invitation acceptée : {_nom} ({email}) — {_entreprise or 'sans entreprise'}",
                    html=email_nouveau_client(_nom, email, _entreprise, lien_admin)
                )
                conn2 = get_db_connection()
                push_admin_notif(
                    conn2,
                    titre=f"Nouveau client activé — {_nom}",
                    message=f"Invitation acceptée : {_nom} ({email}) — {_entreprise or 'sans entreprise'}",
                    type='info',
                    lien=f"/admin/client/{_uid}",
                )
                conn2.commit()
                conn2.close()
            except Exception as e:
                print(f"[MAIL] Notif admin invitation échouée: {e}")

    import threading
    threading.Thread(target=_post_invitation_emails, daemon=True).start()
    return jsonify({'success': True, 'nom': user['nom_complet']})

# ───────────────────────────────────────────────────────────
# API v1 — Profile client
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/profile', methods=['GET'])
@login_required
def api_profile():
    user_id = session['user_id']
    conn = get_db_connection()
    client = conn.execute(
        'SELECT * FROM clients WHERE id = ?', (user_id,)
    ).fetchone()
    projets = conn.execute("""
        SELECT p.*, s.nom_service
        FROM projets p
        LEFT JOIN services s ON s.id = p.id_service
        WHERE p.id_client = ?
        ORDER BY p.created_at DESC
    """, (user_id,)).fetchall()
    factures = conn.execute(
        "SELECT * FROM factures WHERE id_client = ? AND statut != 'annulee' ORDER BY date_emission DESC",
        (user_id,)
    ).fetchall()
    conn.close()
    return jsonify({
        'id': client['id'],
        'nom_complet': client['nom_complet'],
        'email': client['email'],
        'nom_entreprise': client['nom_entreprise'],
        'telephone': client['telephone'],
        'adresse_facturation': client['adresse_facturation'],
        'ville_facturation': client['ville_facturation'],
        'province_facturation': client['province_facturation'],
        'code_postal_facturation': client['code_postal_facturation'],
        'logo_url': client['logo_url'],
        'favicon_url': client['favicon_url'],
        'couleur_primaire': client['couleur_primaire'],
        'couleur_secondaire': client['couleur_secondaire'],
        'projets': [{
            'id': p['id'],
            'nom_projet': p['nom_projet'],
            'statut': p['statut'],
            'nom_service': p['nom_service'],
            'date_livraison_estimee': p['date_livraison_estimee'],
            'is_archived': p['is_archived'],
        } for p in projets],
        'factures': [{
            'id': f['id'],
            'numero': f['numero'],
            'statut': f['statut'],
            'total': f['total'],
            'date_emission': f['date_emission'],
            'date_echeance': f['date_echeance'],
            'stripe_payment_url': f['stripe_payment_url'],
        } for f in factures],
    })

@app.route('/api/v1/profile/update', methods=['POST'])
@login_required
def api_profile_update():
    data = request.get_json() or {}
    user_id = session['user_id']
    conn = get_db_connection()
    conn.execute("""
        UPDATE clients SET
            nom_complet = ?,
            telephone = ?,
            nom_entreprise = ?,
            adresse_facturation = ?,
            ville_facturation = ?,
            province_facturation = ?,
            code_postal_facturation = ?
        WHERE id = ?
    """, (
        data.get('nom_complet'),
        data.get('telephone'),
        data.get('nom_entreprise'),
        data.get('adresse_facturation'),
        data.get('ville_facturation'),
        data.get('province_facturation'),
        data.get('code_postal_facturation'),
        user_id
    ))
    conn.commit()
    conn.close()
    session['user_name'] = data.get('nom_complet', session['user_name'])
    return jsonify({'success': True})


@app.route('/api/v1/profile/brand', methods=['POST'])
@login_required
def api_profile_brand():
    data = request.get_json() or {}
    user_id = session['user_id']
    conn = get_db_connection()
    conn.execute("""
        UPDATE clients SET couleur_primaire = ?, couleur_secondaire = ? WHERE id = ?
    """, (data.get('couleur_primaire'), data.get('couleur_secondaire'), user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/profile/upload-asset', methods=['POST'])
@login_required
def api_profile_upload_asset():
    user_id = session['user_id']
    asset_type = request.form.get('type', 'document')
    if asset_type not in ('logo', 'favicon', 'document'):
        return jsonify({'error': 'Type invalide'}), 400
    file = request.files.get('file')
    if not file or file.filename == '':
        return jsonify({'error': 'Fichier requis'}), 400

    conn = get_db_connection()
    client = conn.execute("SELECT drive_folder_id, nom_entreprise, nom_complet, email FROM clients WHERE id = ?", (user_id,)).fetchone()
    drive_folder_id = client['drive_folder_id'] if client else None
    if not drive_folder_id:
        nom_dossier = client['nom_entreprise'] or client['nom_complet']
        drive_folder_id = create_folder(nom_dossier, parent_id=os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID'))
        conn.execute("UPDATE clients SET drive_folder_id = ? WHERE id = ?", (drive_folder_id, user_id))
        conn.commit()
    if client and client['email']:
        try:
            share_folder_with_user(drive_folder_id, client['email'])
        except Exception:
            pass

    try:
        import tempfile
        marque_folder_id = create_folder("Marque", parent_id=drive_folder_id)
        safe_name = secure_filename(file.filename)
        ext = safe_name.rsplit('.', 1)[-1].lower() if '.' in safe_name else 'bin'
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
        file.save(tmp.name)
        tmp.close()
        file_id, _ = upload_file(tmp.name, safe_name, marque_folder_id)
        os.unlink(tmp.name)
        public_url = make_file_public(file_id)
    except Exception as e:
        conn.close()
        print(f"[DRIVE] Upload asset profil échoué: {e}")
        return jsonify({'error': "Échec de l'envoi du fichier"}), 500

    if asset_type == 'logo':
        conn.execute("UPDATE clients SET logo_url = ? WHERE id = ?", (public_url, user_id))
        conn.commit()
    elif asset_type == 'favicon':
        conn.execute("UPDATE clients SET favicon_url = ? WHERE id = ?", (public_url, user_id))
        conn.commit()
    conn.close()
    return jsonify({'success': True, 'url': public_url})


# ───────────────────────────────────────────────────────────
# API v1 — Projet détail client
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/projet/<int:projet_id>', methods=['GET'])
@login_required
def api_projet_detail(projet_id):
    user_id = session['user_id']
    conn = get_db_connection()
    p = conn.execute("""
        SELECT p.*, s.nom_service, s.icon as service_icon,
               s.documents_requis as svc_documents_requis,
               s.appel_exploratoire_requis as svc_appel_exploratoire_requis
        FROM projets p
        LEFT JOIN services s ON s.id = p.id_service
        WHERE p.id = ? AND p.id_client = ?
    """, (projet_id, user_id)).fetchone()
    if not p:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    checklist = conn.execute(
        'SELECT * FROM checklistes WHERE id_projet = ?', (projet_id,)
    ).fetchone()
    items = []
    if checklist:
        items = conn.execute(
            'SELECT * FROM checklist_items WHERE id_checklist = ? ORDER BY position',
            (checklist['id'],)
        ).fetchall()
    dossiers = []
    if p['drive_folder_id']:
        try:
            from drive_service import get_drive_service
            drive = get_drive_service()
            result = drive.files().list(
                q=f"'{p['drive_folder_id']}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                fields='files(id,name,webViewLink)'
            ).execute()
            dossiers = result.get('files', [])
        except Exception:
            dossiers = []
    iv = conn.execute(
        "SELECT id FROM identite_visuelle WHERE id_projet = ? AND is_complete = 1",
        (projet_id,)
    ).fetchone()
    board = conn.execute(
        "SELECT id FROM decision_boards WHERE id_projet = ? AND is_active = 1",
        (projet_id,)
    ).fetchone()
    logo_fichiers = conn.execute(
        "SELECT id, filename FROM projet_logo_fichiers WHERE id_projet=? ORDER BY created_at", (projet_id,)
    ).fetchall()
    conn.close()
    pipeline_steps = pipeline_for_service({
        'documents_requis': p['svc_documents_requis'] if p['svc_documents_requis'] is not None else 1,
        'appel_exploratoire_requis': p['svc_appel_exploratoire_requis'] or 0,
    })
    pipeline_steps = pipeline_with_current_statut(pipeline_steps, p['statut'])
    return jsonify({
        'id': p['id'],
        'nom_projet': p['nom_projet'],
        'statut': p['statut'],
        'pipeline_steps': pipeline_steps,
        'progress_pct': phase_progress(p['statut']),
        'is_archived': bool(p['is_archived']),
        'lien_gdrive': p['lien_gdrive'],
        'lien_site_test': p['lien_site_test'],
        'drive_folder_id': p['drive_folder_id'],
        'date_livraison_estimee': p['date_livraison_estimee'],
        'nom_service': p['nom_service'],
        'service_icon': p['service_icon'],
        'has_identite_visuelle': iv is not None,
        'has_decision_board': board is not None,
        'logo_fichiers': [{'id': f['id'], 'filename': f['filename']} for f in logo_fichiers],
        'items': [{
            'id': i['id'],
            'nom_item': i['nom_item'],
            'est_coche': bool(i['est_coche']),
            'requires_file': bool(i['requires_file']),
            'is_required': bool(i['is_required']),
            'is_revision': bool(i['is_revision']),
            'file_path': i['file_path'],
            'item_type': i['item_type'],
            'video_url': i['video_url'],
            'field_type': i['field_type'] if i['field_type'] else 'check',
            'text_value': i['text_value'],
        } for i in items],
        'dossiers_drive': dossiers,
    })


@app.route('/api/v1/projet/<int:projet_id>/identite', methods=['GET'])
@login_required
def api_identite_visuelle_client(projet_id):
    user_id = session['user_id']
    conn = get_db_connection()
    projet = conn.execute(
        "SELECT * FROM projets WHERE id = ? AND id_client = ?", (projet_id, user_id)
    ).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    iv = conn.execute(
        "SELECT * FROM identite_visuelle WHERE id_projet = ? AND is_complete = 1", (projet_id,)
    ).fetchone()
    if not iv:
        conn.close()
        return jsonify({'error': 'Identité visuelle non disponible'}), 404
    logos_rows = conn.execute("SELECT * FROM iv_logos WHERE id_iv = ?", (iv['id'],)).fetchall()
    fonts = conn.execute("SELECT * FROM iv_fonts WHERE id_iv = ?", (iv['id'],)).fetchall()
    declinaisons = conn.execute(
        "SELECT * FROM iv_declinaisons WHERE id_iv = ? ORDER BY position, id", (iv['id'],)
    ).fetchall()
    mockups = conn.execute(
        "SELECT * FROM iv_mockups WHERE id_iv = ? ORDER BY position, id", (iv['id'],)
    ).fetchall()
    conn.close()
    logos = {}
    for l in logos_rows:
        logos[l['variante']] = {
            'public_url': l['public_url'],
            'preview_url': l['preview_url'] if 'preview_url' in l.keys() else None,
            'filename': l['filename'],
        }
    import json as _json
    palette = []
    if iv['palette_json']:
        try:
            palette = _json.loads(iv['palette_json'])
        except Exception:
            palette = []
    return jsonify({
        'id': iv['id'],
        'contexte': iv['contexte'] if 'contexte' in iv.keys() else None,
        'nom_projet': projet['nom_projet'],
        'palette': palette,
        'logos': logos,
        'fonts': [{'nom_font': f['nom_font'], 'google_font_url': f['google_font_url'], 'usage': f['usage']} for f in fonts],
        'declinaisons': [{'id': d['id'], 'public_url': d['public_url'], 'label': d['label'], 'filename': d['filename']} for d in declinaisons],
        'mockups': [{'id': m['id'], 'public_url': m['public_url'], 'label': m['label'], 'filename': m['filename']} for m in mockups],
        'zip_url': f'/projet/{projet_id}/identite/zip',
    })


# ───────────────────────────────────────────────────────────
# Routes Auth
# ───────────────────────────────────────────────────────────
@app.route('/')
def accueil():
    return render_template('login.html')

@app.route('/connexion')
def connexion():
    return render_template('login.html')

@app.route('/login', methods=['GET','POST'])
@limiter.limit("5 per minute")
def login():
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower()
        password = request.form.get('password','')

        conn = get_db_connection()
        user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
        conn.close()

        if user and user['auth_provider'] == 'password' and bcrypt.check_password_hash(user['mot_de_passe_hash'], password):
            if not int(user['is_email_confirmed'] or 0):
                flash("Veuillez confirmer votre adresse email avant de vous connecter.", "error")
                return redirect(url_for('accueil'))
            session['user_id'] = user['id']
            session['user_name'] = user['nom_complet']
            session['is_admin'] = bool(user['is_admin'])
            return redirect(url_for('admin_dashboard' if session['is_admin'] else 'dashboard'))

        flash("Email ou mot de passe incorrect.", "error")
        return redirect(url_for('accueil'))
    return render_template('login.html')

@app.route('/register', methods=['GET','POST'])
@limiter.limit("5 per minute")
def register():
    if request.method == 'POST':
        nom = request.form.get('nom_complet','').strip()
        email = request.form.get('email','').strip().lower()
        password = request.form.get('password','')
        password2 = request.form.get('password2','')
        nom_entreprise = request.form.get('nom_entreprise','').strip()
        telephone = request.form.get('telephone','').strip()

        if not nom_entreprise:
            flash("Le nom d'entreprise est obligatoire. Pour un travailleur autonome, inscrivez votre nom complet.", "error")
            return redirect(url_for('register'))

        if password != password2:
            flash('Les mots de passe ne correspondent pas.', 'error')
            return redirect(url_for('register'))
        if not is_password_strong(password):
            flash('Le mot de passe doit contenir au moins 8 caractères, une majuscule, une minuscule, un chiffre et un caractère spécial.', 'error')
            return redirect(url_for('register'))

        conn = get_db_connection()
        exists = conn.execute("SELECT is_email_confirmed FROM clients WHERE email = ?", (email,)).fetchone()
        if exists:
            conn.close()
            if not int(exists['is_email_confirmed'] or 0):
                # Compte non confirmé déjà existant — ne pas renvoyer de confirmation pour éviter le relay d'emails
                flash("Un email de confirmation a déjà été envoyé à cette adresse.", "info")
            else:
                flash("Cette adresse email est déjà utilisée.", "error")
            return redirect(url_for('register'))

        match = find_matching_company(nom_entreprise, conn)
        # S'assurer que le match a bien un dossier Drive existant
        drive_folder_id = match["drive_folder_id"] if match and match["drive_folder_id"] else None
        hashed = bcrypt.generate_password_hash(password).decode('utf-8')
        conn.execute("""
            INSERT INTO clients (nom_complet, email, nom_entreprise, telephone, mot_de_passe_hash, auth_provider, is_email_confirmed, is_admin, drive_folder_id)
            VALUES (?, ?, ?, ?, ?, 'password', 0, 0, ?)
        """, (nom, email, nom_entreprise, telephone, hashed, drive_folder_id))
        conn.commit()

        if not drive_folder_id:
            try:
                new_folder_id = create_folder(
                    nom_entreprise,
                    parent_id=os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
                )
                factures_folder_id = create_folder("Factures", parent_id=new_folder_id)
                conn.execute("UPDATE clients SET drive_folder_id = ?, factures_folder_id = ? WHERE email = ?", (new_folder_id, factures_folder_id, email))

                conn.commit()
                # Mettre à jour tous les autres comptes de la même entreprise sans dossier
                conn.execute("""
                    UPDATE clients SET drive_folder_id = ?
                    WHERE drive_folder_id IS NULL AND nom_entreprise = ? AND email != ?
                """, (new_folder_id, nom_entreprise, email))
                conn.commit()
            except Exception as e:
                print(f"[DRIVE] Création dossier client échouée: {e}")
        conn.close()

        try:
            token = s.dumps(email, salt='email-confirm-salt')
            confirm_url = url_for('confirm_email', token=token, _external=True)
            html_confirm = _base_confirm(nom, confirm_url)
            send_email(
                email,
                "Confirmez votre compte — Cocktail Média",
                f"Bonjour {nom}, veuillez confirmer votre compte : {confirm_url}",
                html=html_confirm
            )
        except Exception as e:
            print(f"[MAIL] Confirmation échouée: {e}")
        return render_template('register.html', show_success_popup=True)

    return render_template('register.html', show_success_popup=False)
@app.route('/confirm/<token>')
def confirm_email(token):
    try:
        email = s.loads(token, salt='email-confirm-salt', max_age=3600)
    except (SignatureExpired, BadSignature):
        flash("Le lien de confirmation est invalide ou a expiré.", "error")
        return redirect(url_for('accueil'))

    conn = get_db_connection()
    conn.execute("UPDATE clients SET is_email_confirmed = 1 WHERE email = ?", (email,))
    conn.commit()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    conn.close()
    try:
        send_email(
            email,
            "Bienvenue chez Cocktail Média !",
            f"Bonjour {user['nom_complet']}, votre compte est maintenant actif.",
            html=email_bienvenue(user['nom_complet'], email)
        )
    except Exception as e:
        print(f"[MAIL] Email bienvenue échoué: {e}")
    try:
        lien_admin = url_for('admin_dashboard', _external=True)
        send_email(
            ["marie-christine.blanchette@cocktailmedia.ca", "felix.dumont@cocktailmedia.ca"],
            f"Nouveau client inscrit — {user['nom_complet']}",
            f"Nouveau client : {user['nom_complet']} ({email}) — {user['nom_entreprise'] or 'sans entreprise'}",
            html=email_nouveau_client(user['nom_complet'], email, user['nom_entreprise'], lien_admin)
        )
    except Exception as e:
        print(f"[MAIL] Notif nouveau client échouée: {e}")
    flash("Votre compte a été confirmé avec succès !", "success")
    return redirect(url_for('accueil'))
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('accueil'))

@app.route('/forgot-password', methods=['GET','POST'])
@limiter.limit("5 per minute")
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower()
        if not _redis_rate_limit(f"reset_rate:{email}", 900):
            flash("Si votre adresse email est dans notre système, vous recevrez un lien de réinitialisation.", "success")
            return redirect(url_for('forgot_password'))
        conn = get_db_connection()
        user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
        conn.close()
        if user:
            try:
                token = s.dumps([email, user['mot_de_passe_hash']], salt='password-reset-salt')
                long_reset_url = f"{PORTAIL_URL}/reset-password?token={token}"
                short_url = create_short_link(long_reset_url, expires_in=3600)
                body = (
                    f"Bonjour {user['nom_complet']},\n\n"
                    f"Pour réinitialiser votre mot de passe, cliquez (valide 1 heure) :\n{short_url}\n\n"
                    f"Si vous n'avez pas demandé cette réinitialisation, ignorez cet email."
                )
                send_email(email, "Réinitialisation de votre mot de passe - Portail Client", body)
            except Exception as e:
                print(f"[MAIL] Reset échoué: {e}")
        flash("Si votre adresse email est dans notre système, vous recevrez un lien de réinitialisation.", "success")
        return redirect(url_for('forgot_password'))
    return render_template('forgot_password.html')

@app.route('/invitation/<token>', methods=['GET','POST'])
def accepter_invitation(token):
    try:
        email = s.loads(token, salt='invitation-client-salt', max_age=604800)  # 7 jours
    except (SignatureExpired, BadSignature):
        flash("Ce lien d'invitation est invalide ou a expiré.", "error")
        return redirect(url_for('accueil'))

    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    if not user:
        conn.close()
        flash("Compte introuvable.", "error")
        return redirect(url_for('accueil'))

    if int(user['is_email_confirmed'] or 0):
        conn.close()
        flash("Ce lien d'invitation a déjà été utilisé.", "error")
        return redirect(url_for('accueil'))

    if request.method == 'POST':
        password = request.form.get('password','')
        confirm = request.form.get('confirm_password','')
        if password != confirm:
            conn.close()
            flash("Les mots de passe ne correspondent pas.", "error")
            return render_template('invitation.html', email=email, token=token)
        if not is_password_strong(password):
            conn.close()
            flash("Le mot de passe doit contenir au moins 8 caractères, une majuscule, une minuscule, un chiffre et un caractère spécial.", "error")
            return render_template('invitation.html', email=email, token=token)
        nom_entreprise = request.form.get('nom_entreprise','').strip() or None
        telephone = request.form.get('telephone','').strip() or None
        adresse = request.form.get('adresse_facturation','').strip() or None
        ville = request.form.get('ville_facturation','').strip() or None
        province = request.form.get('province_facturation','Québec').strip()
        code_postal = request.form.get('code_postal_facturation','').strip() or None

        hashed = bcrypt.generate_password_hash(password).decode('utf-8')
        conn.execute("""
            UPDATE clients SET mot_de_passe_hash=?, is_email_confirmed=1,
            nom_entreprise=COALESCE(?, nom_entreprise),
            telephone=COALESCE(?, telephone),
            adresse_facturation=?, ville_facturation=?,
            province_facturation=?, code_postal_facturation=?
            WHERE email=?
        """, (hashed, nom_entreprise, telephone, adresse, ville, province, code_postal, email))

        conn.commit()
        conn.close()
        flash("Votre compte est activé ! Vous pouvez maintenant vous connecter.", "success")
        return redirect(url_for('accueil'))

    conn.close()
    return render_template('invitation.html', email=email, nom=user['nom_complet'], token=token)

@app.route('/reset-password/<token>', methods=['GET','POST'])
def reset_password(token):
    try:
        payload = s.loads(token, salt='password-reset-salt', max_age=3600)
        email, token_hash = payload
    except (SignatureExpired, BadTimeSignature, ValueError, TypeError):
        flash("Le lien de réinitialisation est invalide ou a expiré.", "error")
        return redirect(url_for('forgot_password'))

    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    conn.close()
    if not user or user['mot_de_passe_hash'] != token_hash:
        flash("Ce lien a déjà été utilisé ou est invalide.", "error")
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        password = request.form.get('password','')
        confirm_password = request.form.get('confirm_password','')
        if password != confirm_password:
            flash('Les mots de passe ne correspondent pas.', 'error')
            return redirect(url_for('reset_password', token=token))
        if not is_password_strong(password):
            flash('Le mot de passe doit contenir au moins 8 caractères, une majuscule, une minuscule, un chiffre et un caractère spécial.', 'error')
            return redirect(url_for('reset_password', token=token))

        hashed = bcrypt.generate_password_hash(password).decode('utf-8')
        conn = get_db_connection()
        conn.execute("UPDATE clients SET mot_de_passe_hash = ? WHERE email = ?", (hashed, email))
        conn.commit()
        conn.close()
        flash('Votre mot de passe a été mis à jour avec succès !', 'success')
        return redirect(url_for('accueil'))
    return render_template('reset_password.html')

# ───────────────────────────────────────────────────────────
# Routes Client
# ───────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    if session.get('is_admin'):
        return redirect(url_for('admin_dashboard'))
    user_id = session['user_id']
    conn = get_db_connection()
    projets_actifs = conn.execute("""
        SELECT p.*, s.icon as service_icon, s.nom_service as service_nom
        FROM projets p
        LEFT JOIN services s ON s.nom_service = SUBSTR(p.nom_projet, INSTR(p.nom_projet, ' — ') + 3)
        WHERE p.id_client = ? AND (p.is_archived = 0 OR p.is_archived IS NULL)
        ORDER BY p.created_at DESC
    """, (user_id,)).fetchall()
    projets_archives = conn.execute("""
        SELECT p.*, s.icon as service_icon, s.nom_service as service_nom
        FROM projets p
        LEFT JOIN services s ON s.nom_service = SUBSTR(p.nom_projet, INSTR(p.nom_projet, ' — ') + 3)
        WHERE p.id_client = ? AND p.is_archived = 1
        ORDER BY p.created_at DESC
    """, (user_id,)).fetchall()

    actifs_with_pastille = []
    for p in projets_actifs:
        ready, done, total = compute_checklist_readiness(p['id'])
        actifs_with_pastille.append(
            {"projet": p, "pastille": pastille_color(ready), "done": done, "total": total}
        )
    archives_with_pastille = []
    for p in projets_archives:
        ready, done, total = compute_checklist_readiness(p['id'])
        archives_with_pastille.append(
            {"projet": p, "pastille": pastille_color(ready), "done": done, "total": total}
        )
    client = conn.execute("SELECT drive_folder_id FROM clients WHERE id=?", (user_id,)).fetchone()
    conn.close()
    conn2 = get_db_connection()
    notifications = conn2.execute(
        "SELECT * FROM notifications WHERE id_client=? AND is_read=0 ORDER BY created_at DESC",
        (user_id,)
    ).fetchall()
    conn2.close()
    notif_par_projet = {}
    for n in notifications:
        notif_par_projet.setdefault(n['id_projet'], []).append(n)
    resp = make_response(render_template('dashboard.html',
        projets=actifs_with_pastille,
        projets_archives=archives_with_pastille,
        notifications=notifications,
        notif_par_projet=notif_par_projet,
        client=client))
    resp.headers['Cache-Control'] = 'no-store'
    return resp

@app.route('/notifications/<int:notif_id>/read', methods=['POST'])
@login_required
def mark_notification_read(notif_id):
    user_id = session['user_id']
    conn = get_db_connection()
    conn.execute(
        "UPDATE notifications SET is_read=1 WHERE id=? AND id_client=?",
        (notif_id, user_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/projet/<int:project_id>')
@login_required
def project_detail(project_id):
    conn = get_db_connection()
    projet = conn.execute("""
        SELECT p.*, s.icon as service_icon, s.nom_service as service_nom
        FROM projets p
        LEFT JOIN services s ON s.nom_service = SUBSTR(p.nom_projet, INSTR(p.nom_projet, ' — ') + 3)
        WHERE p.id = ?
    """, (project_id,)).fetchone()
    if not projet:
        conn.close()
        flash("Projet introuvable.", "error")
        return redirect(url_for('dashboard'))

    is_owner = (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not (is_owner or is_admin):
        conn.close()
        flash("Accès non autorisé à ce projet.", "error")
        return redirect(url_for('dashboard'))

    checklist = conn.execute("SELECT * FROM checklistes WHERE id_projet = ?", (project_id,)).fetchone()
    items = []
    if checklist:
        items = conn.execute("""
            SELECT * FROM checklist_items
            WHERE id_checklist = ?
            ORDER BY id ASC
        """, (checklist['id'],)).fetchall()

    ready, done, total = compute_checklist_readiness(project_id)
    color = pastille_color(ready)

    # Lister les sous-dossiers Drive du projet
    drive_folders = []
    if projet['drive_folder_id']:
        try:
            drive_folders = list_subfolders(projet['drive_folder_id'])
        except Exception as e:
            print(f"[DRIVE] Liste sous-dossiers échouée: {e}")
    board = conn.execute("SELECT * FROM decision_boards WHERE id_projet=? AND is_active=1", (project_id,)).fetchone()
    choices = conn.execute("SELECT id FROM decision_board_choices WHERE id_projet=?", (project_id,)).fetchone()
    iv = conn.execute("SELECT is_complete FROM identite_visuelle WHERE id_projet=? AND is_complete=1", (project_id,)).fetchone()
    has_identite_visuelle = iv is not None
    iv_assigned = conn.execute("SELECT id FROM identite_visuelle WHERE id_projet=?", (project_id,)).fetchone()
    has_identite_assigned = iv_assigned is not None
    client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
    conn.close()
    resp = make_response(render_template('project_detail.html',
                           projet=projet, checklist=checklist, items=items,
                           drive_folders=drive_folders,
                           readiness={"ready": ready, "done": done, "total": total, "color": color},
                           board=board, already_submitted=choices,
                           has_identite_visuelle=has_identite_visuelle,
                           has_identite_assigned=has_identite_assigned,
                           client=client))
    resp.headers['Cache-Control'] = 'no-store'
    return resp



@app.route('/item/toggle/<int:item_id>')
@login_required
def toggle_checklist_item(item_id):
    conn = get_db_connection()
    item = conn.execute("SELECT * FROM checklist_items WHERE id = ?", (item_id,)).fetchone()
    if not item:
        conn.close()
        return redirect(url_for('dashboard'))

    checklist = conn.execute("SELECT id_projet FROM checklistes WHERE id = ?", (item['id_checklist'],)).fetchone()
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (checklist['id_projet'],)).fetchone()
    client = conn.execute("SELECT * FROM clients WHERE id = ?", (projet['id_client'],)).fetchone()

    is_owner = (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not (is_owner or is_admin):
        conn.close()
        flash("Action non autorisée.", "error")
        return redirect(url_for('dashboard'))

    requires_file = int(item['requires_file'] or 0) == 1
    has_file = bool(item['file_path'])
    current = int(item['est_coche'] or 0)
    new_status = 1 - current

    if requires_file and new_status == 1 and not has_file:
        conn.close()
        flash("Vous devez téléverser le fichier demandé avant de cocher cette tâche.", "error")
        return redirect(url_for('project_detail', project_id=projet['id']))

    conn.execute("UPDATE checklist_items SET est_coche = ? WHERE id = ?", (new_status, item_id))
    conn.commit()
    conn.close()

    if not is_admin and int(item['important'] or 0) == 1:
        settings = get_notification_settings()
        if settings["admin_updates"] == 1 and settings["admin_emails"]:
            status_txt = "complétée" if new_status == 1 else "décochée"
            subject = f"[Portail] {client['nom_complet']} a {status_txt} une tâche importante"
            body = (
                f"Client: {client['nom_complet']} ({client['email']})\n"
                f"Projet: {projet['nom_projet']} (ID {projet['id']})\n"
                f"Tâche: {item['nom_item']}\n"
                f"Statut: {status_txt}\n"
                f"Accéder: {url_for('project_detail', project_id=projet['id'], _external=True)}"
            )
            send_email(settings["admin_emails"], subject, body)
    if new_status == 1 and normalize_status(projet['statut']) == "Documents à donner":
        ready, done, total = compute_checklist_readiness(projet['id'])
        if ready:
            conn2 = get_db_connection()
            conn2.execute("UPDATE projets SET statut='Documents reçus' WHERE id=?", (projet['id'],))
            conn2.commit()
            try:
                lien = url_for('project_detail', project_id=projet['id'], _external=True)
                send_email_client(client,
                    f"Documents reçus — {projet['nom_projet']}",
                    f"Bonjour {client['nom_complet']}, nous avons bien reçu tous vos documents.",
                    html=email_documents_recus(client['nom_complet'], projet['nom_projet'], lien)
                )

            except Exception as e:
                print(f"[MAIL] Email auto-statut échoué: {e}")
            conn2.close()

    return redirect(url_for('project_detail', project_id=projet['id']))

# ───────────────────────────────────────────────────────────
# Uploads de fichiers liés aux items
# ───────────────────────────────────────────────────────────
@app.route('/item/upload/<int:item_id>', methods=['POST'])
@login_required
def upload_item_file(item_id):
    file = request.files.get('document')
    if not file or file.filename == "":
        flash("Aucun fichier reçu.", "error")
        return redirect(request.referrer or url_for('dashboard'))
    if not allowed(file.filename):
        flash("Extension de fichier non autorisée.", "error")
        return redirect(request.referrer or url_for('dashboard'))

    conn = get_db_connection()
    item = conn.execute("SELECT * FROM checklist_items WHERE id = ?", (item_id,)).fetchone()
    if not item:
        conn.close()
        flash("Élément introuvable.", "error")
        return redirect(url_for('dashboard'))

    if item['file_category'] and not allowed_for_category(file.filename, item['file_category']):
        conn.close()
        cat_label = FILE_CATEGORIES.get(item['file_category'], {}).get('label', item['file_category'])
        flash(f"Ce champ attend un fichier de type : {cat_label}", "error")
        return redirect(request.referrer or url_for('dashboard'))

    checklist = conn.execute("SELECT id_projet FROM checklistes WHERE id = ?", (item['id_checklist'],)).fetchone()

    projet   = conn.execute("SELECT * FROM projets WHERE id = ?", (checklist['id_projet'],)).fetchone()

    is_owner = (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not (is_owner or is_admin):
        conn.close()
        flash("Action non autorisée.", "error")
        return redirect(url_for('dashboard'))

    safe_name = secure_filename(file.filename)
    if not safe_name:
        conn.close()
        flash("Nom de fichier invalide.", "error")
        return redirect(request.referrer or url_for('dashboard'))
    # Préfixe unique pour éviter l'écrasement entre uploads successifs
    safe_name = f"{uuid.uuid4().hex[:8]}_{safe_name}"
    base_dir = os.path.join(app.config["UPLOAD_ROOT"], f"projet_{projet['id']}", f"item_{item_id}")
    pathlib.Path(base_dir).mkdir(parents=True, exist_ok=True)
    save_path = os.path.join(base_dir, safe_name)

    file.save(save_path)

    # Upload vers Google Drive dans le dossier du projet
    drive_file_id = None
    try:
        # Utiliser le dossier "Dépôt de fichiers" si disponible, sinon le dossier projet
        target_folder_id = projet['depot_folder_id'] if projet['depot_folder_id'] else projet['drive_folder_id'] if projet['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
        drive_file_id, _ = upload_file(save_path, safe_name, target_folder_id)
    except Exception as e:
        print(f"[DRIVE] Upload fichier échoué: {e}")

    conn.execute("""
        INSERT INTO uploads (id_item, filename, filepath, uploaded_by)
        VALUES (?, ?, ?, ?)
    """, (item_id, safe_name, save_path, 'admin' if is_admin else 'client'))

    if int(item['requires_file'] or 0) == 1:
        conn.execute("UPDATE checklist_items SET file_path = ?, est_coche = 1 WHERE id = ?",
                     (save_path, item_id))
    else:
        conn.execute("UPDATE checklist_items SET file_path = ? WHERE id = ?", (save_path, item_id))

    conn.commit()
    conn.close()

    flash("Fichier téléversé.", "success")

    # Auto-changement de statut si tous les documents sont reçus
    if normalize_status(projet['statut']) == "Documents à donner":
        ready, done, total = compute_checklist_readiness(projet['id'])
        if ready:
            conn3 = get_db_connection()
            conn3.execute("UPDATE projets SET statut='Documents reçus' WHERE id=?", (projet['id'],))
            conn3.commit()
            try:
                client_notif = conn3.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
                lien = url_for('project_detail', project_id=projet['id'], _external=True)
                send_email_client(client_notif,
                    f"Documents reçus — {projet['nom_projet']}",
                    f"Bonjour {client_notif['nom_complet']}, nous avons reçu tous vos documents.",
                    html=email_documents_recus(client_notif['nom_complet'], projet['nom_projet'], lien)
                )
            except Exception as e:
                print(f"[MAIL] Email auto-statut upload échoué: {e}")
            conn3.close()

    return redirect(url_for('project_detail', project_id=projet['id']))


@app.route('/api/v1/item/toggle/<int:item_id>', methods=['POST'])
@login_required
def api_toggle_checklist_item(item_id):
    conn = get_db_connection()
    item = conn.execute("SELECT * FROM checklist_items WHERE id = ?", (item_id,)).fetchone()
    if not item:
        conn.close()
        return jsonify({'error': 'Item introuvable'}), 404
    checklist = conn.execute("SELECT id_projet FROM checklistes WHERE id = ?", (item['id_checklist'],)).fetchone()
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (checklist['id_projet'],)).fetchone()
    is_owner = (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not (is_owner or is_admin):
        conn.close()
        return jsonify({'error': 'Non autorisé'}), 403
    requires_file = int(item['requires_file'] or 0) == 1
    has_file = bool(item['file_path'])
    current = int(item['est_coche'] or 0)
    new_status = 1 - current
    if requires_file and new_status == 1 and not has_file:
        conn.close()
        return jsonify({'error': 'Fichier requis avant de cocher'}), 400
    conn.execute("UPDATE checklist_items SET est_coche = ? WHERE id = ?", (new_status, item_id))
    conn.commit()
    if new_status == 1 and normalize_status(projet['statut']) == "Documents à donner":
        ready, done, total = compute_checklist_readiness(projet['id'])
        if ready:
            conn.execute("UPDATE projets SET statut='Documents reçus' WHERE id=?", (projet['id'],))
            conn.commit()
            _notify_checklist_complete(conn, projet)
    if int(item['is_revision'] or 0) == 1:
        _check_revision_auto_transition(conn, projet['id'])
    conn.close()
    return jsonify({'success': True, 'est_coche': bool(new_status)})


@app.route('/api/v1/item/upload/<int:item_id>', methods=['POST'])
@login_required
def api_upload_item_file(item_id):
    file = request.files.get('file')
    if not file or file.filename == "":
        return jsonify({'error': 'Aucun fichier reçu'}), 400
    if not allowed(file.filename):
        return jsonify({'error': 'Extension non autorisée'}), 400
    conn = get_db_connection()
    item = conn.execute("SELECT * FROM checklist_items WHERE id = ?", (item_id,)).fetchone()
    if not item:
        conn.close()
        return jsonify({'error': 'Item introuvable'}), 404
    checklist = conn.execute("SELECT id_projet FROM checklistes WHERE id = ?", (item['id_checklist'],)).fetchone()
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (checklist['id_projet'],)).fetchone()
    is_owner = (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not (is_owner or is_admin):
        conn.close()
        return jsonify({'error': 'Non autorisé'}), 403
    safe_name = f"{uuid.uuid4().hex[:8]}_{secure_filename(file.filename)}"
    base_dir = os.path.join(app.config["UPLOAD_ROOT"], f"projet_{projet['id']}", f"item_{item_id}")
    pathlib.Path(base_dir).mkdir(parents=True, exist_ok=True)
    save_path = os.path.join(base_dir, safe_name)
    file.save(save_path)
    try:
        target_folder_id = projet['depot_folder_id'] if projet['depot_folder_id'] else projet['drive_folder_id'] if projet['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
        from drive_service import upload_file as drive_upload
        drive_upload(save_path, safe_name, target_folder_id)
    except Exception as e:
        print(f"[DRIVE] Upload API échoué: {e}")
    conn.execute("INSERT INTO uploads (id_item, filename, filepath, uploaded_by) VALUES (?, ?, ?, ?)",
                 (item_id, safe_name, save_path, 'admin' if is_admin else 'client'))
    conn.execute("UPDATE checklist_items SET file_path = ?, est_coche = 1 WHERE id = ?", (save_path, item_id))
    conn.commit()
    if normalize_status(projet['statut']) == "Documents à donner":
        ready, _, _ = compute_checklist_readiness(projet['id'])
        if ready:
            conn.execute("UPDATE projets SET statut='Documents reçus' WHERE id=?", (projet['id'],))
            conn.commit()
            _notify_checklist_complete(conn, projet)
    if int(item['is_revision'] or 0) == 1:
        _check_revision_auto_transition(conn, projet['id'])
    conn.close()
    return jsonify({'success': True, 'file_path': save_path})


@app.route('/api/v1/projet/<int:projet_id>/upload-fichier', methods=['POST'])
@login_required
def api_projet_upload_fichier(projet_id):
    client_id = session['user_id']
    file = request.files.get('file')
    if not file or file.filename == '':
        return jsonify({'error': 'Fichier requis'}), 400
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id = ? AND id_client = ?", (projet_id, client_id)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    safe_name = f"{uuid.uuid4().hex[:8]}_{secure_filename(file.filename)}"
    base_dir = os.path.join(app.config["UPLOAD_ROOT"], f"projet_{projet_id}", "depot_client")
    pathlib.Path(base_dir).mkdir(parents=True, exist_ok=True)
    save_path = os.path.join(base_dir, safe_name)
    file.save(save_path)
    try:
        target_folder_id = projet['depot_folder_id'] or projet['drive_folder_id'] or os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
        upload_file(save_path, file.filename, target_folder_id)
    except Exception as e:
        print(f"[DRIVE] Upload fichier client échoué: {e}")
    push_admin_notif(
        conn,
        titre=f"Fichier reçu — {projet['nom_projet']}",
        message=file.filename,
        type='info',
        lien=f"/admin/projet/{projet_id}",
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/projet/<int:projet_id>/contact', methods=['POST'])
@login_required
def api_projet_contact(projet_id):
    client_id = session['user_id']
    message = (request.get_json(silent=True) or {}).get('message', '').strip()
    if not message:
        return jsonify({'error': 'Message requis'}), 400
    conn = get_db_connection()
    projet = conn.execute("SELECT nom_projet FROM projets WHERE id = ? AND id_client = ?", (projet_id, client_id)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    client = conn.execute("SELECT nom_complet FROM clients WHERE id = ?", (client_id,)).fetchone()
    nom_client = client['nom_complet'] if client else 'Client'
    push_admin_notif(
        conn,
        titre=f"Question — {projet['nom_projet']}",
        message=f"{nom_client} : {message}",
        type='info',
        lien=f"/admin/projet/{projet_id}",
    )
    conn.commit()
    conn.close()
    try:
        send_email(
            'felix.dumont@cocktailmedia.ca',
            f"Question — {projet['nom_projet']}",
            f"{nom_client} : {message}\n\nVoir le projet : {PORTAIL_URL}/admin/projet/{projet_id}"
        )
    except Exception as e:
        print(f"[MAIL] api_projet_contact: {e}")
    return jsonify({'success': True})


@app.route('/api/v1/projet/<int:projet_id>/ajouter_upload', methods=['POST'])
@login_required
def api_projet_ajouter_upload(projet_id):
    user_id = session['user_id']
    conn = get_db_connection()
    projet = conn.execute(
        "SELECT * FROM projets WHERE id = ? AND id_client = ?", (projet_id, user_id)
    ).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    data = request.get_json() or {}
    nom_base = (data.get('nom_base') or 'Fichier').strip()
    checklist = conn.execute(
        "SELECT * FROM checklistes WHERE id_projet = ?", (projet_id,)
    ).fetchone()
    if not checklist:
        conn.close()
        return jsonify({'error': 'Checklist introuvable'}), 404
    count = conn.execute(
        "SELECT COUNT(*) FROM checklist_items WHERE id_checklist = ? AND nom_item LIKE ?",
        (checklist['id'], f'{nom_base}%')
    ).fetchone()[0]
    max_pos = conn.execute(
        "SELECT COALESCE(MAX(position), 0) FROM checklist_items WHERE id_checklist = ?",
        (checklist['id'],)
    ).fetchone()[0]
    nouveau_nom = f"{nom_base} — {count + 1}"
    cur = conn.execute(
        "INSERT INTO checklist_items (id_checklist, nom_item, requires_file, is_required, position) VALUES (?, ?, 1, 0, ?)",
        (checklist['id'], nouveau_nom, max_pos + 1)
    )
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({
        'id': new_id,
        'nom_item': nouveau_nom,
        'est_coche': False,
        'requires_file': True,
        'is_required': False,
        'file_path': None,
        'item_type': None,
        'video_url': None,
        'field_type': 'check',
        'text_value': None,
    }), 201


@app.route('/api/v1/item/text/<int:item_id>', methods=['POST'])
@login_required
def api_save_item_text(item_id):
    data = request.get_json(force=True)
    text_value = (data.get('text_value') or '').strip()
    conn = get_db_connection()
    item = conn.execute("SELECT * FROM checklist_items WHERE id = ?", (item_id,)).fetchone()
    if not item:
        conn.close()
        return jsonify({'error': 'Item introuvable'}), 404
    checklist = conn.execute("SELECT id_projet FROM checklistes WHERE id = ?", (item['id_checklist'],)).fetchone()
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (checklist['id_projet'],)).fetchone()
    is_owner = (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not (is_owner or is_admin):
        conn.close()
        return jsonify({'error': 'Non autorisé'}), 403
    est_coche = 1 if text_value else 0
    conn.execute("UPDATE checklist_items SET text_value = ?, est_coche = ? WHERE id = ?",
                 (text_value if text_value else None, est_coche, item_id))
    conn.commit()
    if est_coche and normalize_status(projet['statut']) == "Documents à donner":
        ready, _, _ = compute_checklist_readiness(projet['id'])
        if ready:
            conn.execute("UPDATE projets SET statut='Documents reçus' WHERE id=?", (projet['id'],))
            conn.commit()
            _notify_checklist_complete(conn, projet)
    if int(item['is_revision'] or 0) == 1:
        _check_revision_auto_transition(conn, projet['id'])
    conn.close()
    return jsonify({'success': True, 'est_coche': bool(est_coche), 'text_value': text_value or None})


@app.route('/api/v1/item/commentaire/<int:item_id>', methods=['POST'])
@login_required
def api_save_item_commentaire(item_id):
    """Enregistre le commentaire d'un item de révision (« ce qui doit changer »).
    Contrairement à /item/text/, ne touche jamais est_coche : un commentaire décrit
    une demande de changement, il ne vaut pas approbation de l'item."""
    data = request.get_json(force=True)
    commentaire = (data.get('commentaire') or '').strip()
    conn = get_db_connection()
    item = conn.execute("SELECT * FROM checklist_items WHERE id = ?", (item_id,)).fetchone()
    if not item:
        conn.close()
        return jsonify({'error': 'Item introuvable'}), 404
    checklist = conn.execute("SELECT id_projet FROM checklistes WHERE id = ?", (item['id_checklist'],)).fetchone()
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (checklist['id_projet'],)).fetchone()
    is_owner = (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not (is_owner or is_admin):
        conn.close()
        return jsonify({'error': 'Non autorisé'}), 403
    conn.execute("UPDATE checklist_items SET commentaire = ? WHERE id = ?",
                 (commentaire if commentaire else None, item_id))
    conn.commit()
    if commentaire and int(item['is_revision'] or 0) == 1:
        _check_revision_auto_transition(conn, projet['id'])
    conn.close()
    return jsonify({'success': True, 'commentaire': commentaire or None})


@app.route('/item/file/<int:upload_id>')
@login_required
def download_uploaded_file(upload_id):
    conn = get_db_connection()
    up = conn.execute("SELECT * FROM uploads WHERE id = ?", (upload_id,)).fetchone()
    if not up:
        conn.close()
        flash("Fichier introuvable.", "error")
        return redirect(url_for('dashboard'))

    item = conn.execute("SELECT * FROM checklist_items WHERE id = ?", (up['id_item'],)).fetchone()
    checklist = conn.execute("SELECT id_projet FROM checklistes WHERE id = ?", (item['id_checklist'],)).fetchone()
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (checklist['id_projet'],)).fetchone()

    is_owner = (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not (is_owner or is_admin):
        conn.close()
        flash("Accès non autorisé à ce fichier.", "error")
        return redirect(url_for('dashboard'))
    conn.close()

    upload_root = os.path.realpath(app.config["UPLOAD_ROOT"])
    filepath = os.path.realpath(up['filepath'])
    if not filepath.startswith(upload_root + os.sep):
        flash("Fichier introuvable.", "error")
        return redirect(url_for('dashboard'))

    return send_file(filepath, as_attachment=True, download_name=up['filename'])

# ───────────────────────────────────────────────────────────
# Profil / MDP
# ───────────────────────────────────────────────────────────
@app.route('/profile')
@login_required
def profile():
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE id = ?", (session['user_id'],)).fetchone()
    projets_actifs = conn.execute("""
        SELECT * FROM projets WHERE id_client = ? AND (is_archived = 0 OR is_archived IS NULL)
        ORDER BY created_at DESC
    """, (session['user_id'],)).fetchall()
    projets_archives = conn.execute("""
        SELECT * FROM projets WHERE id_client = ? AND is_archived = 1
        ORDER BY created_at DESC
    """, (session['user_id'],)).fetchall()

    # Factures depuis la DB
    factures_db = conn.execute("""
        SELECT * FROM factures
        WHERE id_client = ? AND statut NOT IN ('ouverte', 'annulee')
        ORDER BY created_at DESC
    """, (session['user_id'],)).fetchall()
    conn.close()

    return render_template('profile.html', user=user,
                           projets_actifs=projets_actifs,
                           projets_archives=projets_archives,
                           factures_db=factures_db)

@app.route('/update-profile', methods=['POST'])
@login_required
def update_profile():
    user_id = session['user_id']
    nom = request.form.get('nom_complet','').strip()
    entreprise = request.form.get('nom_entreprise','')
    telephone = request.form.get('telephone','')
    adresse = request.form.get('adresse','').strip()
    ville = request.form.get('ville','').strip()
    code_postal = request.form.get('code_postal','').strip()
    province = request.form.get('province','').strip()

    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE id = ?", (user_id,)).fetchone()
    if user['auth_provider'] == 'password':
        email = request.form.get('email','').strip().lower()
        conn.execute("""
            UPDATE clients SET nom_complet=?, nom_entreprise=?, telephone=?,
            adresse=?, ville=?, code_postal=?, province=?, email=? WHERE id=?
        """, (nom, entreprise, telephone, adresse, ville, code_postal, province, email, user_id))
    else:
        conn.execute("""
            UPDATE clients SET nom_complet=?, nom_entreprise=?, telephone=?,
            adresse=?, ville=?, code_postal=?, province=? WHERE id=?
        """, (nom, entreprise, telephone, adresse, ville, code_postal, province, user_id))
    conn.commit()
    conn.close()

    session['user_name'] = nom
    flash("Vos informations ont été mises à jour avec succès !", "success")
    return redirect(url_for('profile') + '#infos')
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE id = ?", (user_id,)).fetchone()
    if user['auth_provider'] == 'password':
        email = request.form.get('email','').strip().lower()
        conn.execute("""
            UPDATE clients SET nom_complet=?, nom_entreprise=?, telephone=?, email=? WHERE id=?
        """, (nom, entreprise, telephone, email, user_id))
    else:
        conn.execute("""
            UPDATE clients SET nom_complet=?, nom_entreprise=?, telephone=? WHERE id=?
        """, (nom, entreprise, telephone, user_id))
    conn.commit()
    conn.close()

    session['user_name'] = nom
    flash("Vos informations ont été mises à jour avec succès !", "success")
    return redirect(url_for('profile'))

@app.route('/update-password', methods=['POST'])
@login_required
def update_password():
    user_id = session['user_id']
    current = request.form.get('current_password','')
    new = request.form.get('new_password','')
    confirm = request.form.get('confirm_password','')

    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE id = ?", (user_id,)).fetchone()

    if not bcrypt.check_password_hash(user['mot_de_passe_hash'], current):
        conn.close()
        flash("Le mot de passe actuel est incorrect.", "error")
        return redirect(url_for('profile'))

    if new != confirm:
        conn.close()
        flash("Les nouveaux mots de passe ne correspondent pas.", "error")
        return redirect(url_for('profile'))

    if not is_password_strong(new):
        conn.close()
        flash("Le mot de passe doit contenir au moins 8 caractères, une majuscule, une minuscule, un chiffre et un caractère spécial.", "error")
        return redirect(url_for('profile'))

    hashed = bcrypt.generate_password_hash(new).decode('utf-8')
    conn.execute("UPDATE clients SET mot_de_passe_hash = ? WHERE id = ?", (hashed, user_id))
    conn.commit()
    conn.close()

    flash("Votre mot de passe a été mis à jour avec succès !", "success")
    return redirect(url_for('profile'))

# ───────────────────────────────────────────────────────────
# Admin
# ───────────────────────────────────────────────────────────
# Section "Aujourd'hui" du dashboard admin
# ───────────────────────────────────────────────────────────
def build_aujourd_hui(conn):
    from datetime import date, timedelta
    today = date.today()
    weekday = today.weekday()  # 0=Lun 1=Mar 2=Mer 3=Jeu 4=Ven 5=Sam 6=Dim
    day = today.day

    JOURS = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']

    def _projets_en_cours_par_mots(mots):
        rows = conn.execute("""
            SELECT p.id, p.nom_projet, p.statut, p.date_livraison_estimee,
                   c.nom_complet AS client_nom, s.nom_service
            FROM projets p
            JOIN clients c ON c.id = p.id_client
            LEFT JOIN services s ON s.id = p.id_service
            WHERE p.is_archived = 0 AND p.statut = 'Travaux en cours'
            ORDER BY p.date_livraison_estimee ASC NULLS LAST
        """).fetchall()
        return [
            dict(r) for r in rows
            if any(m in (r['nom_service'] or '').lower() for m in mots)
        ]

    projets_du_jour = []
    taches_admin = []

    if weekday == 0:  # Lundi — graphisme
        projets_du_jour = _projets_en_cours_par_mots([
            'logo', 'identité', 'identite', 'graphisme', 'design', 'branding',
            'refonte', 'imprimé', 'imprime', 'visuel', 'support',
        ])
        taches_admin = ["Préparer posts du mois", "Vérifier factures impayées"]

    elif weekday in (1, 3):  # Mardi / Jeudi — photo & vidéo
        projets_du_jour = _projets_en_cours_par_mots([
            'photo', 'vidéo', 'video', 'reel', 'short',
        ])

    elif weekday == 2:  # Mercredi — web
        projets_du_jour = _projets_en_cours_par_mots([
            'web', 'site', 'wordpress', 'landing',
        ])

    elif weekday == 4:  # Vendredi — tous les projets actifs en retard
        rows = conn.execute("""
            SELECT p.id, p.nom_projet, p.statut, p.date_livraison_estimee,
                   c.nom_complet AS client_nom, s.nom_service
            FROM projets p
            JOIN clients c ON c.id = p.id_client
            LEFT JOIN services s ON s.id = p.id_service
            WHERE p.is_archived = 0
              AND p.date_livraison_estimee IS NOT NULL
              AND p.date_livraison_estimee < ?
              AND p.statut NOT IN ('Complété', 'Livré')
            ORDER BY p.date_livraison_estimee ASC
        """, (str(today),)).fetchall()
        projets_du_jour = [dict(r) for r in rows]

    if 25 <= day <= 31:
        taches_admin += ["Factures fin de mois à envoyer", "Vérifier taxes TPS/TVQ"]
    if 1 <= day <= 5:
        taches_admin.append("Vérifier paiements reçus du mois précédent")

    # ── À ne pas oublier ────────────────────────────────────
    a_ne_pas_oublier = []
    ids_vus = set()

    # 1. En révision depuis 3+ jours
    seuil_revision = str(today - timedelta(days=3))
    rows = conn.execute("""
        SELECT p.id, p.nom_projet, p.statut, p.date_livraison_estimee,
               c.nom_complet AS client_nom, p.statut_updated_at
        FROM projets p
        JOIN clients c ON c.id = p.id_client
        WHERE p.is_archived = 0
          AND p.statut = 'En révision'
          AND (p.statut_updated_at IS NULL OR p.statut_updated_at <= ?)
        ORDER BY p.statut_updated_at ASC NULLS FIRST
    """, (seuil_revision,)).fetchall()
    for r in rows:
        a_ne_pas_oublier.append({**dict(r), 'raison': 'En révision depuis 3+ jours'})
        ids_vus.add(r['id'])

    # 2. Documents à donner depuis 5+ jours sans action
    seuil_docs = str(today - timedelta(days=5))
    rows = conn.execute("""
        SELECT p.id, p.nom_projet, p.statut, p.date_livraison_estimee,
               c.nom_complet AS client_nom, p.statut_updated_at
        FROM projets p
        JOIN clients c ON c.id = p.id_client
        WHERE p.is_archived = 0
          AND p.statut = 'Documents à donner'
          AND (p.statut_updated_at IS NULL OR p.statut_updated_at <= ?)
        ORDER BY p.statut_updated_at ASC NULLS FIRST
    """, (seuil_docs,)).fetchall()
    for r in rows:
        if r['id'] not in ids_vus:
            a_ne_pas_oublier.append({**dict(r), 'raison': 'Documents à donner depuis 5+ jours'})
            ids_vus.add(r['id'])

    # 3. Date de livraison estimée dépassée
    rows = conn.execute("""
        SELECT p.id, p.nom_projet, p.statut, p.date_livraison_estimee,
               c.nom_complet AS client_nom
        FROM projets p
        JOIN clients c ON c.id = p.id_client
        WHERE p.is_archived = 0
          AND p.date_livraison_estimee IS NOT NULL
          AND p.date_livraison_estimee < ?
          AND p.statut NOT IN ('Complété', 'Livré')
        ORDER BY p.date_livraison_estimee ASC
    """, (str(today),)).fetchall()
    for r in rows:
        if r['id'] not in ids_vus:
            a_ne_pas_oublier.append({**dict(r), 'raison': 'Date de livraison dépassée'})
            ids_vus.add(r['id'])

    return {
        'weekday_label': JOURS[weekday],
        'projets_du_jour': projets_du_jour,
        'taches_admin': taches_admin,
        'a_ne_pas_oublier': a_ne_pas_oublier,
    }


# ───────────────────────────────────────────────────────────
@app.route('/admin')
@admin_required
def admin_dashboard():
    from datetime import date
    today = str(date.today())
    conn = get_db_connection()
    clients = conn.execute("SELECT id, nom_complet, email, nom_entreprise FROM clients").fetchall()
    projets = conn.execute("""
        SELECT p.id, p.nom_projet, p.statut, c.nom_complet AS nom_client
        FROM projets p JOIN clients c ON p.id_client = c.id
    """).fetchall()
    services = conn.execute("SELECT * FROM services ORDER BY nom_service").fetchall()
    services_localisation = {str(s['id']): bool(s['localisation_requise']) for s in services}

    # Factures ouvertes (clients mensuels + non confirmés)
    factures_ouvertes = conn.execute("""
        SELECT f.*, c.nom_complet, c.nom_entreprise, c.is_email_confirmed,
               COUNT(fl.id) as nb_lignes
        FROM factures f
        JOIN clients c ON c.id = f.id_client
        LEFT JOIN facture_lignes fl ON fl.id_facture = f.id
        WHERE f.statut = 'ouverte'
        GROUP BY f.id
        ORDER BY c.is_email_confirmed ASC, f.created_at DESC
    """).fetchall()

    # Toutes les factures
    toutes_factures = conn.execute("""
        SELECT f.*, c.nom_complet, c.nom_entreprise
        FROM factures f
        JOIN clients c ON c.id = f.id_client
        WHERE f.statut != 'ouverte'
        ORDER BY f.created_at DESC
    """).fetchall()

    from datetime import date
    mois_actuel = date.today().strftime('%Y-%m')
    marketing_posts_todo = conn.execute("""
        SELECT * FROM marketing_posts
        WHERE strftime('%Y-%m', date_publication) = ?
        AND todo_felix_done = 0
        ORDER BY date_publication ASC
    """, (mois_actuel,)).fetchall()
    marketing_posts_complets = conn.execute("""
        SELECT * FROM marketing_posts
        WHERE strftime('%Y-%m', date_publication) = ?
        AND todo_felix_done = 1
        ORDER BY date_publication ASC
    """, (mois_actuel,)).fetchall()
    todos_roadmap = conn.execute("""
        SELECT rt.id, rt.texte, rp.titre as phase_titre,
               r.titre as roadmap_titre, r.id as roadmap_id
        FROM roadmap_todos rt
        JOIN roadmap_phases rp ON rt.id_phase = rp.id
        JOIN roadmaps r ON rp.id_roadmap = r.id
        WHERE rt.est_coche = 0
        AND rp.badge = 'En cours'
        AND r.is_archived = 0
        ORDER BY rt.id DESC
        LIMIT 3
    """).fetchall()

    aujourd_hui = build_aujourd_hui(conn)
    conn.close()
    projets_with_pastille = []
    for p in projets:
        ready, done, total = compute_checklist_readiness(p['id'])
        projets_with_pastille.append({"p": p, "pastille": pastille_color(ready), "done": done, "total": total})

    return render_template('admin_dashboard.html',
        today=today,
        services_localisation=services_localisation,
        clients=clients, projets=projets_with_pastille, services=services,
        factures_ouvertes=factures_ouvertes,
        toutes_factures=toutes_factures,
        marketing_posts_todo=marketing_posts_todo,
        marketing_posts_complets=marketing_posts_complets,
        todos_roadmap=todos_roadmap,
        aujourd_hui=aujourd_hui)

@app.route('/admin/add_client', methods=['POST'])
@admin_required
def add_client():
    nom = request.form.get('nom_complet','').strip()
    email = request.form.get('email','').strip().lower()
    entreprise = request.form.get('nom_entreprise','').strip() or None
    telephone = request.form.get('telephone','').strip() or None

    if not nom or not email:
        flash("Nom et email obligatoires.", "error")
        return redirect(url_for('admin_dashboard'))

    conn = get_db_connection()
    try:
        # Compte créé sans mot de passe, non confirmé
        conn.execute("""
            INSERT INTO clients (nom_complet, email, nom_entreprise, telephone, mot_de_passe_hash, auth_provider, is_email_confirmed, is_admin)
            VALUES (?, ?, ?, ?, '', 'password', 0, 0)
        """, (nom, email, entreprise, telephone))
        conn.commit()

        # Dossier Drive
        try:
            dossier_nom = entreprise if entreprise else nom
            drive_folder_id = create_folder(dossier_nom, parent_id=os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID'))
            factures_folder_id = create_folder("Factures", parent_id=drive_folder_id)
            conn.execute("UPDATE clients SET drive_folder_id=?, factures_folder_id=? WHERE email=?", (drive_folder_id, factures_folder_id, email))
            conn.commit()
        except Exception as e:
            print(f"[DRIVE] Création dossier client échouée: {e}")

        # Email d'invitation
        try:
            token = s.dumps(email, salt='invitation-client-salt')
            long_invite_url = f"{PORTAIL_URL}/invitation/{token}"
            invite_url = create_short_link(long_invite_url, expires_in=604800)
            html_invite = _invitation_client(nom, invite_url)
            send_email(
                email,
                "Bienvenue chez Cocktail Média — Créez votre accès",
                f"Bonjour {nom}, créez votre accès au portail : {invite_url}",
                html=html_invite
            )
        except Exception as e:
            print(f"[MAIL] Email invitation échoué: {e}")

        flash(f"Client '{nom}' ajouté et invitation envoyée.", "success")
    except sqlite3.IntegrityError:
        flash(f"Erreur : L'email '{email}' existe déjà.", "error")
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/add_project', methods=['POST'])
@admin_required
def add_project():
    from datetime import date
    id_client = request.form.get('id_client')
    id_service = request.form.get('id_service')
    date_seance = request.form.get('date_seance') or str(date.today())
    heure_seance = request.form.get('heure_seance', '').strip() or None
    generer_meet = request.form.get('generer_meet') == '1'
    date_reunion = request.form.get('date_reunion', '').strip() or None
    heure_reunion = request.form.get('heure_reunion', '').strip() or None
    lien_reunion = request.form.get('lien_reunion', '').strip() or None
    titre_projet = request.form.get('titre_projet', '').strip() or None
    localisation = request.form.get('localisation', '').strip() or None
    lien_gdrive = None

    # Charger le service EN PREMIER
    conn_tmp = get_db_connection()
    service_row = conn_tmp.execute("SELECT * FROM services WHERE id=?", (id_service,)).fetchone()
    conn_tmp.close()

    nom_service = service_row['nom_service'] if service_row else "Projet"
    documents_requis = bool(service_row['documents_requis']) if service_row else True
    statut = "Documents à donner" if documents_requis else "En attente de rendez-vous"

    # Heure de séance par défaut depuis le service si non fournie
    if not heure_seance and service_row and service_row['heure_seance_defaut']:
        heure_seance = service_row['heure_seance_defaut']

    # Durée de séance depuis le service
    duree_seance_minutes = int(service_row['duree_seance_minutes'] or 60) if service_row else 60

    # Meet auto si service l'exige
    appel_requis = bool(service_row['appel_exploratoire_requis']) if service_row else False
    if appel_requis and date_reunion and heure_reunion and not lien_reunion:
        generer_meet = True

    # Générer Meet AVANT l'INSERT pour avoir le lien
    if generer_meet and date_reunion and heure_reunion:
        try:
            from calendar_service import create_meet_event
            lien_reunion = create_meet_event(nom_projet if 'nom_projet' in dir() else titre_projet or nom_service, date_reunion, heure_reunion)
        except Exception as e:
            print(f"[MEET] Génération lien échouée: {e}")

    nom_projet = f"{date_seance} — {nom_service} — {localisation}" if localisation else f"{date_seance} — {nom_service}"

    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO projets (nom_projet, titre_projet, heure_seance, lien_reunion, duree_seance_minutes, statut, lien_gdrive, id_client, localisation, id_service)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (nom_projet, titre_projet, heure_seance, lien_reunion, duree_seance_minutes, statut, lien_gdrive, id_client, localisation, id_service))
        id_projet = cur.lastrowid
        if id_service:
            cur.execute("INSERT INTO checklistes (id_projet) VALUES (?)", (id_projet,))
            id_checklist = cur.lastrowid
            items_modele = conn.execute("""
                SELECT nom_item, requires_file, is_required, item_type, video_url, is_revision_item, file_category
                FROM checklist_model_items WHERE id_service = ?
            """, (id_service,)).fetchall()
            for m in items_modele:
                if not int(m['is_revision_item'] or 0):
                    cur.execute("""
                        INSERT INTO checklist_items (id_checklist, nom_item, requires_file, is_required, item_type, video_url, is_revision, file_category)
                        VALUES (?, ?, ?, ?, ?, ?, 0, ?)
                    """, (id_checklist, m['nom_item'], int(m['requires_file'] or 0), int(m['is_required'] or 1), m['item_type'] or 'document', m['video_url'], m['file_category'] or 'autre'))

        conn.commit()

        # Créer dossier Drive
        try:
            client = conn.execute("SELECT * FROM clients WHERE id=?", (id_client,)).fetchone()
            parent = client['drive_folder_id'] if client and client['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            projet_folder_id = create_folder(nom_projet, parent_id=parent)
            make_folder_public(projet_folder_id)
            # Créer "Dépôt de fichiers" seulement si documents_requis
            depot_folder_id = None
            if documents_requis:
                depot_folder_id = create_folder("Dépôt de fichiers", parent_id=projet_folder_id)
                make_folder_public(depot_folder_id)
            lien_gdrive_new = get_folder_link(projet_folder_id)
            conn.execute("UPDATE projets SET lien_gdrive=?, drive_folder_id=?, depot_folder_id=? WHERE id=?", (lien_gdrive_new, projet_folder_id, depot_folder_id, id_projet))
            # Créer les sous-dossiers Drive selon le service
            if service_row and service_row['drive_subfolders']:
                for nom_sous_dossier in service_row['drive_subfolders'].split('|'):
                    nom_sous_dossier = nom_sous_dossier.strip()
                    if nom_sous_dossier:
                        sous_folder_id = create_folder(nom_sous_dossier, parent_id=projet_folder_id)
                        make_folder_public(sous_folder_id)

            conn.commit()
        except Exception as drive_e:
            print(f"[DRIVE] Création dossier projet échouée: {drive_e}")

        # Invitation Calendar séance (si date + heure fournies)
        if heure_seance and localisation:
            try:
                from calendar_service import create_seance_event
                client_row = conn.execute("SELECT * FROM clients WHERE id=?", (id_client,)).fetchone()
                create_seance_event(nom_projet, date_seance, heure_seance, duree_seance_minutes, localisation, client_row['email'])
            except Exception as e:
                print(f"[CALENDAR] Invitation séance échouée: {e}")

        # Email création projet
        try:
            client_notif = conn.execute("SELECT * FROM clients WHERE id=?", (id_client,)).fetchone()
            if client_notif and client_notif['email']:
                lien_projet = url_for('project_detail', project_id=id_projet, _external=True)
                send_email_client(client_notif,
                    f"Nouveau projet — {nom_projet}",
                    f"Bonjour {client_notif['nom_complet']}, un nouveau projet vous a été assigné : {nom_projet}",
                    html=email_projet_cree(client_notif['nom_complet'], nom_projet, lien_projet)
                )
        except Exception as e:
            print(f"[MAIL] Email projet créé échoué: {e}")

        flash(f"Le projet '{nom_projet}' a été ajouté avec succès.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Erreur lors de la création du projet : {e}", "error")
    finally:
        conn.close()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/edit_client/<int:client_id>', methods=['GET','POST'])
@admin_required
def edit_client(client_id):
    conn = get_db_connection()
    if request.method == 'POST':
        nom = request.form.get('nom_complet','').strip()
        email = request.form.get('email','').strip().lower()
        entreprise = request.form.get('nom_entreprise','')
        mode_facturation = request.form.get('mode_facturation', 'projet')
        adresse_facturation = request.form.get('adresse_facturation','').strip()
        ville_facturation = request.form.get('ville_facturation','').strip()
        province_facturation = request.form.get('province_facturation','Québec').strip()
        code_postal_facturation = request.form.get('code_postal_facturation','').strip()
        pays_facturation = request.form.get('pays_facturation','Canada').strip()
        conn.execute("""
            UPDATE clients SET nom_complet=?, email=?, nom_entreprise=?,
            mode_facturation=?, adresse_facturation=?, ville_facturation=?,
            province_facturation=?, code_postal_facturation=?, pays_facturation=?
            WHERE id=?
        """, (nom, email, entreprise, mode_facturation, adresse_facturation,
               ville_facturation, province_facturation, code_postal_facturation,
               pays_facturation, client_id))
        conn.commit()
        conn.close()
        flash(f"Le client '{nom}' a été mis à jour.", "success")
        return redirect(url_for('admin_dashboard'))

    client = conn.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    conn.close()
    if not client:
        return redirect(url_for('admin_dashboard'))
    return render_template('edit_client.html', client=client)

@app.route('/admin/delete_client/<int:client_id>')
@admin_required
def delete_client(client_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM clients WHERE id = ?", (client_id,))
    conn.commit()
    conn.close()
    flash("Client supprimé.", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/edit_project/<int:project_id>', methods=['GET','POST'])
@admin_required
def edit_project(project_id):
    conn = get_db_connection()
    if request.method == 'POST':
        nom_projet = request.form.get('nom_projet','').strip()

        # Normaliser le statut (4 phases)
        statut_raw = request.form.get('statut','').strip()
        statut = normalize_status(statut_raw)

        lien_gdrive = request.form.get('lien_gdrive','')
        id_client = request.form.get('id_client')
        titre_affiche = request.form.get('titre_affiche','').strip() or None

        old = conn.execute("SELECT statut, id_client FROM projets WHERE id = ?", (project_id,)).fetchone()

        conn.execute("""
            UPDATE projets SET nom_projet=?, statut=?, lien_gdrive=?, id_client=?, titre_affiche=? WHERE id=?
        """, (nom_projet, statut, lien_gdrive, id_client, titre_affiche, project_id))
        conn.commit()

        # Notification client si passage à "Travaux terminés"
        try:
            if old and (old['statut'] != statut):
                client = conn.execute("SELECT * FROM clients WHERE id = ?", (id_client,)).fetchone()
                if client and client['email']:
                    lien = url_for('project_detail', project_id=project_id, _external=True)
                    statut_norm = normalize_status(statut)
                    subject = None
                    body_txt = None
                    body_html = None
                    if statut_norm == "Documents à donner":
                        subject = f"Documents requis — {nom_projet}"
                        body_txt = f"Bonjour {client['nom_complet']}, nous avons besoin de vos documents pour le projet : {nom_projet}"
                        body_html = email_documents_requis(client['nom_complet'], nom_projet, lien)
                        push_notification(conn, client['id'], project_id, f"Documents requis pour votre projet « {nom_projet} ».", type='documents_requis')
                    elif statut_norm == "Travaux en cours":
                        subject = f"Les travaux sont en cours — {nom_projet}"
                        body_txt = f"Bonjour {client['nom_complet']}, les travaux sont maintenant en cours sur votre projet : {nom_projet}"
                        body_html = email_travaux_en_cours(client['nom_complet'], nom_projet, lien)
                        push_notification(conn, client['id'], project_id, f"Les travaux sont en cours sur votre projet « {nom_projet} ».", type='travaux_en_cours')
                    elif statut_norm == "En révision":
                        subject = f"Votre projet est en révision — {nom_projet}"
                        body_txt = f"Bonjour {client['nom_complet']}, votre projet est en révision : {nom_projet}"
                        body_html = email_en_revision(client['nom_complet'], nom_projet, lien)
                        push_notification(conn, client['id'], project_id, f"Votre projet « {nom_projet} » est en révision.", type='revision')
                    elif statut_norm in ["Travaux terminés", "Complété"]:
                        subject = f"Votre projet est terminé — {nom_projet}"
                        body_txt = f"Bonjour {client['nom_complet']}, votre projet est terminé : {nom_projet}"
                        lien_drive = conn.execute("SELECT lien_gdrive FROM projets WHERE id=?", (project_id,)).fetchone()
                        body_html = email_livraison(client['nom_complet'], nom_projet, lien, lien_drive['lien_gdrive'] if lien_drive else None)
                        push_notification(conn, client['id'], project_id, f"Votre projet « {nom_projet} » est terminé !", type='termine')
                    if subject and body_txt:
                       send_email_client(client, subject, body_txt, html=body_html)
                    conn.commit()
        except Exception as e:
            print(f"[MAIL] Notification client échouée: {e}")
        conn.close()
        flash(f"Le projet '{nom_projet}' a été mis à jour.", "success")
        return redirect(url_for('admin_dashboard'))

    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
    clients = conn.execute("SELECT id, nom_complet FROM clients").fetchall()
    conn.close()
    if not projet:
        return redirect(url_for('admin_dashboard'))
    return render_template('edit_project.html', projet=projet, clients=clients)

@app.route('/admin/delete_project/<int:project_id>')
@admin_required
def delete_project(project_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM projets WHERE id = ?", (project_id,))
    conn.commit()
    conn.close()
    flash("Projet supprimé.", "success")
    return redirect(url_for('admin_dashboard'))

# Services & modèle checklist (admin)
@app.route('/admin/services')
@admin_required
def admin_services():
    conn = get_db_connection()
    les_services = conn.execute("SELECT * FROM services ORDER BY nom_service").fetchall()
    services_avec_items = []
    for service in les_services:
        items = conn.execute("""
            SELECT * FROM checklist_model_items
            WHERE id_service = ?
            ORDER BY is_revision_item ASC, id ASC
        """, (service['id'],)).fetchall()
        services_avec_items.append({'service': service, 'items': items})
    conn.close()
    return render_template('admin_services.html', services=services_avec_items, file_categories=FILE_CATEGORIES)

@app.route('/admin/add_service', methods=['POST'])
@admin_required
def add_service():
    nom_service = request.form.get('nom_service','').strip()
    description = request.form.get('description','')
    localisation_requise = int(request.form.get('localisation_requise', 0))
    documents_requis = int(request.form.get('documents_requis', 0))
    conn = get_db_connection()
    try:
        icon = request.form.get('icon', 'default')
        heure_seance_defaut = request.form.get('heure_seance_defaut', '').strip() or None
        duree_seance_minutes = int(request.form.get('duree_seance_minutes', 60))
        appel_exploratoire_requis = int(request.form.get('appel_exploratoire_requis', 0))        
        duree_heures = int(request.form.get('duree_heures', 1))
        duree_minutes_form = int(request.form.get('duree_minutes', 0))
        duree_production_minutes = (duree_heures * 60) + duree_minutes_form
        delai_fixe_heures = int(request.form.get('delai_fixe_heures', 0))
        prix = float(request.form.get('prix', 0) or 0)
        exonere_taxes = 1 if request.form.get('exonere_taxes') else 0
        exonere_taxes = 1 if request.form.get('exonere_taxes') else 0
        decision_board_requis = 1 if request.form.get('decision_board_requis') else 0
        drive_subfolders = request.form.get('drive_subfolders', '').strip()
        conn.execute("""
            INSERT INTO services (nom_service, description, localisation_requise, documents_requis, icon, duree_production_minutes, delai_fixe_heures, drive_subfolders)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (nom_service, description, localisation_requise, documents_requis, icon, duree_production_minutes, delai_fixe_heures, drive_subfolders))
        conn.commit()
        flash(f"Le service '{nom_service}' a été ajouté.", "success")
    except sqlite3.IntegrityError:
        flash(f"Erreur : Le service '{nom_service}' existe déjà.", "error")
    finally:
        conn.close()
    return redirect(url_for('admin_services'))

@app.route('/admin/delete_service/<int:service_id>')
@admin_required
def delete_service(service_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM checklist_model_items WHERE id_service = ?", (service_id,))
    conn.execute("DELETE FROM services WHERE id = ?", (service_id,))
    conn.commit()
    conn.close()
    flash("Service supprimé (et items de modèle associés).", "success")
    return redirect(url_for('admin_services'))

@app.route('/admin/add_checklist_item/<int:service_id>', methods=['POST'])
@admin_required
def add_checklist_item(service_id):
    nom_item = request.form.get('nom_item','').strip()
    is_required = int(request.form.get('is_required', 1))
    is_revision_item = int(request.form.get('is_revision_item', 0))
    video_url = request.form.get('video_url', '').strip() or None
    type_unified = request.form.get('type_unified', 'document')

    # Déduction item_type + file_category + requires_file selon choix unique
    TYPE_MAP = {
        'photo':     ('document', 'photo',    1),
        'vecteur':   ('document', 'vecteur',  1),
        'video_file':('document', 'video',    1),
        'document':  ('document', 'document', 1),
        'donnees':   ('document', 'donnees',  1),
        'archive':   ('document', 'archive',  1),
        'video_url': ('video',    'autre',    0),
        'autre':     ('document', 'autre',    0),
    }
    item_type, file_category, requires_file = TYPE_MAP.get(type_unified, ('document', 'autre', 0))
    conn = get_db_connection()
    conn.execute("""
        INSERT INTO checklist_model_items (id_service, nom_item, requires_file, is_required, item_type, video_url, is_revision_item, file_category)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (service_id, nom_item, requires_file, is_required, item_type, video_url, is_revision_item, file_category))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_services'))

@app.route('/admin/delete_checklist_item/<int:item_id>')
@admin_required
def delete_checklist_item(item_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM checklist_model_items WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_services'))

def update_project_status(project_id):
    """Met à jour le statut du projet selon l'état de la checklist."""
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return
    checklist = conn.execute("SELECT id FROM checklistes WHERE id_projet=?", (project_id,)).fetchone()
    if not checklist:
        conn.close()
        return
    items = conn.execute("SELECT * FROM checklist_items WHERE id_checklist=?", (checklist['id'],)).fetchall()
    if not items:
        conn.close()
        return
    statut_actuel = projet['statut'].lower() if projet['statut'] else ''
    # Si déjà en révision ou complété, on ne touche pas au statut automatiquement
    if statut_actuel in ['en révision', 'complété', 'complete']:
        conn.close()
        return
    items_requis = [i for i in items if int(i['is_required'] or 0) == 1 and int(i['is_revision'] or 0) == 0]
    tous_coches = all(int(i['est_coche'] or 0) == 1 for i in items_requis)
    if tous_coches and items_requis:
        conn.execute("UPDATE projets SET statut='Travaux en cours' WHERE id=?", (project_id,))
        conn.commit()
    conn.close()

def _envoyer_courriel_revision(conn, projet, client):
    """Envoie le courriel « projet en révision », en utilisant le template spécifique
    Site Web Vitrine (lien du site + pistes de révision cochables) quand applicable,
    sinon le template générique. Centralisé ici pour que tous les points d'entrée
    (panneau admin, bouton « Notifier le client », API JSON) se comportent pareil."""
    service_row = conn.execute("SELECT nom_service FROM services WHERE id=?", (projet['id_service'],)).fetchone() if projet['id_service'] else None
    nom_service = service_row['nom_service'] if service_row else None
    lien_portail = url_for('project_detail', project_id=projet['id'], _external=True)
    if nom_service == 'Site Web Vitrine' and projet['lien_site_test']:
        items_rows = conn.execute("""
            SELECT ci.nom_item FROM checklist_items ci
            JOIN checklistes c ON c.id = ci.id_checklist
            WHERE c.id_projet = ? AND ci.is_revision = 1
            ORDER BY ci.position, ci.id
        """, (projet['id'],)).fetchall()
        items_revision = [r['nom_item'] for r in items_rows]
        html = email_revision_site_web(client['nom_complet'], projet['nom_projet'], projet['lien_site_test'], lien_portail, items_revision)
        subject = f"Votre site web est prêt pour révision — {projet['nom_projet']}"
    else:
        html = email_en_revision(client['nom_complet'], projet['nom_projet'], lien_portail)
        subject = f"Votre projet est en révision — {projet['nom_projet']}"
    send_email_client(client, subject, f"Bonjour {client['nom_complet']}, votre projet est en révision.", html=html)

@app.route('/admin/projet/<int:project_id>/force_status', methods=['POST'])
@admin_required
def force_status(project_id):
    statut = request.form.get('statut', '').strip()
    if statut:
        conn = get_db_connection()
        projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
        client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
        conn.execute("UPDATE projets SET statut=? WHERE id=?", (statut, project_id))
        conn.commit()
        try:
            if client and client['email']:
                lien = url_for('project_detail', project_id=project_id, _external=True)
                nom_projet = projet['nom_projet']
                statut_norm = normalize_status(statut)
                subject = None
                body_txt = None
                body_html = None
                if statut_norm == "Documents à donner":
                    subject = f"Documents requis — {nom_projet}"
                    body_txt = f"Bonjour {client['nom_complet']}, nous avons besoin de vos documents."
                    body_html = email_documents_requis(client['nom_complet'], nom_projet, lien)
                    push_notification(conn, client['id'], project_id, f"Documents requis pour votre projet « {nom_projet} ».", type='documents_requis')
                elif statut_norm == "Travaux en cours":
                    subject = f"Les travaux sont en cours — {nom_projet}"
                    body_txt = f"Bonjour {client['nom_complet']}, les travaux sont en cours."
                    body_html = email_travaux_en_cours(client['nom_complet'], nom_projet, lien)
                    push_notification(conn, client['id'], project_id, f"Les travaux sont en cours sur votre projet « {nom_projet} ».", type='travaux_en_cours')
                elif statut_norm in ["Travaux terminés", "Complété"]:
                    subject = f"Votre projet est terminé — {nom_projet}"
                    body_txt = f"Bonjour {client['nom_complet']}, votre projet est terminé."
                    lien_drive = projet['lien_gdrive'] if projet['lien_gdrive'] else None
                    body_html = email_livraison(client['nom_complet'], nom_projet, lien, lien_drive)
                    push_notification(conn, client['id'], project_id, f"Votre projet « {nom_projet} » est terminé !", type='termine')
                elif statut_norm == "Annulé":
                    subject = f"Projet annulé — {nom_projet}"
                    body_txt = f"Bonjour {client['nom_complet']}, votre projet a été annulé."
                    body_html = email_annulation(client['nom_complet'], nom_projet)
                elif statut_norm == "En révision":
                    push_notification(conn, client['id'], project_id, f"Votre projet « {nom_projet} » est en révision.", type='revision')
                if subject and body_txt:
                   send_email_client(client, subject, body_txt, html=body_html)
                conn.commit()
        except Exception as e:
            print(f"[MAIL] Email force_status échoué: {e}")
        conn.close()
        flash(f"Statut mis à jour : {statut}", "success")
    return redirect(url_for('project_detail', project_id=project_id))

# ───────────────────────────────────────────────────────────
# Marketing — Calendrier
# ───────────────────────────────────────────────────────────

@app.route('/admin/marketing')
@admin_required
def marketing_calendrier():
    from datetime import date
    mois = request.args.get('mois', date.today().strftime('%Y-%m'))
    conn = get_db_connection()
    posts = conn.execute("""
        SELECT * FROM marketing_posts
        WHERE strftime('%Y-%m', date_publication) = ?
        ORDER BY date_publication ASC
    """, (mois,)).fetchall()
    conn.close()
    return render_template('marketing.html', posts=posts, mois=mois)


@app.route('/admin/marketing/nouveau', methods=['GET', 'POST'])
@admin_required
def marketing_nouveau_post():
    import json
    from datetime import datetime
    from email_templates import email_nouveau_post_marketing

    if request.method == 'POST':
        titre = request.form.get('titre', '').strip()
        description = request.form.get('description', '').strip()
        date_publication = request.form.get('date_publication', '').strip()
        plateformes = request.form.getlist('plateformes')

        if not titre or not date_publication or not plateformes:
            flash("Titre, date et au moins une plateforme sont requis.", "error")
            return redirect(url_for('marketing_nouveau_post'))

        plateformes_json = json.dumps(plateformes)

        # Créer dossier Drive Marketing/YYYY-MM/Titre
        drive_folder_id = None
        try:
            root_id = os.getenv('MARKETING_DRIVE_FOLDER_ID')
            mois_folder = create_folder(date_publication[:7], parent_id=root_id)
            post_folder = create_folder(f"{date_publication} — {titre}", parent_id=mois_folder)
            drive_folder_id = post_folder

        except Exception as e:
            print(f"[DRIVE] Dossier marketing échoué: {e}")

        conn = get_db_connection()
        conn.execute("""
            INSERT INTO marketing_posts (titre, description, date_publication, plateformes, statut, drive_folder_id, created_by)
            VALUES (?, ?, ?, ?, 'planifié', ?, ?)
        """, (titre, description, date_publication, plateformes_json, drive_folder_id, session.get('user_id')))
        conn.commit()
        conn.close()

        # Email à Félix
        try:
            send_email(
                'felix.dumont@cocktailmedia.ca',
                f"📅 Nouveau post planifié — {titre}",
                f"Nouveau post marketing planifié : {titre} le {date_publication}",
                html=email_nouveau_post_marketing(titre, date_publication, plateformes, description)
            )
        except Exception as e:
            print(f"[MAIL] Email marketing échoué: {e}")

        flash("Post planifié avec succès.", "success")
        return redirect(url_for('marketing_calendrier', mois=date_publication[:7], action='post_cree'))
    return render_template('marketing_nouveau.html')

@app.route('/admin/marketing/notifier-felix/<mois>', methods=['POST'])
@admin_required
def marketing_notifier_felix(mois):
    import json
    from email_templates import email_nouveau_post_marketing

    conn = get_db_connection()
    posts = conn.execute("""
        SELECT * FROM marketing_posts
        WHERE strftime('%Y-%m', date_publication) = ?
        AND demande_envoyee = 0
        ORDER BY date_publication ASC
    """, (mois,)).fetchall()

    if not posts:
        flash("Aucun nouveau post à envoyer.", "info")
        conn.close()
        return redirect(url_for('marketing_calendrier', mois=mois))

    # Construire le récap HTML
    lignes_html = ""
    lignes_txt = ""
    for post in posts:
        plateformes = json.loads(post['plateformes']) if post['plateformes'] else []
        plates_str = ', '.join(plateformes)
        lignes_html += f"""
        <tr>
            <td style="padding:8px;border-bottom:1px solid #eee;font-weight:600;">{post['date_publication']}</td>
            <td style="padding:8px;border-bottom:1px solid #eee;">{post['titre']}</td>
            <td style="padding:8px;border-bottom:1px solid #eee;">{plates_str}</td>
            <td style="padding:8px;border-bottom:1px solid #eee;color:#666;">{post['description'] or '—'}</td>
        </tr>
        """
        lignes_txt += f"- {post['date_publication']} | {post['titre']} | {plates_str}\n"

    html_body = f"""
    <div style="font-family:Montserrat,sans-serif;max-width:700px;margin:auto;background:#fff;border-radius:12px;overflow:hidden;">
      <div style="background:#c0321a;padding:32px;text-align:center;">
        <h1 style="color:#fff;margin:0;font-size:22px;">📋 Demande de visuels — {mois}</h1>
      </div>
      <div style="padding:32px;">
        <p style="color:#2b2b2b;">Marie-Christine a planifié <strong>{len(posts)} post(s)</strong> pour le mois de {mois}. Voici les visuels à créer :</p>
        <table style="width:100%;border-collapse:collapse;margin-top:16px;">
          <tr style="background:#fdecea;">
            <th style="padding:8px;text-align:left;font-size:12px;color:#888;">DATE</th>
            <th style="padding:8px;text-align:left;font-size:12px;color:#888;">TITRE</th>
            <th style="padding:8px;text-align:left;font-size:12px;color:#888;">PLATEFORMES</th>
            <th style="padding:8px;text-align:left;font-size:12px;color:#888;">DESCRIPTION</th>
          </tr>
          {lignes_html}
        </table>
        <p style="margin-top:24px;color:#888;font-size:13px;">Connecte-toi au portail pour déposer les visuels.</p>
      </div>
    </div>
    """

    try:
        send_email(
            'felix.dumont@cocktailmedia.ca',
            f"📋 Demande de visuels — {mois} ({len(posts)} posts)",
            f"Demande de visuels pour {mois}:\n{lignes_txt}",
            html=html_body
        )
        # Marquer tous comme envoyés
        for post in posts:
            conn.execute("UPDATE marketing_posts SET demande_envoyee=1 WHERE id=?", (post['id'],))
        conn.commit()
        flash(f"Demande envoyée à Félix — {len(posts)} post(s) inclus.", "success")
    except Exception as e:
        print(f"[MAIL] Demande marketing échouée: {e}")
        flash("Erreur lors de l'envoi. Vérifie les logs.", "error")

    conn.close()
    return redirect(url_for('marketing_calendrier', mois=mois))

@app.route('/admin/marketing/<int:post_id>/supprimer', methods=['POST'])
@admin_required
def marketing_supprimer_post(post_id):
    conn = get_db_connection()
    post = conn.execute("SELECT * FROM marketing_posts WHERE id=?", (post_id,)).fetchone()
    if not post:
        conn.close()
        flash("Post introuvable.", "error")
        return redirect(url_for('marketing_calendrier'))
    mois = post['date_publication'][:7]
    conn.execute("DELETE FROM marketing_posts WHERE id=?", (post_id,))
    conn.commit()
    conn.close()
    flash("Post supprimé.", "success")
    return redirect(url_for('marketing_calendrier', mois=mois))

@app.route('/admin/marketing/post/<int:post_id>/todo-toggle', methods=['POST'])
@admin_required
def marketing_todo_toggle(post_id):
    conn = get_db_connection()
    post = conn.execute("SELECT * FROM marketing_posts WHERE id=?", (post_id,)).fetchone()
    if not post:
        conn.close()
        return jsonify({'error': 'Post introuvable'}), 404
    new_val = 0 if int(post['todo_felix_done'] or 0) else 1
    conn.execute("UPDATE marketing_posts SET todo_felix_done=? WHERE id=?", (new_val, post_id))
    conn.commit()
    conn.close()
    return jsonify({'done': bool(new_val)})

@app.route('/admin/marketing/notifier-marie/<mois>', methods=['POST'])
@admin_required
def marketing_notifier_marie(mois):
    import json
    from email_templates import email_visuel_depose

    conn = get_db_connection()
    posts = conn.execute("""
        SELECT * FROM marketing_posts
        WHERE strftime('%Y-%m', date_publication) = ?
        AND statut IN ('visuels prêts', 'planifié', 'Planifié')
        AND todo_marie_done = 0
        ORDER BY date_publication ASC
    """, (mois,)).fetchall()

    if not posts:
        flash("Aucun visuel à notifier.", "info")
        conn.close()
        return redirect(url_for('marketing_calendrier', mois=mois))

    lignes_html = ""
    lignes_txt = ""
    for post in posts:
        plateformes = json.loads(post['plateformes']) if post['plateformes'] else []
        plates_str = ', '.join(plateformes)
        lignes_html += f"""
        <tr>
            <td style="padding:8px;border-bottom:1px solid #eee;font-weight:600;">{post['date_publication']}</td>
            <td style="padding:8px;border-bottom:1px solid #eee;">{post['titre']}</td>
            <td style="padding:8px;border-bottom:1px solid #eee;">{plates_str}</td>
        </tr>
        """
        lignes_txt += f"- {post['date_publication']} | {post['titre']} | {plates_str}\n"

    html_body = f"""
    <div style="font-family:Montserrat,sans-serif;max-width:700px;margin:auto;background:#fff;border-radius:12px;overflow:hidden;">
      <div style="background:#c0321a;padding:32px;text-align:center;">
        <h1 style="color:#fff;margin:0;font-size:22px;">🎨 Visuels prêts — {mois}</h1>
      </div>
      <div style="padding:32px;">
        <p style="color:#2b2b2b;">Félix a déposé les visuels pour <strong>{len(posts)} post(s)</strong>. Ils sont prêts à être publiés !</p>
        <table style="width:100%;border-collapse:collapse;margin-top:16px;">
          <tr style="background:#fdecea;">
            <th style="padding:8px;text-align:left;font-size:12px;color:#888;">DATE</th>
            <th style="padding:8px;text-align:left;font-size:12px;color:#888;">TITRE</th>
            <th style="padding:8px;text-align:left;font-size:12px;color:#888;">PLATEFORMES</th>
          </tr>
          {lignes_html}
        </table>
        <p style="margin-top:24px;color:#888;font-size:13px;">Connecte-toi au portail pour télécharger les visuels.</p>
      </div>
    </div>
    """

    try:
        send_email(
            'marie-christine.blanchette@cocktailmedia.ca',
            f"🎨 Visuels prêts à publier — {mois} ({len(posts)} posts)",
            f"Visuels prêts pour {mois}:\n{lignes_txt}",
            html=html_body
        )
        for post in posts:
            conn.execute("UPDATE marketing_posts SET todo_marie_done=1 WHERE id=?", (post['id'],))
        conn.commit()
        flash(f"Marie notifiée — {len(posts)} post(s) inclus.", "success")
    except Exception as e:
        print(f"[MAIL] Notif Marie échouée: {e}")
        flash("Erreur lors de l'envoi. Vérifie les logs.", "error")

    conn.close()
    return redirect(url_for('marketing_calendrier', mois=mois))


@app.route('/admin/marketing/<int:post_id>/deposer', methods=['POST'])
@admin_required
def marketing_deposer_visuel(post_id):
    import json
    from email_templates import email_visuel_depose

    conn = get_db_connection()
    post = conn.execute("SELECT * FROM marketing_posts WHERE id=?", (post_id,)).fetchone()
    if not post:
        conn.close()
        flash("Post introuvable.", "error")
        return redirect(url_for('marketing_calendrier'))

    fichiers = request.files.getlist('visuels')
    fichiers = [f for f in fichiers if f and f.filename]

    if not fichiers:
        flash("Aucun fichier sélectionné.", "error")
        return redirect(url_for('marketing_calendrier', mois=post['date_publication'][:7]))

    if len(fichiers) > 4:
        flash("Maximum 4 visuels par post.", "error")
        return redirect(url_for('marketing_calendrier', mois=post['date_publication'][:7]))

    nb_uploades = 0
    for f in fichiers:
        try:
            fname = secure_filename(f.filename)
            tmp_path = os.path.join('/tmp', fname)
            f.save(tmp_path)
            upload_file(tmp_path, fname, folder_id=post['drive_folder_id'])
            os.remove(tmp_path)
            nb_uploades += 1
        except Exception as e:
            print(f"[DRIVE] Upload visuel échoué: {e}")

    if nb_uploades > 0:
        conn.execute("UPDATE marketing_posts SET statut='visuels prêts' WHERE id=?", (post_id,))
        conn.commit()

        flash(f"{nb_uploades} visuel(s) déposé(s) avec succès.", "success")
        return redirect(url_for('marketing_calendrier', mois=post['date_publication'][:7], action='visuel_depose'))
    else:
        flash("Échec de l'upload. Vérifie les fichiers.", "error")

    conn.close()
    return redirect(url_for('marketing_calendrier', mois=post['date_publication'][:7], action='visuel_depose'))

# ───────────────────────────────────────────────────────────
# Main
# ───────────────────────────────────────────────────────────

# ───────────────────────────────────────────────────────────
# Création manuelle d'un compte admin (exécuter une seule fois)
# ───────────────────────────────────────────────────────────
@app.route('/admin/projet/<int:project_id>/notifier_revision', methods=['POST'])
@admin_required
def notifier_revision(project_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
    iv = conn.execute("SELECT is_complete FROM identite_visuelle WHERE id_projet=?", (project_id,)).fetchone()
    try:
        if iv and int(iv['is_complete'] or 0) == 1:
            lien = url_for('projet_identite', project_id=project_id, _external=True)
            send_email_client(
                client,
                f"Votre identité visuelle est prête — {projet['nom_projet']}",
                f"Bonjour {client['nom_complet']}, votre identité visuelle est prête pour révision.",
                html=email_identite_visuelle_prete(client['nom_complet'], projet['nom_projet'], lien)
            )
        else:
            _envoyer_courriel_revision(conn, projet, client)
    except Exception as e:
        print(f"[MAIL] notifier_revision échoué: {e}")
    conn.close()
    return redirect(url_for('project_detail', project_id=project_id))

@app.route('/admin/edit_service/<int:service_id>', methods=['POST'])
@admin_required
def edit_service(service_id):
    conn = get_db_connection()
    documents_requis = int(request.form.get('documents_requis', 0))
    localisation_requise = int(request.form.get('localisation_requise', 0))
    icon = request.form.get('icon', 'default')
    duree_seance_minutes = int(request.form.get('duree_seance_minutes', 60))
    appel_exploratoire_requis = int(request.form.get('appel_exploratoire_requis', 0))    
    heure_seance_defaut = request.form.get('heure_seance_defaut', '').strip() or None    
    duree_heures = int(request.form.get('duree_heures', 1))
    duree_minutes_form = int(request.form.get('duree_minutes', 0))
    duree_production_minutes = (duree_heures * 60) + duree_minutes_form
    delai_fixe_heures = int(request.form.get('delai_fixe_heures', 0))
    prix = float(request.form.get('prix', 0) or 0)
    exonere_taxes = 1 if request.form.get('exonere_taxes') else 0
    decision_board_requis = 1 if request.form.get('decision_board_requis') else 0
    drive_subfolders = request.form.get('drive_subfolders', '').strip()
    conn.execute("""
        UPDATE services SET documents_requis=?, localisation_requise=?, icon=?,
        duree_production_minutes=?, delai_fixe_heures=?, heure_seance_defaut=?,
        duree_seance_minutes=?, appel_exploratoire_requis=?, prix=?, exonere_taxes=?,
        decision_board_requis=?, drive_subfolders=? WHERE id=?
    """, (documents_requis, localisation_requise, icon, duree_production_minutes, delai_fixe_heures, heure_seance_defaut, duree_seance_minutes, appel_exploratoire_requis, prix, exonere_taxes, decision_board_requis, drive_subfolders, service_id))

    conn.commit()
    conn.close()
    flash("Service mis à jour.", "success")
    return redirect(url_for('admin_services'))

def create_admin():
    pass

@app.route('/admin/projet/<int:project_id>/edit_items', methods=['POST'])
@admin_required
def edit_checklist_items(project_id):
    conn = get_db_connection()
    checklist = conn.execute("SELECT id FROM checklistes WHERE id_projet = ?", (project_id,)).fetchone()
    if not checklist:
        conn.close()
        flash("Checklist introuvable.", "error")
        return redirect(url_for('project_detail', project_id=project_id))

    item_ids = request.form.getlist('item_id[]')
    for iid in item_ids:
        nom = request.form.get(f'nom_{iid}', '').strip()
        type_unified = request.form.get(f'type_{iid}', 'document')
        is_required = 1 if request.form.get(f'required_{iid}') else 0

        TYPE_MAP = {
            'photo':      ('document', 'photo',    1),
            'vecteur':    ('document', 'vecteur',  1),
            'video_file': ('document', 'video',    1),
            'document':   ('document', 'document', 1),
            'donnees':    ('document', 'donnees',  1),
            'archive':    ('document', 'archive',  1),
            'video_url':  ('video',    'autre',    0),
            'autre':      ('document', 'autre',    0),
        }
        item_type, file_category, requires_file = TYPE_MAP.get(type_unified, ('document', 'autre', 0))

        if nom:
            conn.execute("""
                UPDATE checklist_items
                SET nom_item=?, item_type=?, file_category=?, requires_file=?, is_required=?
                WHERE id=? AND id_checklist=?
            """, (nom, item_type, file_category, requires_file, is_required, iid, checklist['id']))

    conn.commit()
    conn.close()
    flash("Items mis à jour.", "success")
    return redirect(url_for('project_detail', project_id=project_id))

# ───────────────────────────────────────────────────────────
# API v1 — Factures
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/factures', methods=['GET'])
@admin_required
def api_admin_list_factures():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT f.*, c.nom_complet AS client_nom, c.nom_entreprise AS client_entreprise,
               c.is_test_client AS client_is_test
        FROM factures f
        JOIN clients c ON c.id = f.id_client
        ORDER BY f.created_at DESC
    """).fetchall()
    conn.close()
    return jsonify([{
        'id': f['id'],
        'numero': f['numero'],
        'statut': f['statut'],
        'client_nom': f['client_entreprise'] or f['client_nom'],
        'date_emission': f['date_emission'],
        'date_echeance': f['date_echeance'],
        'total': float(f['total'] or 0),
        'is_test_client': bool(f['client_is_test']),
    } for f in rows])


@app.route('/api/v1/admin/factures', methods=['POST'])
@admin_required
def api_admin_create_facture():
    # Facture manuelle — créée directement pour un client choisi, sans passer par un
    # projet tarifé. Démarre vide ("ouverte", 0$) ; les lignes s'ajoutent ensuite via
    # /api/v1/admin/facture/<id>/ajouter_ligne (page de détail existante, inchangée).
    data = request.get_json(force=True) or {}
    id_client = data.get('id_client')
    if not id_client:
        return jsonify({'error': 'Client requis.'}), 400
    conn = get_db_connection()
    try:
        client = conn.execute("SELECT id FROM clients WHERE id=?", (id_client,)).fetchone()
        if not client:
            return jsonify({'error': 'Client introuvable.'}), 404
        numero = generer_numero_facture(id_client, conn)
        cur = conn.execute("""
            INSERT INTO factures (numero, id_client, statut, type_facturation, sous_total, tps, tvq, total)
            VALUES (?, ?, 'ouverte', 'manuelle', 0, 0, 0, 0)
        """, (numero, id_client))
        facture_id = cur.lastrowid
        conn.commit()
        return jsonify({'success': True, 'id': facture_id, 'numero': numero}), 201
    finally:
        conn.close()


@app.route('/api/v1/admin/facture/<int:facture_id>', methods=['GET'])
@admin_required
def api_admin_get_facture(facture_id):
    conn = get_db_connection()
    facture = conn.execute("SELECT * FROM factures WHERE id=?", (facture_id,)).fetchone()
    if not facture:
        conn.close()
        return jsonify({'error': 'Facture introuvable'}), 404
    client = conn.execute("SELECT * FROM clients WHERE id=?", (facture['id_client'],)).fetchone()
    lignes = conn.execute(
        "SELECT * FROM facture_lignes WHERE id_facture=? ORDER BY date_service, id", (facture_id,)
    ).fetchall()
    conn.close()
    return jsonify({
        'id': facture['id'],
        'numero': facture['numero'],
        'statut': facture['statut'],
        'periode_mois': facture['periode_mois'],
        'date_emission': facture['date_emission'],
        'date_echeance': facture['date_echeance'],
        'sous_total': float(facture['sous_total'] or 0),
        'tps': float(facture['tps'] or 0),
        'tvq': float(facture['tvq'] or 0),
        'total': float(facture['total'] or 0),
        'pdf_path': facture['pdf_path'],
        'stripe_payment_url': facture['stripe_payment_url'],
        'client': {
            'id': client['id'],
            'nom_complet': client['nom_complet'],
            'nom_entreprise': client['nom_entreprise'],
            'email': client['email'],
            'adresse_facturation': client['adresse_facturation'],
            'ville_facturation': client['ville_facturation'],
            'province_facturation': client['province_facturation'],
            'code_postal_facturation': client['code_postal_facturation'],
        } if client else None,
        'lignes': [{
            'id': l['id'],
            'description': l['description'],
            'date_service': l['date_service'],
            'localisation': l['localisation'],
            'quantite': l['quantite'],
            'prix_unitaire': float(l['prix_unitaire'] or 0),
            'total_ligne': float(l['total_ligne'] or 0),
        } for l in lignes],
    })


@app.route('/api/v1/admin/facture/<int:facture_id>/ajouter_ligne', methods=['POST'])
@admin_required
def api_admin_ajouter_ligne(facture_id):
    from datetime import date as date_cls
    data = request.get_json(force=True) or {}
    description  = (data.get('description') or '').strip()
    date_service = data.get('date_service') or str(date_cls.today())
    localisation = (data.get('localisation') or '').strip() or None
    prix         = round(float(data.get('prix_unitaire') or 0), 2)
    quantite     = int(data.get('quantite') or 1)
    total_ligne  = round(prix * quantite, 2)

    if not description:
        return jsonify({'error': 'Description obligatoire.'}), 400

    conn = get_db_connection()
    try:
        conn.execute("""
            INSERT INTO facture_lignes (id_facture, description, date_service, localisation, quantite, prix_unitaire, total_ligne)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (facture_id, description, date_service, localisation, quantite, prix, total_ligne))
        lignes = conn.execute(
            "SELECT prix_unitaire, quantite FROM facture_lignes WHERE id_facture=?", (facture_id,)
        ).fetchall()
        sous_total = sum(float(l['prix_unitaire']) * int(l['quantite']) for l in lignes)
        tps  = round(sous_total * 0.05, 2)
        tvq  = round(sous_total * 0.09975, 2)
        total = round(sous_total + tps + tvq, 2)
        conn.execute("UPDATE factures SET sous_total=?, tps=?, tvq=?, total=? WHERE id=?",
                     (sous_total, tps, tvq, total, facture_id))
        conn.commit()
        regenerer_pdf_facture(facture_id, conn)
        return jsonify({'success': True, 'sous_total': sous_total, 'tps': tps, 'tvq': tvq, 'total': total})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/v1/admin/ligne/<int:ligne_id>/modifier', methods=['POST'])
@admin_required
def api_admin_modifier_ligne(ligne_id):
    data = request.get_json(force=True) or {}
    description  = (data.get('description') or '').strip()
    localisation = (data.get('localisation') or '').strip() or None
    prix         = round(float(data.get('prix_unitaire') or 0), 2)
    quantite     = int(data.get('quantite') or 1)
    total_ligne  = round(prix * quantite, 2)

    conn = get_db_connection()
    try:
        ligne = conn.execute("SELECT id_facture FROM facture_lignes WHERE id=?", (ligne_id,)).fetchone()
        if not ligne:
            return jsonify({'error': 'Ligne introuvable.'}), 404
        conn.execute("""
            UPDATE facture_lignes SET description=?, localisation=?, prix_unitaire=?, quantite=?, total_ligne=? WHERE id=?
        """, (description, localisation, prix, quantite, total_ligne, ligne_id))
        lignes = conn.execute(
            "SELECT prix_unitaire, quantite FROM facture_lignes WHERE id_facture=?", (ligne['id_facture'],)
        ).fetchall()
        sous_total = sum(float(l['prix_unitaire']) * int(l['quantite']) for l in lignes)
        tps  = round(sous_total * 0.05, 2)
        tvq  = round(sous_total * 0.09975, 2)
        total = round(sous_total + tps + tvq, 2)
        conn.execute("UPDATE factures SET sous_total=?, tps=?, tvq=?, total=? WHERE id=?",
                     (sous_total, tps, tvq, total, ligne['id_facture']))
        conn.commit()
        regenerer_pdf_facture(ligne['id_facture'], conn)
        return jsonify({'success': True, 'sous_total': sous_total, 'tps': tps, 'tvq': tvq, 'total': total})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/v1/admin/ligne/<int:ligne_id>/supprimer', methods=['POST'])
@admin_required
def api_admin_supprimer_ligne(ligne_id):
    conn = get_db_connection()
    try:
        ligne = conn.execute("SELECT id_facture FROM facture_lignes WHERE id=?", (ligne_id,)).fetchone()
        if not ligne:
            return jsonify({'error': 'Ligne introuvable.'}), 404
        facture_id = ligne['id_facture']
        conn.execute("DELETE FROM facture_lignes WHERE id=?", (ligne_id,))
        lignes = conn.execute(
            "SELECT prix_unitaire, quantite FROM facture_lignes WHERE id_facture=?", (facture_id,)
        ).fetchall()
        sous_total = sum(float(l['prix_unitaire']) * int(l['quantite']) for l in lignes)
        tps  = round(sous_total * 0.05, 2)
        tvq  = round(sous_total * 0.09975, 2)
        total = round(sous_total + tps + tvq, 2)
        conn.execute("UPDATE factures SET sous_total=?, tps=?, tvq=?, total=? WHERE id=?",
                     (sous_total, tps, tvq, total, facture_id))
        conn.commit()
        regenerer_pdf_facture(facture_id, conn)
        return jsonify({'success': True, 'sous_total': sous_total, 'tps': tps, 'tvq': tvq, 'total': total})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/v1/admin/facture/<int:facture_id>/fermer', methods=['POST'])
@admin_required
def api_admin_fermer_facture(facture_id):
    from invoice_service import generer_pdf_facture as gen_pdf, calculer_taxes
    from datetime import date, timedelta
    import pathlib
    conn = get_db_connection()
    facture = conn.execute("SELECT * FROM factures WHERE id=?", (facture_id,)).fetchone()
    if not facture or facture['statut'] != 'ouverte':
        conn.close()
        return jsonify({'error': 'Facture introuvable ou déjà fermée.'}), 400
    client = conn.execute("SELECT * FROM clients WHERE id=?", (facture['id_client'],)).fetchone()
    lignes = conn.execute("SELECT * FROM facture_lignes WHERE id_facture=? ORDER BY date_service", (facture_id,)).fetchall()
    today = date.today()
    date_str = today.strftime("%Y-%m-%d")
    upload_root  = os.getenv("UPLOAD_ROOT", "/data/uploads")
    factures_dir = os.path.join(upload_root, "factures", f"client_{client['id']}")
    pathlib.Path(factures_dir).mkdir(parents=True, exist_ok=True)
    pdf_path = os.path.join(factures_dir, f"{facture['numero']}.pdf")
    lignes_dict = [{"description": l['description'], "date_service": l['date_service'] or '', "localisation": l['localisation'] or '', "quantite": l['quantite'], "prix_unitaire": l['prix_unitaire']} for l in lignes]
    facture_dict = {"numero": facture['numero'], "date_emission": date_str, "date_echeance": "À la réception", "exonere_taxes": False, "stripe_payment_url": facture['stripe_payment_url']}
    try:
        gen_pdf(facture_dict, lignes_dict, dict(client), pdf_path)
    except Exception as e:
        print(f"[INVOICE] Génération PDF échouée: {e}")
    drive_file_id = None
    try:
        client_row = conn.execute("SELECT factures_folder_id FROM clients WHERE id=?", (client['id'],)).fetchone()
        if client_row and client_row['factures_folder_id']:
            drive_file_id, _ = upload_file(pdf_path, f"{facture['numero']}.pdf", client_row['factures_folder_id'])
    except Exception as e:
        print(f"[DRIVE] Upload facture échoué: {e}")
    conn.execute("UPDATE factures SET statut='envoyee', date_emission=?, date_echeance='À la réception', pdf_path=?, drive_file_id=? WHERE id=?",
                 (date_str, pdf_path, drive_file_id, facture_id))
    conn.commit()
    try:
        if client and client['email'] and int(client['is_email_confirmed'] or 0):
            sous_total = sum(float(l['quantite']) * float(l['prix_unitaire']) for l in lignes)
            _, _, total = calculer_taxes(sous_total, False)
            description = lignes[0]['description'] if len(lignes) == 1 else f"{len(lignes)} services"
            from flask_mail import Message as MailMessage
            msg = MailMessage(f"Votre facture {facture['numero']} — Cocktail Média", sender=app.config['MAIL_DEFAULT_SENDER'], recipients=[client['email']])
            msg.body = f"Bonjour {client['nom_complet']}, veuillez trouver ci-joint votre facture {facture['numero']}."
            msg.html = email_nouvelle_facture(client['nom_complet'], facture['numero'], description, total, "À la réception")
            if os.path.exists(pdf_path):
                with open(pdf_path, 'rb') as f:
                    msg.attach(f"{facture['numero']}.pdf", 'application/pdf', f.read())
            mail.send(msg)
    except Exception as e:
        print(f"[MAIL] Email facture échoué: {e}")
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/facture/<int:facture_id>/payee', methods=['POST'])
@admin_required
def api_admin_facture_payee(facture_id):
    data = request.get_json(silent=True) or {}
    date_paiement = (data.get('date_paiement') or '').strip() or datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    conn.execute("UPDATE factures SET statut='payee', date_paiement=? WHERE id=?", (date_paiement, facture_id))
    # Facture payée → entre dans les revenus (grand livre)
    materialiser_revenu_facture(conn, facture_id, date_paiement)
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/facture/<int:facture_id>/supprimer', methods=['POST'])
@admin_required
def api_admin_supprimer_facture(facture_id):
    conn = get_db_connection()
    facture = conn.execute("SELECT * FROM factures WHERE id=?", (facture_id,)).fetchone()
    if not facture:
        conn.close()
        return jsonify({'error': 'Facture introuvable.'}), 404
    if facture['pdf_path'] and os.path.exists(facture['pdf_path']):
        try: os.remove(facture['pdf_path'])
        except Exception as e: print(f"[INVOICE] Suppression PDF échouée: {e}")
    supprimer_revenu_facture(conn, facture_id)
    conn.execute("DELETE FROM facture_lignes WHERE id_facture=?", (facture_id,))
    conn.execute("DELETE FROM factures WHERE id=?", (facture_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/admin/facture/<int:facture_id>')
@admin_required
def admin_facture(facture_id):
    from datetime import date
    conn = get_db_connection()
    facture = conn.execute("SELECT * FROM factures WHERE id=?", (facture_id,)).fetchone()
    if not facture:
        conn.close()
        flash("Facture introuvable.", "error")
        return redirect(url_for('admin_dashboard'))
    client = conn.execute("SELECT * FROM clients WHERE id=?", (facture['id_client'],)).fetchone()
    lignes = conn.execute("""
        SELECT * FROM facture_lignes WHERE id_facture=? ORDER BY date_service, id
    """, (facture_id,)).fetchall()
    conn.close()
    return render_template('admin_facture.html',
        facture=facture, client=client, lignes=lignes,
        today=str(date.today()))


def regenerer_pdf_facture(facture_id, conn):
    """Regénère le PDF d'une facture après modification."""
    from invoice_service import generer_pdf_facture
    import pathlib
    try:
        facture = conn.execute("SELECT * FROM factures WHERE id=?", (facture_id,)).fetchone()
        client  = conn.execute("SELECT * FROM clients WHERE id=?", (facture['id_client'],)).fetchone()
        lignes  = conn.execute("SELECT * FROM facture_lignes WHERE id_facture=? ORDER BY date_service, id", (facture_id,)).fetchall()

        upload_root  = os.getenv("UPLOAD_ROOT", "/data/uploads")
        factures_dir = os.path.join(upload_root, "factures", f"client_{client['id']}")
        pathlib.Path(factures_dir).mkdir(parents=True, exist_ok=True)
        pdf_path = os.path.join(factures_dir, f"{facture['numero']}.pdf")

        # Supprimer ancien PDF
        if os.path.exists(pdf_path):
            os.remove(pdf_path)

        lignes_dict = [{
            "description":   l['description'],
            "date_service":  l['date_service'] or '',
            "localisation":  l['localisation'] or '',
            "quantite":      l['quantite'],
            "prix_unitaire": l['prix_unitaire'],
        } for l in lignes]

        facture_dict = {
            "numero":             facture['numero'],
            "date_emission":      facture['date_emission'] or '',
            "date_echeance":      facture['date_echeance'] or 'À la réception',
            "exonere_taxes":      False,
            "stripe_payment_url": facture['stripe_payment_url'],
        }

        generer_pdf_facture(facture_dict, lignes_dict, dict(client), pdf_path)
        conn.execute("UPDATE factures SET pdf_path=? WHERE id=?", (pdf_path, facture_id))
        conn.commit()

        # Re-upload Drive
        try:
            client_row = conn.execute(
                "SELECT factures_folder_id FROM clients WHERE id=?", (client['id'],)
            ).fetchone()
            if client_row and client_row['factures_folder_id']:
                drive_file_id, _ = upload_file(
                    pdf_path, f"{facture['numero']}.pdf",
                    client_row['factures_folder_id']
                )
                conn.execute("UPDATE factures SET drive_file_id=? WHERE id=?", (drive_file_id, facture_id))
                conn.commit()
        except Exception as e:
            print(f"[DRIVE] Re-upload facture échoué: {e}")

        print(f"[INVOICE] PDF régénéré: {facture['numero']}")
    except Exception as e:
        print(f"[INVOICE] Régénération PDF échouée: {e}")
@app.route('/admin/facture/<int:facture_id>/ajouter_ligne', methods=['POST'])
@admin_required

def ajouter_ligne_facture_admin(facture_id):
    description  = request.form.get('description', '').strip()
    date_service = request.form.get('date_service', '')
    localisation = request.form.get('localisation', '').strip() or None
    prix         = float(request.form.get('prix_unitaire', 0) or 0)
    quantite     = int(request.form.get('quantite', 1) or 1)
    total_ligne  = round(prix * quantite, 2)

    conn = get_db_connection()
    conn.execute("""
        INSERT INTO facture_lignes
        (id_facture, description, date_service, localisation, quantite, prix_unitaire, total_ligne)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (facture_id, description, date_service, localisation, quantite, prix, total_ligne))

    # Recalculer totaux
    lignes = conn.execute(
        "SELECT prix_unitaire, quantite FROM facture_lignes WHERE id_facture=?", (facture_id,)
    ).fetchall()
    sous_total = sum(float(l['prix_unitaire']) * int(l['quantite']) for l in lignes)
    tps  = round(sous_total * 0.05, 2)
    tvq  = round(sous_total * 0.09975, 2)
    total = round(sous_total + tps + tvq, 2)
    conn.execute("""
        UPDATE factures SET sous_total=?, tps=?, tvq=?, total=? WHERE id=?
    """, (sous_total, tps, tvq, total, facture_id))
    regenerer_pdf_facture(facture_id, conn)
    conn.close()
    flash("Ligne ajoutée.", "success")
    return redirect(url_for('admin_facture', facture_id=facture_id))


@app.route('/admin/ligne/<int:ligne_id>/modifier', methods=['POST'])
@admin_required
def modifier_ligne_facture(ligne_id):
    description = request.form.get('description', '').strip()
    prix        = float(request.form.get('prix_unitaire', 0) or 0)
    quantite    = int(request.form.get('quantite', 1) or 1)
    total_ligne = round(prix * quantite, 2)

    conn = get_db_connection()
    ligne = conn.execute("SELECT id_facture FROM facture_lignes WHERE id=?", (ligne_id,)).fetchone()
    conn.execute("""
        UPDATE facture_lignes SET description=?, prix_unitaire=?, quantite=?, total_ligne=?
        WHERE id=?
    """, (description, prix, quantite, total_ligne, ligne_id))

    # Recalculer totaux
    lignes = conn.execute(
        "SELECT prix_unitaire, quantite FROM facture_lignes WHERE id_facture=?", (ligne['id_facture'],)
    ).fetchall()
    sous_total = sum(float(l['prix_unitaire']) * int(l['quantite']) for l in lignes)
    tps  = round(sous_total * 0.05, 2)
    tvq  = round(sous_total * 0.09975, 2)
    total = round(sous_total + tps + tvq, 2)
    conn.execute("""
        UPDATE factures SET sous_total=?, tps=?, tvq=?, total=? WHERE id=?
    """, (sous_total, tps, tvq, total, ligne['id_facture']))
    conn.commit()
    facture_id = ligne['id_facture']
    regenerer_pdf_facture(facture_id, conn)
    conn.close()
    flash("Ligne mise à jour.", "success")
    return redirect(url_for('admin_facture', facture_id=facture_id))


@app.route('/admin/ligne/<int:ligne_id>/supprimer', methods=['POST'])
@admin_required
def supprimer_ligne_facture(ligne_id):
    conn = get_db_connection()
    ligne = conn.execute("SELECT id_facture FROM facture_lignes WHERE id=?", (ligne_id,)).fetchone()
    facture_id = ligne['id_facture']
    conn.execute("DELETE FROM facture_lignes WHERE id=?", (ligne_id,))

    # Recalculer totaux
    lignes = conn.execute(
        "SELECT prix_unitaire, quantite FROM facture_lignes WHERE id_facture=?", (facture_id,)
    ).fetchall()
    sous_total = sum(float(l['prix_unitaire']) * int(l['quantite']) for l in lignes)
    tps  = round(sous_total * 0.05, 2)
    tvq  = round(sous_total * 0.09975, 2)
    total = round(sous_total + tps + tvq, 2)
    conn.execute("""
        UPDATE factures SET sous_total=?, tps=?, tvq=?, total=? WHERE id=?
    """, (sous_total, tps, tvq, total, facture_id))
    conn.commit()
    regenerer_pdf_facture(facture_id, conn)
    conn.close()
    flash("Ligne supprimée.", "success")
    return redirect(url_for('admin_facture', facture_id=facture_id))
@app.route('/admin/projet/<int:project_id>/archive', methods=['POST'])
@admin_required
def archive_project(project_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    conn.execute("UPDATE projets SET is_archived = 1 WHERE id = ?", (project_id,))
    conn.commit()
    push_notification(conn, projet['id_client'], project_id, f"Votre projet « {projet['nom_projet']} » a été archivé.", type='archive')
    conn.commit()
    try:
        client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
        if client and client['email'] and int(client['is_email_confirmed'] or 0):
            send_email(
                client['email'],
                f"Projet archivé — {projet['nom_projet']}",
                f"Bonjour {client['nom_complet']}, votre projet {projet['nom_projet']} a été archivé.",
                html=email_archive(client['nom_complet'], projet['nom_projet'])
            )
    except Exception as e:
        print(f"[MAIL] Email archivage échoué: {e}")
    conn.close()
    flash("Projet archivé.", "success")
    return redirect(url_for('project_detail', project_id=project_id))


@app.route('/admin/projet/<int:project_id>/unarchive', methods=['POST'])
@admin_required
def unarchive_project(project_id):
    conn = get_db_connection()
    conn.execute("UPDATE projets SET is_archived = 0 WHERE id = ?", (project_id,))
    conn.commit()
    conn.close()
    flash("Projet désarchivé.", "success")
    return redirect(url_for('project_detail', project_id=project_id))
    conn.close()
    flash("Projet désarchivé.", "success")
    return redirect(url_for('project_detail', project_id=project_id))

@app.route('/admin/client/<int:client_id>/notifier_facture', methods=['POST'])
@admin_required
def notifier_facture(client_id):
    conn = get_db_connection()
    client = conn.execute("SELECT * FROM clients WHERE id = ?", (client_id,)).fetchone()
    conn.close()
    if client:
        try:
            lien = get_folder_link(client['factures_folder_id']) if client['factures_folder_id'] else '#'
            body = (
                f"Bonjour {client['nom_complet']},\n\n"
                f"Une nouvelle facture est disponible dans votre portail client.\n"
                f"Vous pouvez la consulter ici : {lien}\n\n"
                f"— L'équipe Cocktail Média"
            )
            send_email(
                client['email'],
                "Nouvelle facture disponible — Cocktail Média",
                body,
                html=email_nouvelle_facture(client['nom_complet'])
            )

            flash("Notification envoyée au client.", "success")
        except Exception as e:
            flash(f"Erreur d'envoi: {e}", "error")
    return redirect(url_for('admin_dashboard'))

# ───────────────────────────────────────────────────────────
# API v1 — Notifications email
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/projet/<int:project_id>/notifier-facture', methods=['POST'])
@admin_required
def api_admin_notifier_facture(project_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
    if not client:
        conn.close()
        return jsonify({'error': 'Client introuvable'}), 404
    facture = conn.execute("""
        SELECT f.* FROM factures f
        JOIN facture_lignes fl ON fl.id_facture = f.id
        WHERE fl.id_projet = ? AND f.statut != 'annulee'
        ORDER BY f.created_at DESC LIMIT 1
    """, (project_id,)).fetchone()
    ligne = None
    if facture:
        ligne = conn.execute(
            "SELECT description FROM facture_lignes WHERE id_facture=? AND id_projet=? LIMIT 1",
            (facture['id'], project_id)
        ).fetchone()
    conn.close()
    try:
        if facture:
            send_email(
                client['email'],
                f"Votre facture {facture['numero']} — Cocktail Média",
                f"Bonjour {client['nom_complet']}, veuillez trouver ci-joint votre facture {facture['numero']}.",
                html=email_nouvelle_facture(
                    client['nom_complet'], facture['numero'],
                    ligne['description'] if ligne else None,
                    facture['total'], facture['date_echeance'] or 'À la réception'
                ),
                attachments=[(f"{facture['numero']}.pdf", open(facture['pdf_path'], 'rb').read())]
                    if facture['pdf_path'] and os.path.exists(facture['pdf_path']) else None
            )
        else:
            lien = get_folder_link(client['factures_folder_id']) if client['factures_folder_id'] else '#'
            body = (
                f"Bonjour {client['nom_complet']},\n\n"
                f"Une nouvelle facture est disponible dans votre portail client.\n"
                f"Vous pouvez la consulter ici : {lien}\n\n"
                f"— L'équipe Cocktail Média"
            )
            send_email(
                client['email'],
                "Nouvelle facture disponible — Cocktail Média",
                body,
                html=email_nouvelle_facture(client['nom_complet'])
            )
    except Exception as e:
        print(f"[MAIL] api_notifier_facture: {e}")
        return jsonify({'error': 'Erreur envoi email'}), 500
    return jsonify({'success': True})


@app.route('/api/v1/admin/marketing/notifier-felix/<mois>', methods=['POST'])
@admin_required
def api_marketing_notifier_felix(mois):
    import json as _json
    conn = get_db_connection()
    posts = conn.execute("""
        SELECT * FROM marketing_posts
        WHERE strftime('%Y-%m', date_publication) = ?
        AND demande_envoyee = 0
        ORDER BY date_publication ASC
    """, (mois,)).fetchall()
    if not posts:
        conn.close()
        return jsonify({'success': True, 'count': 0, 'message': 'Aucun nouveau post à envoyer.'})
    lignes_html = ""
    lignes_txt = ""
    for post in posts:
        plateformes = _json.loads(post['plateformes']) if post['plateformes'] else []
        plates_str = ', '.join(plateformes)
        lignes_html += f"<tr><td style='padding:8px;border-bottom:1px solid #eee;font-weight:600;'>{post['date_publication']}</td><td style='padding:8px;border-bottom:1px solid #eee;'>{post['titre']}</td><td style='padding:8px;border-bottom:1px solid #eee;'>{plates_str}</td><td style='padding:8px;border-bottom:1px solid #eee;color:#666;'>{post['description'] or '—'}</td></tr>"
        lignes_txt += f"- {post['date_publication']} | {post['titre']} | {plates_str}\n"
    html_body = f"""<div style="font-family:Montserrat,sans-serif;max-width:700px;margin:auto;background:#fff;border-radius:12px;overflow:hidden;"><div style="background:#c0321a;padding:32px;text-align:center;"><h1 style="color:#fff;margin:0;font-size:22px;">📋 Demande de visuels — {mois}</h1></div><div style="padding:32px;"><p style="color:#2b2b2b;">Marie-Christine a planifié <strong>{len(posts)} post(s)</strong> pour le mois de {mois}. Voici les visuels à créer :</p><table style="width:100%;border-collapse:collapse;margin-top:16px;"><tr style="background:#fdecea;"><th style="padding:8px;text-align:left;font-size:12px;color:#888;">DATE</th><th style="padding:8px;text-align:left;font-size:12px;color:#888;">TITRE</th><th style="padding:8px;text-align:left;font-size:12px;color:#888;">PLATEFORMES</th><th style="padding:8px;text-align:left;font-size:12px;color:#888;">DESCRIPTION</th></tr>{lignes_html}</table><p style="margin-top:24px;color:#888;font-size:13px;">Connecte-toi au portail pour déposer les visuels.</p></div></div>"""
    try:
        send_email(
            'felix.dumont@cocktailmedia.ca',
            f"📋 Demande de visuels — {mois} ({len(posts)} posts)",
            f"Demande de visuels pour {mois}:\n{lignes_txt}",
            html=html_body
        )
        for post in posts:
            conn.execute("UPDATE marketing_posts SET demande_envoyee=1 WHERE id=?", (post['id'],))
        conn.commit()
    except Exception as e:
        print(f"[MAIL] api_notifier_felix: {e}")
        conn.close()
        return jsonify({'error': 'Erreur envoi email'}), 500
    conn.close()
    return jsonify({'success': True, 'count': len(posts)})


@app.route('/api/v1/admin/marketing/notifier-marie/<mois>', methods=['POST'])
@admin_required
def api_marketing_notifier_marie(mois):
    import json as _json
    conn = get_db_connection()
    posts = conn.execute("""
        SELECT * FROM marketing_posts
        WHERE strftime('%Y-%m', date_publication) = ?
        AND statut IN ('visuels prêts', 'planifié', 'Planifié')
        AND todo_marie_done = 0
        ORDER BY date_publication ASC
    """, (mois,)).fetchall()
    if not posts:
        conn.close()
        return jsonify({'success': True, 'count': 0, 'message': 'Aucun visuel à notifier.'})
    lignes_html = ""
    lignes_txt = ""
    for post in posts:
        plateformes = _json.loads(post['plateformes']) if post['plateformes'] else []
        plates_str = ', '.join(plateformes)
        lignes_html += f"<tr><td style='padding:8px;border-bottom:1px solid #eee;font-weight:600;'>{post['date_publication']}</td><td style='padding:8px;border-bottom:1px solid #eee;'>{post['titre']}</td><td style='padding:8px;border-bottom:1px solid #eee;'>{plates_str}</td></tr>"
        lignes_txt += f"- {post['date_publication']} | {post['titre']} | {plates_str}\n"
    html_body = f"""<div style="font-family:Montserrat,sans-serif;max-width:700px;margin:auto;background:#fff;border-radius:12px;overflow:hidden;"><div style="background:#c0321a;padding:32px;text-align:center;"><h1 style="color:#fff;margin:0;font-size:22px;">🎨 Visuels prêts — {mois}</h1></div><div style="padding:32px;"><p style="color:#2b2b2b;">Félix a déposé les visuels pour <strong>{len(posts)} post(s)</strong>. Ils sont prêts à être publiés !</p><table style="width:100%;border-collapse:collapse;margin-top:16px;"><tr style="background:#fdecea;"><th style="padding:8px;text-align:left;font-size:12px;color:#888;">DATE</th><th style="padding:8px;text-align:left;font-size:12px;color:#888;">TITRE</th><th style="padding:8px;text-align:left;font-size:12px;color:#888;">PLATEFORMES</th></tr>{lignes_html}</table><p style="margin-top:24px;color:#888;font-size:13px;">Connecte-toi au portail pour télécharger les visuels.</p></div></div>"""
    try:
        send_email(
            'marie-christine.blanchette@cocktailmedia.ca',
            f"🎨 Visuels prêts à publier — {mois} ({len(posts)} posts)",
            f"Visuels prêts pour {mois}:\n{lignes_txt}",
            html=html_body
        )
        for post in posts:
            conn.execute("UPDATE marketing_posts SET todo_marie_done=1 WHERE id=?", (post['id'],))
        conn.commit()
    except Exception as e:
        print(f"[MAIL] api_notifier_marie: {e}")
        conn.close()
        return jsonify({'error': 'Erreur envoi email'}), 500
    conn.close()
    return jsonify({'success': True, 'count': len(posts)})

# ───────────────────────────────────────────────────────────
# API v1 — Marketing CRUD
# ───────────────────────────────────────────────────────────

def _marketing_post_fichiers(conn, post_id) -> list:
    rows = conn.execute(
        "SELECT id, filename, created_at FROM marketing_post_fichiers WHERE id_post = ? ORDER BY created_at ASC",
        (post_id,)
    ).fetchall()
    return [{'id': r['id'], 'filename': r['filename'], 'created_at': r['created_at']} for r in rows]

@app.route('/api/v1/admin/marketing', methods=['GET'])
@admin_required
def api_admin_marketing_list():
    import json as _json
    from datetime import datetime as _dt
    mois = request.args.get('mois', _dt.now().strftime('%Y-%m'))
    conn = get_db_connection()
    posts = conn.execute("""
        SELECT * FROM marketing_posts
        WHERE strftime('%Y-%m', date_publication) = ?
        ORDER BY date_publication ASC
    """, (mois,)).fetchall()
    result = [{
        'id': p['id'],
        'titre': p['titre'],
        'description': p['description'],
        'date_publication': p['date_publication'],
        'plateformes': _json.loads(p['plateformes']) if p['plateformes'] else [],
        'statut': p['statut'],
        'drive_folder_id': p['drive_folder_id'],
        'demande_envoyee': bool(p['demande_envoyee']),
        'todo_felix_done': bool(p['todo_felix_done']),
        'todo_marie_done': bool(p['todo_marie_done']),
        'created_at': p['created_at'],
        'fichiers': _marketing_post_fichiers(conn, p['id']),
    } for p in posts]
    conn.close()
    return jsonify(result)


@app.route('/api/v1/admin/marketing', methods=['POST'])
@admin_required
def api_admin_marketing_create():
    import json as _json
    data = request.get_json(force=True)
    titre = (data.get('titre') or '').strip()
    description = data.get('description', '')
    date_publication = data.get('date_publication', '')
    plateformes = data.get('plateformes', [])
    if not titre or not date_publication:
        return jsonify({'error': 'titre et date_publication requis'}), 400
    conn = get_db_connection()
    cur = conn.execute("""
        INSERT INTO marketing_posts (titre, description, date_publication, plateformes, statut, created_by)
        VALUES (?, ?, ?, ?, 'planifié', ?)
    """, (titre, description, date_publication, _json.dumps(plateformes), session.get('user_id')))
    conn.commit()
    post_id = cur.lastrowid
    conn.close()
    return jsonify({'success': True, 'id': post_id}), 201


@app.route('/api/v1/admin/marketing/<int:post_id>', methods=['DELETE'])
@admin_required
def api_admin_marketing_delete(post_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM marketing_posts WHERE id = ?", (post_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/marketing/<int:post_id>/upload', methods=['POST'])
@admin_required
def api_admin_marketing_upload(post_id):
    file = request.files.get('file')
    if not file or file.filename == '':
        return jsonify({'error': 'Aucun fichier reçu'}), 400
    if not allowed(file.filename):
        return jsonify({'error': 'Extension non autorisée'}), 400
    conn = get_db_connection()
    post = conn.execute("SELECT * FROM marketing_posts WHERE id = ?", (post_id,)).fetchone()
    if not post:
        conn.close()
        return jsonify({'error': 'Post introuvable'}), 404

    safe_name = f"{uuid.uuid4().hex[:8]}_{secure_filename(file.filename)}"
    base_dir = os.path.join(app.config["UPLOAD_ROOT"], "marketing", f"post_{post_id}")
    pathlib.Path(base_dir).mkdir(parents=True, exist_ok=True)
    save_path = os.path.join(base_dir, safe_name)
    file.save(save_path)

    drive_folder_id = post['drive_folder_id']
    try:
        if not drive_folder_id:
            root_id = os.getenv('MARKETING_DRIVE_FOLDER_ID')
            mois_folder = create_folder(post['date_publication'][:7], parent_id=root_id)
            drive_folder_id = create_folder(f"{post['date_publication']} — {post['titre']}", parent_id=mois_folder)
            conn.execute("UPDATE marketing_posts SET drive_folder_id = ? WHERE id = ?", (drive_folder_id, post_id))
        from drive_service import upload_file as drive_upload
        drive_upload(save_path, safe_name, drive_folder_id)
    except Exception as e:
        print(f"[DRIVE] Upload visuel marketing échoué: {e}")

    conn.execute(
        "INSERT INTO marketing_post_fichiers (id_post, filename, filepath, uploaded_by) VALUES (?, ?, ?, ?)",
        (post_id, file.filename, save_path, session.get('user_id'))
    )
    # Déposer un visuel = la tâche de création est faite — coche automatiquement,
    # et propage vers la tâche perso / le roadmap si ce post y est lié.
    conn.execute("UPDATE marketing_posts SET todo_felix_done = 1 WHERE id = ?", (post_id,))
    conn.commit()
    linked_roadmap_todo_id = post['linked_roadmap_todo_id'] if 'linked_roadmap_todo_id' in post.keys() else None
    if linked_roadmap_todo_id:
        production_id = _admin_id_for_role(conn, 'production')
        is_assigned_to_production = conn.execute(
            "SELECT 1 FROM roadmap_todo_assignees WHERE roadmap_todo_id = ? AND admin_id = ?",
            (linked_roadmap_todo_id, production_id)
        ).fetchone()
        if is_assigned_to_production:
            _sync_roadmap_todo_completion(conn, linked_roadmap_todo_id, True)
    fichiers = _marketing_post_fichiers(conn, post_id)
    conn.close()
    return jsonify({'success': True, 'fichiers': fichiers, 'todo_felix_done': True}), 201


@app.route('/api/v1/admin/marketing/<int:post_id>/fichier/<int:fichier_id>')
@admin_required
def api_admin_marketing_serve_fichier(post_id, fichier_id):
    conn = get_db_connection()
    f = conn.execute(
        "SELECT * FROM marketing_post_fichiers WHERE id = ? AND id_post = ?", (fichier_id, post_id)
    ).fetchone()
    conn.close()
    if not f:
        return jsonify({'error': 'Fichier introuvable'}), 404
    upload_root = os.path.realpath(app.config["UPLOAD_ROOT"])
    filepath = os.path.realpath(f['filepath'])
    if not filepath.startswith(upload_root + os.sep):
        return jsonify({'error': 'Accès non autorisé'}), 403
    if not os.path.exists(filepath):
        return jsonify({'error': 'Fichier manquant sur le disque'}), 404
    return send_file(filepath, as_attachment=False, download_name=f['filename'])


@app.route('/api/v1/admin/marketing/<int:post_id>/fichier/<int:fichier_id>', methods=['DELETE'])
@admin_required
def api_admin_marketing_delete_fichier(post_id, fichier_id):
    conn = get_db_connection()
    f = conn.execute(
        "SELECT * FROM marketing_post_fichiers WHERE id = ? AND id_post = ?", (fichier_id, post_id)
    ).fetchone()
    if not f:
        conn.close()
        return jsonify({'error': 'Fichier introuvable'}), 404
    conn.execute("DELETE FROM marketing_post_fichiers WHERE id = ?", (fichier_id,))
    conn.commit()
    conn.close()
    try:
        if os.path.exists(f['filepath']):
            os.remove(f['filepath'])
    except Exception as e:
        print(f"[MARKETING] Suppression fichier disque échouée: {e}")
    return jsonify({'success': True})


@app.route('/api/v1/admin/marketing/<int:post_id>/todo-toggle', methods=['POST'])
@admin_required
def api_admin_marketing_todo_toggle(post_id):
    conn = get_db_connection()
    post = conn.execute("SELECT * FROM marketing_posts WHERE id = ?", (post_id,)).fetchone()
    if not post:
        conn.close()
        return jsonify({'error': 'Post non trouvé'}), 404
    new_val = 0 if post['todo_felix_done'] else 1
    conn.execute("UPDATE marketing_posts SET todo_felix_done = ? WHERE id = ?", (new_val, post_id))
    conn.commit()
    linked_roadmap_todo_id = post['linked_roadmap_todo_id'] if 'linked_roadmap_todo_id' in post.keys() else None
    if linked_roadmap_todo_id:
        production_id = _admin_id_for_role(conn, 'production')
        is_assigned_to_production = conn.execute(
            "SELECT 1 FROM roadmap_todo_assignees WHERE roadmap_todo_id = ? AND admin_id = ?",
            (linked_roadmap_todo_id, production_id)
        ).fetchone()
        if is_assigned_to_production:
            _sync_roadmap_todo_completion(conn, linked_roadmap_todo_id, bool(new_val))
    conn.close()
    return jsonify({'success': True, 'todo_felix_done': bool(new_val)})


@app.route('/api/v1/admin/marketing/<int:post_id>/publier', methods=['POST'])
@admin_required
def api_admin_marketing_publier(post_id):
    conn = get_db_connection()
    post = conn.execute("SELECT * FROM marketing_posts WHERE id = ?", (post_id,)).fetchone()
    if not post:
        conn.close()
        return jsonify({'error': 'Post non trouvé'}), 404
    conn.execute("UPDATE marketing_posts SET statut = 'publié' WHERE id = ?", (post_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ───────────────────────────────────────────────────────────
# API v1 — Services CRUD
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/services', methods=['POST'])
@admin_required
def api_admin_services_create():
    data = request.get_json(force=True)
    nom_service = (data.get('nom_service') or '').strip()
    if not nom_service:
        return jsonify({'error': 'nom_service requis'}), 400
    description = data.get('description', '')
    icon = data.get('icon', 'default')
    categorie = (data.get('categorie') or '').strip() or None
    try:
        prix = float(data.get('prix') or 0)
    except (TypeError, ValueError):
        prix = 0
    conn = get_db_connection()
    try:
        cur = conn.execute(
            "INSERT INTO services (nom_service, description, icon, categorie, prix, actif) VALUES (?, ?, ?, ?, ?, 1)",
            (nom_service, description, icon, categorie, prix)
        )
        conn.commit()
        service_id = cur.lastrowid
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 409
    conn.close()
    return jsonify({'success': True, 'id': service_id}), 201


@app.route('/api/v1/admin/services/<int:service_id>', methods=['DELETE'])
@admin_required
def api_admin_services_delete(service_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM checklist_model_items WHERE id_service = ?", (service_id,))
    conn.execute("DELETE FROM services WHERE id = ?", (service_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/services/<int:service_id>', methods=['PUT'])
@admin_required
def api_admin_services_update(service_id):
    """Mise à jour partielle : tarification (prix, durée affichée) et bascule actif/inactif."""
    data = request.get_json(silent=True) or {}
    conn = get_db_connection()
    existing = conn.execute("SELECT id FROM services WHERE id = ?", (service_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'Service introuvable'}), 404

    champs, valeurs = [], []
    if 'prix' in data:
        try:
            champs.append('prix = ?'); valeurs.append(float(data.get('prix') or 0))
        except (TypeError, ValueError):
            conn.close()
            return jsonify({'error': 'Prix invalide'}), 400
    if 'duree_affichee' in data:
        champs.append('duree_affichee = ?'); valeurs.append((data.get('duree_affichee') or '').strip() or None)
    if 'actif' in data:
        champs.append('actif = ?'); valeurs.append(1 if data.get('actif') else 0)

    if champs:
        valeurs.append(service_id)
        conn.execute(f"UPDATE services SET {', '.join(champs)} WHERE id = ?", valeurs)
        conn.commit()

    row = conn.execute("SELECT * FROM services WHERE id = ?", (service_id,)).fetchone()
    conn.close()
    return jsonify({
        'success': True,
        'id': row['id'],
        'prix': row['prix'],
        'duree_affichee': row['duree_affichee'],
        'actif': bool(row['actif']) if row['actif'] is not None else True,
    })


@app.route('/api/v1/admin/services/<int:service_id>/extras', methods=['POST'])
@admin_required
def api_admin_services_add_extra(service_id):
    data = request.get_json(silent=True) or {}
    nom = (data.get('nom') or '').strip()
    if not nom:
        return jsonify({'error': 'nom requis'}), 400
    try:
        prix = float(data.get('prix') or 0)
    except (TypeError, ValueError):
        return jsonify({'error': 'Prix invalide'}), 400
    conn = get_db_connection()
    existing = conn.execute("SELECT id FROM services WHERE id = ?", (service_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'Service introuvable'}), 404
    position = conn.execute(
        "SELECT COALESCE(MAX(position), -1) + 1 FROM service_extras WHERE id_service = ?", (service_id,)
    ).fetchone()[0]
    cur = conn.execute(
        "INSERT INTO service_extras (id_service, nom, prix, position) VALUES (?, ?, ?, ?)",
        (service_id, nom, prix, position)
    )
    conn.commit()
    extra_id = cur.lastrowid
    conn.close()
    return jsonify({'success': True, 'id': extra_id, 'nom': nom, 'prix': prix}), 201


@app.route('/api/v1/admin/services/extras/<int:extra_id>', methods=['DELETE'])
@admin_required
def api_admin_services_delete_extra(extra_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM service_extras WHERE id = ?", (extra_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/services/items/<int:item_id>', methods=['PATCH'])
@admin_required
def api_admin_services_update_item(item_id):
    """Bascule obligatoire/optionnel (ou renomme) un checkpoint sans le recréer."""
    data = request.get_json(silent=True) or {}
    conn = get_db_connection()
    existing = conn.execute("SELECT id FROM checklist_model_items WHERE id = ?", (item_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({'error': 'Checkpoint introuvable'}), 404

    champs, valeurs = [], []
    if 'is_required' in data:
        champs.append('is_required = ?'); valeurs.append(1 if data.get('is_required') else 0)
    if 'nom_item' in data:
        nom_item = (data.get('nom_item') or '').strip()
        if not nom_item:
            conn.close()
            return jsonify({'error': 'nom_item requis'}), 400
        champs.append('nom_item = ?'); valeurs.append(nom_item)

    if champs:
        valeurs.append(item_id)
        conn.execute(f"UPDATE checklist_model_items SET {', '.join(champs)} WHERE id = ?", valeurs)
        conn.commit()

    row = conn.execute("SELECT * FROM checklist_model_items WHERE id = ?", (item_id,)).fetchone()
    conn.close()
    return jsonify({'success': True, 'id': row['id'], 'nom_item': row['nom_item'], 'is_required': bool(row['is_required'])})


@app.route('/api/v1/admin/services/<int:service_id>/items', methods=['POST'])
@admin_required
def api_admin_services_add_item(service_id):
    data = request.get_json(force=True)
    nom_item = (data.get('nom_item') or '').strip()
    if not nom_item:
        return jsonify({'error': 'nom_item requis'}), 400
    is_required = int(data.get('is_required', 1))
    conn = get_db_connection()
    cur = conn.execute("""
        INSERT INTO checklist_model_items (id_service, nom_item, requires_file, is_required, item_type, file_category)
        VALUES (?, ?, 0, ?, 'document', 'autre')
    """, (service_id, nom_item, is_required))
    conn.commit()
    item_id = cur.lastrowid
    conn.close()
    return jsonify({'success': True, 'id': item_id}), 201


@app.route('/api/v1/admin/services/items/<int:item_id>', methods=['DELETE'])
@admin_required
def api_admin_services_delete_item(item_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM checklist_model_items WHERE id = ?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/services/with-items', methods=['POST'])
@admin_required
def api_admin_services_create_with_items():
    data = request.get_json(force=True)
    nom_service = (data.get('nom_service') or '').strip()
    if not nom_service:
        return jsonify({'error': 'nom_service requis'}), 400
    slug = (data.get('slug') or '').strip() or None
    categorie = (data.get('categorie') or '').strip() or None
    description = data.get('description', '')
    icon = data.get('icon', 'default')
    prix = float(data.get('prix', 0) or 0)
    exonere_taxes = 1 if data.get('exonere_taxes') else 0
    localisation_requise = 1 if data.get('localisation_requise') else 0
    documents_requis = int(data.get('documents_requis', 1))
    appel_exploratoire_requis = 1 if data.get('appel_exploratoire_requis') else 0
    decision_board_requis = 1 if data.get('decision_board_requis') else 0
    duree_seance_minutes = int(data.get('duree_seance_minutes', 0) or 0)
    duree_tournage_minutes = int(data.get('duree_tournage_minutes', 0) or 0)
    duree_production_minutes = int(data.get('duree_production_minutes', 0) or 0)
    duree_finalisation_minutes = int(data.get('duree_finalisation_minutes', 0) or 0)
    items = data.get('items', [])
    conn = get_db_connection()
    try:
        cur = conn.execute(
            """INSERT INTO services
               (nom_service, description, icon, prix, exonere_taxes,
                localisation_requise, documents_requis, appel_exploratoire_requis,
                decision_board_requis, duree_seance_minutes, duree_tournage_minutes,
                duree_production_minutes, duree_finalisation_minutes, slug, categorie)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (nom_service, description, icon, prix, exonere_taxes,
             localisation_requise, documents_requis, appel_exploratoire_requis,
             decision_board_requis, duree_seance_minutes, duree_tournage_minutes,
             duree_production_minutes, duree_finalisation_minutes, slug, categorie)
        )
        service_id = cur.lastrowid
        for idx, item in enumerate(items):
            nom_item = (item.get('nom_item') or '').strip()
            if not nom_item:
                continue
            requires_file = int(item.get('requires_file', 0))
            is_required = int(item.get('is_required', 1))
            item_type = item.get('item_type', 'document')
            file_category = item.get('file_category', 'autre')
            field_type = item.get('field_type', 'check')
            conn.execute("""
                INSERT INTO checklist_model_items
                    (id_service, nom_item, requires_file, is_required, item_type, file_category, field_type, position)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (service_id, nom_item, requires_file, is_required, item_type, file_category, field_type, idx))
        conn.commit()
        return jsonify({'success': True, 'id': service_id}), 201
    except sqlite3.IntegrityError:
        conn.rollback()
        return jsonify({'error': f"Le service '{nom_service}' existe déjà."}), 409
    finally:
        conn.close()


# ───────────────────────────────────────────────────────────
# Decision Board
# ───────────────────────────────────────────────────────────
# ───────────────────────────────────────────────────────────
# ───────────────────────────────────────────────────────────
# API v1 — Decision Board
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/projet/<int:projet_id>/decision-board', methods=['GET'])
@login_required
def api_client_get_decision_board(projet_id):
    client_id = session['user_id']
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id = ? AND id_client = ?", (projet_id, client_id)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    board = conn.execute("SELECT * FROM decision_boards WHERE id_projet = ? AND is_active = 1", (projet_id,)).fetchone()
    if not board:
        conn.close()
        return jsonify({'error': 'Decision board non disponible'}), 404
    choices = conn.execute("SELECT * FROM decision_board_choices WHERE id_projet = ?", (projet_id,)).fetchone()
    conn.close()
    return jsonify({
        'nom_projet': projet['nom_projet'],
        'config_json': board['config_json'],
        'icons': [{'url': board[f'icon{i}_url'], 'name': board[f'icon{i}_name'] or f'Style {i}'} for i in range(1, 5)],
        'logos': [{'url': board[f'logo{i}_url'], 'name': board[f'logo{i}_name'] or f'Composition {i}'} for i in range(1, 5)],
        'choices': {
            'choix_directions': choices['choix_directions'],
            'choix_noms': choices['choix_noms'],
            'nom_suggestion': choices['nom_suggestion'],
            'choix_icones': choices['choix_icones'],
            'choix_typos': choices['choix_typos'],
            'choix_palettes': choices['choix_palettes'],
            'choix_logos': choices['choix_logos'],
            'commentaires': choices['commentaires'],
        } if choices else None,
    })


@app.route('/api/v1/projet/<int:projet_id>/decision-board', methods=['POST'])
@login_required
def api_client_submit_decision_board(projet_id):
    client_id = session['user_id']
    data = request.get_json(silent=True) or {}
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id = ? AND id_client = ?", (projet_id, client_id)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    board = conn.execute("SELECT id FROM decision_boards WHERE id_projet = ? AND is_active = 1", (projet_id,)).fetchone()
    if not board:
        conn.close()
        return jsonify({'error': 'Decision board non disponible'}), 404

    choix_directions = (data.get('choix_directions') or '').strip()
    choix_noms       = (data.get('choix_noms') or '').strip()
    choix_icones     = (data.get('choix_icones') or '').strip()
    choix_typos      = (data.get('choix_typos') or '').strip()
    choix_palettes   = (data.get('choix_palettes') or '').strip()
    choix_logos      = (data.get('choix_logos') or '').strip()
    nom_suggestion   = (data.get('nom_suggestion') or '').strip()
    commentaires     = (data.get('commentaires') or '').strip()

    already = conn.execute("SELECT id FROM decision_board_choices WHERE id_projet=?", (projet_id,)).fetchone()
    if already:
        conn.execute("""
            UPDATE decision_board_choices SET
                choix_directions=?, choix_noms=?, choix_icones=?, choix_typos=?,
                choix_palettes=?, choix_logos=?, nom_suggestion=?, commentaires=?,
                submitted_at=CURRENT_TIMESTAMP
            WHERE id_projet=?
        """, (choix_directions, choix_noms, choix_icones, choix_typos, choix_palettes, choix_logos, nom_suggestion, commentaires, projet_id))
    else:
        conn.execute("""
            INSERT INTO decision_board_choices
            (id_projet, choix_directions, choix_noms, choix_icones, choix_typos, choix_palettes, choix_logos, nom_suggestion, commentaires)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (projet_id, choix_directions, choix_noms, choix_icones, choix_typos, choix_palettes, choix_logos, nom_suggestion, commentaires))
    conn.commit()

    try:
        client = conn.execute("SELECT * FROM clients WHERE id=?", (client_id,)).fetchone()
        settings = get_notification_settings()
        if settings["admin_emails"]:
            subject = f"[Decision Board] {client['nom_complet']} a soumis ses choix — {projet['nom_projet']}"
            body = (
                f"Client : {client['nom_complet']} ({client['email']})\n"
                f"Projet : {projet['nom_projet']}\n\n"
                f"Directions : {choix_directions or 'Aucun'}\n"
                f"Noms : {choix_noms or 'Aucun'}\n"
                f"Suggestion nom : {nom_suggestion or 'Aucune'}\n"
                f"Icônes : {choix_icones or 'Aucun'}\n"
                f"Typographies : {choix_typos or 'Aucun'}\n"
                f"Palettes : {choix_palettes or 'Aucun'}\n"
                f"Logos : {choix_logos or 'Aucun'}\n\n"
                f"Commentaires : {commentaires or 'Aucun'}"
            )
            send_email(settings["admin_emails"], subject, body)
        push_admin_notif(
            conn,
            titre=f"Decision board soumis — {projet['nom_projet']}",
            message=f"{client['nom_complet']} a soumis ses choix.",
            type='info',
            lien=f"/admin/projet/{projet_id}/decision",
        )
        conn.commit()
    except Exception as e:
        print(f"[MAIL] Email decision board (client JSON) échoué: {e}")

    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/projet/<int:project_id>/decision', methods=['GET'])
@admin_required
def api_admin_get_decision_board(project_id):
    conn = get_db_connection()
    board   = conn.execute("SELECT * FROM decision_boards WHERE id_projet=?", (project_id,)).fetchone()
    choices = conn.execute("SELECT * FROM decision_board_choices WHERE id_projet=?", (project_id,)).fetchone()
    conn.close()

    def _board_field(col):
        return board[col] if board and board[col] else None

    return jsonify({
        'board': {
            'config_json':    _board_field('config_json'),
            'is_active':      bool(board['is_active']) if board else False,
            'icon1_url':  _board_field('icon1_url'),  'icon1_name': _board_field('icon1_name'),
            'icon2_url':  _board_field('icon2_url'),  'icon2_name': _board_field('icon2_name'),
            'icon3_url':  _board_field('icon3_url'),  'icon3_name': _board_field('icon3_name'),
            'icon4_url':  _board_field('icon4_url'),  'icon4_name': _board_field('icon4_name'),
            'logo1_url':  _board_field('logo1_url'),  'logo1_name': _board_field('logo1_name'),
            'logo1_url2': _board_field('logo1_url2'), 'logo1_flat_url': _board_field('logo1_flat_url'),
            'logo2_url':  _board_field('logo2_url'),  'logo2_name': _board_field('logo2_name'),
            'logo2_url2': _board_field('logo2_url2'), 'logo2_flat_url': _board_field('logo2_flat_url'),
            'logo3_url':  _board_field('logo3_url'),  'logo3_name': _board_field('logo3_name'),
            'logo3_url2': _board_field('logo3_url2'), 'logo3_flat_url': _board_field('logo3_flat_url'),
            'logo4_url':  _board_field('logo4_url'),  'logo4_name': _board_field('logo4_name'),
            'logo4_url2': _board_field('logo4_url2'), 'logo4_flat_url': _board_field('logo4_flat_url'),
        } if board else None,
        'choices': {
            'choix_directions': choices['choix_directions'],
            'choix_noms':       choices['choix_noms'],
            'nom_suggestion':   choices['nom_suggestion'],
            'choix_icones':     choices['choix_icones'],
            'choix_typos':      choices['choix_typos'],
            'choix_palettes':   choices['choix_palettes'],
            'choix_logos':      choices['choix_logos'],
            'commentaires':     choices['commentaires'],
            'submitted_at':     choices['submitted_at'],
        } if choices else None,
    })


@app.route('/api/v1/admin/projet/<int:project_id>/decision', methods=['POST'])
@admin_required
def api_admin_save_decision_board(project_id):
    import json as _json, tempfile
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable.'}), 404

    is_active  = bool(request.form.get('is_active') == '1')
    config_raw = request.form.get('config_json', '').strip()
    config_json = None
    if config_raw:
        try:
            _json.loads(config_raw)
            config_json = config_raw
        except Exception:
            conn.close()
            return jsonify({'error': 'JSON invalide. Vérifiez la syntaxe.'}), 400

    existing = conn.execute("SELECT * FROM decision_boards WHERE id_projet=?", (project_id,)).fetchone()
    board_was_active = bool(existing and int(existing['is_active'] or 0))

    assets_folder_id = existing['assets_folder_id'] if existing and existing['assets_folder_id'] else None
    if not assets_folder_id:
        try:
            client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
            parent = client['drive_folder_id'] if client and client['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            assets_folder_id = create_folder("DecisionBoardAssets", parent_id=parent)
        except Exception as e:
            print(f"[DRIVE] Création dossier assets échouée: {e}")

    ALLOWED_ASSETS = {'png', 'svg', 'avif', 'webp', 'jpg', 'jpeg'}

    def upload_asset(field_name, old_url):
        file = request.files.get(field_name)
        if not file or file.filename == '':
            return old_url
        ext = file.filename.rsplit('.', 1)[-1].lower()
        if ext not in ALLOWED_ASSETS or not assets_folder_id:
            return old_url
        try:
            safe_name = secure_filename(file.filename)
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
            file.save(tmp.name); tmp.close()
            file_id, _ = upload_file(tmp.name, safe_name, assets_folder_id)
            os.unlink(tmp.name)
            return make_file_public(file_id)
        except Exception as e:
            print(f"[DRIVE] Upload asset {field_name} échoué: {e}")
            return old_url

    def g(col): return existing[col] if existing and existing[col] else None

    fields = {}
    for i in range(1, 5):
        fields[f'icon{i}_url']      = upload_asset(f'icon{i}', g(f'icon{i}_url'))
        fields[f'icon{i}_name']     = (request.form.get(f'icon{i}_name') or g(f'icon{i}_name') or f'Style {"ABCD"[i-1]}').strip()
        fields[f'logo{i}_url']      = upload_asset(f'logo{i}', g(f'logo{i}_url'))
        fields[f'logo{i}_name']     = (request.form.get(f'logo{i}_name') or g(f'logo{i}_name') or f'Composition {"ABCD"[i-1]}').strip()
        fields[f'logo{i}_url2']     = upload_asset(f'logo{i}_url2', g(f'logo{i}_url2'))
        fields[f'logo{i}_flat_url'] = upload_asset(f'logo{i}_flat', g(f'logo{i}_flat_url'))

    try:
        if existing:
            conn.execute("""
                UPDATE decision_boards SET
                    config_json=?, is_active=?, assets_folder_id=?,
                    icon1_url=?, icon1_name=?, icon2_url=?, icon2_name=?,
                    icon3_url=?, icon3_name=?, icon4_url=?, icon4_name=?,
                    logo1_url=?, logo1_name=?, logo2_url=?, logo2_name=?,
                    logo3_url=?, logo3_name=?, logo4_url=?, logo4_name=?,
                    logo1_flat_url=?, logo2_flat_url=?, logo3_flat_url=?, logo4_flat_url=?,
                    logo1_url2=?, logo2_url2=?, logo3_url2=?, logo4_url2=?,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id_projet=?
            """, (config_json, is_active, assets_folder_id,
                  fields['icon1_url'], fields['icon1_name'], fields['icon2_url'], fields['icon2_name'],
                  fields['icon3_url'], fields['icon3_name'], fields['icon4_url'], fields['icon4_name'],
                  fields['logo1_url'], fields['logo1_name'], fields['logo2_url'], fields['logo2_name'],
                  fields['logo3_url'], fields['logo3_name'], fields['logo4_url'], fields['logo4_name'],
                  fields['logo1_flat_url'], fields['logo2_flat_url'], fields['logo3_flat_url'], fields['logo4_flat_url'],
                  fields['logo1_url2'], fields['logo2_url2'], fields['logo3_url2'], fields['logo4_url2'],
                  project_id))
        else:
            conn.execute("""
                INSERT OR REPLACE INTO decision_boards (
                    id_projet, config_json, is_active, assets_folder_id,
                    icon1_url, icon1_name, icon2_url, icon2_name,
                    icon3_url, icon3_name, icon4_url, icon4_name,
                    logo1_url, logo1_name, logo2_url, logo2_name,
                    logo3_url, logo3_name, logo4_url, logo4_name,
                    logo1_flat_url, logo2_flat_url, logo3_flat_url, logo4_flat_url,
                    logo1_url2, logo2_url2, logo3_url2, logo4_url2
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (project_id, config_json, is_active, assets_folder_id,
                  fields['icon1_url'], fields['icon1_name'], fields['icon2_url'], fields['icon2_name'],
                  fields['icon3_url'], fields['icon3_name'], fields['icon4_url'], fields['icon4_name'],
                  fields['logo1_url'], fields['logo1_name'], fields['logo2_url'], fields['logo2_name'],
                  fields['logo3_url'], fields['logo3_name'], fields['logo4_url'], fields['logo4_name'],
                  fields['logo1_flat_url'], fields['logo2_flat_url'], fields['logo3_flat_url'], fields['logo4_flat_url'],
                  fields['logo1_url2'], fields['logo2_url2'], fields['logo3_url2'], fields['logo4_url2']))

        if is_active and not board_was_active:
            push_notification(conn, projet['id_client'], project_id,
                f"Votre decision board pour « {projet['nom_projet']} » est disponible !", type='decision_board')
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback(); conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/admin/projet/<int:project_id>/decision', methods=['GET', 'POST'])
@admin_required
def admin_decision_board(project_id):
    import json, tempfile
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        flash("Projet introuvable.", "error")
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        is_active  = 1 if request.form.get('is_active') else 0

        # JSON unique — validation
        config_raw = request.form.get('config_json', '').strip()
        config_json = None
        config = {}
        if config_raw:
            try:
                config = json.loads(config_raw)
                config_json = config_raw
            except Exception:
                flash("JSON invalide. Vérifiez la syntaxe.", "error")
                conn.close()
                return redirect(url_for('admin_decision_board', project_id=project_id))

        # Récupérer le board existant
        existing = conn.execute("SELECT * FROM decision_boards WHERE id_projet=?", (project_id,)).fetchone()
        board_was_active = bool(existing and int(existing['is_active'] or 0))

        # Créer le dossier DecisionBoardAssets sur Drive si nécessaire
        assets_folder_id = existing['assets_folder_id'] if existing and existing['assets_folder_id'] else None
        if not assets_folder_id:
            try:
                client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
                parent = client['drive_folder_id'] if client and client['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
                assets_folder_id = create_folder("DecisionBoardAssets", parent_id=parent)
            except Exception as e:
                print(f"[DRIVE] Création dossier assets échouée: {e}")

        # Upload images — icônes et logos
        ALLOWED_ASSETS = {'png', 'svg', 'avif', 'webp', 'jpg', 'jpeg'}

        def upload_asset(field_name, old_url):
            file = request.files.get(field_name)
            if not file or file.filename == '':
                return old_url  # Garde l'ancienne URL
            ext = file.filename.rsplit('.', 1)[-1].lower()
            if ext not in ALLOWED_ASSETS:
                return old_url
            if not assets_folder_id:
                return old_url
            try:
                safe_name = secure_filename(file.filename)
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
                file.save(tmp.name)
                tmp.close()
                file_id, _ = upload_file(tmp.name, safe_name, assets_folder_id)
                os.unlink(tmp.name)
                public_url = make_file_public(file_id)
                return public_url
            except Exception as e:
                print(f"[DRIVE] Upload asset {field_name} échoué: {e}")
                return old_url

        old = existing if existing else {}
        def g(col): return old[col] if old and old[col] else None

        icon1_url = upload_asset('icon1', g('icon1_url'))
        icon2_url = upload_asset('icon2', g('icon2_url'))
        icon3_url = upload_asset('icon3', g('icon3_url'))
        icon4_url = upload_asset('icon4', g('icon4_url'))
        logo1_url = upload_asset('logo1', g('logo1_url'))
        logo2_url = upload_asset('logo2', g('logo2_url'))
        logo3_url = upload_asset('logo3', g('logo3_url'))
        logo4_url = upload_asset('logo4', g('logo4_url'))
        logo1_flat_url = upload_asset('logo1_flat', g('logo1_flat_url'))
        logo2_flat_url = upload_asset('logo2_flat', g('logo2_flat_url'))
        logo3_flat_url = upload_asset('logo3_flat', g('logo3_flat_url'))
        logo4_flat_url = upload_asset('logo4_flat', g('logo4_flat_url'))
        logo1_url2 = upload_asset('logo1_url2', g('logo1_url2'))
        logo2_url2 = upload_asset('logo2_url2', g('logo2_url2'))
        logo3_url2 = upload_asset('logo3_url2', g('logo3_url2'))
        logo4_url2 = upload_asset('logo4_url2', g('logo4_url2'))

        icon1_name = request.form.get('icon1_name', g('icon1_name') or 'Style A').strip()
        icon2_name = request.form.get('icon2_name', g('icon2_name') or 'Style B').strip()
        icon3_name = request.form.get('icon3_name', g('icon3_name') or 'Style C').strip()
        icon4_name = request.form.get('icon4_name', g('icon4_name') or 'Style D').strip()
        logo1_name = request.form.get('logo1_name', g('logo1_name') or 'Composition A').strip()
        logo2_name = request.form.get('logo2_name', g('logo2_name') or 'Composition B').strip()
        logo3_name = request.form.get('logo3_name', g('logo3_name') or 'Composition C').strip()
        logo4_name = request.form.get('logo4_name', g('logo4_name') or 'Composition D').strip()

        if existing:
            conn.execute("""
                UPDATE decision_boards SET
                    config_json=?, is_active=?, assets_folder_id=?,
                    icon1_url=?, icon1_name=?, icon2_url=?, icon2_name=?,
                    icon3_url=?, icon3_name=?, icon4_url=?, icon4_name=?,
                    logo1_url=?, logo1_name=?, logo2_url=?, logo2_name=?,
                    logo3_url=?, logo3_name=?, logo4_url=?, logo4_name=?,
                    logo1_flat_url=?, logo2_flat_url=?, logo3_flat_url=?, logo4_flat_url=?,
                    logo1_url2=?, logo2_url2=?, logo3_url2=?, logo4_url2=?,
                    updated_at=CURRENT_TIMESTAMP
                WHERE id_projet=?
            """, (config_json, is_active, assets_folder_id,
                  icon1_url, icon1_name, icon2_url, icon2_name,
                  icon3_url, icon3_name, icon4_url, icon4_name,
                  logo1_url, logo1_name, logo2_url, logo2_name,
                  logo3_url, logo3_name, logo4_url, logo4_name,
                  logo1_flat_url, logo2_flat_url, logo3_flat_url, logo4_flat_url,
                  logo1_url2, logo2_url2, logo3_url2, logo4_url2,
                  project_id))
        else:
            conn.execute("""
                INSERT OR REPLACE INTO decision_boards (
                    id_projet, config_json, is_active, assets_folder_id,
                    icon1_url, icon1_name, icon2_url, icon2_name,
                    icon3_url, icon3_name, icon4_url, icon4_name,
                    logo1_url, logo1_name, logo2_url, logo2_name,
                    logo3_url, logo3_name, logo4_url, logo4_name,
                    logo1_flat_url, logo2_flat_url, logo3_flat_url, logo4_flat_url,
                    logo1_url2, logo2_url2, logo3_url2, logo4_url2
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (project_id, config_json, is_active, assets_folder_id,
                  icon1_url, icon1_name, icon2_url, icon2_name,
                  icon3_url, icon3_name, icon4_url, icon4_name,
                  logo1_url, logo1_name, logo2_url, logo2_name,
                  logo3_url, logo3_name, logo4_url, logo4_name,
                  logo1_flat_url, logo2_flat_url, logo3_flat_url, logo4_flat_url,
                  logo1_url2, logo2_url2, logo3_url2, logo4_url2))
        if is_active and not board_was_active:
            push_notification(conn, projet['id_client'], project_id,
                f"Votre decision board pour « {projet['nom_projet']} » est disponible !", type='decision_board')
        conn.commit()
        conn.close()
        flash("Decision board sauvegardé.", "success")
        return redirect(url_for('admin_decision_board', project_id=project_id))

    board   = conn.execute("SELECT * FROM decision_boards WHERE id_projet=?", (project_id,)).fetchone()
    choices = conn.execute("SELECT * FROM decision_board_choices WHERE id_projet=?", (project_id,)).fetchone()
    conn.close()
    return render_template('admin_decision_board.html', projet=projet, board=board, choices=choices)

@app.route('/projet/<int:project_id>/decision')
@login_required
def client_decision_board(project_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    is_owner = projet and (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not projet or not (is_owner or is_admin):
        conn.close()
        flash("Accès non autorisé.", "error")
        return redirect(url_for('dashboard'))

    if normalize_status(projet['statut']) != "En révision":
        conn.close()
        flash("Le decision board n'est pas encore disponible.", "error")
        return redirect(url_for('project_detail', project_id=project_id))

    board = conn.execute("SELECT * FROM decision_boards WHERE id_projet=? AND is_active=1", (project_id,)).fetchone()
    if not board:
        conn.close()
        flash("Le decision board n'est pas encore configuré.", "error")
        return redirect(url_for('project_detail', project_id=project_id))

    already_submitted = conn.execute("SELECT id FROM decision_board_choices WHERE id_projet=?", (project_id,)).fetchone()
    conn.close()
    
    conn2 = get_db_connection()
    client = conn2.execute("SELECT nom_complet, nom_entreprise FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
    conn2.close()
    client_name = client['nom_entreprise'] or client['nom_complet'] if client else 'Votre marque'
    return render_template('client_decision_board.html', projet=projet, board=board, already_submitted=already_submitted, client_name=client_name)

@app.route('/projet/<int:project_id>/decision/submit', methods=['POST'])
@login_required
def submit_decision_board(project_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (project_id,)).fetchone()
    is_owner = projet and (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not projet or not (is_owner or is_admin):
        conn.close()
        flash("Accès non autorisé.", "error")
        return redirect(url_for('dashboard'))

    already = conn.execute("SELECT id FROM decision_board_choices WHERE id_projet=?", (project_id,)).fetchone()

    choix_directions = request.form.get('choix_directions', '')
    choix_noms       = request.form.get('choix_noms', '')
    choix_icones     = request.form.get('choix_icones', '')
    choix_typos      = request.form.get('choix_typos', '')
    choix_palettes   = request.form.get('choix_palettes', '')
    choix_logos      = request.form.get('choix_logos', '')
    nom_suggestion   = request.form.get('nom_suggestion', '').strip()
    commentaires     = request.form.get('commentaires', '').strip()

    if already:
        conn.execute("""
            UPDATE decision_board_choices SET
                choix_directions=?, choix_noms=?, choix_icones=?, choix_typos=?,
                choix_palettes=?, choix_logos=?, nom_suggestion=?, commentaires=?,
                submitted_at=CURRENT_TIMESTAMP
            WHERE id_projet=?
        """, (choix_directions, choix_noms, choix_icones, choix_typos, choix_palettes, choix_logos, nom_suggestion, commentaires, project_id))
    else:
        conn.execute("""
            INSERT INTO decision_board_choices
            (id_projet, choix_directions, choix_noms, choix_icones, choix_typos, choix_palettes, choix_logos, nom_suggestion, commentaires)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (project_id, choix_directions, choix_noms, choix_icones, choix_typos, choix_palettes, choix_logos, nom_suggestion, commentaires))
    conn.commit()

    # Statut → Documents reçus + notifications
    conn.execute("UPDATE projets SET statut='Documents reçus' WHERE id=?", (project_id,))
    conn.commit()

    try:
        client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
        settings = get_notification_settings()

        # Email admin avec tous les choix
        if settings["admin_emails"]:
            subject = f"[Decision Board] {client['nom_complet']} a soumis ses choix — {projet['nom_projet']}"
            body = (
                f"Client : {client['nom_complet']} ({client['email']})\n"
                f"Projet : {projet['nom_projet']}\n\n"
                f"Directions : {choix_directions or 'Aucun'}\n"
                f"Noms : {choix_noms or 'Aucun'}\n"
                f"Suggestion nom : {nom_suggestion or 'Aucune'}\n"
                f"Icônes : {choix_icones or 'Aucun'}\n"
                f"Typographies : {choix_typos or 'Aucun'}\n"
                f"Palettes : {choix_palettes or 'Aucun'}\n"
                f"Logos : {choix_logos or 'Aucun'}\n\n"
                f"Commentaires : {commentaires or 'Aucun'}\n\n"
                f"Voir le projet : {url_for('project_detail', project_id=project_id, _external=True)}"
            )
            send_email(settings["admin_emails"], subject, body)

        # Email confirmation au client
        if client and client['email']:
            send_email_client(client,
                f"Vos choix ont été reçus — {projet['nom_projet']}",
                f"Bonjour {client['nom_complet']}, nous avons bien reçu vos choix pour le projet {projet['nom_projet']}. Notre équipe va maintenant analyser vos préférences et démarrer les travaux.",
                html=email_documents_recus(client['nom_complet'], projet['nom_projet'],
                    url_for('project_detail', project_id=project_id, _external=True)))
            push_notification(conn, client['id'], project_id,
                f"Vos choix du decision board ont été reçus pour « {projet['nom_projet']} ».", type='decision_board')
            conn.commit()
    except Exception as e:
        print(f"[MAIL] Email decision board échoué: {e}")

    conn.close()
    flash("Vos choix ont été soumis avec succès. Merci !", "success")
    return redirect(url_for('project_detail', project_id=project_id))

# ───────────────────────────────────────────────────────────
# Routes Factures Admin
# ───────────────────────────────────────────────────────────
@app.route('/admin/facture/<int:facture_id>/fermer', methods=['POST'])
@admin_required
def fermer_facture(facture_id):
    from invoice_service import generer_pdf_facture, calculer_taxes
    from datetime import date, timedelta
    conn = get_db_connection()
    facture = conn.execute("SELECT * FROM factures WHERE id=?", (facture_id,)).fetchone()
    if not facture or facture['statut'] != 'ouverte':
        conn.close()
        flash("Facture introuvable ou déjà fermée.", "error")
        return redirect(url_for('admin_dashboard'))

    client = conn.execute("SELECT * FROM clients WHERE id=?", (facture['id_client'],)).fetchone()
    lignes = conn.execute("SELECT * FROM facture_lignes WHERE id_facture=? ORDER BY date_service", (facture_id,)).fetchall()

    today    = date.today()
    date_str = today.strftime("%Y-%m-%d")

    # Générer PDF
    import pathlib, os
    upload_root  = os.getenv("UPLOAD_ROOT", "/data/uploads")
    factures_dir = os.path.join(upload_root, "factures", f"client_{client['id']}")
    pathlib.Path(factures_dir).mkdir(parents=True, exist_ok=True)
    pdf_path = os.path.join(factures_dir, f"{facture['numero']}.pdf")

    lignes_dict = []
    for l in lignes:
        lignes_dict.append({
            "description":   l['description'],
            "date_service":  l['date_service'] or '',
            "localisation":  l['localisation'] or '',
            "quantite":      l['quantite'],
            "prix_unitaire": l['prix_unitaire'],
        })

    facture_dict = {
        "numero":            facture['numero'],
        "date_emission":     date_str,
        "date_echeance":     "À la réception",
        "exonere_taxes":     False,
        "stripe_payment_url": facture['stripe_payment_url'],
    }

    generer_pdf_facture(facture_dict, lignes_dict, dict(client), pdf_path)

    # Upload Drive
    drive_file_id = None
    try:
        client_row = conn.execute(
            "SELECT factures_folder_id FROM clients WHERE id=?", (client['id'],)
        ).fetchone()
        if client_row and client_row['factures_folder_id']:
            drive_file_id, _ = upload_file(
                pdf_path, f"{facture['numero']}.pdf", client_row['factures_folder_id']
            )
    except Exception as e:
        print(f"[DRIVE] Upload facture mensuelle échoué: {e}")

    # Mettre à jour DB
    conn.execute("""
        UPDATE factures SET statut='envoyee', date_emission=?, date_echeance='À la réception',
        pdf_path=?, drive_file_id=? WHERE id=?
    """, (date_str, pdf_path, drive_file_id, facture_id))
    conn.commit()

    # Envoyer email avec PDF
    try:
        if client and client['email'] and int(client['is_email_confirmed'] or 0):
            from flask_mail import Message as MailMessage
            msg = MailMessage(
                f"Votre facture {facture['numero']} — Cocktail Média",
                sender=app.config['MAIL_DEFAULT_SENDER'],
                recipients=[client['email']]
            )
            msg.body = f"Bonjour {client['nom_complet']}, veuillez trouver ci-joint votre facture {facture['numero']}."
            msg.html = email_nouvelle_facture(client['nom_complet'])
            if os.path.exists(pdf_path):
                with open(pdf_path, 'rb') as f:
                    msg.attach(f"{facture['numero']}.pdf", 'application/pdf', f.read())
            mail.send(msg)
    except Exception as e:
        print(f"[MAIL] Email facture mensuelle échoué: {e}")

    conn.close()
    flash(f"Facture {facture['numero']} fermée et envoyée.", "success")
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/facture/<int:facture_id>/payee', methods=['POST'])
@admin_required
def marquer_facture_payee(facture_id):
    today = datetime.now().strftime('%Y-%m-%d')
    conn = get_db_connection()
    conn.execute("UPDATE factures SET statut='payee', date_paiement=? WHERE id=?", (today, facture_id))
    materialiser_revenu_facture(conn, facture_id, today)
    conn.commit()
    conn.close()
    flash("Facture marquée comme payée.", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/facture/<int:facture_id>/supprimer', methods=['POST'])
@admin_required
def supprimer_facture(facture_id):
    conn = get_db_connection()
    facture = conn.execute("SELECT * FROM factures WHERE id=?", (facture_id,)).fetchone()
    if not facture:
        conn.close()
        flash("Facture introuvable.", "error")
        return redirect(url_for('admin_dashboard'))

    # Supprimer PDF local
    if facture['pdf_path'] and os.path.exists(facture['pdf_path']):
        try:
            os.remove(facture['pdf_path'])
        except Exception as e:
            print(f"[INVOICE] Suppression PDF échouée: {e}")

    # Supprimer lignes + facture en DB (+ revenu lié dans le grand livre)
    supprimer_revenu_facture(conn, facture_id)
    conn.execute("DELETE FROM facture_lignes WHERE id_facture=?", (facture_id,))
    conn.execute("DELETE FROM factures WHERE id=?", (facture_id,))
    conn.commit()
    conn.close()
    flash(f"Facture {facture['numero']} supprimée.", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/facture/<int:facture_id>/download')
@admin_required
def admin_download_facture(facture_id):
    conn = get_db_connection()
    facture = conn.execute("SELECT * FROM factures WHERE id=?", (facture_id,)).fetchone()
    conn.close()
    if not facture or not facture['pdf_path'] or not os.path.exists(facture['pdf_path']):
        flash("PDF introuvable.", "error")
        return redirect(url_for('admin_dashboard'))
    return send_file(facture['pdf_path'], as_attachment=True, download_name=f"{facture['numero']}.pdf")


@app.route('/api/v1/admin/facture/<int:facture_id>/download')
@admin_required
def api_admin_download_facture(facture_id):
    # Équivalent de admin_download_facture ci-dessus, sous /api/v1/ — le portail-next ne
    # proxifie que /api/* vers Flask (voir next.config.*), donc un <a href="/admin/facture/...">
    # tombe dans le vide côté Next.js (404) sans jamais atteindre le backend.
    conn = get_db_connection()
    facture = conn.execute("SELECT * FROM factures WHERE id=?", (facture_id,)).fetchone()
    conn.close()
    if not facture or not facture['pdf_path'] or not os.path.exists(facture['pdf_path']):
        return jsonify({'error': 'PDF introuvable.'}), 404
    resp = send_file(facture['pdf_path'], as_attachment=True, download_name=f"{facture['numero']}.pdf")
    # no-store (pas juste no-cache) : le fichier peut avoir été régénéré entre deux clics sur
    # le même lien — on ne veut jamais qu'un onglet PDF déjà ouvert ou le cache HTTP serve une ancienne version.
    resp.headers['Cache-Control'] = 'no-store, must-revalidate'
    return resp

# ───────────────────────────────────────────────────────────
# Identité visuelle — API JSON (portail-next)
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/projet/<int:projet_id>/identite', methods=['GET'])
@admin_required
def api_admin_identite_get(projet_id):
    import json as _json
    conn = get_db_connection()
    try:
        projet = conn.execute("SELECT * FROM projets WHERE id = ?", (projet_id,)).fetchone()
        if not projet:
            return jsonify({'error': 'Projet introuvable'}), 404
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (projet_id,)).fetchone()
        if not iv:
            return jsonify({'error': 'Identité visuelle non configurée'}), 404
        logos_rows = conn.execute("SELECT * FROM iv_logos WHERE id_iv = ?", (iv['id'],)).fetchall()
        fonts     = conn.execute("SELECT * FROM iv_fonts WHERE id_iv = ?", (iv['id'],)).fetchall()
        decls     = conn.execute("SELECT * FROM iv_declinaisons WHERE id_iv = ? ORDER BY position, id", (iv['id'],)).fetchall()
        mockups   = conn.execute("SELECT * FROM iv_mockups WHERE id_iv = ? ORDER BY position, id", (iv['id'],)).fetchall()
        logos = {}
        for l in logos_rows:
            logos[l['variante']] = {
                'public_url':  l['public_url'],
                'preview_url': l['preview_url'] if 'preview_url' in l.keys() else None,
                'filename':    l['filename'],
            }
        palette = []
        if iv['palette_json']:
            try:
                palette = _json.loads(iv['palette_json'])
            except Exception:
                palette = []
        return jsonify({
            'id':                  iv['id'],
            'nom_projet':          projet['nom_projet'],
            'statut_publication':  'publie' if iv['is_complete'] else 'brouillon',
            'contexte':            iv['contexte'] if 'contexte' in iv.keys() else None,
            'palette':             palette,
            'logos':               logos,
            'fonts':               [{'nom_font': f['nom_font'], 'google_font_url': f['google_font_url'], 'usage': f['usage']} for f in fonts],
            'declinaisons':        [{'id': d['id'], 'public_url': d['public_url'], 'label': d['label'], 'filename': d['filename']} for d in decls],
            'mockups':             [{'id': m['id'], 'public_url': m['public_url'], 'label': m['label'], 'filename': m['filename']} for m in mockups],
            'zip_url':             f'/projet/{projet_id}/identite/zip',
        })
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:projet_id>/identite/publier', methods=['POST'])
@admin_required
def api_admin_identite_publier(projet_id):
    conn = get_db_connection()
    try:
        projet = conn.execute("SELECT * FROM projets WHERE id = ?", (projet_id,)).fetchone()
        if not projet:
            return jsonify({'error': 'Projet introuvable'}), 404
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (projet_id,)).fetchone()
        if not iv:
            return jsonify({'error': 'Identité visuelle non configurée'}), 404
        conn.execute("UPDATE identite_visuelle SET is_complete = 1 WHERE id = ?", (iv['id'],))
        conn.execute("UPDATE projets SET statut='En révision' WHERE id = ?", (projet_id,))
        push_notification(conn, projet['id_client'], projet_id,
            f"Votre identité visuelle pour « {projet['nom_projet']} » est prête !", type='identite_visuelle')
        conn.commit()
        try:
            client = conn.execute("SELECT * FROM clients WHERE id = ?", (projet['id_client'],)).fetchone()
            if client and client['email']:
                lien = url_for('projet_identite', project_id=projet_id, _external=True)
                send_email_client(client,
                    f"Votre identité visuelle est prête — {projet['nom_projet']}",
                    f"Bonjour {client['nom_complet']}, votre identité visuelle est prête pour révision.",
                    html=email_identite_visuelle_prete(client['nom_complet'], projet['nom_projet'], lien))
        except Exception as e:
            print(f"[MAIL] api_admin_identite_publier: {e}")
        return jsonify({'ok': True})
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:projet_id>/identite/logo', methods=['POST'])
@admin_required
def api_admin_identite_logo(projet_id):
    import tempfile
    conn = get_db_connection()
    try:
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (projet_id,)).fetchone()
        if not iv:
            return jsonify({'ok': False, 'error': 'Identité visuelle non trouvée'}), 404
        variante = request.form.get('variant', '').strip()
        import re as _re
        if not variante or not _re.match(r'^[a-z][a-z0-9_]{0,49}$', variante):
            return jsonify({'ok': False, 'error': 'Variante invalide'}), 400
        fichier = request.files.get('file')
        if not fichier or fichier.filename == '':
            return jsonify({'ok': False, 'error': 'Fichier manquant'}), 400
        ext = fichier.filename.rsplit('.', 1)[-1].lower()
        if ext not in {'png', 'webp', 'jpg', 'jpeg', 'svg'}:
            return jsonify({'ok': False, 'error': 'Format non supporté'}), 400
        iv_folder_id = iv['iv_folder_id']
        if not iv_folder_id:
            projet = conn.execute("SELECT p.*, c.drive_folder_id as client_folder_id, c.nom_entreprise, c.nom_complet, c.email FROM projets p JOIN clients c ON c.id = p.id_client WHERE p.id = ?", (projet_id,)).fetchone()
            proj_folder = projet['drive_folder_id'] if projet else None
            if not proj_folder:
                # Créer le dossier projet sous le dossier client
                client_folder = projet['client_folder_id'] if projet else None
                if not client_folder:
                    nom_client = (projet['nom_entreprise'] or projet['nom_complet']) if projet else 'Client'
                    client_folder = create_folder(nom_client, parent_id=os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID'))
                    conn.execute("UPDATE clients SET drive_folder_id = ? WHERE id = (SELECT id_client FROM projets WHERE id = ?)", (client_folder, projet_id))
                proj_folder = create_folder(projet['nom_projet'], parent_id=client_folder)
                conn.execute("UPDATE projets SET drive_folder_id = ? WHERE id = ?", (proj_folder, projet_id))
                conn.commit()
            iv_folder_id = create_folder("Identité visuelle", parent_id=proj_folder)
            conn.execute("UPDATE identite_visuelle SET iv_folder_id = ? WHERE id = ?", (iv_folder_id, iv['id']))
            conn.commit()
            if projet and projet['email']:
                try:
                    share_folder_with_user(iv_folder_id, projet['email'])
                except Exception:
                    pass
        safe_name = f"logo_{variante}.{ext}"
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
        fichier.save(tmp.name)
        tmp.close()
        try:
            file_id, _ = upload_file(tmp.name, safe_name, iv_folder_id)
            public_url = make_file_public(file_id)
        finally:
            os.unlink(tmp.name)
        existing = conn.execute(
            "SELECT id FROM iv_logos WHERE id_iv = ? AND variante = ?", (iv['id'], variante)
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE iv_logos SET drive_file_id=?, public_url=?, filename=? WHERE id=?",
                (file_id, public_url, fichier.filename, existing['id'])
            )
        else:
            conn.execute(
                "INSERT INTO iv_logos (id_iv, variante, drive_file_id, public_url, filename) VALUES (?,?,?,?,?)",
                (iv['id'], variante, file_id, public_url, fichier.filename)
            )
        conn.commit()
        return jsonify({'ok': True, 'public_url': public_url, 'filename': fichier.filename})
    except Exception as e:
        print(f"[API] Upload logo {variante} échoué: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


def _get_or_create_iv(conn, projet_id):
    iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (projet_id,)).fetchone()
    if not iv:
        conn.execute("INSERT INTO identite_visuelle (id_projet, is_complete) VALUES (?, 0)", (projet_id,))
        conn.commit()
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (projet_id,)).fetchone()
    return iv


def _ensure_iv_folder(conn, iv, projet_id):
    if iv['iv_folder_id']:
        return iv['iv_folder_id']
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (projet_id,)).fetchone()
    parent = projet['drive_folder_id'] if projet and projet['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
    folder_id = create_folder("Identité visuelle", parent_id=parent)
    conn.execute("UPDATE identite_visuelle SET iv_folder_id = ? WHERE id = ?", (folder_id, iv['id']))
    conn.commit()
    return folder_id


@app.route('/api/v1/admin/projet/<int:projet_id>/identite', methods=['POST'])
@admin_required
def api_admin_identite_init(projet_id):
    conn = get_db_connection()
    try:
        if not conn.execute("SELECT id FROM projets WHERE id = ?", (projet_id,)).fetchone():
            return jsonify({'error': 'Projet introuvable'}), 404
        iv = _get_or_create_iv(conn, projet_id)
        return jsonify({'ok': True, 'id': iv['id']})
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:projet_id>/identite/palette', methods=['POST'])
@admin_required
def api_admin_palette_save(projet_id):
    import json as _json
    conn = get_db_connection()
    try:
        iv = _get_or_create_iv(conn, projet_id)
        palette = request.get_json(force=True).get('palette', [])
        conn.execute("UPDATE identite_visuelle SET palette_json = ? WHERE id = ?",
                     (_json.dumps(palette), iv['id']))
        conn.commit()
        return jsonify({'ok': True})
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:projet_id>/identite/fonts', methods=['POST'])
@admin_required
def api_admin_fonts_save(projet_id):
    conn = get_db_connection()
    try:
        iv = _get_or_create_iv(conn, projet_id)
        fonts = request.get_json(force=True).get('fonts', [])
        conn.execute("DELETE FROM iv_fonts WHERE id_iv = ?", (iv['id'],))
        for f in fonts:
            nom = f.get('nom_font', '').strip()
            if not nom:
                continue
            usage = f.get('usage', '').strip()
            gurl = f.get('google_font_url') or f"https://fonts.google.com/specimen/{nom.replace(' ', '+')}"
            conn.execute("INSERT INTO iv_fonts (id_iv, nom_font, google_font_url, usage) VALUES (?,?,?,?)",
                         (iv['id'], nom, gurl, usage))
        conn.commit()
        return jsonify({'ok': True})
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:projet_id>/identite/declinaison', methods=['POST'])
@admin_required
def api_admin_declinaison_upload(projet_id):
    import tempfile
    conn = get_db_connection()
    try:
        iv = _get_or_create_iv(conn, projet_id)
        fichier = request.files.get('file')
        label = request.form.get('label', '').strip()
        if not fichier:
            return jsonify({'ok': False, 'error': 'Fichier manquant'}), 400
        ext = fichier.filename.rsplit('.', 1)[-1].lower()
        if ext not in {'png', 'webp', 'jpg', 'jpeg', 'svg'}:
            return jsonify({'ok': False, 'error': 'Format non supporté'}), 400
        folder_id = _ensure_iv_folder(conn, iv, projet_id)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
        fichier.save(tmp.name); tmp.close()
        try:
            file_id, _ = upload_file(tmp.name, fichier.filename, folder_id)
            public_url = make_file_public(file_id)
        finally:
            os.unlink(tmp.name)
        conn.execute("INSERT INTO iv_declinaisons (id_iv, drive_file_id, public_url, filename, label) VALUES (?,?,?,?,?)",
                     (iv['id'], file_id, public_url, fichier.filename, label or None))
        conn.commit()
        row_id = conn.execute("SELECT last_insert_rowid() as id").fetchone()['id']
        return jsonify({'ok': True, 'id': row_id, 'public_url': public_url, 'filename': fichier.filename, 'label': label or None})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:projet_id>/identite/declinaison/<int:decl_id>', methods=['DELETE'])
@admin_required
def api_admin_declinaison_delete(projet_id, decl_id):
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM iv_declinaisons WHERE id = ?", (decl_id,))
        conn.commit()
        return jsonify({'ok': True})
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:projet_id>/identite/mockup', methods=['POST'])
@admin_required
def api_admin_mockup_upload(projet_id):
    import tempfile
    conn = get_db_connection()
    try:
        iv = _get_or_create_iv(conn, projet_id)
        fichier = request.files.get('file')
        label = request.form.get('label', '').strip()
        if not fichier:
            return jsonify({'ok': False, 'error': 'Fichier manquant'}), 400
        ext = fichier.filename.rsplit('.', 1)[-1].lower()
        if ext not in {'png', 'webp', 'jpg', 'jpeg'}:
            return jsonify({'ok': False, 'error': 'Format non supporté'}), 400
        folder_id = _ensure_iv_folder(conn, iv, projet_id)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
        fichier.save(tmp.name); tmp.close()
        try:
            file_id, _ = upload_file(tmp.name, fichier.filename, folder_id)
            public_url = make_file_public(file_id)
        finally:
            os.unlink(tmp.name)
        conn.execute("INSERT INTO iv_mockups (id_iv, drive_file_id, public_url, filename, label) VALUES (?,?,?,?,?)",
                     (iv['id'], file_id, public_url, fichier.filename, label or None))
        conn.commit()
        row_id = conn.execute("SELECT last_insert_rowid() as id").fetchone()['id']
        return jsonify({'ok': True, 'id': row_id, 'public_url': public_url, 'filename': fichier.filename, 'label': label or None})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:projet_id>/identite/mockup/<int:mockup_id>', methods=['DELETE'])
@admin_required
def api_admin_mockup_delete(projet_id, mockup_id):
    conn = get_db_connection()
    try:
        conn.execute("DELETE FROM iv_mockups WHERE id = ?", (mockup_id,))
        conn.commit()
        return jsonify({'ok': True})
    finally:
        conn.close()


# ───────────────────────────────────────────────────────────
# Identité visuelle — Interface admin (HTML legacy)
# ───────────────────────────────────────────────────────────

@app.route('/admin/projet/<int:project_id>/identite', methods=['GET', 'POST'])
@admin_required
def admin_identite_visuelle(project_id):
    import tempfile
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        flash("Projet introuvable.", "error")
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        is_complete = 1 if request.form.get('is_complete') else 0
        nom_compagnie = request.form.get('nom_compagnie', '').strip()
        contexte = request.form.get('contexte', '').strip()
        sous_titre = request.form.get('sous_titre', '').strip()
        palette_json = request.form.get('palette_json', '').strip()

        # Récupérer ou créer l'entrée identite_visuelle
        iv = conn.execute(
            "SELECT * FROM identite_visuelle WHERE id_projet = ?",
            (project_id,)
        ).fetchone()
        iv_was_complete = bool(iv and int(iv['is_complete'] or 0))

        # Créer le dossier Drive "Identité visuelle" si nécessaire
        iv_folder_id = iv['iv_folder_id'] if iv and iv['iv_folder_id'] else None
        if not iv_folder_id:
            try:
                parent = projet['drive_folder_id'] if projet['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
                iv_folder_id = create_folder("Identité visuelle", parent_id=parent)
            except Exception as e:
                print(f"[DRIVE] Création dossier IV échouée: {e}")

        if not iv:
            conn.execute("""
                INSERT INTO identite_visuelle (
                    id_projet,
                    is_complete,
                    iv_folder_id,
                    nom_compagnie,
                    sous_titre,
                    palette_json
                )
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                project_id,
                is_complete,
                iv_folder_id,
                nom_compagnie,
                sous_titre,
                palette_json
            ))
            conn.commit()
            iv = conn.execute(
                "SELECT * FROM identite_visuelle WHERE id_projet = ?",
                (project_id,)
            ).fetchone()
        else:
            conn.execute("""
                UPDATE identite_visuelle
                SET is_complete = ?,
                    iv_folder_id = ?,
                    nom_compagnie = ?,
                    sous_titre = ?,
                    palette_json = ?,
                    contexte = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (
                is_complete,
                iv_folder_id,
                nom_compagnie,
                sous_titre,
                palette_json,
                contexte,
                iv['id']
            ))
            conn.commit()
            iv = conn.execute(
                "SELECT * FROM identite_visuelle WHERE id_projet = ?",
                (project_id,)
            ).fetchone()

        if is_complete and not iv_was_complete:
            push_notification(conn, projet['id_client'], project_id,
                f"Votre identité visuelle pour « {projet['nom_projet']} » est prête !", type='identite_visuelle')
            conn.commit()

        id_iv = iv['id']
        ALLOWED_IV = {'png', 'webp', 'jpg', 'jpeg', 'avif', 'svg', 'pdf', 'ai', 'eps', 'zip'}

        def upload_logo(field_name):
            file = request.files.get(field_name)
            if not file or file.filename == '':
                return None, None
            ext = file.filename.rsplit('.', 1)[-1].lower()
            if ext not in ALLOWED_IV:
                return None, None
            if not iv_folder_id:
                return None, None
            try:
                safe_name = secure_filename(file.filename)
                if not safe_name:
                    return None, None
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f'.{ext}')
                file.save(tmp.name)
                tmp.close()
                file_id, _ = upload_file(tmp.name, safe_name, iv_folder_id)
                os.unlink(tmp.name)
                public_url = make_file_public(file_id)
                return file_id, public_url
            except Exception as e:
                print(f"[DRIVE] Upload logo {field_name} échoué: {e}")
                return None, None

        # Upload les 3 variantes
        for variante in ['principal', 'icone', 'variante']:
            file_id, public_url = upload_logo(f'logo_{variante}')
            prev_file_id, prev_url = upload_logo(f'logo_{variante}_preview')
            existing = conn.execute(
                "SELECT id, drive_file_id, public_url FROM iv_logos WHERE id_iv = ? AND variante = ?",
                (id_iv, variante)
            ).fetchone()

            if file_id and prev_file_id:
                if existing:
                    conn.execute("""
                        UPDATE iv_logos SET drive_file_id=?, public_url=?, filename=?, preview_file_id=?, preview_url=?
                        WHERE id=?
                    """, (file_id, public_url, request.files[f'logo_{variante}'].filename, prev_file_id, prev_url, existing['id']))
                else:
                    conn.execute("""
                        INSERT INTO iv_logos (id_iv, variante, drive_file_id, public_url, filename, preview_file_id, preview_url)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (id_iv, variante, file_id, public_url, request.files[f'logo_{variante}'].filename, prev_file_id, prev_url))
            elif file_id:
                if existing:
                    conn.execute("""
                        UPDATE iv_logos SET drive_file_id=?, public_url=?, filename=? WHERE id=?
                    """, (file_id, public_url, request.files[f'logo_{variante}'].filename, existing['id']))
                else:
                    conn.execute("""
                        INSERT INTO iv_logos (id_iv, variante, drive_file_id, public_url, filename)
                        VALUES (?, ?, ?, ?, ?)
                    """, (id_iv, variante, file_id, public_url, request.files[f'logo_{variante}'].filename))
            elif prev_file_id and existing:
                conn.execute("""
                    UPDATE iv_logos SET preview_file_id=?, preview_url=? WHERE id=?
                """, (prev_file_id, prev_url, existing['id']))
            elif not existing:
                conn.execute("""
                    INSERT INTO iv_logos (id_iv, variante)
                    VALUES (?, ?)
                """, (id_iv, variante))
        conn.commit()

        # Fonts — efface et recrée
        conn.execute("DELETE FROM iv_fonts WHERE id_iv = ?", (id_iv,))
        noms_fonts = request.form.getlist('font_nom')
        usages_fonts = request.form.getlist('font_usage')
        for nom, usage in zip(noms_fonts, usages_fonts):
            nom = nom.strip()
            if not nom:
                continue
            google_url = f"https://fonts.google.com/specimen/{nom.replace(' ', '+')}"
            conn.execute("""
                INSERT INTO iv_fonts (id_iv, nom_font, google_font_url, usage)
                VALUES (?, ?, ?, ?)
            """, (id_iv, nom, google_url, usage.strip()))

        conn.commit()
        conn.close()
        flash("Identité visuelle sauvegardée.", "success")
        return redirect(url_for('admin_identite_visuelle', project_id=project_id))

    # GET
    iv = conn.execute(
        "SELECT * FROM identite_visuelle WHERE id_projet = ?",
        (project_id,)
    ).fetchone()

    logos = []
    fonts = []

    if iv:
        logos = conn.execute(
            "SELECT * FROM iv_logos WHERE id_iv = ?",
            (iv['id'],)
        ).fetchall()

        fonts = conn.execute(
            "SELECT * FROM iv_fonts WHERE id_iv = ?",
            (iv['id'],)
        ).fetchall()

    logos_dict = {l['variante']: l for l in logos}
    declinaisons = []
    mockups = []

    if iv:
        declinaisons = conn.execute(
            "SELECT * FROM iv_declinaisons WHERE id_iv = ? ORDER BY position, id",
            (iv['id'],)
        ).fetchall()

        mockups = conn.execute(
            "SELECT * FROM iv_mockups WHERE id_iv = ? ORDER BY position, id",
            (iv['id'],)
        ).fetchall()

    conn.close()

    return render_template(
        'admin_identite_visuelle.html',
        projet=projet,
        iv=iv,
        logos=logos_dict,
        fonts=fonts,
        declinaisons=declinaisons,
        mockups=mockups
    )

@app.route('/projet/<int:project_id>/identite')
@login_required
def client_identite_visuelle(project_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        flash("Projet introuvable.", "error")
        return redirect(url_for('dashboard'))

    is_owner = (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not (is_owner or is_admin):
        conn.close()
        flash("Accès non autorisé.", "error")
        return redirect(url_for('dashboard'))

    iv = conn.execute("""
        SELECT * FROM identite_visuelle WHERE id_projet = ? AND is_complete = 1
    """, (project_id,)).fetchone()

    if not iv:
        conn.close()
        flash("L'identité visuelle n'est pas encore disponible.", "error")
        return redirect(url_for('project_detail', project_id=project_id))

    logos = conn.execute("SELECT * FROM iv_logos WHERE id_iv = ?", (iv['id'],)).fetchall()
    fonts = conn.execute("SELECT * FROM iv_fonts WHERE id_iv = ?", (iv['id'],)).fetchall()
    declinaisons = conn.execute("SELECT * FROM iv_declinaisons WHERE id_iv = ? ORDER BY position, id", (iv['id'],)).fetchall()
    mockups = conn.execute("SELECT * FROM iv_mockups WHERE id_iv = ? ORDER BY position, id", (iv['id'],)).fetchall()

    logos_dict = {l['variante']: l for l in logos}
    conn.close()

    return render_template(
        'identite_visuelle.html',
        projet=projet,
        iv=iv,
        logos=logos_dict,
        fonts=fonts,
        declinaisons=declinaisons,
        mockups=mockups
    )
@app.route('/admin/projet/<int:project_id>/identite/upload_declinaison', methods=['POST'])
@admin_required
def upload_declinaison(project_id):
    import tempfile
    conn = get_db_connection()
    try:
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)).fetchone()
        if not iv:
            return jsonify({'ok': False, 'error': 'IV non trouvée'})
        fichier = request.files.get('fichier')
        label = request.form.get('label', '').strip()
        if not fichier:
            return jsonify({'ok': False, 'error': 'Fichier manquant'})
        iv_folder_id = iv['iv_folder_id']
        if not iv_folder_id:
            projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
            parent = projet['drive_folder_id'] if projet and projet['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            iv_folder_id = create_folder("Identité visuelle", parent_id=parent)
            conn.execute("UPDATE identite_visuelle SET iv_folder_id = ? WHERE id = ?", (iv_folder_id, iv['id']))
            conn.commit()
        ext = fichier.filename.rsplit('.', 1)[-1].lower()
        safe_name = secure_filename(fichier.filename)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.'+ext)
        fichier.save(tmp.name)
        tmp.close()
        file_id, _ = upload_file(tmp.name, safe_name, iv_folder_id)
        os.unlink(tmp.name)
        public_url = make_file_public(file_id)
        conn.execute("""
            INSERT INTO iv_declinaisons (id_iv, drive_file_id, public_url, filename, label)
            VALUES (?, ?, ?, ?, ?)
        """, (iv['id'], file_id, public_url, safe_name, label))
        conn.commit()
        return jsonify({'ok': True, 'public_url': public_url, 'filename': safe_name})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        conn.close()


@app.route('/admin/projet/<int:project_id>/identite/delete_declinaison/<int:decl_id>', methods=['POST'])
@admin_required
def delete_declinaison(project_id, decl_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM iv_declinaisons WHERE id = ?", (decl_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/admin/projet/<int:project_id>/identite/upload_mockup', methods=['POST'])
@admin_required
def upload_mockup(project_id):
    import tempfile
    conn = get_db_connection()
    try:
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)).fetchone()
        if not iv:
            return jsonify({'ok': False, 'error': 'IV non trouvée'})
        fichier = request.files.get('fichier')
        label = request.form.get('label', '').strip()
        if not fichier:
            return jsonify({'ok': False, 'error': 'Fichier manquant'})
        iv_folder_id = iv['iv_folder_id']
        if not iv_folder_id:
            projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
            parent = projet['drive_folder_id'] if projet and projet['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            iv_folder_id = create_folder("Identité visuelle", parent_id=parent)
            conn.execute("UPDATE identite_visuelle SET iv_folder_id = ? WHERE id = ?", (iv_folder_id, iv['id']))
            conn.commit()
        ext = fichier.filename.rsplit('.', 1)[-1].lower()
        safe_name = secure_filename(fichier.filename)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.'+ext)
        fichier.save(tmp.name)
        tmp.close()
        file_id, _ = upload_file(tmp.name, safe_name, iv_folder_id)
        os.unlink(tmp.name)
        public_url = make_file_public(file_id)
        conn.execute("""
            INSERT INTO iv_mockups (id_iv, drive_file_id, public_url, filename, label)
            VALUES (?, ?, ?, ?, ?)
        """, (iv['id'], file_id, public_url, safe_name, label))
        conn.commit()
        return jsonify({'ok': True, 'public_url': public_url, 'filename': safe_name})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        conn.close()


@app.route('/admin/projet/<int:project_id>/identite/delete_mockup/<int:mockup_id>', methods=['POST'])
@admin_required
def delete_mockup(project_id, mockup_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM iv_mockups WHERE id = ?", (mockup_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/admin/projet/<int:project_id>/identite/save_palette', methods=['POST'])
@admin_required
def save_palette(project_id):
    import json
    conn = get_db_connection()
    try:
        palette_json = request.form.get('palette_json', '')
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)).fetchone()
        if not iv:
            conn.execute("INSERT INTO identite_visuelle (id_projet, is_complete, palette_json) VALUES (?, 0, ?)", (project_id, palette_json))
        else:
            conn.execute("UPDATE identite_visuelle SET palette_json = ? WHERE id_projet = ?", (palette_json, project_id))
        conn.commit()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        conn.close()

@app.route('/admin/projet/<int:project_id>/identite/upload_svg', methods=['POST'])
@admin_required
def upload_svg_genere(project_id):
    import tempfile
    conn = get_db_connection()
    try:
        iv = conn.execute(
            "SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)
        ).fetchone()

        # Créer le dossier IV si pas encore fait
        iv_folder_id = iv['iv_folder_id'] if iv and iv['iv_folder_id'] else None
        if not iv_folder_id:
            projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
            parent = projet['drive_folder_id'] if projet and projet['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            iv_folder_id = create_folder("Identité visuelle", parent_id=parent)
            if iv:
                conn.execute("UPDATE identite_visuelle SET iv_folder_id = ? WHERE id_projet = ?", (iv_folder_id, project_id))
            conn.commit()

        fichier = request.files.get('fichier')
        nom = request.form.get('nom', 'logo.svg')
        if not fichier or not iv_folder_id:
            return jsonify({'ok': False, 'error': 'Fichier ou dossier manquant'})

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.svg')
        fichier.save(tmp.name)
        tmp.close()
        file_id, _ = upload_file(tmp.name, nom, iv_folder_id)
        os.unlink(tmp.name)
        make_file_public(file_id)
        return jsonify({'ok': True, 'file_id': file_id})
    except Exception as e:
        print(f"[DRIVE] Upload SVG généré échoué: {e}")
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        conn.close()

@app.route('/projet/<int:project_id>/identite/zip')
@login_required
def telecharger_zip_identite(project_id):
    import io, zipfile

    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        flash("Projet introuvable.", "error")
        return redirect(url_for('dashboard'))

    is_owner = (projet['id_client'] == session['user_id'])
    is_admin = bool(session.get('is_admin', False))
    if not (is_owner or is_admin):
        conn.close()
        flash("Accès non autorisé.", "error")
        return redirect(url_for('dashboard'))

    iv = conn.execute("""
        SELECT * FROM identite_visuelle WHERE id_projet = ? AND is_complete = 1
    """, (project_id,)).fetchone()

    if not iv or not iv['iv_folder_id']:
        conn.close()
        flash("Aucun fichier disponible.", "error")
        return redirect(url_for('client_identite_visuelle', project_id=project_id))

    logos = conn.execute("SELECT * FROM iv_logos WHERE id_iv = ?", (iv['id'],)).fetchall()
    declinaisons = conn.execute("SELECT * FROM iv_declinaisons WHERE id_iv = ?", (iv['id'],)).fetchall()
    mockups = conn.execute("SELECT * FROM iv_mockups WHERE id_iv = ?", (iv['id'],)).fetchall()
    nom_projet = projet['nom_projet']
    conn.close()

    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build

        SERVICE_ACCOUNT_FILE = os.path.join(os.path.dirname(__file__), 'service_account.json')
        SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
        creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        service = build('drive', 'v3', credentials=creds)

        def download_file(file_id):
            try:
                return service.files().get_media(
                    fileId=file_id,
                    supportsAllDrives=True
                ).execute()
            except Exception as e:
                print(f"[ZIP] Download échoué {file_id}: {e}")
                return None

        # Mapping clé → (groupe, dossier)
        NOIR_MAP = {
            'principal': 'Logo principal',
            'icone':     'Icône',
            'variante':  'Variante',
        }
        BLANC_MAP = {
            'principal_blanc': 'Logo principal',
            'icone_blanc':     'Icône',
            'variante_blanc':  'Variante',
        }

        def logo_path(variante):
            if variante in NOIR_MAP:
                return f"01 - Logos/Version noire/{NOIR_MAP[variante]}"
            if variante in BLANC_MAP:
                return f"01 - Logos/Version blanche/{BLANC_MAP[variante]}"
            if variante.startswith('noir_'):
                label = variante[5:].replace('_', ' ').title()
                return f"01 - Logos/Version noire/{label}"
            if variante.startswith('blanc_'):
                label = variante[6:].replace('_', ' ').title()
                return f"01 - Logos/Version blanche/{label}"
            return f"01 - Logos/Autres/{variante}"

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:

            # 01 — Logos (organisés par groupe et dossier lisible)
            for logo in logos:
                folder = logo_path(logo['variante'])
                if logo['drive_file_id'] and logo['filename']:
                    data = download_file(logo['drive_file_id'])
                    if data:
                        zf.writestr(f"{folder}/{logo['filename']}", data)
                if logo['preview_file_id'] and logo['filename']:
                    data = download_file(logo['preview_file_id'])
                    if data:
                        base = logo['filename'].rsplit('.', 1)[0]
                        zf.writestr(f"{folder}/{base}_aperçu.png", data)

            # 02 — Déclinaisons
            for d in declinaisons:
                if d['drive_file_id'] and d['filename']:
                    data = download_file(d['drive_file_id'])
                    if data:
                        zf.writestr(f"02 - Déclinaisons/{d['filename']}", data)

            # 03 — Mockups
            for m in mockups:
                if m['drive_file_id'] and m['filename']:
                    data = download_file(m['drive_file_id'])
                    if data:
                        zf.writestr(f"03 - Mockups/{m['filename']}", data)

        zip_buffer.seek(0)
        nom_zip = f"identite_visuelle_{nom_projet.replace(' ', '_')}.zip"
        return send_file(zip_buffer, mimetype='application/zip',
                        as_attachment=True, download_name=nom_zip)

    except Exception as e:
        flash(f"Erreur ZIP : {e}", "error")
        return redirect(url_for('client_identite_visuelle', project_id=project_id))

def _todo_assignees(conn, todo_id) -> list:
    rows = conn.execute(
        "SELECT c.id, c.nom_complet FROM todo_assignees ta JOIN clients c ON c.id = ta.admin_id WHERE ta.todo_id = ?",
        (todo_id,)
    ).fetchall()
    return [{'id': r['id'], 'nom_complet': r['nom_complet']} for r in rows]

def _roadmap_todo_assignees(conn, roadmap_todo_id) -> list:
    rows = conn.execute(
        "SELECT c.id, c.nom_complet FROM roadmap_todo_assignees ra JOIN clients c ON c.id = ra.admin_id WHERE ra.roadmap_todo_id = ?",
        (roadmap_todo_id,)
    ).fetchall()
    return [{'id': r['id'], 'nom_complet': r['nom_complet']} for r in rows]

def _admin_id_for_role(conn, role):
    row = conn.execute("SELECT id FROM clients WHERE role = ? AND is_admin = 1 LIMIT 1", (role,)).fetchone()
    return row['id'] if row else None

def _sync_roadmap_todo_completion(conn, roadmap_todo_id, done):
    """Propage l'état coché/pas-coché d'un item de checklist roadmap vers sa tâche
    perso liée et son post marketing lié (si présents) — appelée peu importe lequel
    des 3 endroits (roadmap, Mes tâches, calendrier marketing) a été coché, pour que
    les trois restent synchronisés."""
    rt = conn.execute("SELECT * FROM roadmap_todos WHERE id = ?", (roadmap_todo_id,)).fetchone()
    if not rt:
        return
    val = 1 if done else 0
    conn.execute("UPDATE roadmap_todos SET est_coche = ? WHERE id = ?", (val, roadmap_todo_id))
    if rt['linked_todo_perso_id']:
        conn.execute("UPDATE todos_perso SET est_coche = ? WHERE id = ?", (val, rt['linked_todo_perso_id']))
    if rt['linked_marketing_post_id']:
        assignee_ids = {r['admin_id'] for r in conn.execute(
            "SELECT admin_id FROM roadmap_todo_assignees WHERE roadmap_todo_id = ?", (roadmap_todo_id,)
        ).fetchall()}
        production_id = _admin_id_for_role(conn, 'production')
        gestion_id = _admin_id_for_role(conn, 'gestion')
        if production_id in assignee_ids:
            conn.execute("UPDATE marketing_posts SET todo_felix_done = ? WHERE id = ?", (val, rt['linked_marketing_post_id']))
        if gestion_id in assignee_ids:
            conn.execute("UPDATE marketing_posts SET todo_marie_done = ? WHERE id = ?", (val, rt['linked_marketing_post_id']))
    conn.commit()

@app.route('/admin/roadmap')
@admin_required
def admin_roadmap():
    return render_template('admin_roadmap.html')

# ───────────────────────────────────────────────────────────
# Roadmaps (gestion de projets internes)
# ───────────────────────────────────────────────────────────
@app.route('/admin/roadmaps')
@admin_required
def admin_roadmaps():
    conn = get_db_connection()
    roadmaps_actives = conn.execute("SELECT * FROM roadmaps WHERE is_archived = 0 ORDER BY created_at DESC").fetchall()
    roadmaps_archivees = conn.execute("SELECT * FROM roadmaps WHERE is_archived = 1 ORDER BY created_at DESC").fetchall()
    conn.close()
    return render_template('admin_roadmaps.html', roadmaps_actives=roadmaps_actives, roadmaps_archivees=roadmaps_archivees)

@app.route('/admin/roadmaps/new', methods=['GET', 'POST'])
@admin_required
def admin_roadmap_new():
    if request.method == 'POST':
        titre = request.form.get('titre', '').strip()
        description = request.form.get('description', '').strip()
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("INSERT INTO roadmaps (titre, description) VALUES (?, ?)", (titre, description))
        id_roadmap = cur.lastrowid
        conn.commit()
        conn.close()
        flash(f"Roadmap '{titre}' créée.", "success")
        return redirect(url_for('admin_roadmap_detail', roadmap_id=id_roadmap))
    return render_template('admin_roadmap_new.html')

@app.route('/admin/roadmaps/<int:roadmap_id>')
@admin_required
def admin_roadmap_detail(roadmap_id):
    conn = get_db_connection()
    roadmap = conn.execute("SELECT * FROM roadmaps WHERE id = ?", (roadmap_id,)).fetchone()
    if not roadmap:
        conn.close()
        flash("Roadmap introuvable.", "error")
        return redirect(url_for('admin_roadmaps'))
    phases = conn.execute("""
        SELECT * FROM roadmap_phases WHERE id_roadmap = ? ORDER BY position ASC
    """, (roadmap_id,)).fetchall()
    phases_avec_todos = []
    for phase in phases:
        todos = conn.execute("""
            SELECT * FROM roadmap_todos WHERE id_phase = ? ORDER BY position ASC
        """, (phase['id'],)).fetchall()
        notes = conn.execute("""
            SELECT * FROM roadmap_phase_notes WHERE id_phase = ? ORDER BY created_at ASC
        """, (phase['id'],)).fetchall()
        phases_avec_todos.append({'phase': phase, 'todos': todos, 'notes': notes})
    conn.close()
    return render_template('admin_roadmap_detail.html', roadmap=roadmap, phases=phases_avec_todos)

@app.route('/admin/roadmaps/<int:roadmap_id>/add_phase', methods=['POST'])
@admin_required
def admin_roadmap_add_phase(roadmap_id):
    titre = request.form.get('titre', '').strip()
    date_debut = request.form.get('date_debut', '').strip()
    date_fin = request.form.get('date_fin', '').strip()
    couleur = request.form.get('couleur', '#3498db').strip()
    badge = request.form.get('badge', 'Planifiee').strip()
    description = request.form.get('description', '').strip()
    conn = get_db_connection()
    max_pos = conn.execute("SELECT MAX(position) FROM roadmap_phases WHERE id_roadmap = ?", (roadmap_id,)).fetchone()[0] or 0
    conn.execute("""
        INSERT INTO roadmap_phases (id_roadmap, titre, date_debut, date_fin, couleur, badge, position, description)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (roadmap_id, titre, date_debut, date_fin, couleur, badge, max_pos + 1, description))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_roadmap_detail', roadmap_id=roadmap_id))

@app.route('/admin/roadmaps/phase/<int:phase_id>/add_todo', methods=['POST'])
@admin_required
def admin_roadmap_add_todo(phase_id):
    texte = request.form.get('texte', '').strip()
    conn = get_db_connection()
    phase = conn.execute("SELECT * FROM roadmap_phases WHERE id = ?", (phase_id,)).fetchone()
    if texte:
        max_pos = conn.execute("SELECT MAX(position) FROM roadmap_todos WHERE id_phase = ?", (phase_id,)).fetchone()[0] or 0
        conn.execute("INSERT INTO roadmap_todos (id_phase, texte, position) VALUES (?, ?, ?)", (phase_id, texte, max_pos + 1))
        conn.commit()
    conn.close()
    return redirect(url_for('admin_roadmap_detail', roadmap_id=phase['id_roadmap']))

@app.route('/admin/roadmaps/todo/<int:todo_id>/toggle', methods=['POST'])
@admin_required
def admin_roadmap_toggle_todo(todo_id):
    conn = get_db_connection()
    todo = conn.execute("SELECT * FROM roadmap_todos WHERE id = ?", (todo_id,)).fetchone()
    phase = conn.execute("SELECT * FROM roadmap_phases WHERE id = ?", (todo['id_phase'],)).fetchone()
    new_val = 0 if int(todo['est_coche']) else 1
    conn.execute("UPDATE roadmap_todos SET est_coche = ? WHERE id = ?", (new_val, todo_id))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_roadmap_detail', roadmap_id=phase['id_roadmap']))

@app.route('/admin/roadmaps/todo/<int:todo_id>/delete', methods=['POST'])
@admin_required
def admin_roadmap_delete_todo(todo_id):
    conn = get_db_connection()
    todo = conn.execute("SELECT * FROM roadmap_todos WHERE id = ?", (todo_id,)).fetchone()
    phase = conn.execute("SELECT * FROM roadmap_phases WHERE id = ?", (todo['id_phase'],)).fetchone()
    conn.execute("DELETE FROM roadmap_todos WHERE id = ?", (todo_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_roadmap_detail', roadmap_id=phase['id_roadmap']))

@app.route('/admin/roadmaps/<int:roadmap_id>/delete', methods=['POST'])
@admin_required
def admin_roadmap_delete(roadmap_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM roadmaps WHERE id = ?", (roadmap_id,))
    conn.commit()
    conn.close()
    flash("Roadmap supprimée.", "success")
    return redirect(url_for('admin_roadmaps'))

@app.route('/admin/roadmaps/<int:roadmap_id>/archive', methods=['POST'])
@admin_required
def admin_roadmap_archive(roadmap_id):
    conn = get_db_connection()
    conn.execute("UPDATE roadmaps SET is_archived = 1 WHERE id = ?", (roadmap_id,))
    conn.commit()
    conn.close()
    flash("Roadmap archivée.", "success")
    return redirect(url_for('admin_roadmaps'))

@app.route('/admin/roadmaps/<int:roadmap_id>/unarchive', methods=['POST'])
@admin_required
def admin_roadmap_unarchive(roadmap_id):
    conn = get_db_connection()
    conn.execute("UPDATE roadmaps SET is_archived = 0 WHERE id = ?", (roadmap_id,))
    conn.commit()
    conn.close()
    flash("Roadmap désarchivée.", "success")
    return redirect(url_for('admin_roadmaps'))


@app.route('/admin/roadmaps/<int:roadmap_id>/edit', methods=['POST'])
@admin_required
def admin_roadmap_edit(roadmap_id):
    titre = request.form.get('titre', '').strip()
    description = request.form.get('description', '').strip()
    notes = request.form.get('notes', '').strip()
    conn = get_db_connection()
    conn.execute("UPDATE roadmaps SET titre=?, description=?, notes=? WHERE id=?", (titre, description, notes, roadmap_id))
    conn.commit()
    conn.close()
    flash("Projet mis à jour.", "success")
    return redirect(url_for('admin_roadmap_detail', roadmap_id=roadmap_id))

@app.route('/admin/roadmaps/phase/<int:phase_id>/edit', methods=['POST'])
@admin_required
def admin_roadmap_edit_phase(phase_id):
    titre = request.form.get('titre', '').strip()
    description = request.form.get('description', '').strip()
    date_debut = request.form.get('date_debut', '').strip()
    date_fin = request.form.get('date_fin', '').strip()
    badge = request.form.get('badge', 'Planifiee').strip()
    couleur = request.form.get('couleur', '#3498db').strip()
    conn = get_db_connection()
    phase = conn.execute("SELECT * FROM roadmap_phases WHERE id=?", (phase_id,)).fetchone()
    conn.execute("UPDATE roadmap_phases SET titre=?, description=?, date_debut=?, date_fin=?, badge=?, couleur=? WHERE id=?",
        (titre, description, date_debut, date_fin, badge, couleur, phase_id))
    conn.commit()
    conn.close()
    flash("Phase mise à jour.", "success")
    return redirect(url_for('admin_roadmap_detail', roadmap_id=phase['id_roadmap']))

@app.route('/admin/roadmaps/phase/<int:phase_id>/delete', methods=['POST'])
@admin_required
def admin_roadmap_delete_phase(phase_id):
    conn = get_db_connection()
    phase = conn.execute("SELECT * FROM roadmap_phases WHERE id=?", (phase_id,)).fetchone()
    conn.execute("DELETE FROM roadmap_phases WHERE id=?", (phase_id,))
    conn.commit()
    conn.close()
    flash("Phase supprimée.", "success")
    return redirect(url_for('admin_roadmap_detail', roadmap_id=phase['id_roadmap']))


@app.route('/admin/roadmaps/phase/<int:phase_id>/notes', methods=['POST'])
@admin_required
def admin_roadmap_save_notes(phase_id):
    notes = request.form.get('notes', '').strip()
    conn = get_db_connection()
    phase = conn.execute("SELECT * FROM roadmap_phases WHERE id=?", (phase_id,)).fetchone()
    conn.execute("UPDATE roadmap_phases SET notes=? WHERE id=?", (notes, phase_id))
    conn.commit()
    conn.close()
    flash("Notes sauvegardées.", "success")
    return redirect(url_for('admin_roadmap_detail', roadmap_id=phase['id_roadmap']))


@app.route('/admin/roadmaps/phase/<int:phase_id>/note/add', methods=['POST'])
@admin_required
def admin_roadmap_add_note(phase_id):
    texte = request.form.get('texte', '').strip()
    conn = get_db_connection()
    phase = conn.execute("SELECT * FROM roadmap_phases WHERE id=?", (phase_id,)).fetchone()
    if texte:
        conn.execute("INSERT INTO roadmap_phase_notes (id_phase, texte) VALUES (?, ?)", (phase_id, texte))
        conn.commit()
    conn.close()
    return redirect(url_for('admin_roadmap_detail', roadmap_id=phase['id_roadmap']))

@app.route('/admin/roadmaps/note/<int:note_id>/delete', methods=['POST'])
@admin_required
def admin_roadmap_delete_note(note_id):
    conn = get_db_connection()
    note = conn.execute("SELECT * FROM roadmap_phase_notes WHERE id=?", (note_id,)).fetchone()
    phase = conn.execute("SELECT * FROM roadmap_phases WHERE id=?", (note['id_phase'],)).fetchone()
    conn.execute("DELETE FROM roadmap_phase_notes WHERE id=?", (note_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('admin_roadmap_detail', roadmap_id=phase['id_roadmap']))


# ── API v1 Roadmaps ──────────────────────────────────────
@app.route('/api/v1/admin/roadmaps')
@admin_required
def api_roadmaps_list():
    conn = get_db_connection()
    actives = conn.execute("SELECT * FROM roadmaps WHERE is_archived=0 ORDER BY created_at DESC").fetchall()
    archivees = conn.execute("SELECT * FROM roadmaps WHERE is_archived=1 ORDER BY created_at DESC").fetchall()
    conn.close()
    return jsonify({'actives': [dict(r) for r in actives], 'archivees': [dict(r) for r in archivees]})

@app.route('/api/v1/admin/roadmaps/new', methods=['POST'])
@admin_required
def api_roadmap_new():
    data = request.get_json() or {}
    titre = data.get('titre','').strip()
    description = data.get('description','').strip()
    if not titre:
        return jsonify({'error': 'Titre obligatoire'}), 400
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("INSERT INTO roadmaps (titre, description) VALUES (?,?)", (titre, description))
    rid = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': rid})

@app.route('/api/v1/admin/roadmaps/<int:roadmap_id>')
@admin_required
def api_roadmap_detail(roadmap_id):
    conn = get_db_connection()
    roadmap = conn.execute("SELECT * FROM roadmaps WHERE id=?", (roadmap_id,)).fetchone()
    if not roadmap:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    phases = conn.execute("SELECT * FROM roadmap_phases WHERE id_roadmap=? ORDER BY position ASC", (roadmap_id,)).fetchall()
    result = []
    for phase in phases:
        todos = conn.execute("SELECT * FROM roadmap_todos WHERE id_phase=? ORDER BY position ASC", (phase['id'],)).fetchall()
        journal = conn.execute("SELECT * FROM roadmap_phase_notes WHERE id_phase=? ORDER BY created_at ASC", (phase['id'],)).fetchall()
        todos_out = []
        for t in todos:
            td = dict(t)
            td['assignees'] = _roadmap_todo_assignees(conn, t['id'])
            todos_out.append(td)
        result.append({**dict(phase), 'todos': todos_out, 'journal': [dict(n) for n in journal]})
    conn.close()
    return jsonify({'roadmap': dict(roadmap), 'phases': result})

@app.route('/api/v1/admin/roadmaps/<int:roadmap_id>/edit', methods=['POST'])
@admin_required
def api_roadmap_edit(roadmap_id):
    data = request.get_json() or {}
    conn = get_db_connection()
    conn.execute("UPDATE roadmaps SET titre=?,description=?,notes=? WHERE id=?",
        (data.get('titre',''), data.get('description',''), data.get('notes',''), roadmap_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/v1/admin/roadmaps/<int:roadmap_id>/archive', methods=['POST'])
@admin_required
def api_roadmap_archive(roadmap_id):
    conn = get_db_connection()
    conn.execute("UPDATE roadmaps SET is_archived=1 WHERE id=?", (roadmap_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/v1/admin/roadmaps/<int:roadmap_id>/unarchive', methods=['POST'])
@admin_required
def api_roadmap_unarchive(roadmap_id):
    conn = get_db_connection()
    conn.execute("UPDATE roadmaps SET is_archived=0 WHERE id=?", (roadmap_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/v1/admin/roadmaps/<int:roadmap_id>/delete', methods=['POST'])
@admin_required
def api_roadmap_delete(roadmap_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM roadmaps WHERE id=?", (roadmap_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/v1/admin/roadmaps/<int:roadmap_id>/add_phase', methods=['POST'])
@admin_required
def api_roadmap_add_phase(roadmap_id):
    data = request.get_json() or {}
    conn = get_db_connection()
    max_pos = conn.execute("SELECT MAX(position) FROM roadmap_phases WHERE id_roadmap=?", (roadmap_id,)).fetchone()[0] or 0
    cur = conn.cursor()
    sql = "INSERT INTO roadmap_phases (id_roadmap,titre,description,date_debut,date_fin,couleur,badge,position) VALUES (?,?,?,?,?,?,?,?)"
    cur.execute(sql, (roadmap_id, data.get('titre',''), data.get('description',''), data.get('date_debut',''), data.get('date_fin',''), data.get('couleur','#3498db'), data.get('badge','Planifiee'), max_pos+1))
    pid = cur.lastrowid
    conn.commit()
    phase = conn.execute("SELECT * FROM roadmap_phases WHERE id=?", (pid,)).fetchone()
    conn.close()
    return jsonify({'phase': dict(phase)})

@app.route('/api/v1/admin/roadmaps/phase/<int:phase_id>/delete', methods=['POST'])
@admin_required
def api_roadmap_phase_delete(phase_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM roadmap_phases WHERE id=?", (phase_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/v1/admin/roadmaps/phase/<int:phase_id>/add_todo', methods=['POST'])
@admin_required
def api_roadmap_add_todo(phase_id):
    import json as _json
    data = request.get_json() or {}
    texte = data.get('texte','').strip()
    assigne_admin_ids = [i for i in (data.get('assigne_admin_ids') or []) if i]
    marketing = data.get('marketing') or None  # {date_publication, plateformes}
    conn = get_db_connection()
    phase = conn.execute("SELECT * FROM roadmap_phases WHERE id = ?", (phase_id,)).fetchone()
    roadmap = conn.execute("SELECT * FROM roadmaps WHERE id = ?", (phase['id_roadmap'],)).fetchone() if phase else None
    max_pos = conn.execute("SELECT MAX(position) FROM roadmap_todos WHERE id_phase=?", (phase_id,)).fetchone()[0] or 0
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO roadmap_todos (id_phase,texte,position) VALUES (?,?,?)",
        (phase_id, texte, max_pos+1)
    )
    tid = cur.lastrowid
    for admin_id in assigne_admin_ids:
        conn.execute("INSERT OR IGNORE INTO roadmap_todo_assignees (roadmap_todo_id, admin_id) VALUES (?, ?)", (tid, admin_id))
    conn.commit()

    if assigne_admin_ids and texte:
        contexte = " › ".join(filter(None, [roadmap['titre'] if roadmap else None, phase['titre'] if phase else None]))
        texte_tache = f"[{contexte}] {texte}" if contexte else texte
        cur2 = conn.execute(
            "INSERT INTO todos_perso (texte, date_echeance, linked_roadmap_todo_id) VALUES (?, ?, ?)",
            (texte_tache, phase['date_fin'] if phase else None, tid)
        )
        perso_id = cur2.lastrowid
        for admin_id in assigne_admin_ids:
            conn.execute("INSERT OR IGNORE INTO todo_assignees (todo_id, admin_id) VALUES (?, ?)", (perso_id, admin_id))
        conn.execute("UPDATE roadmap_todos SET linked_todo_perso_id = ? WHERE id = ?", (perso_id, tid))
        conn.commit()

    if marketing and marketing.get('date_publication') and texte:
        cur3 = conn.execute(
            "INSERT INTO marketing_posts (titre, description, date_publication, plateformes, statut, created_by, linked_roadmap_todo_id) "
            "VALUES (?, ?, ?, ?, 'planifié', ?, ?)",
            (texte, roadmap['titre'] if roadmap else '', marketing['date_publication'],
             _json.dumps(marketing.get('plateformes') or []), session.get('user_id'), tid)
        )
        conn.execute("UPDATE roadmap_todos SET linked_marketing_post_id = ? WHERE id = ?", (cur3.lastrowid, tid))
        conn.commit()

    todo = conn.execute("SELECT * FROM roadmap_todos WHERE id=?", (tid,)).fetchone()
    d = dict(todo)
    d['assignees'] = _roadmap_todo_assignees(conn, tid)
    conn.close()
    return jsonify({'todo': d})

@app.route('/api/v1/admin/roadmaps/todo/<int:todo_id>/toggle', methods=['POST'])
@admin_required
def api_roadmap_toggle_todo(todo_id):
    conn = get_db_connection()
    todo = conn.execute("SELECT * FROM roadmap_todos WHERE id=?", (todo_id,)).fetchone()
    if not todo:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    _sync_roadmap_todo_completion(conn, todo_id, not todo['est_coche'])
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/v1/admin/roadmaps/todo/<int:todo_id>/delete', methods=['POST'])
@admin_required
def api_roadmap_delete_todo(todo_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM roadmap_todos WHERE id=?", (todo_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ── Todos de la roadmap produit CocktailOS (onglet Vision, page /admin/roadmaps) ──
# Convergence Phase 5 du module Tâches : remplace le localStorage isolé
# (cocktailos_todos_phase_<id>) par un stockage serveur partagé entre appareils/admins.

@app.route('/api/v1/admin/roadmaps/vision-todos', methods=['GET'])
@admin_required
def api_vision_todos_list():
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT id, phase_id, texte, est_coche FROM cocktailos_vision_todos ORDER BY phase_id ASC, position ASC, id ASC"
    ).fetchall()
    conn.close()
    return jsonify([{'id': r['id'], 'phase_id': r['phase_id'], 'texte': r['texte'], 'done': bool(r['est_coche'])} for r in rows])

@app.route('/api/v1/admin/roadmaps/vision-todos', methods=['POST'])
@admin_required
def api_vision_todos_create():
    data = request.get_json() or {}
    phase_id = data.get('phase_id')
    texte = (data.get('texte') or '').strip()
    if phase_id is None or not texte:
        return jsonify({'error': 'phase_id et texte requis'}), 400
    conn = get_db_connection()
    max_pos = conn.execute("SELECT MAX(position) FROM cocktailos_vision_todos WHERE phase_id=?", (phase_id,)).fetchone()[0] or 0
    cur = conn.execute(
        "INSERT INTO cocktailos_vision_todos (phase_id, texte, position) VALUES (?, ?, ?)",
        (phase_id, texte, max_pos + 1)
    )
    todo_id = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({'id': todo_id, 'phase_id': phase_id, 'texte': texte, 'done': False}), 201

@app.route('/api/v1/admin/roadmaps/vision-todos/<int:todo_id>/toggle', methods=['POST'])
@admin_required
def api_vision_todos_toggle(todo_id):
    conn = get_db_connection()
    todo = conn.execute("SELECT est_coche FROM cocktailos_vision_todos WHERE id=?", (todo_id,)).fetchone()
    if not todo:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    new_val = 0 if todo['est_coche'] else 1
    conn.execute("UPDATE cocktailos_vision_todos SET est_coche=? WHERE id=?", (new_val, todo_id))
    conn.commit()
    conn.close()
    return jsonify({'done': bool(new_val)})

@app.route('/api/v1/admin/roadmaps/vision-todos/<int:todo_id>', methods=['DELETE'])
@admin_required
def api_vision_todos_delete(todo_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM cocktailos_vision_todos WHERE id=?", (todo_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/roadmaps/phase/<int:phase_id>/note/add', methods=['POST'])
@admin_required
def api_roadmap_add_note(phase_id):
    data = request.get_json() or {}
    texte = data.get('texte','').strip()
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("INSERT INTO roadmap_phase_notes (id_phase,texte) VALUES (?,?)", (phase_id, texte))
    nid = cur.lastrowid
    conn.commit()
    note = conn.execute("SELECT * FROM roadmap_phase_notes WHERE id=?", (nid,)).fetchone()
    conn.close()
    return jsonify({'note': dict(note)})

@app.route('/api/v1/admin/roadmaps/note/<int:note_id>/delete', methods=['POST'])
@admin_required
def api_roadmap_delete_note(note_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM roadmap_phase_notes WHERE id=?", (note_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

# ───────────────────────────────────────────────────────────
# API v1 — Identité Visuelle (admin)
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/projet/<int:project_id>/identite', methods=['GET'])
@admin_required
def api_admin_get_identite(project_id):
    import json as _json
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)).fetchone()
    result = {
        'id': iv['id'] if iv else None,
        'is_complete': bool(iv['is_complete']) if iv else False,
        'nom_compagnie': iv['nom_compagnie'] if iv else '',
        'sous_titre': iv['sous_titre'] if iv else '',
        'contexte': iv['contexte'] if iv else '',
        'palette': [],
        'logos': {},
        'fonts': [],
        'declinaisons': [],
        'mockups': [],
    }
    if iv:
        if iv['palette_json']:
            try:
                result['palette'] = _json.loads(iv['palette_json'])
            except Exception:
                result['palette'] = []
        logos_rows = conn.execute("SELECT * FROM iv_logos WHERE id_iv = ?", (iv['id'],)).fetchall()
        for l in logos_rows:
            result['logos'][l['variante']] = {
                'public_url': l['public_url'],
                'preview_url': l['preview_url'] if 'preview_url' in l.keys() else None,
                'filename': l['filename'],
            }
        fonts = conn.execute("SELECT * FROM iv_fonts WHERE id_iv = ?", (iv['id'],)).fetchall()
        result['fonts'] = [{'nom_font': f['nom_font'], 'usage': f['usage']} for f in fonts]
        declinaisons = conn.execute("SELECT * FROM iv_declinaisons WHERE id_iv = ? ORDER BY position, id", (iv['id'],)).fetchall()
        result['declinaisons'] = [{'id': d['id'], 'public_url': d['public_url'], 'label': d['label'], 'filename': d['filename']} for d in declinaisons]
        mockups = conn.execute("SELECT * FROM iv_mockups WHERE id_iv = ? ORDER BY position, id", (iv['id'],)).fetchall()
        result['mockups'] = [{'id': m['id'], 'public_url': m['public_url'], 'label': m['label'], 'filename': m['filename']} for m in mockups]
    conn.close()
    return jsonify(result)


@app.route('/api/v1/admin/projet/<int:project_id>/identite/save', methods=['POST'])
@admin_required
def api_admin_save_identite(project_id):
    import json as _json
    data = request.get_json(force=True, silent=True) or {}
    is_complete = int(bool(data.get('is_complete', False)))
    nom_compagnie = (data.get('nom_compagnie') or '').strip()
    sous_titre = (data.get('sous_titre') or '').strip()
    contexte = (data.get('contexte') or '').strip()
    palette = data.get('palette', [])
    fonts = data.get('fonts', [])
    palette_json = _json.dumps(palette)
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)).fetchone()
    iv_was_complete = bool(iv and int(iv['is_complete'] or 0))
    iv_folder_id = iv['iv_folder_id'] if iv and iv['iv_folder_id'] else None
    if not iv_folder_id:
        try:
            parent = projet['drive_folder_id'] if projet['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            iv_folder_id = create_folder("Identité visuelle", parent_id=parent)
        except Exception as e:
            print(f"[DRIVE] Création dossier IV: {e}")
    if not iv:
        conn.execute("""
            INSERT INTO identite_visuelle (id_projet, is_complete, iv_folder_id, nom_compagnie, sous_titre, palette_json, contexte)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (project_id, is_complete, iv_folder_id, nom_compagnie, sous_titre, palette_json, contexte))
        conn.commit()
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)).fetchone()
    else:
        conn.execute("""
            UPDATE identite_visuelle
            SET is_complete = ?, iv_folder_id = ?, nom_compagnie = ?, sous_titre = ?, palette_json = ?, contexte = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        """, (is_complete, iv_folder_id, nom_compagnie, sous_titre, palette_json, contexte, iv['id']))
        conn.commit()
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)).fetchone()
    if is_complete and not iv_was_complete:
        push_notification(conn, projet['id_client'], project_id,
            f"Votre identité visuelle pour « {projet['nom_projet']} » est prête !", type='identite_visuelle')
        conn.commit()
    id_iv = iv['id']
    conn.execute("DELETE FROM iv_fonts WHERE id_iv = ?", (id_iv,))
    for font in fonts:
        nom = (font.get('nom_font') or '').strip()
        if not nom:
            continue
        usage = (font.get('usage') or '').strip()
        google_url = f"https://fonts.google.com/specimen/{nom.replace(' ', '+')}"
        conn.execute("INSERT INTO iv_fonts (id_iv, nom_font, google_font_url, usage) VALUES (?, ?, ?, ?)",
                     (id_iv, nom, google_url, usage))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'is_complete': bool(is_complete)})


@app.route('/api/v1/admin/projet/<int:project_id>/identite/upload_logo', methods=['POST'])
@admin_required
def api_admin_upload_logo(project_id):
    import tempfile
    conn = get_db_connection()
    try:
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)).fetchone()
        if not iv:
            conn.execute("INSERT INTO identite_visuelle (id_projet, is_complete) VALUES (?, 0)", (project_id,))
            conn.commit()
            iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)).fetchone()
        iv_folder_id = iv['iv_folder_id']
        if not iv_folder_id:
            projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
            parent = projet['drive_folder_id'] if projet and projet['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            iv_folder_id = create_folder("Identité visuelle", parent_id=parent)
            conn.execute("UPDATE identite_visuelle SET iv_folder_id = ? WHERE id = ?", (iv_folder_id, iv['id']))
            conn.commit()
        variante = request.form.get('variante', 'principal')
        fichier = request.files.get('fichier')
        if not fichier:
            return jsonify({'ok': False, 'error': 'Fichier manquant'}), 400
        ext = fichier.filename.rsplit('.', 1)[-1].lower()
        safe_name = secure_filename(fichier.filename)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.' + ext)
        fichier.save(tmp.name)
        tmp.close()
        file_id, _ = upload_file(tmp.name, safe_name, iv_folder_id)
        os.unlink(tmp.name)
        public_url = make_file_public(file_id)
        existing = conn.execute("SELECT id FROM iv_logos WHERE id_iv = ? AND variante = ?", (iv['id'], variante)).fetchone()
        if existing:
            conn.execute("UPDATE iv_logos SET drive_file_id=?, public_url=?, filename=? WHERE id=?",
                         (file_id, public_url, safe_name, existing['id']))
        else:
            conn.execute("INSERT INTO iv_logos (id_iv, variante, drive_file_id, public_url, filename) VALUES (?, ?, ?, ?, ?)",
                         (iv['id'], variante, file_id, public_url, safe_name))
        conn.commit()
        return jsonify({'ok': True, 'public_url': public_url, 'filename': safe_name, 'variante': variante})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:project_id>/identite/upload_declinaison', methods=['POST'])
@admin_required
def api_admin_upload_declinaison(project_id):
    import tempfile
    conn = get_db_connection()
    try:
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)).fetchone()
        if not iv:
            return jsonify({'ok': False, 'error': 'IV non trouvée'}), 404
        fichier = request.files.get('fichier')
        label = request.form.get('label', '').strip()
        if not fichier:
            return jsonify({'ok': False, 'error': 'Fichier manquant'}), 400
        iv_folder_id = iv['iv_folder_id']
        if not iv_folder_id:
            projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
            parent = projet['drive_folder_id'] if projet and projet['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            iv_folder_id = create_folder("Identité visuelle", parent_id=parent)
            conn.execute("UPDATE identite_visuelle SET iv_folder_id = ? WHERE id = ?", (iv_folder_id, iv['id']))
            conn.commit()
        ext = fichier.filename.rsplit('.', 1)[-1].lower()
        safe_name = secure_filename(fichier.filename)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.' + ext)
        fichier.save(tmp.name)
        tmp.close()
        file_id, _ = upload_file(tmp.name, safe_name, iv_folder_id)
        os.unlink(tmp.name)
        public_url = make_file_public(file_id)
        cur = conn.execute("INSERT INTO iv_declinaisons (id_iv, drive_file_id, public_url, filename, label) VALUES (?, ?, ?, ?, ?)",
                           (iv['id'], file_id, public_url, safe_name, label))
        conn.commit()
        return jsonify({'ok': True, 'id': cur.lastrowid, 'public_url': public_url, 'filename': safe_name, 'label': label})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:project_id>/identite/delete_declinaison/<int:decl_id>', methods=['POST'])
@admin_required
def api_admin_delete_declinaison(project_id, decl_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM iv_declinaisons WHERE id = ?", (decl_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/v1/admin/projet/<int:project_id>/identite/upload_mockup', methods=['POST'])
@admin_required
def api_admin_upload_mockup(project_id):
    import tempfile
    conn = get_db_connection()
    try:
        iv = conn.execute("SELECT * FROM identite_visuelle WHERE id_projet = ?", (project_id,)).fetchone()
        if not iv:
            return jsonify({'ok': False, 'error': 'IV non trouvée'}), 404
        fichier = request.files.get('fichier')
        label = request.form.get('label', '').strip()
        if not fichier:
            return jsonify({'ok': False, 'error': 'Fichier manquant'}), 400
        iv_folder_id = iv['iv_folder_id']
        if not iv_folder_id:
            projet = conn.execute("SELECT * FROM projets WHERE id = ?", (project_id,)).fetchone()
            parent = projet['drive_folder_id'] if projet and projet['drive_folder_id'] else os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            iv_folder_id = create_folder("Identité visuelle", parent_id=parent)
            conn.execute("UPDATE identite_visuelle SET iv_folder_id = ? WHERE id = ?", (iv_folder_id, iv['id']))
            conn.commit()
        ext = fichier.filename.rsplit('.', 1)[-1].lower()
        safe_name = secure_filename(fichier.filename)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.' + ext)
        fichier.save(tmp.name)
        tmp.close()
        file_id, _ = upload_file(tmp.name, safe_name, iv_folder_id)
        os.unlink(tmp.name)
        public_url = make_file_public(file_id)
        cur = conn.execute("INSERT INTO iv_mockups (id_iv, drive_file_id, public_url, filename, label) VALUES (?, ?, ?, ?, ?)",
                           (iv['id'], file_id, public_url, safe_name, label))
        conn.commit()
        return jsonify({'ok': True, 'id': cur.lastrowid, 'public_url': public_url, 'filename': safe_name, 'label': label})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/v1/admin/projet/<int:project_id>/identite/delete_mockup/<int:mockup_id>', methods=['POST'])
@admin_required
def api_admin_delete_mockup(project_id, mockup_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM iv_mockups WHERE id = ?", (mockup_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ═══════════════════════════════════════════════════════════
# MODULE PIGISTES
# ═══════════════════════════════════════════════════════════

# ── Helpers ─────────────────────────────────────────────────

def _generer_numero_facture_pigiste(conn) -> str:
    from datetime import datetime as _dt
    year = _dt.now().year
    row = conn.execute(
        "SELECT COUNT(*) as n FROM factures_pigiste WHERE numero LIKE ?",
        (f"FP-{year}-%",)
    ).fetchone()
    seq = (row['n'] if row else 0) + 1
    numero = f"FP-{year}-{str(seq).zfill(3)}"
    while conn.execute("SELECT 1 FROM factures_pigiste WHERE numero = ?", (numero,)).fetchone():
        seq += 1
        numero = f"FP-{year}-{str(seq).zfill(3)}"
    return numero

def _pigiste_est_producteur_principal(conn, id_pigiste) -> bool:
    row = conn.execute("SELECT is_producteur_principal FROM pigistes WHERE id = ?", (id_pigiste,)).fetchone()
    return bool(row and int(row['is_producteur_principal'] or 0) == 1)

def _pigiste_to_dict(p) -> dict:
    return {
        'id': p['id'],
        'nom_complet': p['nom_complet'],
        'email': p['email'],
        'telephone': p['telephone'],
        'adresse': p['adresse'],
        'ville': p['ville'],
        'province': p['province'],
        'code_postal': p['code_postal'],
        'numero_tps': p['numero_tps'],
        'numero_tvq': p['numero_tvq'],
        'is_active': bool(p['is_active']),
        'is_producteur_principal': bool(p['is_producteur_principal']) if 'is_producteur_principal' in p.keys() else False,
        'created_at': p['created_at'],
    }

def _mandat_to_dict(m) -> dict:
    return {
        'id': m['id'],
        'id_pigiste': m['id_pigiste'],
        'id_projet': m['id_projet'],
        'titre': m['titre'],
        'description': m['description'],
        'date_debut': m['date_debut'],
        'date_echeance': m['date_echeance'],
        'montant_convenu': m['montant_convenu'],
        'statut': m['statut'],
        'notes_admin': m['notes_admin'],
        'created_at': m['created_at'],
        'updated_at': m['updated_at'],
        'nom_projet': m['nom_projet'] if 'nom_projet' in m.keys() else None,
        'nom_pigiste': m['nom_pigiste'] if 'nom_pigiste' in m.keys() else None,
        'type_prestation': m['type_prestation'] if 'type_prestation' in m.keys() else None,
        'quantite': m['quantite'] if 'quantite' in m.keys() else 1,
    }

def _facture_pigiste_to_dict(f, lignes=None) -> dict:
    d = {
        'id': f['id'],
        'id_pigiste': f['id_pigiste'],
        'numero': f['numero'],
        'date_emission': f['date_emission'],
        'date_echeance': f['date_echeance'],
        'statut': f['statut'],
        'montant_ht': f['montant_ht'],
        'tps': f['tps'],
        'tvq': f['tvq'],
        'montant_total': f['montant_total'],
        'notes': f['notes'],
        'created_at': f['created_at'],
        'nom_pigiste': f['nom_pigiste'] if 'nom_pigiste' in f.keys() else None,
        'pdf_path': f['pdf_path'] if 'pdf_path' in f.keys() else None,
        'has_pdf': bool(f['pdf_path'] if 'pdf_path' in f.keys() else None),
    }
    if lignes is not None:
        d['lignes'] = [{
            'id': l['id'],
            'id_mandat': l['id_mandat'],
            'description': l['description'],
            'quantite': l['quantite'],
            'taux': l['taux'],
            'montant': l['montant'],
        } for l in lignes]
    return d


# ── API Pigiste (espace pigiste) ─────────────────────────────

@app.route('/api/v1/pigiste/me', methods=['GET'])
@pigiste_required
def api_pigiste_me():
    conn = get_db_connection()
    p = conn.execute("SELECT * FROM pigistes WHERE id = ?", (session['pigiste_id'],)).fetchone()
    conn.close()
    if not p:
        return jsonify({'error': 'Introuvable'}), 404
    return jsonify(_pigiste_to_dict(p))


@app.route('/api/v1/pigiste/dashboard', methods=['GET'])
@pigiste_required
def api_pigiste_dashboard():
    pid = session['pigiste_id']
    conn = get_db_connection()
    mandats_actifs = conn.execute("""
        SELECT m.*, p.nom_projet
        FROM mandats m
        LEFT JOIN projets p ON p.id = m.id_projet
        WHERE m.id_pigiste = ? AND m.statut NOT IN ('annulé','approuvé')
        ORDER BY m.date_echeance ASC
    """, (pid,)).fetchall()
    factures_en_attente = conn.execute("""
        SELECT * FROM factures_pigiste
        WHERE id_pigiste = ? AND statut IN ('brouillon','soumise')
        ORDER BY created_at DESC
    """, (pid,)).fetchall()
    conn.close()
    return jsonify({
        'mandats_actifs': [_mandat_to_dict(m) for m in mandats_actifs],
        'factures_en_attente': [_facture_pigiste_to_dict(f) for f in factures_en_attente],
        'nb_mandats_actifs': len(mandats_actifs),
        'nb_factures_en_attente': len(factures_en_attente),
    })


@app.route('/api/v1/pigiste/mandats', methods=['GET'])
@pigiste_required
def api_pigiste_mandats():
    pid = session['pigiste_id']
    conn = get_db_connection()
    mandats = conn.execute("""
        SELECT m.*, p.nom_projet
        FROM mandats m
        LEFT JOIN projets p ON p.id = m.id_projet
        WHERE m.id_pigiste = ?
        ORDER BY m.created_at DESC
    """, (pid,)).fetchall()
    conn.close()
    return jsonify([_mandat_to_dict(m) for m in mandats])


@app.route('/api/v1/pigiste/mandats/<int:mandat_id>', methods=['GET'])
@pigiste_required
def api_pigiste_mandat_detail(mandat_id):
    pid = session['pigiste_id']
    conn = get_db_connection()
    m = conn.execute("""
        SELECT m.*, p.nom_projet, p.statut as statut_projet
        FROM mandats m
        LEFT JOIN projets p ON p.id = m.id_projet
        WHERE m.id = ? AND m.id_pigiste = ?
    """, (mandat_id, pid)).fetchone()
    if not m:
        conn.close()
        return jsonify({'error': 'Mandat introuvable'}), 404
    livrables = conn.execute(
        "SELECT * FROM mandats_livrables WHERE id_mandat = ? ORDER BY uploaded_at DESC",
        (mandat_id,)
    ).fetchall()
    conn.close()
    d = _mandat_to_dict(m)
    d['statut_projet'] = m['statut_projet']
    d['livrables'] = [{'id': l['id'], 'filename': l['filename'], 'public_url': l['public_url'], 'drive_file_id': l['drive_file_id'], 'uploaded_at': l['uploaded_at']} for l in livrables]
    return jsonify(d)


@app.route('/api/v1/pigiste/mandats/<int:mandat_id>/remettre', methods=['POST'])
@pigiste_required
def api_pigiste_remettre_mandat(mandat_id):
    import tempfile
    pid = session['pigiste_id']
    conn = get_db_connection()
    m = conn.execute("SELECT * FROM mandats WHERE id = ? AND id_pigiste = ?", (mandat_id, pid)).fetchone()
    if not m:
        conn.close()
        return jsonify({'error': 'Mandat introuvable'}), 404

    uploaded = []
    for fichier in request.files.getlist('fichiers'):
        if not fichier or not fichier.filename:
            continue
        safe_name = secure_filename(fichier.filename)
        ext = safe_name.rsplit('.', 1)[-1].lower() if '.' in safe_name else 'bin'
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.' + ext)
        fichier.save(tmp.name)
        tmp.close()
        drive_file_id = None
        public_url = None
        try:
            parent = m['drive_folder_id'] or os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            drive_file_id, _ = upload_file(tmp.name, safe_name, parent)
            public_url = make_file_public(drive_file_id)
        except Exception as e:
            print(f"[DRIVE] Upload livrable pigiste: {e}")
        finally:
            try:
                os.unlink(tmp.name)
            except Exception:
                pass
        conn.execute("""
            INSERT INTO mandats_livrables (id_mandat, filename, drive_file_id, public_url)
            VALUES (?, ?, ?, ?)
        """, (mandat_id, safe_name, drive_file_id, public_url))
        uploaded.append({'filename': safe_name, 'public_url': public_url})

    if uploaded:
        conn.execute(
            "UPDATE mandats SET statut = 'remis', updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (mandat_id,)
        )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'livrables': uploaded})


@app.route('/api/v1/pigiste/factures', methods=['GET'])
@pigiste_required
def api_pigiste_factures():
    pid = session['pigiste_id']
    conn = get_db_connection()
    factures = conn.execute(
        "SELECT * FROM factures_pigiste WHERE id_pigiste = ? ORDER BY created_at DESC",
        (pid,)
    ).fetchall()
    conn.close()
    return jsonify([_facture_pigiste_to_dict(f) for f in factures])


@app.route('/api/v1/pigiste/factures/<int:facture_id>', methods=['GET'])
@pigiste_required
def api_pigiste_facture_detail(facture_id):
    pid = session['pigiste_id']
    conn = get_db_connection()
    f = conn.execute(
        "SELECT * FROM factures_pigiste WHERE id = ? AND id_pigiste = ?",
        (facture_id, pid)
    ).fetchone()
    if not f:
        conn.close()
        return jsonify({'error': 'Facture introuvable'}), 404
    lignes = conn.execute(
        "SELECT * FROM factures_pigiste_lignes WHERE id_facture = ? ORDER BY id",
        (facture_id,)
    ).fetchall()
    conn.close()
    return jsonify(_facture_pigiste_to_dict(f, lignes))


@app.route('/api/v1/pigiste/factures', methods=['POST'])
@pigiste_required
def api_pigiste_creer_facture():
    from datetime import datetime as _dt, timedelta as _td
    pid = session['pigiste_id']
    data = request.get_json(force=True)
    lignes_data = data.get('lignes', [])
    notes = (data.get('notes') or '').strip()
    date_emission = _dt.now().strftime('%Y-%m-%d')
    date_echeance_dt = _dt.now() + _td(days=30)
    date_echeance = data.get('date_echeance') or date_echeance_dt.strftime('%Y-%m-%d')

    # Calcul totaux
    montant_ht = sum(float(l.get('montant', 0)) for l in lignes_data)

    conn = get_db_connection()
    pigiste = conn.execute("SELECT * FROM pigistes WHERE id = ?", (pid,)).fetchone()
    taux_tps = 0.05 if pigiste['numero_tps'] else 0.0
    taux_tvq = 0.09975 if pigiste['numero_tvq'] else 0.0
    tps = round(montant_ht * taux_tps, 2)
    tvq = round(montant_ht * taux_tvq, 2)
    montant_total = round(montant_ht + tps + tvq, 2)

    numero = _generer_numero_facture_pigiste(conn)
    cur = conn.execute("""
        INSERT INTO factures_pigiste (id_pigiste, numero, date_emission, date_echeance, statut, montant_ht, tps, tvq, montant_total, notes)
        VALUES (?, ?, ?, ?, 'brouillon', ?, ?, ?, ?, ?)
    """, (pid, numero, date_emission, date_echeance, montant_ht, tps, tvq, montant_total, notes))
    facture_id = cur.lastrowid

    for l in lignes_data:
        desc = (l.get('description') or '').strip()
        if not desc:
            continue
        qte = float(l.get('quantite', 1))
        taux = float(l.get('taux', 0))
        montant = float(l.get('montant', qte * taux))
        id_mandat = l.get('id_mandat') or None
        conn.execute("""
            INSERT INTO factures_pigiste_lignes (id_facture, id_mandat, description, quantite, taux, montant)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (facture_id, id_mandat, desc, qte, taux, montant))

    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': facture_id, 'numero': numero}), 201


def _generer_et_sauver_pdf_pigiste(facture_id: int, conn) -> str | None:
    """Génère le PDF d'une facture pigiste et met à jour pdf_path en DB. Retourne le chemin ou None."""
    import pathlib
    from invoice_service import generer_pdf_facture_pigiste
    try:
        f       = conn.execute("SELECT * FROM factures_pigiste WHERE id = ?", (facture_id,)).fetchone()
        pigiste = conn.execute("SELECT * FROM pigistes WHERE id = ?", (f['id_pigiste'],)).fetchone()
        lignes  = conn.execute("""
            SELECT fl.*,
                   m.titre       AS mandat_titre,
                   m.date_debut  AS mandat_date_debut,
                   m.date_echeance AS mandat_date_echeance,
                   p.nom_projet,
                   c.nom_complet AS nom_client
            FROM factures_pigiste_lignes fl
            LEFT JOIN mandats m  ON m.id  = fl.id_mandat
            LEFT JOIN projets p  ON p.id  = m.id_projet
            LEFT JOIN clients c  ON c.id  = p.id_client
            WHERE fl.id_facture = ?
            ORDER BY fl.id
        """, (facture_id,)).fetchall()

        upload_root  = os.getenv("UPLOAD_ROOT", "/data/uploads")
        factures_dir = os.path.join(upload_root, "factures_pigiste", f"pigiste_{f['id_pigiste']}")
        pathlib.Path(factures_dir).mkdir(parents=True, exist_ok=True)
        pdf_path = os.path.join(factures_dir, f"{f['numero']}.pdf")

        facture_dict = {
            "numero":        f["numero"],
            "date_emission": f["date_emission"],
            "date_echeance": f["date_echeance"] or "À la réception",
            "tps":           f["tps"],
            "tvq":           f["tvq"],
            "montant_total": f["montant_total"],
        }

        def _fmt_date_mandat(debut, echeance):
            if debut and echeance and debut != echeance:
                return f"{debut} → {echeance}"
            return debut or echeance or ""

        lignes_dict = [{
            "description":   l["description"],
            "quantite":      l["quantite"],
            "taux":          l["taux"],
            "montant":       l["montant"],
            "mandat_titre":  l["mandat_titre"] or "",
            "nom_projet":    l["nom_projet"] or "",
            "nom_client":    l["nom_client"] or "",
            "date_mandat":   _fmt_date_mandat(l["mandat_date_debut"], l["mandat_date_echeance"]),
        } for l in lignes]

        generer_pdf_facture_pigiste(facture_dict, lignes_dict, dict(pigiste), pdf_path)
        conn.execute("UPDATE factures_pigiste SET pdf_path = ? WHERE id = ?", (pdf_path, facture_id))
        return pdf_path
    except Exception as e:
        print(f"[INVOICE] Génération PDF pigiste: {e}")
        return None


@app.route('/api/v1/pigiste/factures/<int:facture_id>/soumettre', methods=['POST'])
@pigiste_required
def api_pigiste_soumettre_facture(facture_id):
    pid = session['pigiste_id']
    conn = get_db_connection()
    f = conn.execute(
        "SELECT * FROM factures_pigiste WHERE id = ? AND id_pigiste = ? AND statut = 'brouillon'",
        (facture_id, pid)
    ).fetchone()
    if not f:
        conn.close()
        return jsonify({'error': 'Facture introuvable ou déjà soumise'}), 404

    conn.execute("UPDATE factures_pigiste SET statut = 'soumise' WHERE id = ?", (facture_id,))
    conn.commit()
    _generer_et_sauver_pdf_pigiste(facture_id, conn)
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/v1/pigiste/factures/<int:facture_id>/lignes', methods=['POST'])
@pigiste_required
def api_pigiste_ajouter_ligne(facture_id):
    pid = session['pigiste_id']
    data = request.get_json(force=True)
    conn = get_db_connection()
    f = conn.execute(
        "SELECT * FROM factures_pigiste WHERE id = ? AND id_pigiste = ? AND statut = 'brouillon'",
        (facture_id, pid)
    ).fetchone()
    if not f:
        conn.close()
        return jsonify({'error': 'Facture introuvable ou non modifiable'}), 404
    desc = (data.get('description') or '').strip()
    qte = float(data.get('quantite', 1))
    taux = float(data.get('taux', 0))
    montant = round(qte * taux, 2)
    id_mandat = data.get('id_mandat') or None
    cur = conn.execute("""
        INSERT INTO factures_pigiste_lignes (id_facture, id_mandat, description, quantite, taux, montant)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (facture_id, id_mandat, desc, qte, taux, montant))
    # Recalcul totaux
    lignes = conn.execute("SELECT montant FROM factures_pigiste_lignes WHERE id_facture = ?", (facture_id,)).fetchall()
    pigiste = conn.execute("SELECT * FROM pigistes WHERE id = ?", (pid,)).fetchone()
    montant_ht = sum(l['montant'] for l in lignes)
    taux_tps = 0.05 if pigiste['numero_tps'] else 0.0
    taux_tvq = 0.09975 if pigiste['numero_tvq'] else 0.0
    tps = round(montant_ht * taux_tps, 2)
    tvq = round(montant_ht * taux_tvq, 2)
    conn.execute(
        "UPDATE factures_pigiste SET montant_ht=?, tps=?, tvq=?, montant_total=? WHERE id=?",
        (montant_ht, tps, tvq, round(montant_ht + tps + tvq, 2), facture_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': cur.lastrowid, 'montant': montant})


@app.route('/api/v1/pigiste/factures/<int:facture_id>/lignes/<int:ligne_id>', methods=['DELETE'])
@pigiste_required
def api_pigiste_supprimer_ligne(facture_id, ligne_id):
    pid = session['pigiste_id']
    conn = get_db_connection()
    f = conn.execute(
        "SELECT * FROM factures_pigiste WHERE id = ? AND id_pigiste = ? AND statut = 'brouillon'",
        (facture_id, pid)
    ).fetchone()
    if not f:
        conn.close()
        return jsonify({'error': 'Non modifiable'}), 403
    conn.execute("DELETE FROM factures_pigiste_lignes WHERE id = ? AND id_facture = ?", (ligne_id, facture_id))
    lignes = conn.execute("SELECT montant FROM factures_pigiste_lignes WHERE id_facture = ?", (facture_id,)).fetchall()
    pigiste = conn.execute("SELECT * FROM pigistes WHERE id = ?", (pid,)).fetchone()
    montant_ht = sum(l['montant'] for l in lignes)
    taux_tps = 0.05 if pigiste['numero_tps'] else 0.0
    taux_tvq = 0.09975 if pigiste['numero_tvq'] else 0.0
    tps = round(montant_ht * taux_tps, 2)
    tvq = round(montant_ht * taux_tvq, 2)
    conn.execute(
        "UPDATE factures_pigiste SET montant_ht=?, tps=?, tvq=?, montant_total=? WHERE id=?",
        (montant_ht, tps, tvq, round(montant_ht + tps + tvq, 2), facture_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ── API Admin — Pigistes ─────────────────────────────────────

@app.route('/api/v1/admin/pigistes', methods=['GET'])
@admin_required
def api_admin_pigistes_list():
    conn = get_db_connection()
    pigistes = conn.execute("""
        SELECT p.*, COUNT(m.id) as nb_mandats
        FROM pigistes p
        LEFT JOIN mandats m ON m.id_pigiste = p.id
        GROUP BY p.id
        ORDER BY p.nom_complet
    """).fetchall()
    conn.close()
    result = []
    for p in pigistes:
        d = _pigiste_to_dict(p)
        d['nb_mandats'] = p['nb_mandats']
        result.append(d)
    return jsonify(result)


@app.route('/api/v1/admin/pigistes', methods=['POST'])
@admin_required
def api_admin_pigistes_create():
    data = request.get_json(force=True)
    nom = (data.get('nom_complet') or '').strip()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''
    if not nom or not email or not password:
        return jsonify({'error': 'nom_complet, email et password requis'}), 400
    hashed = bcrypt.generate_password_hash(password).decode('utf-8')
    conn = get_db_connection()
    try:
        cur = conn.execute("""
            INSERT INTO pigistes (nom_complet, email, mot_de_passe_hash, telephone, adresse, ville, province, code_postal, numero_tps, numero_tvq)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (nom, email, hashed,
              data.get('telephone', ''), data.get('adresse', ''), data.get('ville', ''),
              data.get('province', 'Québec'), data.get('code_postal', ''),
              data.get('numero_tps', ''), data.get('numero_tvq', '')))
        conn.commit()
        pid = cur.lastrowid
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)}), 409
    conn.close()
    return jsonify({'ok': True, 'id': pid}), 201


@app.route('/api/v1/admin/pigistes/<int:pigiste_id>', methods=['GET'])
@admin_required
def api_admin_pigiste_detail(pigiste_id):
    conn = get_db_connection()
    p = conn.execute("SELECT * FROM pigistes WHERE id = ?", (pigiste_id,)).fetchone()
    if not p:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    mandats = conn.execute("""
        SELECT m.*, p.nom_projet
        FROM mandats m
        LEFT JOIN projets p ON p.id = m.id_projet
        WHERE m.id_pigiste = ?
        ORDER BY m.created_at DESC
    """, (pigiste_id,)).fetchall()
    factures = conn.execute(
        "SELECT * FROM factures_pigiste WHERE id_pigiste = ? ORDER BY created_at DESC",
        (pigiste_id,)
    ).fetchall()
    conn.close()
    d = _pigiste_to_dict(p)
    d['mandats'] = [_mandat_to_dict(m) for m in mandats]
    d['factures'] = [_facture_pigiste_to_dict(f) for f in factures]
    return jsonify(d)


@app.route('/api/v1/admin/pigistes/<int:pigiste_id>', methods=['PUT'])
@admin_required
def api_admin_pigiste_update(pigiste_id):
    data = request.get_json(force=True)
    conn = get_db_connection()
    p = conn.execute("SELECT * FROM pigistes WHERE id = ?", (pigiste_id,)).fetchone()
    if not p:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    fields = ['nom_complet', 'email', 'telephone', 'adresse', 'ville', 'province', 'code_postal', 'numero_tps', 'numero_tvq', 'is_active']
    updates = {f: data.get(f, p[f]) for f in fields}
    if data.get('password'):
        updates['mot_de_passe_hash'] = bcrypt.generate_password_hash(data['password']).decode('utf-8')
        conn.execute("""
            UPDATE pigistes SET nom_complet=?, email=?, telephone=?, adresse=?, ville=?, province=?, code_postal=?, numero_tps=?, numero_tvq=?, is_active=?, mot_de_passe_hash=?
            WHERE id=?
        """, (*[updates[f] for f in fields], updates['mot_de_passe_hash'], pigiste_id))
    else:
        conn.execute("""
            UPDATE pigistes SET nom_complet=?, email=?, telephone=?, adresse=?, ville=?, province=?, code_postal=?, numero_tps=?, numero_tvq=?, is_active=?
            WHERE id=?
        """, (*[updates[f] for f in fields], pigiste_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/v1/admin/pigistes/<int:pigiste_id>', methods=['DELETE'])
@admin_required
def api_admin_pigiste_delete(pigiste_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM pigistes WHERE id = ?", (pigiste_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/v1/admin/projet/<int:projet_id>/mandats', methods=['GET'])
@admin_required
def api_admin_projet_mandats(projet_id):
    conn = get_db_connection()
    mandats = conn.execute("""
        SELECT m.*, pg.nom_complet as nom_pigiste
        FROM mandats m
        LEFT JOIN pigistes pg ON pg.id = m.id_pigiste
        WHERE m.id_projet = ?
        ORDER BY m.created_at DESC
    """, (projet_id,)).fetchall()
    conn.close()
    return jsonify([_mandat_to_dict(m) for m in mandats])


@app.route('/api/v1/admin/mandats', methods=['GET'])
@admin_required
def api_admin_mandats_list():
    conn = get_db_connection()
    mandats = conn.execute("""
        SELECT m.*, p.nom_projet, pg.nom_complet as nom_pigiste
        FROM mandats m
        LEFT JOIN projets p ON p.id = m.id_projet
        LEFT JOIN pigistes pg ON pg.id = m.id_pigiste
        ORDER BY m.created_at DESC
    """).fetchall()
    conn.close()
    return jsonify([_mandat_to_dict(m) for m in mandats])


@app.route('/api/v1/admin/mandats', methods=['POST'])
@admin_required
def api_admin_mandats_create():
    data = request.get_json(force=True)
    id_pigiste = data.get('id_pigiste')
    id_projet = data.get('id_projet') or None
    titre = (data.get('titre') or '').strip()
    if not id_pigiste or not titre:
        return jsonify({'error': 'id_pigiste et titre requis'}), 400
    conn = get_db_connection()
    if not _pigiste_est_producteur_principal(conn, id_pigiste):
        conn.close()
        return jsonify({'error': "Ce pigiste ne peut pas recevoir de mandat (producteur principal requis)"}), 403
    cur = conn.execute("""
        INSERT INTO mandats (id_pigiste, id_projet, titre, description, date_debut, date_echeance, montant_convenu, statut, notes_admin)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'en_attente', ?)
    """, (id_pigiste, id_projet, titre,
          data.get('description', ''), data.get('date_debut'), data.get('date_echeance'),
          float(data.get('montant_convenu', 0)), data.get('notes_admin', '')))
    conn.commit()
    mandat_id = cur.lastrowid
    # Mettre à jour id_pigiste sur le projet si fourni
    if id_projet:
        conn.execute("UPDATE projets SET id_pigiste = ? WHERE id = ?", (id_pigiste, id_projet))
        conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': mandat_id}), 201


@app.route('/api/v1/admin/mandats/<int:mandat_id>', methods=['GET'])
@admin_required
def api_admin_mandat_detail(mandat_id):
    conn = get_db_connection()
    m = conn.execute("""
        SELECT m.*, p.nom_projet, pg.nom_complet as nom_pigiste
        FROM mandats m
        LEFT JOIN projets p ON p.id = m.id_projet
        LEFT JOIN pigistes pg ON pg.id = m.id_pigiste
        WHERE m.id = ?
    """, (mandat_id,)).fetchone()
    if not m:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    livrables = conn.execute(
        "SELECT * FROM mandats_livrables WHERE id_mandat = ? ORDER BY uploaded_at DESC",
        (mandat_id,)
    ).fetchall()
    conn.close()
    d = _mandat_to_dict(m)
    d['livrables'] = [{'id': l['id'], 'filename': l['filename'], 'public_url': l['public_url'], 'drive_file_id': l['drive_file_id'], 'uploaded_at': l['uploaded_at']} for l in livrables]
    return jsonify(d)


@app.route('/api/v1/admin/mandats/<int:mandat_id>', methods=['PUT'])
@admin_required
def api_admin_mandat_update(mandat_id):
    data = request.get_json(force=True)
    conn = get_db_connection()
    m = conn.execute("SELECT * FROM mandats WHERE id = ?", (mandat_id,)).fetchone()
    if not m:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    conn.execute("""
        UPDATE mandats SET titre=?, description=?, date_debut=?, date_echeance=?, montant_convenu=?, statut=?, notes_admin=?, updated_at=CURRENT_TIMESTAMP
        WHERE id=?
    """, (
        data.get('titre', m['titre']),
        data.get('description', m['description']),
        data.get('date_debut', m['date_debut']),
        data.get('date_echeance', m['date_echeance']),
        float(data.get('montant_convenu', m['montant_convenu'])),
        data.get('statut', m['statut']),
        data.get('notes_admin', m['notes_admin']),
        mandat_id
    ))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/v1/admin/mandats/<int:mandat_id>/approuver', methods=['POST'])
@admin_required
def api_admin_mandat_approuver(mandat_id):
    conn = get_db_connection()
    conn.execute(
        "UPDATE mandats SET statut='approuvé', updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (mandat_id,)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/v1/admin/factures-pigiste', methods=['GET'])
@admin_required
def api_admin_factures_pigiste_list():
    conn = get_db_connection()
    factures = conn.execute("""
        SELECT f.*, pg.nom_complet as nom_pigiste
        FROM factures_pigiste f
        JOIN pigistes pg ON pg.id = f.id_pigiste
        ORDER BY f.created_at DESC
    """).fetchall()
    conn.close()
    return jsonify([_facture_pigiste_to_dict(f) for f in factures])


@app.route('/api/v1/admin/factures-pigiste/<int:facture_id>', methods=['GET'])
@admin_required
def api_admin_facture_pigiste_detail(facture_id):
    conn = get_db_connection()
    f = conn.execute("""
        SELECT f.*, pg.nom_complet as nom_pigiste
        FROM factures_pigiste f
        JOIN pigistes pg ON pg.id = f.id_pigiste
        WHERE f.id = ?
    """, (facture_id,)).fetchone()
    if not f:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    lignes = conn.execute(
        "SELECT * FROM factures_pigiste_lignes WHERE id_facture = ? ORDER BY id",
        (facture_id,)
    ).fetchall()
    conn.close()
    return jsonify(_facture_pigiste_to_dict(f, lignes))


@app.route('/api/v1/admin/factures-pigiste/<int:facture_id>/payer', methods=['POST'])
@admin_required
def api_admin_facture_pigiste_payer(facture_id):
    conn = get_db_connection()
    conn.execute("UPDATE factures_pigiste SET statut='payée' WHERE id=?", (facture_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/v1/admin/factures-pigiste/<int:facture_id>/approuver', methods=['POST'])
@admin_required
def api_admin_facture_pigiste_approuver(facture_id):
    conn = get_db_connection()
    conn.execute("UPDATE factures_pigiste SET statut='approuvée' WHERE id=?", (facture_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ── Admin: mandats-pigistes ──────────────────────────────────

@app.route('/api/v1/admin/mandats-pigistes', methods=['GET'])
@admin_required
def api_admin_mandats_pigistes_list():
    conn = get_db_connection()
    mandats = conn.execute("""
        SELECT m.*, p.nom_projet, pg.nom_complet as nom_pigiste
        FROM mandats m
        LEFT JOIN projets p ON p.id = m.id_projet
        LEFT JOIN pigistes pg ON pg.id = m.id_pigiste
        ORDER BY m.created_at DESC
    """).fetchall()
    conn.close()
    return jsonify([_mandat_to_dict(m) for m in mandats])


@app.route('/api/v1/admin/tarifs-pigiste', methods=['GET'])
@admin_required
def api_admin_tarifs_pigiste():
    return jsonify(TARIFS_PIGISTE)


@app.route('/api/v1/admin/mandats-pigistes', methods=['POST'])
@admin_required
def api_admin_mandats_pigistes_create():
    data = request.get_json(force=True)
    id_pigiste       = data.get('id_pigiste')
    id_projet        = data.get('id_projet') or None
    titre            = (data.get('titre') or '').strip()
    type_prestation  = data.get('type_prestation') or None
    quantite         = int(data.get('quantite') or 1)
    if not id_pigiste or not titre:
        return jsonify({'error': 'id_pigiste et titre requis'}), 400
    # Prix automatique si type_prestation reconnu
    if type_prestation and type_prestation in _TARIFS_MAP:
        montant = _TARIFS_MAP[type_prestation]['prix'] * quantite
    else:
        montant = float(data.get('montant_convenu') or 0)
    conn = get_db_connection()
    if not _pigiste_est_producteur_principal(conn, id_pigiste):
        conn.close()
        return jsonify({'error': "Ce pigiste ne peut pas recevoir de mandat (producteur principal requis)"}), 403
    cur = conn.execute("""
        INSERT INTO mandats (id_pigiste, id_projet, titre, description, date_debut, date_echeance,
                             montant_convenu, statut, notes_admin, type_prestation, quantite)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'en_attente', ?, ?, ?)
    """, (id_pigiste, id_projet, titre,
          data.get('description', ''), data.get('date_debut'), data.get('date_echeance'),
          montant, data.get('notes_admin', ''), type_prestation, quantite))
    conn.commit()
    mandat_id = cur.lastrowid
    if id_projet:
        conn.execute("UPDATE projets SET id_pigiste = ? WHERE id = ?", (id_pigiste, id_projet))
        conn.commit()
    conn.close()
    return jsonify({'ok': True, 'id': mandat_id}), 201


@app.route('/api/v1/admin/mandats-pigistes/<int:mandat_id>', methods=['GET'])
@admin_required
def api_admin_mandats_pigistes_detail(mandat_id):
    conn = get_db_connection()
    m = conn.execute("""
        SELECT m.*, p.nom_projet, pg.nom_complet as nom_pigiste
        FROM mandats m
        LEFT JOIN projets p ON p.id = m.id_projet
        LEFT JOIN pigistes pg ON pg.id = m.id_pigiste
        WHERE m.id = ?
    """, (mandat_id,)).fetchone()
    if not m:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    livrables = conn.execute(
        "SELECT * FROM mandats_livrables WHERE id_mandat = ? ORDER BY uploaded_at DESC",
        (mandat_id,)
    ).fetchall()
    conn.close()
    d = _mandat_to_dict(m)
    d['livrables'] = [{'id': l['id'], 'filename': l['filename'], 'public_url': l['public_url'], 'drive_file_id': l['drive_file_id'], 'uploaded_at': l['uploaded_at']} for l in livrables]
    return jsonify(d)


@app.route('/api/v1/admin/mandats-pigistes/<int:mandat_id>', methods=['PUT'])
@admin_required
def api_admin_mandats_pigistes_update(mandat_id):
    data = request.get_json(force=True)
    conn = get_db_connection()
    m = conn.execute("SELECT * FROM mandats WHERE id = ?", (mandat_id,)).fetchone()
    if not m:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    conn.execute("""
        UPDATE mandats SET titre=?, description=?, date_debut=?, date_echeance=?, montant_convenu=?, statut=?, notes_admin=?, updated_at=CURRENT_TIMESTAMP
        WHERE id=?
    """, (
        data.get('titre', m['titre']),
        data.get('description', m['description']),
        data.get('date_debut', m['date_debut']),
        data.get('date_echeance', m['date_echeance']),
        float(data.get('montant_convenu', m['montant_convenu'])),
        data.get('statut', m['statut']),
        data.get('notes_admin', m['notes_admin']),
        mandat_id
    ))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/v1/admin/mandats-pigistes/<int:mandat_id>/assigner', methods=['POST'])
@admin_required
def api_admin_mandats_pigistes_assigner(mandat_id):
    """Passe le mandat à 'en_cours' (travaux assignés et démarrés)."""
    conn = get_db_connection()
    m = conn.execute("SELECT id FROM mandats WHERE id = ?", (mandat_id,)).fetchone()
    if not m:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    conn.execute(
        "UPDATE mandats SET statut='en_cours', updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (mandat_id,)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'statut': 'en_cours'})


@app.route('/api/v1/admin/mandats-pigistes/<int:mandat_id>/approuver', methods=['POST'])
@admin_required
def api_admin_mandats_pigistes_approuver(mandat_id):
    """Approuve les livrables remis — statut → 'approuvé' + crée la facture pigiste automatiquement."""
    import datetime as _dt
    conn = get_db_connection()
    m = conn.execute("""
        SELECT m.*, pg.numero_tps, pg.numero_tvq
        FROM mandats m
        JOIN pigistes pg ON pg.id = m.id_pigiste
        WHERE m.id = ?
    """, (mandat_id,)).fetchone()
    if not m:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404

    # Marquer le mandat approuvé
    conn.execute(
        "UPDATE mandats SET statut='approuvé', updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (mandat_id,)
    )

    # ── Construire la ligne de facture ──────────────────────────
    type_prestation = m['type_prestation'] if 'type_prestation' in m.keys() else None
    quantite        = (m['quantite'] if 'quantite' in m.keys() else None) or 1
    tarif           = _TARIFS_MAP.get(type_prestation) if type_prestation else None

    if tarif:
        taux        = tarif['prix']
        description = f"{tarif['label']} × {quantite} {tarif['unite']}"
        montant_ht  = round(taux * quantite, 2)
    else:
        taux        = m['montant_convenu'] or 0
        description = m['titre']
        montant_ht  = round(float(taux), 2)
        quantite    = 1

    tps            = round(montant_ht * 0.05,    2) if m['numero_tps'] else 0.0
    tvq            = round(montant_ht * 0.09975, 2) if m['numero_tvq'] else 0.0
    montant_total  = round(montant_ht + tps + tvq, 2)

    # ── Numéro de facture unique ─────────────────────────────────
    now     = _dt.datetime.now()
    mois    = now.strftime('%Y%m')
    count   = conn.execute(
        "SELECT COUNT(*) FROM factures_pigiste WHERE strftime('%Y%m', created_at) = ?", (mois,)
    ).fetchone()[0] + 1
    numero  = f"FPIG-{mois}-{count:04d}"

    # ── Insérer la facture ───────────────────────────────────────
    cur = conn.execute("""
        INSERT INTO factures_pigiste
            (id_pigiste, numero, date_emission, statut, montant_ht, tps, tvq, montant_total, notes)
        VALUES (?, ?, ?, 'soumise', ?, ?, ?, ?, ?)
    """, (m['id_pigiste'], numero, now.strftime('%Y-%m-%d'),
          montant_ht, tps, tvq, montant_total,
          f"Générée automatiquement — mandat #{mandat_id}"))
    facture_id = cur.lastrowid

    conn.execute("""
        INSERT INTO factures_pigiste_lignes (id_facture, id_mandat, description, quantite, taux, montant)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (facture_id, mandat_id, description, quantite, taux, montant_ht))

    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'statut': 'approuvé', 'facture_id': facture_id, 'facture_numero': numero})


@app.route('/api/v1/admin/mandats-pigistes/<int:mandat_id>/corrections', methods=['POST'])
@admin_required
def api_admin_mandats_pigistes_corrections(mandat_id):
    """Demande des corrections — repasse à 'en_cours' avec une note admin mise à jour."""
    data = request.get_json(force=True)
    note = (data.get('note') or '').strip()
    conn = get_db_connection()
    m = conn.execute("SELECT * FROM mandats WHERE id = ?", (mandat_id,)).fetchone()
    if not m:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    conn.execute(
        "UPDATE mandats SET statut='en_cours', notes_admin=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
        (note if note else m['notes_admin'], mandat_id)
    )
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'statut': 'en_cours'})


# ── Admin: factures-pigistes ─────────────────────────────────

@app.route('/api/v1/admin/factures-pigistes', methods=['GET'])
@admin_required
def api_admin_factures_pigistes_list():
    conn = get_db_connection()
    factures = conn.execute("""
        SELECT f.*, pg.nom_complet as nom_pigiste
        FROM factures_pigiste f
        JOIN pigistes pg ON pg.id = f.id_pigiste
        ORDER BY f.created_at DESC
    """).fetchall()
    conn.close()
    return jsonify([_facture_pigiste_to_dict(f) for f in factures])


@app.route('/api/v1/admin/factures-pigistes/<int:facture_id>', methods=['GET'])
@admin_required
def api_admin_factures_pigistes_detail(facture_id):
    conn = get_db_connection()
    f = conn.execute("""
        SELECT f.*, pg.nom_complet as nom_pigiste
        FROM factures_pigiste f
        JOIN pigistes pg ON pg.id = f.id_pigiste
        WHERE f.id = ?
    """, (facture_id,)).fetchone()
    if not f:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    lignes = conn.execute(
        "SELECT * FROM factures_pigiste_lignes WHERE id_facture = ? ORDER BY id",
        (facture_id,)
    ).fetchall()
    conn.close()
    return jsonify(_facture_pigiste_to_dict(f, lignes))


@app.route('/api/v1/admin/factures-pigistes/<int:facture_id>/pdf', methods=['GET'])
@admin_required
def api_admin_factures_pigistes_pdf(facture_id):
    conn = get_db_connection()
    f = conn.execute("SELECT pdf_path, numero FROM factures_pigiste WHERE id = ?", (facture_id,)).fetchone()
    conn.close()
    if not f or not f['pdf_path'] or not os.path.exists(f['pdf_path']):
        return jsonify({'error': 'PDF non disponible'}), 404
    from flask import send_file
    return send_file(f['pdf_path'], mimetype='application/pdf',
                     as_attachment=True, download_name=f"{f['numero']}.pdf")


@app.route('/api/v1/pigiste/factures/<int:facture_id>/pdf', methods=['GET'])
@pigiste_required
def api_pigiste_facture_pdf(facture_id):
    pid = session['pigiste_id']
    conn = get_db_connection()
    f = conn.execute(
        "SELECT pdf_path, numero FROM factures_pigiste WHERE id = ? AND id_pigiste = ?",
        (facture_id, pid)
    ).fetchone()
    conn.close()
    if not f or not f['pdf_path'] or not os.path.exists(f['pdf_path']):
        return jsonify({'error': 'PDF non disponible'}), 404
    from flask import send_file
    return send_file(f['pdf_path'], mimetype='application/pdf',
                     as_attachment=True, download_name=f"{f['numero']}.pdf")


@app.route('/api/v1/admin/factures-pigistes/<int:facture_id>/approuver', methods=['POST'])
@admin_required
def api_admin_factures_pigistes_approuver(facture_id):
    conn = get_db_connection()
    f = conn.execute("SELECT id, pdf_path FROM factures_pigiste WHERE id = ?", (facture_id,)).fetchone()
    if not f:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    conn.execute("UPDATE factures_pigiste SET statut='approuvée' WHERE id=?", (facture_id,))
    conn.commit()
    if not f['pdf_path'] or not os.path.exists(f['pdf_path'] or ''):
        _generer_et_sauver_pdf_pigiste(facture_id, conn)
    conn.close()
    return jsonify({'ok': True, 'statut': 'approuvée'})


@app.route('/api/v1/admin/factures-pigistes/<int:facture_id>/payer', methods=['POST'])
@admin_required
def api_admin_factures_pigistes_payer(facture_id):
    conn = get_db_connection()
    f = conn.execute("SELECT id FROM factures_pigiste WHERE id = ?", (facture_id,)).fetchone()
    if not f:
        conn.close()
        return jsonify({'error': 'Introuvable'}), 404
    conn.execute("UPDATE factures_pigiste SET statut='payée' WHERE id=?", (facture_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'statut': 'payée'})


# ── Projet: exposer id_pigiste dans l'API admin projet ──────

# Patch: l'API admin/projet/<id> GET inclut déjà tout le row projets,
# mais on s'assure que le PUT accepte id_pigiste.

# ── Générateur de documents — sauvegarde via Google Drive ────────────────────

from drive_service import (
    upload_json_content, update_json_content, get_json_content, delete_drive_file,
    get_file_bytes, get_file_meta, upload_bytes,
)

# ── Médiathèque — stockage Google Drive ─────────────────────────────────────

_MEDIA_CATS = ['canva', 'psd', 'photos', 'polices', 'logos']

def _mediatheque_root():
    root = os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
    return create_folder('Médiathèque', parent_id=root)

def _media_cat_folder(root_id, category):
    return create_folder(category, parent_id=root_id)

@app.route('/api/v1/tools/assets', methods=['GET'])
def api_tools_assets():
    if 'user_id' not in session and 'pigiste_id' not in session:
        return jsonify({'error': 'Non authentifié'}), 401
    try:
        cat_filter = request.args.get('category')
        root_id = _mediatheque_root()
        cats = [cat_filter] if cat_filter and cat_filter in _MEDIA_CATS else _MEDIA_CATS
        results = []
        for cat in cats:
            folder_id = _media_cat_folder(root_id, cat)
            files = list_files_in_folder(folder_id)
            for f in files:
                results.append({
                    'id':       f['id'],
                    'filename': f['name'],
                    'category': cat,
                    'size':     int(f.get('size') or 0),
                    'uploaded': f.get('createdTime', ''),
                    'url':      f'/api/v1/tools/file/{f["id"]}',
                })
        return jsonify(results)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/tools/upload', methods=['POST'])
def api_tools_upload():
    if 'user_id' not in session:
        return jsonify({'error': 'Non authentifié'}), 401
    try:
        file = request.files.get('file')
        category = request.form.get('category', 'photos')
        if not file or not file.filename:
            return jsonify({'error': 'Fichier requis'}), 400
        if category not in _MEDIA_CATS:
            category = 'photos'
        root_id = _mediatheque_root()
        folder_id = _media_cat_folder(root_id, category)
        content = file.read()
        mimetype = file.content_type or 'application/octet-stream'
        result = upload_bytes(folder_id, file.filename, content, mimetype)
        return jsonify({
            'ok':  True,
            'id':  result['id'],
            'url': f'/api/v1/tools/file/{result["id"]}',
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/tools/delete', methods=['DELETE'])
def api_tools_delete():
    if 'user_id' not in session:
        return jsonify({'error': 'Non authentifié'}), 401
    try:
        data = request.get_json(force=True)
        file_id = data.get('id')
        if not file_id:
            return jsonify({'error': 'ID requis'}), 400
        delete_drive_file(file_id)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/v1/tools/file/<file_id>')
def api_tools_serve_file(file_id):
    if 'user_id' not in session and 'pigiste_id' not in session:
        return jsonify({'error': 'Non authentifié'}), 401
    try:
        meta    = get_file_meta(file_id)
        content = get_file_bytes(file_id)
        resp = make_response(content)
        resp.headers['Content-Type']        = meta.get('mimeType', 'application/octet-stream')
        resp.headers['Content-Disposition'] = f'inline; filename="{meta["name"]}"'
        resp.headers['Cache-Control']       = 'private, max-age=3600'
        return resp
    except Exception as e:
        return jsonify({'error': str(e)}), 404


# ── Ressources clients — guides/documents livrés (généraux ou ciblés) ───────

def _ressources_clients_root():
    return create_folder('Ressources Clients', parent_id=os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID'))

def _ressource_to_dict(r, sections=None):
    keys = r.keys()
    return {
        'id': r['id'],
        'id_client': r['id_client'],
        'titre': r['titre'],
        'description': r['description'],
        'categorie': r['categorie'],
        'type_source': r['type_source'],
        'url': (f"/api/v1/tools/file/{r['drive_file_id']}" if r['type_source'] == 'upload' and r['drive_file_id'] else r['url']),
        'sections': sections if sections is not None else [],
        'created_at': r['created_at'],
        'bundle_id': r['bundle_id'] if 'bundle_id' in keys else None,
    }

def _guide_sections_brief(conn, ressource_id):
    rows = conn.execute(
        "SELECT id, titre FROM guide_sections WHERE id_ressource=? ORDER BY ordre ASC, id ASC",
        (ressource_id,)
    ).fetchall()
    return [{'id': str(s['id']), 'label': s['titre']} for s in rows]

def _guide_section_to_dict(s):
    try:
        etapes = json.loads(s['etapes_json']) if s['etapes_json'] else []
    except Exception:
        etapes = []
    return {
        'id': s['id'],
        'id_ressource': s['id_ressource'],
        'ordre': s['ordre'],
        'titre': s['titre'],
        'intro': s['intro'],
        'astuce': s['astuce'],
        'etapes': etapes,
        'created_at': s['created_at'],
    }

def _parse_etapes(raw):
    """Valide et normalise une liste d'étapes envoyée par l'éditeur de guide."""
    if not isinstance(raw, list):
        return []
    etapes = []
    for e in raw:
        if not isinstance(e, dict):
            continue
        titre = (e.get('titre') or '').strip()
        texte = (e.get('texte') or '').strip()
        image_url = (e.get('image_url') or '').strip() or None
        if titre or texte:
            etapes.append({'titre': titre, 'texte': texte, 'image_url': image_url})
    return etapes

@app.route('/api/v1/admin/ressources/<int:ressource_id>/sections', methods=['GET'])
@admin_required
def api_admin_guide_sections_list(ressource_id):
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT * FROM guide_sections WHERE id_ressource=? ORDER BY ordre ASC, id ASC",
        (ressource_id,)
    ).fetchall()
    conn.close()
    return jsonify([_guide_section_to_dict(s) for s in rows])

@app.route('/api/v1/admin/ressources/<int:ressource_id>/sections', methods=['POST'])
@admin_required
def api_admin_guide_sections_create(ressource_id):
    conn = get_db_connection()
    ressource = conn.execute("SELECT id FROM client_ressources WHERE id=?", (ressource_id,)).fetchone()
    if not ressource:
        conn.close()
        return jsonify({'error': 'Ressource introuvable'}), 404

    data = request.get_json(silent=True) or {}
    titre = (data.get('titre') or '').strip()
    if not titre:
        conn.close()
        return jsonify({'error': 'Titre de section requis'}), 400
    intro = (data.get('intro') or '').strip() or None
    astuce = (data.get('astuce') or '').strip() or None
    etapes = _parse_etapes(data.get('etapes'))

    max_ordre = conn.execute(
        "SELECT COALESCE(MAX(ordre), -1) AS m FROM guide_sections WHERE id_ressource=?",
        (ressource_id,)
    ).fetchone()['m']

    cur = conn.execute("""
        INSERT INTO guide_sections (id_ressource, ordre, titre, intro, astuce, etapes_json)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (ressource_id, max_ordre + 1, titre, intro, astuce, json.dumps(etapes)))
    section_id = cur.lastrowid
    conn.commit()
    row = conn.execute("SELECT * FROM guide_sections WHERE id=?", (section_id,)).fetchone()
    conn.close()
    return jsonify(_guide_section_to_dict(row))

@app.route('/api/v1/admin/ressources/<int:ressource_id>/sections/<int:section_id>', methods=['PUT'])
@admin_required
def api_admin_guide_sections_update(ressource_id, section_id):
    conn = get_db_connection()
    row = conn.execute(
        "SELECT * FROM guide_sections WHERE id=? AND id_ressource=?",
        (section_id, ressource_id)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Section introuvable'}), 404

    data = request.get_json(silent=True) or {}
    titre = (data.get('titre') or '').strip() or row['titre']
    intro = data.get('intro', row['intro'])
    astuce = data.get('astuce', row['astuce'])
    if isinstance(intro, str):
        intro = intro.strip() or None
    if isinstance(astuce, str):
        astuce = astuce.strip() or None
    etapes = _parse_etapes(data['etapes']) if 'etapes' in data else json.loads(row['etapes_json'] or '[]')
    ordre = data.get('ordre', row['ordre'])
    try:
        ordre = int(ordre)
    except (TypeError, ValueError):
        ordre = row['ordre']

    conn.execute("""
        UPDATE guide_sections SET titre=?, intro=?, astuce=?, etapes_json=?, ordre=?
        WHERE id=?
    """, (titre, intro, astuce, json.dumps(etapes), ordre, section_id))
    conn.commit()
    row = conn.execute("SELECT * FROM guide_sections WHERE id=?", (section_id,)).fetchone()
    conn.close()
    return jsonify(_guide_section_to_dict(row))

@app.route('/api/v1/admin/ressources/<int:ressource_id>/sections/<int:section_id>', methods=['DELETE'])
@admin_required
def api_admin_guide_sections_delete(ressource_id, section_id):
    conn = get_db_connection()
    row = conn.execute(
        "SELECT * FROM guide_sections WHERE id=? AND id_ressource=?",
        (section_id, ressource_id)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Section introuvable'}), 404
    conn.execute("DELETE FROM guide_sections WHERE id=?", (section_id,))
    conn.execute(
        "UPDATE ressource_images SET section_id=NULL WHERE id_ressource=? AND section_id=?",
        (ressource_id, str(section_id))
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/ressources/<int:ressource_id>/sections/reorder', methods=['POST'])
@admin_required
def api_admin_guide_sections_reorder(ressource_id):
    data = request.get_json(silent=True) or {}
    ordre_ids = data.get('ordre')
    if not isinstance(ordre_ids, list):
        return jsonify({'error': 'Liste d’ordre requise'}), 400
    conn = get_db_connection()
    for index, sid in enumerate(ordre_ids):
        conn.execute(
            "UPDATE guide_sections SET ordre=? WHERE id=? AND id_ressource=?",
            (index, sid, ressource_id)
        )
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/ressources/<int:ressource_id>/sections', methods=['GET'])
def api_guide_sections_public(ressource_id):
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT * FROM guide_sections WHERE id_ressource=? ORDER BY ordre ASC, id ASC",
        (ressource_id,)
    ).fetchall()
    conn.close()
    return jsonify([_guide_section_to_dict(s) for s in rows])


@app.route('/api/v1/admin/ressources', methods=['GET'])
@admin_required
def api_admin_ressources_list():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT r.*, c.nom_complet AS client_nom
        FROM client_ressources r
        LEFT JOIN clients c ON c.id = r.id_client
        ORDER BY r.created_at DESC
    """).fetchall()
    result = []
    for r in rows:
        d = _ressource_to_dict(r, sections=_guide_sections_brief(conn, r['id']))
        d['client_nom'] = r['client_nom']
        result.append(d)
    conn.close()
    return jsonify(result)

@app.route('/api/v1/admin/ressources', methods=['POST'])
@admin_required
def api_admin_ressources_create():
    titre = (request.form.get('titre') or '').strip()
    description = (request.form.get('description') or '').strip() or None
    categorie = (request.form.get('categorie') or 'guide').strip()
    type_source = (request.form.get('type_source') or 'lien').strip()
    id_client_raw = (request.form.get('id_client') or '').strip()
    id_client = int(id_client_raw) if id_client_raw else None
    url = (request.form.get('url') or '').strip() or None

    if not titre:
        return jsonify({'error': 'Titre requis'}), 400
    if type_source not in ('upload', 'lien'):
        return jsonify({'error': 'Type de source invalide'}), 400

    drive_file_id = None
    if type_source == 'upload':
        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({'error': 'Fichier requis'}), 400
        try:
            content = file.read()
            mimetype = file.content_type or 'application/octet-stream'
            uploaded = upload_bytes(_ressources_clients_root(), secure_filename(file.filename), content, mimetype)
            drive_file_id = uploaded['id']
            make_file_public(drive_file_id, secure_filename(file.filename))
        except Exception as e:
            return jsonify({'error': f"Échec du téléversement : {e}"}), 500
    else:
        if not url:
            return jsonify({'error': 'Lien requis'}), 400

    conn = get_db_connection()
    cur = conn.execute("""
        INSERT INTO client_ressources (id_client, titre, description, categorie, type_source, drive_file_id, url)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (id_client, titre, description, categorie, type_source, drive_file_id, url))
    ressource_id = cur.lastrowid
    conn.commit()
    row = conn.execute("SELECT * FROM client_ressources WHERE id=?", (ressource_id,)).fetchone()
    conn.close()
    return jsonify(_ressource_to_dict(row))

@app.route('/api/v1/admin/ressources/<int:ressource_id>', methods=['DELETE'])
@admin_required
def api_admin_ressources_delete(ressource_id):
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM client_ressources WHERE id=?", (ressource_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Ressource introuvable'}), 404
    if row['type_source'] == 'upload' and row['drive_file_id']:
        try:
            delete_drive_file(row['drive_file_id'])
        except Exception as e:
            print(f"[RESSOURCES] Suppression fichier Drive échouée: {e}")
    conn.execute("DELETE FROM client_ressources WHERE id=?", (ressource_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/ressources/<int:ressource_id>/bundle', methods=['PATCH'])
@admin_required
def api_admin_ressources_set_bundle(ressource_id):
    data = request.get_json(silent=True) or {}
    bundle_id = data.get('bundle_id')  # None = retirer du bundle
    conn = get_db_connection()
    if not conn.execute("SELECT id FROM client_ressources WHERE id=?", (ressource_id,)).fetchone():
        conn.close()
        return jsonify({'error': 'Ressource introuvable'}), 404
    conn.execute("UPDATE client_ressources SET bundle_id=? WHERE id=?", (bundle_id, ressource_id))
    conn.commit()
    row = conn.execute("SELECT * FROM client_ressources WHERE id=?", (ressource_id,)).fetchone()
    conn.close()
    return jsonify(_ressource_to_dict(row))

@app.route('/api/v1/admin/projet/<int:project_id>/ressources-disponibles', methods=['GET'])
@admin_required
def api_admin_projet_ressources_disponibles(project_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT id_client FROM projets WHERE id=?", (project_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    id_client = projet['id_client']
    rows = conn.execute("""
        SELECT r.id, r.titre, r.categorie, r.bundle_id, r.id_client,
               EXISTS(
                   SELECT 1 FROM ressource_assignations a
                   WHERE a.id_ressource = r.id AND a.id_client = ?
               ) AS deja_assignee
        FROM client_ressources r
        WHERE r.id_client IS NULL OR r.id_client = ?
        ORDER BY r.created_at DESC
    """, (id_client, id_client)).fetchall()
    bundles = conn.execute("SELECT id, nom, icone FROM ressource_bundles ORDER BY ordre ASC, id ASC").fetchall()
    conn.close()
    return jsonify({
        'ressources': [{
            'id': r['id'],
            'titre': r['titre'],
            'categorie': r['categorie'],
            'bundle_id': r['bundle_id'],
            'is_global': r['id_client'] is None,
            'already_assigned': bool(r['deja_assignee']) or r['id_client'] == id_client,
        } for r in rows],
        'bundles': [{'id': b['id'], 'nom': b['nom'], 'icone': b['icone']} for b in bundles],
    })

@app.route('/api/v1/admin/ressource-bundles', methods=['GET'])
@admin_required
def api_admin_bundles_list():
    conn = get_db_connection()
    bundles = conn.execute("SELECT * FROM ressource_bundles ORDER BY ordre ASC, id ASC").fetchall()
    conn.close()
    return jsonify([{'id': b['id'], 'nom': b['nom'], 'description': b['description'], 'icone': b['icone'], 'ordre': b['ordre']} for b in bundles])

@app.route('/api/v1/admin/ressource-bundles', methods=['POST'])
@admin_required
def api_admin_bundles_create():
    data = request.get_json(silent=True) or {}
    nom = (data.get('nom') or '').strip()
    if not nom:
        return jsonify({'error': 'Nom requis'}), 400
    description = (data.get('description') or '').strip() or None
    icone = (data.get('icone') or 'folder').strip()
    conn = get_db_connection()
    max_ordre = conn.execute("SELECT COALESCE(MAX(ordre), -1) FROM ressource_bundles").fetchone()[0]
    cur = conn.execute("INSERT INTO ressource_bundles (nom, description, icone, ordre) VALUES (?, ?, ?, ?)",
        (nom, description, icone, max_ordre + 1))
    bundle_id = cur.lastrowid
    conn.commit()
    row = conn.execute("SELECT * FROM ressource_bundles WHERE id=?", (bundle_id,)).fetchone()
    conn.close()
    return jsonify({'id': row['id'], 'nom': row['nom'], 'description': row['description'], 'icone': row['icone'], 'ordre': row['ordre']})

@app.route('/api/v1/admin/ressource-bundles/<int:bundle_id>', methods=['PUT'])
@admin_required
def api_admin_bundles_update(bundle_id):
    data = request.get_json(silent=True) or {}
    nom = (data.get('nom') or '').strip()
    if not nom:
        return jsonify({'error': 'Nom requis'}), 400
    description = (data.get('description') or '').strip() or None
    icone = (data.get('icone') or 'folder').strip()
    conn = get_db_connection()
    if not conn.execute("SELECT id FROM ressource_bundles WHERE id=?", (bundle_id,)).fetchone():
        conn.close()
        return jsonify({'error': 'Bundle introuvable'}), 404
    conn.execute("UPDATE ressource_bundles SET nom=?, description=?, icone=? WHERE id=?",
        (nom, description, icone, bundle_id))
    conn.commit()
    row = conn.execute("SELECT * FROM ressource_bundles WHERE id=?", (bundle_id,)).fetchone()
    conn.close()
    return jsonify({'id': row['id'], 'nom': row['nom'], 'description': row['description'], 'icone': row['icone'], 'ordre': row['ordre']})

@app.route('/api/v1/admin/ressource-bundles/<int:bundle_id>', methods=['DELETE'])
@admin_required
def api_admin_bundles_delete(bundle_id):
    conn = get_db_connection()
    if not conn.execute("SELECT id FROM ressource_bundles WHERE id=?", (bundle_id,)).fetchone():
        conn.close()
        return jsonify({'error': 'Bundle introuvable'}), 404
    conn.execute("UPDATE client_ressources SET bundle_id=NULL WHERE bundle_id=?", (bundle_id,))
    conn.execute("DELETE FROM ressource_bundles WHERE id=?", (bundle_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/client/factures', methods=['GET'])
@login_required
def api_client_factures_list():
    client_id = session['user_id']
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT id, numero, statut, total, date_emission, date_echeance "
        "FROM factures WHERE id_client = ? AND statut NOT IN ('ouverte', 'annulee') "
        "ORDER BY date_emission DESC",
        (client_id,)
    ).fetchall()
    conn.close()
    from datetime import date
    today = date.today().isoformat()
    factures, total_paye, total_attente = [], 0.0, 0.0
    for f in rows:
        montant = f['total'] or 0
        statut = f['statut']
        en_retard = statut != 'payee' and bool(f['date_echeance']) and f['date_echeance'] < today
        if statut == 'payee':
            total_paye += montant
        else:
            total_attente += montant
        factures.append({
            'id': f['id'],
            'numero': f['numero'],
            'statut': statut,
            'en_retard': bool(en_retard),
            'total': montant,
            'date_emission': f['date_emission'],
            'date_echeance': f['date_echeance'],
        })
    return jsonify({
        'factures': factures,
        'total_paye': round(total_paye, 2),
        'total_attente': round(total_attente, 2),
    })


@app.route('/api/v1/client/factures/<int:facture_id>/pdf', methods=['GET'])
@login_required
def api_client_facture_pdf(facture_id):
    client_id = session['user_id']
    conn = get_db_connection()
    f = conn.execute(
        "SELECT pdf_path, numero FROM factures WHERE id = ? AND id_client = ?",
        (facture_id, client_id)
    ).fetchone()
    conn.close()
    if not f or not f['pdf_path'] or not os.path.exists(f['pdf_path']):
        return jsonify({'error': 'PDF non disponible'}), 404
    inline = request.args.get('view') == '1'
    return send_file(f['pdf_path'], mimetype='application/pdf',
                     as_attachment=not inline, download_name=f"{f['numero']}.pdf")


@app.route('/api/v1/client/factures/<int:facture_id>/correction', methods=['POST'])
@login_required
def api_client_facture_correction(facture_id):
    client_id = session['user_id']
    message = (request.get_json(silent=True) or {}).get('message', '').strip()
    if not message:
        return jsonify({'error': 'Message requis'}), 400
    conn = get_db_connection()
    f = conn.execute(
        "SELECT numero FROM factures WHERE id = ? AND id_client = ?",
        (facture_id, client_id)
    ).fetchone()
    if not f:
        conn.close()
        return jsonify({'error': 'Facture introuvable'}), 404
    client = conn.execute("SELECT nom_complet FROM clients WHERE id = ?", (client_id,)).fetchone()
    nom_client = client['nom_complet'] if client else 'Client'
    push_admin_notif(
        conn,
        titre=f"Correction demandée — Facture {f['numero']}",
        message=f"{nom_client} : {message}",
        type='facture',
        lien=f"/admin/factures/{facture_id}",
    )
    conn.commit()
    conn.close()
    try:
        send_email(
            'felix.dumont@cocktailmedia.ca',
            f"Correction demandée — Facture {f['numero']}",
            f"{nom_client} : {message}\n\nVoir : {PORTAIL_URL}/admin/factures/{facture_id}"
        )
    except Exception as e:
        print(f"[MAIL] api_client_facture_correction: {e}")
    return jsonify({'success': True})


@app.route('/api/v1/notifications', methods=['GET'])
@login_required
def api_client_notifications_list():
    user_id = session['user_id']
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT * FROM notifications WHERE id_client = ? ORDER BY created_at DESC LIMIT 50",
        (user_id,)
    ).fetchall()
    unread = conn.execute(
        "SELECT COUNT(*) FROM notifications WHERE id_client = ? AND is_read = 0", (user_id,)
    ).fetchone()[0]
    conn.close()
    return jsonify({
        'notifications': [{
            'id': r['id'], 'message': r['message'], 'id_projet': r['id_projet'],
            'type': r['type'] if 'type' in r.keys() else 'info',
            'is_read': bool(r['is_read']), 'created_at': r['created_at'],
        } for r in rows],
        'unread': unread,
    })


@app.route('/api/v1/notifications/<int:notif_id>/read', methods=['POST'])
@login_required
def api_client_notification_read(notif_id):
    user_id = session['user_id']
    conn = get_db_connection()
    conn.execute("UPDATE notifications SET is_read=1 WHERE id=? AND id_client=?", (notif_id, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/notifications/read-all', methods=['POST'])
@login_required
def api_client_notifications_read_all():
    user_id = session['user_id']
    conn = get_db_connection()
    conn.execute("UPDATE notifications SET is_read=1 WHERE id_client=? AND is_read=0", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/client/ressources', methods=['GET'])
def api_client_ressources_list():
    if 'user_id' not in session:
        return jsonify({'error': 'Non authentifié'}), 401
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT * FROM client_ressources
        WHERE id_client IS NULL OR id_client = ?
           OR id IN (SELECT id_ressource FROM ressource_assignations WHERE id_client = ?)
        ORDER BY created_at DESC
    """, (session['user_id'], session['user_id'])).fetchall()
    conn.close()
    return jsonify([_ressource_to_dict(r) for r in rows])


# ── Captures d'écran liées à une ressource (Drive, dossier au nom de la ressource) ──

def _ressource_images_folder(ressource_titre):
    return create_folder(ressource_titre, parent_id=_ressources_clients_root())

def _ressource_image_to_dict(img):
    return {
        'id': img['id'],
        'id_ressource': img['id_ressource'],
        'nom_fichier': img['nom_fichier'],
        'legende': img['legende'],
        'section_id': img['section_id'] if 'section_id' in img.keys() else None,
        'ordre': img['ordre'],
        'url': f"https://drive.google.com/thumbnail?id={img['drive_file_id']}&sz=w1600",
        'created_at': img['created_at'],
    }

@app.route('/api/v1/admin/ressources/<int:ressource_id>/images', methods=['GET'])
@admin_required
def api_admin_ressource_images_list(ressource_id):
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT * FROM ressource_images WHERE id_ressource=? ORDER BY ordre ASC, id ASC",
        (ressource_id,)
    ).fetchall()
    conn.close()
    return jsonify([_ressource_image_to_dict(r) for r in rows])

@app.route('/api/v1/admin/ressources/<int:ressource_id>/images', methods=['POST'])
@admin_required
def api_admin_ressource_images_create(ressource_id):
    conn = get_db_connection()
    ressource = conn.execute("SELECT * FROM client_ressources WHERE id=?", (ressource_id,)).fetchone()
    if not ressource:
        conn.close()
        return jsonify({'error': 'Ressource introuvable'}), 404

    file = request.files.get('file')
    if not file or not file.filename:
        conn.close()
        return jsonify({'error': 'Image requise'}), 400
    legende = (request.form.get('legende') or '').strip() or None
    section_id = (request.form.get('section_id') or '').strip() or None
    try:
        ordre = int(request.form.get('ordre') or 0)
    except ValueError:
        ordre = 0

    try:
        content = file.read()
        mimetype = file.content_type or 'application/octet-stream'
        nom_fichier = secure_filename(file.filename)
        folder_id = _ressource_images_folder(ressource['titre'])
        uploaded = upload_bytes(folder_id, nom_fichier, content, mimetype)
        drive_file_id = uploaded['id']
        make_file_public(drive_file_id, nom_fichier)
    except Exception as e:
        conn.close()
        return jsonify({'error': f"Échec du téléversement : {e}"}), 500

    cur = conn.execute("""
        INSERT INTO ressource_images (id_ressource, drive_file_id, nom_fichier, legende, section_id, ordre)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (ressource_id, drive_file_id, nom_fichier, legende, section_id, ordre))
    image_id = cur.lastrowid
    conn.commit()
    row = conn.execute("SELECT * FROM ressource_images WHERE id=?", (image_id,)).fetchone()
    conn.close()
    return jsonify(_ressource_image_to_dict(row))

@app.route('/api/v1/admin/ressources/<int:ressource_id>/images/<int:image_id>', methods=['PATCH'])
@admin_required
def api_admin_ressource_images_update(ressource_id, image_id):
    conn = get_db_connection()
    row = conn.execute(
        "SELECT * FROM ressource_images WHERE id=? AND id_ressource=?",
        (image_id, ressource_id)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Image introuvable'}), 404
    data = request.get_json(silent=True) or {}
    legende = data.get('legende', row['legende'])
    section_id = data.get('section_id', row['section_id'] if 'section_id' in row.keys() else None)
    if isinstance(legende, str):
        legende = legende.strip() or None
    if isinstance(section_id, str):
        section_id = section_id.strip() or None
    conn.execute(
        "UPDATE ressource_images SET legende=?, section_id=? WHERE id=?",
        (legende, section_id, image_id)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM ressource_images WHERE id=?", (image_id,)).fetchone()
    conn.close()
    return jsonify(_ressource_image_to_dict(row))

@app.route('/api/v1/admin/ressources/<int:ressource_id>/images/<int:image_id>', methods=['DELETE'])
@admin_required
def api_admin_ressource_images_delete(ressource_id, image_id):
    conn = get_db_connection()
    row = conn.execute(
        "SELECT * FROM ressource_images WHERE id=? AND id_ressource=?",
        (image_id, ressource_id)
    ).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Image introuvable'}), 404
    if row['drive_file_id']:
        try:
            delete_drive_file(row['drive_file_id'])
        except Exception as e:
            print(f"[RESSOURCES] Suppression image Drive échouée: {e}")
    conn.execute("DELETE FROM ressource_images WHERE id=?", (image_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/ressources/<int:ressource_id>/images', methods=['GET'])
def api_ressource_images_public(ressource_id):
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT * FROM ressource_images WHERE id_ressource=? ORDER BY ordre ASC, id ASC",
        (ressource_id,)
    ).fetchall()
    conn.close()
    return jsonify([_ressource_image_to_dict(r) for r in rows])


# ── Médiathèque — Gabarits : pages HTML ─────────────────────────────────────

@app.route('/admin/gabarits')
@admin_required
def mediatech_admin():
    conn = get_db_connection()
    rows = conn.execute("SELECT * FROM mediatech_gabarits ORDER BY id ASC").fetchall()
    conn.close()
    gabarits = []
    for r in rows:
        files = []
        if r['drive_folder_id']:
            try:
                files = list_files_in_folder(r['drive_folder_id'])
            except Exception:
                files = []
        gabarits.append(dict(r) | {'files': files})
    return render_template('mediatech_admin.html', gabarits=gabarits)

@app.route('/admin/gabarits/creer', methods=['POST'])
@admin_required
def mediatech_admin_creer():
    nom = request.form.get('nom', '').strip()
    description = request.form.get('description', '').strip()
    if not nom:
        flash("Le nom est requis.", "error")
        return redirect(url_for('mediatech_admin'))
    try:
        drive_folder_id = create_folder(nom, parent_id=_gabarits_root())
    except Exception:
        drive_folder_id = None
    conn = get_db_connection()
    conn.execute("INSERT INTO mediatech_gabarits (nom, description, drive_folder_id) VALUES (?, ?, ?)",
                 (nom, description, drive_folder_id))
    conn.commit()
    gabarit_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    if 'preview' in request.files and drive_folder_id:
        f = request.files['preview']
        if f and f.filename:
            try:
                content = f.read()
                mimetype = f.content_type or 'image/jpeg'
                uploaded = upload_bytes(drive_folder_id, f'_preview_{secure_filename(f.filename)}', content, mimetype)
                conn.execute("UPDATE mediatech_gabarits SET preview_drive_id = ? WHERE id = ?",
                             (uploaded['id'], gabarit_id))
                conn.commit()
            except Exception as e:
                flash(f"Dossier créé mais l'aperçu n'a pas pu être uploadé : {e}", "warning")
    conn.close()
    flash(f"Dossier « {nom} » créé.", "success")
    return redirect(url_for('mediatech_admin'))

@app.route('/admin/gabarits/<int:gabarit_id>/supprimer', methods=['POST'])
@admin_required
def mediatech_admin_supprimer(gabarit_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM mediatech_gabarits WHERE id = ?", (gabarit_id,))
    conn.commit()
    conn.close()
    flash("Dossier supprimé.", "success")
    return redirect(url_for('mediatech_admin'))

@app.route('/admin/gabarits/<int:gabarit_id>/fichier', methods=['POST'])
@admin_required
def mediatech_admin_upload_fichier(gabarit_id):
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM mediatech_gabarits WHERE id = ?", (gabarit_id,)).fetchone()
    conn.close()
    if not row:
        flash("Dossier introuvable.", "error")
        return redirect(url_for('mediatech_admin'))
    f = request.files.get('fichier')
    if not f or not f.filename:
        flash("Aucun fichier sélectionné.", "error")
        return redirect(url_for('mediatech_admin'))
    try:
        folder_id = row['drive_folder_id']
        if not folder_id:
            folder_id = create_folder(row['nom'], parent_id=_gabarits_root())
            conn = get_db_connection()
            conn.execute("UPDATE mediatech_gabarits SET drive_folder_id = ? WHERE id = ?",
                         (folder_id, gabarit_id))
            conn.commit()
            conn.close()
        content = f.read()
        mimetype = f.content_type or 'application/octet-stream'
        upload_bytes(folder_id, secure_filename(f.filename), content, mimetype)
        flash(f"Fichier « {f.filename} » ajouté.", "success")
    except Exception as e:
        flash(f"Erreur upload : {e}", "error")
    return redirect(url_for('mediatech_admin'))

@app.route('/admin/gabarits/<int:gabarit_id>/fichier/<file_id>/supprimer', methods=['POST'])
@admin_required
def mediatech_admin_supprimer_fichier(gabarit_id, file_id):
    try:
        delete_drive_file(file_id)
        conn = get_db_connection()
        row = conn.execute("SELECT preview_drive_id FROM mediatech_gabarits WHERE id = ?", (gabarit_id,)).fetchone()
        if row and row['preview_drive_id'] == file_id:
            conn.execute("UPDATE mediatech_gabarits SET preview_drive_id = NULL WHERE id = ?", (gabarit_id,))
            conn.commit()
        conn.close()
        flash("Fichier supprimé.", "success")
    except Exception as e:
        flash(f"Erreur : {e}", "error")
    return redirect(url_for('mediatech_admin'))

@app.route('/admin/gabarits/<int:gabarit_id>/preview', methods=['POST'])
@admin_required
def mediatech_admin_set_preview(gabarit_id):
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM mediatech_gabarits WHERE id = ?", (gabarit_id,)).fetchone()
    conn.close()
    if not row:
        flash("Dossier introuvable.", "error")
        return redirect(url_for('mediatech_admin'))
    f = request.files.get('preview')
    if not f or not f.filename:
        flash("Aucun fichier sélectionné.", "error")
        return redirect(url_for('mediatech_admin'))
    try:
        folder_id = row['drive_folder_id']
        if not folder_id:
            folder_id = create_folder(row['nom'], parent_id=_gabarits_root())
            conn = get_db_connection()
            conn.execute("UPDATE mediatech_gabarits SET drive_folder_id = ? WHERE id = ?",
                         (folder_id, gabarit_id))
            conn.commit()
            conn.close()
        if row['preview_drive_id']:
            try:
                delete_drive_file(row['preview_drive_id'])
            except Exception:
                pass
        content = f.read()
        mimetype = f.content_type or 'image/jpeg'
        uploaded = upload_bytes(folder_id, f'_preview_{secure_filename(f.filename)}', content, mimetype)
        conn = get_db_connection()
        conn.execute("UPDATE mediatech_gabarits SET preview_drive_id = ? WHERE id = ?",
                     (uploaded['id'], gabarit_id))
        conn.commit()
        conn.close()
        flash("Aperçu mis à jour.", "success")
    except Exception as e:
        flash(f"Erreur : {e}", "error")
    return redirect(url_for('mediatech_admin'))

@app.route('/gabarits')
@login_required
def mediatech_galerie():
    conn = get_db_connection()
    rows = conn.execute("SELECT * FROM mediatech_gabarits ORDER BY id ASC").fetchall()
    conn.close()
    return render_template('mediatech_galerie.html', gabarits=[dict(r) for r in rows])

@app.route('/gabarits/<int:gabarit_id>')
@login_required
def mediatech_dossier(gabarit_id):
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM mediatech_gabarits WHERE id = ?", (gabarit_id,)).fetchone()
    conn.close()
    if not row:
        flash("Dossier introuvable.", "error")
        return redirect(url_for('mediatech_galerie'))
    fichiers = []
    if row['drive_folder_id']:
        try:
            raw = list_files_in_folder(row['drive_folder_id'])
            for f in raw:
                fichiers.append({
                    'id': f['id'], 'name': f['name'],
                    'size': int(f.get('size') or 0),
                    'is_preview': f['id'] == row['preview_drive_id'],
                })
        except Exception:
            pass
    return render_template('mediatech_dossier.html', gabarit=dict(row), fichiers=fichiers)

@app.route('/gabarits/telecharger/<file_id>/<path:filename>')
@login_required
def mediatech_telecharger(file_id, filename):
    try:
        content = get_file_bytes(file_id)
        resp = make_response(content)
        resp.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        resp.headers['Content-Type'] = 'application/octet-stream'
        return resp
    except Exception as e:
        flash(f"Fichier indisponible : {e}", "error")
        return redirect(request.referrer or url_for('mediatech_galerie'))


# ── Médiathèque — Gabarits (dossiers avec preview) ──────────────────────────

def _gabarits_root():
    return create_folder('Gabarits', parent_id=_mediatheque_root())

@app.route('/api/v1/tools/gabarits', methods=['GET'])
def api_gabarits_list():
    if 'user_id' not in session and 'pigiste_id' not in session:
        return jsonify({'error': 'Non authentifié'}), 401
    conn = get_db_connection()
    rows = conn.execute("SELECT * FROM mediatech_gabarits ORDER BY id ASC").fetchall()
    conn.close()
    result = []
    for r in rows:
        result.append({
            'id': r['id'],
            'nom': r['nom'],
            'description': r['description'] or '',
            'preview_url': f'/api/v1/tools/file/{r["preview_drive_id"]}' if r['preview_drive_id'] else None,
            'created_at': r['created_at'],
        })
    return jsonify({'gabarits': result})

@app.route('/api/v1/tools/gabarits', methods=['POST'])
def api_gabarits_create():
    if not session.get('user_id'):
        return jsonify({'error': 'Non authentifié'}), 401
    nom = request.form.get('nom', '').strip()
    description = request.form.get('description', '').strip()
    if not nom:
        return jsonify({'error': 'Le nom est requis'}), 400
    try:
        drive_folder_id = create_folder(nom, parent_id=_gabarits_root())
    except Exception:
        drive_folder_id = None
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO mediatech_gabarits (nom, description, drive_folder_id) VALUES (?, ?, ?)",
        (nom, description, drive_folder_id)
    )
    conn.commit()
    gabarit_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    preview_drive_id = None
    if 'preview' in request.files:
        f = request.files['preview']
        if f and f.filename and drive_folder_id:
            try:
                content = f.read()
                mimetype = f.content_type or 'image/jpeg'
                uploaded = upload_bytes(drive_folder_id, f'_preview_{secure_filename(f.filename)}', content, mimetype)
                preview_drive_id = uploaded['id']
                conn.execute("UPDATE mediatech_gabarits SET preview_drive_id = ? WHERE id = ?",
                             (preview_drive_id, gabarit_id))
                conn.commit()
            except Exception:
                pass
    conn.close()
    return jsonify({'ok': True, 'id': gabarit_id, 'nom': nom,
                    'preview_url': f'/api/v1/tools/file/{preview_drive_id}' if preview_drive_id else None})

@app.route('/api/v1/tools/gabarits/<int:gabarit_id>', methods=['DELETE'])
def api_gabarits_delete(gabarit_id):
    if not session.get('user_id'):
        return jsonify({'error': 'Non authentifié'}), 401
    conn = get_db_connection()
    conn.execute("DELETE FROM mediatech_gabarits WHERE id = ?", (gabarit_id,))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/v1/tools/gabarits/<int:gabarit_id>/files', methods=['GET'])
def api_gabarits_files(gabarit_id):
    if 'user_id' not in session and 'pigiste_id' not in session:
        return jsonify({'error': 'Non authentifié'}), 401
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM mediatech_gabarits WHERE id = ?", (gabarit_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Gabarit introuvable'}), 404
    if not row['drive_folder_id']:
        return jsonify({'files': []})
    try:
        files = list_files_in_folder(row['drive_folder_id'])
        result = []
        for f in files:
            is_preview = (f['id'] == row['preview_drive_id'])
            result.append({
                'id': f['id'],
                'filename': f['name'],
                'size': int(f.get('size') or 0),
                'mimeType': f.get('mimeType', ''),
                'url': f'/api/v1/tools/file/{f["id"]}',
                'is_preview': is_preview,
            })
        return jsonify({'files': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/tools/gabarits/<int:gabarit_id>/files', methods=['POST'])
def api_gabarits_upload_file(gabarit_id):
    if not session.get('user_id'):
        return jsonify({'error': 'Non authentifié'}), 401
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM mediatech_gabarits WHERE id = ?", (gabarit_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Gabarit introuvable'}), 404
    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'Fichier requis'}), 400
    try:
        folder_id = row['drive_folder_id']
        if not folder_id:
            folder_id = create_folder(row['nom'], parent_id=_gabarits_root())
            conn = get_db_connection()
            conn.execute("UPDATE mediatech_gabarits SET drive_folder_id = ? WHERE id = ?", (folder_id, gabarit_id))
            conn.commit()
            conn.close()
        content = file.read()
        mimetype = file.content_type or 'application/octet-stream'
        uploaded = upload_bytes(folder_id, secure_filename(file.filename), content, mimetype)
        return jsonify({'ok': True, 'id': uploaded['id'], 'filename': file.filename,
                        'url': f'/api/v1/tools/file/{uploaded["id"]}'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/tools/gabarits/<int:gabarit_id>/files/<file_id>', methods=['DELETE'])
def api_gabarits_delete_file(gabarit_id, file_id):
    if not session.get('user_id'):
        return jsonify({'error': 'Non authentifié'}), 401
    try:
        delete_drive_file(file_id)
        conn = get_db_connection()
        row = conn.execute("SELECT preview_drive_id FROM mediatech_gabarits WHERE id = ?", (gabarit_id,)).fetchone()
        if row and row['preview_drive_id'] == file_id:
            conn.execute("UPDATE mediatech_gabarits SET preview_drive_id = NULL WHERE id = ?", (gabarit_id,))
            conn.commit()
        conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/v1/tools/gabarits/<int:gabarit_id>/preview', methods=['POST'])
def api_gabarits_set_preview(gabarit_id):
    if not session.get('user_id'):
        return jsonify({'error': 'Non authentifié'}), 401
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM mediatech_gabarits WHERE id = ?", (gabarit_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Gabarit introuvable'}), 404
    # Cas 1 : désigner un fichier Drive existant comme preview
    data = request.get_json(force=True, silent=True) or {}
    file_id = data.get('file_id') or request.form.get('file_id')
    if file_id:
        conn = get_db_connection()
        conn.execute("UPDATE mediatech_gabarits SET preview_drive_id = ? WHERE id = ?", (file_id, gabarit_id))
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'preview_url': f'/api/v1/tools/file/{file_id}'})
    # Cas 2 : uploader une nouvelle image preview
    file = request.files.get('preview')
    if not file or not file.filename:
        return jsonify({'error': 'Fichier ou file_id requis'}), 400
    try:
        folder_id = row['drive_folder_id']
        if not folder_id:
            folder_id = create_folder(row['nom'], parent_id=_gabarits_root())
            conn = get_db_connection()
            conn.execute("UPDATE mediatech_gabarits SET drive_folder_id = ? WHERE id = ?", (folder_id, gabarit_id))
            conn.commit()
            conn.close()
        # Supprimer l'ancienne preview si elle existe
        if row['preview_drive_id']:
            try:
                delete_drive_file(row['preview_drive_id'])
            except Exception:
                pass
        content = file.read()
        mimetype = file.content_type or 'image/jpeg'
        uploaded = upload_bytes(folder_id, f'_preview_{secure_filename(file.filename)}', content, mimetype)
        conn = get_db_connection()
        conn.execute("UPDATE mediatech_gabarits SET preview_drive_id = ? WHERE id = ?",
                     (uploaded['id'], gabarit_id))
        conn.commit()
        conn.close()
        return jsonify({'ok': True, 'preview_url': f'/api/v1/tools/file/{uploaded["id"]}'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/v1/tools/config')
def api_tools_config():
    import json as _json
    if session.get('user_id'):
        if session.get('is_admin'):
            return jsonify({'role': 'admin', 'social': None, 'pdf': None})
        conn = get_db_connection()
        row = conn.execute("SELECT outils_config FROM clients WHERE id = ?",
                           (session['user_id'],)).fetchone()
        conn.close()
        cfg = {}
        if row and row['outils_config']:
            try:
                cfg = _json.loads(row['outils_config'])
            except Exception:
                cfg = {}
        return jsonify({'role': 'client',
                        'social': cfg.get('social', None),
                        'pdf':    cfg.get('pdf',    None)})
    if session.get('pigiste_id'):
        conn = get_db_connection()
        pig = conn.execute("SELECT tools_config FROM pigistes WHERE id = ?",
                           (session['pigiste_id'],)).fetchone()
        conn.close()
        cfg = {}
        if pig and pig['tools_config']:
            try:
                cfg = _json.loads(pig['tools_config'])
            except Exception:
                cfg = {}
        return jsonify({'role': 'pigiste',
                        'social': cfg.get('social', None),
                        'pdf':    cfg.get('pdf',    None)})
    return jsonify({'error': 'Non authentifié'}), 401

@app.route('/api/v1/admin/pigistes/<int:pigiste_id>/tools', methods=['GET'])
def api_admin_get_pigiste_tools(pigiste_id):
    import json as _json
    if not session.get('user_id'):
        return jsonify({'error': 'Non autorisé'}), 403
    conn = get_db_connection()
    row = conn.execute("SELECT tools_config FROM pigistes WHERE id = ?", (pigiste_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Pigiste introuvable'}), 404
    cfg = {}
    if row['tools_config']:
        try:
            cfg = _json.loads(row['tools_config'])
        except Exception:
            cfg = {}
    return jsonify({'social': cfg.get('social', None), 'pdf': cfg.get('pdf', None)})

@app.route('/api/v1/admin/pigistes/<int:pigiste_id>/tools', methods=['PUT'])
def api_admin_pigiste_tools(pigiste_id):
    import json as _json
    if not session.get('user_id'):
        return jsonify({'error': 'Non autorisé'}), 403
    data = request.get_json(force=True) or {}
    conn = get_db_connection()
    conn.execute("UPDATE pigistes SET tools_config = ? WHERE id = ?",
                 (_json.dumps(data), pigiste_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


def _current_doc_actor():
    """Return (type, id, folder_id) for the current session user."""
    if not session.get('user_id') and not session.get('pigiste_id'):
        return None, None, None
    root = os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
    gen_root = create_folder('Générateur de documents', parent_id=root)
    if session.get('user_id'):
        return 'admin', session['user_id'], create_folder('admin', parent_id=gen_root)
    pid = session['pigiste_id']
    pig_root = create_folder('pigistes', parent_id=gen_root)
    return 'pigiste', pid, create_folder(str(pid), parent_id=pig_root)

@app.route('/api/v1/documents', methods=['GET'])
def api_list_documents():
    actor_type, actor_id, folder_id = _current_doc_actor()
    if not actor_type:
        return jsonify({'error': 'Non authentifié'}), 401
    try:
        results = []
        if actor_type == 'admin':
            # Admin sees their own docs + all pigiste docs
            files = list_files_in_folder(folder_id)
            results += [f for f in files if f['name'].endswith('.json')]
            root = os.getenv('GOOGLE_DRIVE_ROOT_FOLDER_ID')
            gen_root = create_folder('Générateur de documents', parent_id=root)
            pig_root = create_folder('pigistes', parent_id=gen_root)
            for sub in list_subfolders(pig_root):
                pfiles = list_files_in_folder(sub['id'])
                results += [dict(f, _subfolder=sub['name']) for f in pfiles if f['name'].endswith('.json')]
        else:
            files = list_files_in_folder(folder_id)
            results = [f for f in files if f['name'].endswith('.json')]
        return jsonify({'documents': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/documents', methods=['POST'])
def api_save_document():
    actor_type, actor_id, folder_id = _current_doc_actor()
    if not actor_type:
        return jsonify({'error': 'Non authentifié'}), 401
    data = request.get_json(force=True)
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Nom requis'}), 400
    try:
        file = upload_json_content(folder_id, f"{name}.json", data)
        return jsonify({'id': file['id'], 'name': file['name'], 'createdTime': file.get('createdTime', '')})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/documents/<file_id>', methods=['GET'])
def api_get_document(file_id):
    actor_type, _, _ = _current_doc_actor()
    if not actor_type:
        return jsonify({'error': 'Non authentifié'}), 401
    try:
        data = get_json_content(file_id)
        return jsonify(data)
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/api/v1/documents/<file_id>', methods=['PUT'])
def api_update_document(file_id):
    actor_type, _, _ = _current_doc_actor()
    if not actor_type:
        return jsonify({'error': 'Non authentifié'}), 401
    data = request.get_json(force=True)
    try:
        file = update_json_content(file_id, data)
        return jsonify({'id': file['id'], 'name': file['name'], 'modifiedTime': file.get('modifiedTime', '')})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/v1/documents/<file_id>', methods=['DELETE'])
def api_delete_document(file_id):
    actor_type, _, _ = _current_doc_actor()
    if not actor_type:
        return jsonify({'error': 'Non authentifié'}), 401
    try:
        delete_drive_file(file_id)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 400


# ───────────────────────────────────────────────────────────
# Todos personnels (Marie-Christine)
# ───────────────────────────────────────────────────────────
def _current_admin_email(conn):
    """Email de l'admin connecté (via session), ou None."""
    uid = session.get('user_id')
    if not uid:
        return None
    row = conn.execute("SELECT email FROM clients WHERE id=?", (uid,)).fetchone()
    return row['email'] if row else None

@app.route('/api/v1/admin/notifications', methods=['GET'])
@admin_required
def api_admin_notifications_list():
    conn = get_db_connection()
    email = _current_admin_email(conn)
    rows = conn.execute("""
        SELECT * FROM admin_notifications
        WHERE destinataire IS NULL OR destinataire = '' OR destinataire = ?
        ORDER BY is_read ASC, created_at DESC
        LIMIT 50
    """, (email,)).fetchall()
    unread = conn.execute("""
        SELECT COUNT(*) FROM admin_notifications
        WHERE is_read = 0 AND (destinataire IS NULL OR destinataire = '' OR destinataire = ?)
    """, (email,)).fetchone()[0]
    conn.close()
    return jsonify({'items': [dict(r) for r in rows], 'unread': unread})

@app.route('/api/v1/admin/notifications/<int:notif_id>/read', methods=['POST'])
@admin_required
def api_admin_notifications_read(notif_id):
    conn = get_db_connection()
    conn.execute("UPDATE admin_notifications SET is_read=1 WHERE id=?", (notif_id,))
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/notifications/read-all', methods=['POST'])
@admin_required
def api_admin_notifications_read_all():
    conn = get_db_connection()
    email = _current_admin_email(conn)
    conn.execute("""
        UPDATE admin_notifications SET is_read=1
        WHERE is_read=0 AND (destinataire IS NULL OR destinataire = '' OR destinataire = ?)
    """, (email,))
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/push/vapid-public-key', methods=['GET'])
@admin_required
def api_push_vapid_key():
    return jsonify({'key': VAPID_PUBLIC_KEY})

@app.route('/api/v1/admin/push/subscribe', methods=['POST'])
@admin_required
def api_push_subscribe():
    data = request.get_json() or {}
    endpoint = data.get('endpoint')
    keys = data.get('keys') or {}
    p256dh, auth = keys.get('p256dh'), keys.get('auth')
    if not (endpoint and p256dh and auth):
        return jsonify({'error': 'Abonnement invalide'}), 400
    conn = get_db_connection()
    email = _current_admin_email(conn)
    conn.execute(
        "INSERT INTO push_subscriptions (email, endpoint, p256dh, auth) VALUES (?, ?, ?, ?) "
        "ON CONFLICT(endpoint) DO UPDATE SET email=excluded.email, p256dh=excluded.p256dh, auth=excluded.auth",
        (email, endpoint, p256dh, auth)
    )
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/push/unsubscribe', methods=['POST'])
@admin_required
def api_push_unsubscribe():
    data = request.get_json() or {}
    endpoint = data.get('endpoint')
    if endpoint:
        conn = get_db_connection()
        conn.execute("DELETE FROM push_subscriptions WHERE endpoint=?", (endpoint,))
        conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/push/test', methods=['POST'])
@admin_required
def api_push_test():
    conn = get_db_connection()
    push_admin_notif(conn, "🔔 Test de notification", "Si tu vois ça, les notifs push fonctionnent !", type='info', lien='/admin')
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/team', methods=['GET'])
@admin_required
def api_admin_team():
    """Liste des comptes admin avec un rôle (gestion/production) — pour l'assignation
    de tâches. Les comptes admin sans rôle défini (ex. compte générique historique)
    n'apparaissent pas ici : ce ne sont pas des personnes assignables."""
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT id, nom_complet, email, role FROM clients WHERE is_admin = 1 AND role IS NOT NULL ORDER BY nom_complet"
    ).fetchall()
    conn.close()
    return jsonify([{
        'id': r['id'], 'nom_complet': r['nom_complet'], 'email': r['email'], 'role': r['role'],
    } for r in rows])

def _todos_query(user_id, view='mine'):
    """Requête partagée des tâches — utilisée par la session CRM (/admin/todos) et par
    le jeton d'appareil de la PWA Tâches (/taches/todos), pour ne pas dupliquer le SQL."""
    conn = get_db_connection()
    where_assign = ""
    params: list = []
    if view != 'all':
        # "Mine" = strictement ce qui m'est assigné (les tâches partagées/non-assignées
        # ne s'y affichent plus — elles restent visibles dans "Toutes"). Une section-titre
        # reste incluse si elle contient au moins une tâche qui m'est assignée, même si le
        # titre lui-même ne l'est pas, pour ne pas orpheliner ses enfants dans l'UI groupée.
        where_assign = """WHERE EXISTS (SELECT 1 FROM todo_assignees ta WHERE ta.todo_id = t.id AND ta.admin_id = ?)
           OR (t.is_titre = 1 AND EXISTS (
                 SELECT 1 FROM todos_perso c
                 JOIN todo_assignees ca ON ca.todo_id = c.id AND ca.admin_id = ?
                 WHERE c.parent_titre_id = t.id
               ))"""
        params.extend([user_id, user_id])
    todos = conn.execute(f"""
        SELECT t.*,
               COALESCE(pc.nom_complet, dc.nom_complet) AS client_nom,
               COALESCE(t.client_id, p.id_client) AS client_id_effectif
        FROM todos_perso t
        LEFT JOIN projets p  ON p.id = t.projet_id
        LEFT JOIN clients pc ON pc.id = p.id_client
        LEFT JOIN clients dc ON dc.id = t.client_id
        {where_assign}
        ORDER BY t.est_coche ASC,
                 CASE t.priorite WHEN 'haute' THEN 0 WHEN 'normale' THEN 1 ELSE 2 END ASC,
                 t.date_echeance ASC,
                 t.created_at DESC
    """, params).fetchall()
    todo_ids = [t['id'] for t in todos]
    assignees_by_todo: dict = {}
    if todo_ids:
        placeholders = ','.join('?' * len(todo_ids))
        rows = conn.execute(
            f"SELECT ta.todo_id, c.id, c.nom_complet FROM todo_assignees ta "
            f"JOIN clients c ON c.id = ta.admin_id WHERE ta.todo_id IN ({placeholders})",
            todo_ids
        ).fetchall()
        for r in rows:
            assignees_by_todo.setdefault(r['todo_id'], []).append({'id': r['id'], 'nom_complet': r['nom_complet']})
    conn.close()
    result = []
    for t in todos:
        d = dict(t)
        d['assignees'] = assignees_by_todo.get(t['id'], [])
        result.append(d)
    return result


def _create_todo(default_assignee_id, data):
    """Création de tâche partagée — utilisée par la session CRM et par le jeton d'appareil
    de la PWA Tâches. `default_assignee_id` : assigné par défaut si assigne_admin_ids absent."""
    texte = (data.get('texte') or '').strip()
    if not texte:
        return {'error': 'Texte requis'}, 400
    priorite = data.get('priorite', 'normale')
    date_echeance = data.get('date_echeance') or None
    is_titre = 1 if data.get('is_titre') else 0
    parent_titre_id = data.get('parent_titre_id') or None
    # Assignée au créateur par défaut (comme une todo-list perso classique) — envoyer
    # explicitement assigne_admin_ids=[] pour créer une tâche partagée/visible à tous.
    # Une tâche peut être assignée à plusieurs personnes en même temps.
    if 'assigne_admin_ids' in data:
        assigne_admin_ids = [i for i in (data.get('assigne_admin_ids') or []) if i]
    else:
        assigne_admin_ids = [default_assignee_id] if default_assignee_id else []
    agenda_date  = (data.get('agenda_date')  or '').strip() or None
    agenda_heure = (data.get('agenda_heure') or '').strip() or None
    agenda_duree = int(data.get('agenda_duree') or 60)
    # Quick-add langage naturel : "appeler comptable demain 14h" → texte + date + heure
    # Actif par défaut pour les tâches (pas les titres), sauf si une date explicite est fournie.
    if not is_titre and data.get('parse_nl', True) and not date_echeance and not agenda_date:
        try:
            from date_parser_fr import parse_todo
            texte_net, d_parsed, h_parsed = parse_todo(texte)
            if d_parsed:
                texte = texte_net or texte
                if h_parsed:
                    agenda_date, agenda_heure = d_parsed, h_parsed
                else:
                    date_echeance = d_parsed
        except Exception as e:
            print(f"[QUICKADD] parse échoué: {e}")
    conn = get_db_connection()
    cur = conn.execute(
        "INSERT INTO todos_perso (texte, priorite, date_echeance, is_titre, parent_titre_id) VALUES (?, ?, ?, ?, ?)",
        (texte, priorite, date_echeance, is_titre, parent_titre_id)
    )
    todo_id = cur.lastrowid
    for admin_id in assigne_admin_ids:
        conn.execute("INSERT OR IGNORE INTO todo_assignees (todo_id, admin_id) VALUES (?, ?)", (todo_id, admin_id))
    conn.commit()
    cal_event_id = ''
    if agenda_date and agenda_heure and not is_titre:
        try:
            from calendar_service import create_task_block
            cal_event_id = create_task_block(texte, agenda_date, agenda_heure, agenda_duree, priorite)
            if cal_event_id:
                conn.execute("UPDATE todos_perso SET calendar_event_id=?, date_echeance=? WHERE id=?",
                             (cal_event_id, agenda_date, todo_id))
        except Exception as e:
            print(f"[CALENDAR] Task block: {e}")
    elif date_echeance and not is_titre:
        try:
            from calendar_service import create_todo_reminder
            cal_event_id = create_todo_reminder(texte, date_echeance)
            if cal_event_id:
                conn.execute("UPDATE todos_perso SET calendar_event_id=? WHERE id=?", (cal_event_id, todo_id))
        except Exception as e:
            print(f"[CALENDAR] Todo reminder: {e}")
    if not is_titre:
        push_admin_notif(conn, "Nouvelle tâche", texte, type='todo', lien='/admin')
    conn.commit()
    todo = conn.execute("SELECT * FROM todos_perso WHERE id=?", (todo_id,)).fetchone()
    d = dict(todo)
    d['assignees'] = _todo_assignees(conn, todo_id)
    conn.close()
    # Projet/client/contact fournis dès la création (ex. feuille d'ajout enrichie de la PWA
    # Tâches) — réutilise _update_todo plutôt que dupliquer cette logique d'assignation.
    extra = {k: data[k] for k in ('projet_id', 'client_id', 'contact_nom', 'contact_telephone', 'contact_courriel') if k in data}
    if extra:
        d, _status = _update_todo(todo_id, extra)
    return d, 201


def _toggle_todo(todo_id):
    """Bascule coché/décoché — utilisée par la session CRM et par le jeton d'appareil de
    la PWA Tâches. Répercute sur la roadmap liée si applicable."""
    conn = get_db_connection()
    todo = conn.execute("SELECT * FROM todos_perso WHERE id=?", (todo_id,)).fetchone()
    if not todo:
        conn.close()
        return {'error': 'Todo introuvable'}, 404
    new_val = 0 if todo['est_coche'] else 1
    conn.execute("UPDATE todos_perso SET est_coche=? WHERE id=?", (new_val, todo_id))
    conn.commit()
    linked_roadmap_todo_id = todo['linked_roadmap_todo_id'] if 'linked_roadmap_todo_id' in todo.keys() else None
    if linked_roadmap_todo_id:
        _sync_roadmap_todo_completion(conn, linked_roadmap_todo_id, bool(new_val))
    conn.close()
    return {'success': True, 'est_coche': bool(new_val)}, 200


@app.route('/api/v1/admin/todos', methods=['GET'])
@admin_required
def api_todos_list():
    view = request.args.get('view', 'mine')
    return jsonify(_todos_query(session['user_id'], view))

@app.route('/api/v1/admin/todos', methods=['POST'])
@admin_required
def api_todos_create():
    body, status = _create_todo(session.get('user_id'), request.get_json() or {})
    return jsonify(body), status

@app.route('/api/v1/admin/todos/<int:todo_id>/toggle', methods=['POST'])
@admin_required
def api_todos_toggle(todo_id):
    body, status = _toggle_todo(todo_id)
    return jsonify(body), status

def _update_todo(todo_id, data):
    """Mise à jour partagée d'une tâche — utilisée par la session CRM et par le jeton
    d'appareil de la PWA Tâches."""
    conn = get_db_connection()
    todo = conn.execute("SELECT * FROM todos_perso WHERE id=?", (todo_id,)).fetchone()
    if not todo:
        conn.close()
        return {'error': 'Todo introuvable'}, 404
    if 'priorite' in data:
        conn.execute("UPDATE todos_perso SET priorite=? WHERE id=?", (data['priorite'], todo_id))
    if 'texte' in data:
        txt = (data.get('texte') or '').strip()
        if txt:
            conn.execute("UPDATE todos_perso SET texte=? WHERE id=?", (txt, todo_id))
    if 'date_echeance' in data:
        conn.execute("UPDATE todos_perso SET date_echeance=? WHERE id=?", (data.get('date_echeance') or None, todo_id))
    if 'contact_nom' in data:
        conn.execute("UPDATE todos_perso SET contact_nom=? WHERE id=?", (data.get('contact_nom') or None, todo_id))
    if 'contact_telephone' in data:
        conn.execute("UPDATE todos_perso SET contact_telephone=? WHERE id=?", (data.get('contact_telephone') or None, todo_id))
    if 'contact_courriel' in data:
        conn.execute("UPDATE todos_perso SET contact_courriel=? WHERE id=?", (data.get('contact_courriel') or None, todo_id))
    # Assignation à un projet (le client est déduit du projet). On sort la tâche de sa section.
    if 'projet_id' in data:
        pid = data['projet_id']
        if pid:
            proj = conn.execute("SELECT id, nom_projet, id_client FROM projets WHERE id=?", (pid,)).fetchone()
            if proj:
                conn.execute(
                    "UPDATE todos_perso SET projet_id=?, projet_nom=?, client_id=?, parent_titre_id=NULL WHERE id=?",
                    (proj['id'], proj['nom_projet'], proj['id_client'], todo_id)
                )
        else:
            conn.execute("UPDATE todos_perso SET projet_id=NULL, projet_nom=NULL WHERE id=?", (todo_id,))
    # Assignation à un client seul (retire le projet éventuel). On sort la tâche de sa section.
    if 'client_id' in data:
        cid = data['client_id']
        if cid:
            conn.execute(
                "UPDATE todos_perso SET client_id=?, projet_id=NULL, projet_nom=NULL, parent_titre_id=NULL WHERE id=?",
                (cid, todo_id)
            )
        else:
            conn.execute("UPDATE todos_perso SET client_id=NULL WHERE id=?", (todo_id,))
    # Déplacer sous une section existante (titre)
    if 'parent_titre_id' in data:
        conn.execute("UPDATE todos_perso SET parent_titre_id=? WHERE id=?",
                     (data['parent_titre_id'] or None, todo_id))
    # Assignation à un ou plusieurs comptes admin (liste vide = tâche partagée,
    # visible par toute l'équipe) — remplace le set complet des assigné·es. Notifie
    # (push) uniquement les personnes nouvellement ajoutées, pas tout le monde à
    # chaque modification de la tâche.
    if 'assigne_admin_ids' in data:
        old_ids = {a['id'] for a in _todo_assignees(conn, todo_id)}
        new_ids = {i for i in (data.get('assigne_admin_ids') or []) if i}
        conn.execute("DELETE FROM todo_assignees WHERE todo_id=?", (todo_id,))
        for admin_id in new_ids:
            conn.execute("INSERT OR IGNORE INTO todo_assignees (todo_id, admin_id) VALUES (?, ?)", (todo_id, admin_id))
        added = new_ids - old_ids
        if added:
            texte_notif = (data.get('texte') or todo['texte'] or '').strip()
            for row in conn.execute(f"SELECT email FROM clients WHERE id IN ({','.join('?' * len(added))})", tuple(added)).fetchall():
                push_admin_notif(conn, "Tâche assignée", texte_notif, type='assignation', lien='/taches', destinataire=row['email'])
    conn.commit()
    todo = conn.execute("SELECT * FROM todos_perso WHERE id=?", (todo_id,)).fetchone()
    d = dict(todo)
    d['assignees'] = _todo_assignees(conn, todo_id)
    conn.close()
    return d, 200


def _delete_todo(todo_id):
    """Suppression partagée — utilisée par la session CRM et par le jeton d'appareil de
    la PWA Tâches."""
    conn = get_db_connection()
    todo = conn.execute("SELECT * FROM todos_perso WHERE id=?", (todo_id,)).fetchone()
    if not todo:
        conn.close()
        return {'error': 'Todo introuvable'}, 404
    if todo['calendar_event_id']:
        try:
            from calendar_service import delete_calendar_event
            delete_calendar_event(todo['calendar_event_id'])
        except Exception as e:
            print(f"[CALENDAR] Delete todo event: {e}")
    conn.execute("DELETE FROM todos_perso WHERE id=?", (todo_id,))
    conn.commit()
    conn.close()
    return {'success': True}, 200


def _planifier_todo(todo_id, data):
    """Planification calendrier partagée — utilisée par la session CRM et par le jeton
    d'appareil de la PWA Tâches."""
    date_str = (data.get('date') or '').strip()
    heure = (data.get('heure') or '09:00').strip()
    try:
        duree = int(data.get('duree') or 60)
    except (TypeError, ValueError):
        duree = 60
    if not date_str:
        return {'error': 'Date requise'}, 400
    conn = get_db_connection()
    todo = conn.execute("SELECT * FROM todos_perso WHERE id=?", (todo_id,)).fetchone()
    if not todo:
        conn.close()
        return {'error': 'Todo introuvable'}, 404
    # Re-planification : on retire l'ancien événement d'abord
    if todo['calendar_event_id']:
        try:
            from calendar_service import delete_calendar_event
            delete_calendar_event(todo['calendar_event_id'])
        except Exception as e:
            print(f"[CALENDAR] suppression ancien event: {e}")
    try:
        from calendar_service import create_task_block
        cal_id = create_task_block(todo['texte'], date_str, heure, duree, todo['priorite'] or 'normale')
    except Exception as e:
        conn.close()
        return {'error': f'Calendrier: {e}'}, 500
    conn.execute("UPDATE todos_perso SET calendar_event_id=?, date_echeance=? WHERE id=?",
                 (cal_id, date_str, todo_id))
    conn.commit()
    conn.close()
    return {'success': True, 'calendar_event_id': cal_id, 'date_echeance': date_str}, 200


def _deplanifier_todo(todo_id):
    """Retrait du calendrier partagé — utilisée par la session CRM et par le jeton
    d'appareil de la PWA Tâches."""
    conn = get_db_connection()
    todo = conn.execute("SELECT * FROM todos_perso WHERE id=?", (todo_id,)).fetchone()
    if not todo:
        conn.close()
        return {'error': 'Todo introuvable'}, 404
    if todo['calendar_event_id']:
        try:
            from calendar_service import delete_calendar_event
            delete_calendar_event(todo['calendar_event_id'])
        except Exception as e:
            print(f"[CALENDAR] déplanification: {e}")
    conn.execute("UPDATE todos_perso SET calendar_event_id=NULL WHERE id=?", (todo_id,))
    conn.commit()
    conn.close()
    return {'success': True}, 200


@app.route('/api/v1/admin/todos/<int:todo_id>', methods=['PATCH'])
@admin_required
def api_todos_update(todo_id):
    body, status = _update_todo(todo_id, request.get_json() or {})
    return jsonify(body), status

@app.route('/api/v1/admin/todos/<int:todo_id>', methods=['DELETE'])
@admin_required
def api_todos_delete(todo_id):
    body, status = _delete_todo(todo_id)
    return jsonify(body), status

@app.route('/api/v1/admin/todos/<int:todo_id>/planifier', methods=['POST'])
@admin_required
def api_todos_planifier(todo_id):
    body, status = _planifier_todo(todo_id, request.get_json() or {})
    return jsonify(body), status

@app.route('/api/v1/admin/todos/<int:todo_id>/deplanifier', methods=['POST'])
@admin_required
def api_todos_deplanifier(todo_id):
    body, status = _deplanifier_todo(todo_id)
    return jsonify(body), status

@app.route('/api/v1/admin/projet/<int:projet_id>/rename', methods=['POST'])
@admin_required
def api_projet_rename(projet_id):
    """Renomme un projet (depuis le groupe todo) + répercute sur les todos liés."""
    data = request.get_json() or {}
    nom = (data.get('nom') or '').strip()
    if not nom:
        return jsonify({'error': 'Nom requis'}), 400
    conn = get_db_connection()
    proj = conn.execute("SELECT id FROM projets WHERE id=?", (projet_id,)).fetchone()
    if not proj:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404
    conn.execute("UPDATE projets SET nom_projet=? WHERE id=?", (nom, projet_id))
    conn.execute("UPDATE todos_perso SET projet_nom=? WHERE projet_id=?", (nom, projet_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'nom_projet': nom})

@app.route('/api/v1/admin/todos/bulk-assign', methods=['POST'])
@admin_required
def api_todos_bulk_assign():
    """Assigne un même client/projet à plusieurs todos (assignation de catégorie)."""
    data = request.get_json() or {}
    ids = data.get('ids') or []
    if not isinstance(ids, list) or not ids:
        return jsonify({'error': 'ids requis'}), 400
    ids = [int(i) for i in ids]
    conn = get_db_connection()
    placeholders = ",".join("?" * len(ids))
    if 'projet_id' in data:
        pid = data['projet_id']
        if pid:
            proj = conn.execute("SELECT id, nom_projet, id_client FROM projets WHERE id=?", (pid,)).fetchone()
            if proj:
                conn.execute(
                    f"UPDATE todos_perso SET projet_id=?, projet_nom=?, client_id=?, parent_titre_id=NULL WHERE id IN ({placeholders})",
                    (proj['id'], proj['nom_projet'], proj['id_client'], *ids))
        else:
            conn.execute(f"UPDATE todos_perso SET projet_id=NULL, projet_nom=NULL WHERE id IN ({placeholders})", ids)
    if 'client_id' in data:
        cid = data['client_id']
        if cid:
            conn.execute(
                f"UPDATE todos_perso SET client_id=?, projet_id=NULL, projet_nom=NULL, parent_titre_id=NULL WHERE id IN ({placeholders})",
                (cid, *ids))
        else:
            conn.execute(f"UPDATE todos_perso SET client_id=NULL WHERE id IN ({placeholders})", ids)
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ═══════════════════════════════════════════════════════════════
# TEMPLATE BUILDER — Sites clients
# ═══════════════════════════════════════════════════════════════
import requests as _req
import subprocess as _sub
import shutil as _shutil
import tempfile as _tempfile
import threading as _threading

_GITHUB_TOKEN   = os.getenv('GITHUB_TOKEN', '')
_VERCEL_TOKEN   = os.getenv('VERCEL_TOKEN', '')
_VERCEL_TEAM_ID = os.getenv('VERCEL_TEAM_ID', '')
_SANITY_TOKEN   = os.getenv('SANITY_TOKEN', '')
_TEMPLATES_PATH = '/app/site_templates'

def _github_username():
    r = _req.get('https://api.github.com/user',
                 headers={'Authorization': f'token {_GITHUB_TOKEN}', 'Accept': 'application/vnd.github.v3+json'})
    return r.json().get('login', '')

def _slugify_site(text: str) -> str:
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = text.lower()
    text = re.sub(r'[^a-z0-9\s-]', '', text)
    text = re.sub(r'\s+', '-', text.strip())
    return re.sub(r'-+', '-', text)[:40]

def _ts_str(val) -> str:
    escaped = (str(val or '')
               .replace('\\', '\\\\')
               .replace("'", "\\'")
               .replace('\r\n', '\\n')
               .replace('\n', '\\n')
               .replace('\r', '\\n'))
    return f"'{escaped}'"

def _ts_list(lst: list) -> str:
    return '[' + ', '.join(_ts_str(v) for v in lst) + ']'

def _generate_config_ts(template: str, d: dict) -> str:
    if template == 'reservation':
        kw_fr = [k.strip() for k in d.get('seo_keywords_fr', '').split(',') if k.strip()][:3]
        kw_en = [k.strip() for k in d.get('seo_keywords_en', '').split(',') if k.strip()][:3]
        hero  = d.get('hero_style', 'A')
        return f"""const clientConfig = {{
  businessName: {_ts_str(d.get('business_name'))},
  ownerName:    {_ts_str(d.get('owner_name'))},
  ownerTitle:   {{ fr: {_ts_str(d.get('owner_title_fr'))}, en: {_ts_str(d.get('owner_title_en'))} }},
  tagline:      {{ fr: {_ts_str(d.get('tagline_fr'))},     en: {_ts_str(d.get('tagline_en'))} }},
  description:  {{
    fr: {_ts_str(d.get('description_fr'))},
    en: {_ts_str(d.get('description_en'))},
  }},

  address:    {_ts_str(d.get('address'))},
  city:       {_ts_str(d.get('city'))},
  province:   {_ts_str(d.get('province', 'QC'))},
  postalCode: {_ts_str(d.get('postal_code'))},
  phone:      {_ts_str(d.get('phone'))},
  email:      {_ts_str(d.get('email'))},

  acuityUrl: process.env.NEXT_PUBLIC_ACUITY_URL ?? '',

  social: {{
    instagram: {_ts_str(d.get('instagram'))},
    facebook:  {_ts_str(d.get('facebook'))},
  }},

  heroStyle: {_ts_str(hero)} as 'A' | 'B' | 'C',

  siteUrl: process.env.NEXT_PUBLIC_SITE_URL ?? 'https://example.com',
  seo: {{
    metaTitle:       {{ fr: {_ts_str(d.get('seo_meta_title_fr'))},       en: {_ts_str(d.get('seo_meta_title_en'))} }},
    metaDescription: {{ fr: {_ts_str(d.get('seo_meta_description_fr'))}, en: {_ts_str(d.get('seo_meta_description_en'))} }},
    keywords: {{
      fr: {_ts_list(kw_fr)},
      en: {_ts_list(kw_en)},
    }},
    ogImage:       {_ts_str(d.get('seo_og_image'))},
    twitterHandle: {_ts_str(d.get('seo_twitter_handle'))},
    logoUrl:       {_ts_str(d.get('seo_logo_url'))},
    businessType:  {_ts_str(d.get('seo_business_type', 'BeautySalon'))},
    priceRange:    {_ts_str(d.get('seo_price_range', '$$'))},
  }},
}} as const;

export default clientConfig;
"""
    elif template == 'vitrine':
        kw = [k.strip() for k in d.get('seo_keywords', '').split(',') if k.strip()][:3]
        direction = d.get('direction') or DEFAULT_DIRECTION
        return f"""import type {{ Direction }} from "@/lib/directions";

const clientConfig = {{
  businessName: {_ts_str(d.get('business_name'))},
  ownerName:    {_ts_str(d.get('owner_name'))},
  ownerTitle:   {_ts_str(d.get('owner_title'))},
  tagline:      {_ts_str(d.get('tagline'))},
  description:  {_ts_str(d.get('description'))},

  address:    {_ts_str(d.get('address'))},
  city:       {_ts_str(d.get('city'))},
  province:   {_ts_str(d.get('province', 'QC'))},
  postalCode: {_ts_str(d.get('postal_code'))},
  phone:      {_ts_str(d.get('phone'))},
  email:      {_ts_str(d.get('email'))},

  bookingUrl:      {_ts_str(d.get('booking_url'))},
  sanityStudioUrl: '',

  direction: {_ts_str(direction)} as Direction,

  social: {{
    instagram: {_ts_str(d.get('instagram'))},
    facebook:  {_ts_str(d.get('facebook'))},
    linkedin:  {_ts_str(d.get('linkedin'))},
  }},

  siteUrl: process.env.NEXT_PUBLIC_SITE_URL ?? 'https://example.com',
  seo: {{
    metaTitle:       {_ts_str(d.get('seo_meta_title'))},
    metaDescription: {_ts_str(d.get('seo_meta_description'))},
    keywords: {_ts_list(kw)},
    ogImage:       {_ts_str(d.get('seo_og_image'))},
    twitterHandle: {_ts_str(d.get('seo_twitter_handle'))},
    logoUrl:       {_ts_str(d.get('seo_logo_url'))},
    businessType:  {_ts_str(d.get('seo_business_type', 'LocalBusiness'))},
    priceRange:    {_ts_str(d.get('seo_price_range', '$$'))},
  }},
}};

export default clientConfig;
"""
    else:  # sante
        kw = [k.strip() for k in d.get('seo_keywords', '').split(',') if k.strip()][:3]
        return f"""const clientConfig = {{
  businessName: {_ts_str(d.get('business_name'))},
  ownerName:    {_ts_str(d.get('owner_name'))},
  ownerTitle:   {_ts_str(d.get('owner_title'))},
  tagline:      {_ts_str(d.get('tagline'))},
  description:  {_ts_str(d.get('description'))},

  address:    {_ts_str(d.get('address'))},
  city:       {_ts_str(d.get('city'))},
  province:   {_ts_str(d.get('province', 'QC'))},
  postalCode: {_ts_str(d.get('postal_code'))},
  phone:      {_ts_str(d.get('phone'))},
  email:      {_ts_str(d.get('email'))},

  bookingUrl:      {_ts_str(d.get('booking_url') or d.get('acuity_url'))},
  sanityStudioUrl: '',

  social: {{
    instagram: {_ts_str(d.get('instagram'))},
    facebook:  {_ts_str(d.get('facebook'))},
    linkedin:  {_ts_str(d.get('linkedin'))},
  }},

  siteUrl: process.env.NEXT_PUBLIC_SITE_URL ?? 'https://example.com',
  seo: {{
    metaTitle:       {_ts_str(d.get('seo_meta_title'))},
    metaDescription: {_ts_str(d.get('seo_meta_description'))},
    keywords: {_ts_list(kw)},
    ogImage:       {_ts_str(d.get('seo_og_image'))},
    twitterHandle: {_ts_str(d.get('seo_twitter_handle'))},
    logoUrl:       {_ts_str(d.get('seo_logo_url'))},
    businessType:  {_ts_str(d.get('seo_business_type', 'LocalBusiness'))},
    priceRange:    {_ts_str(d.get('seo_price_range', '$$'))},
  }},
}} as const;

export default clientConfig;
"""

# ─── Gabarit vitrine — 8 directions artistiques Baseline (fichiers figés, ────
# aucune génération à la volée : chaque direction est copiée telle quelle,
# seul l'id de direction change dans client.config.ts). Miroir de
# templates/vitrine/lib/directions.ts — garder synchronisé si une direction
# est ajoutée/renommée côté template.
DEFAULT_DIRECTION = 'editorial'
DIRECTIONS = {
    'editorial': 'Éditorial',
    'manifesto': 'Manifesto',
    'feutre':    'Feutré',
    'ludique':   'Ludique',
    'clean':     'Clean',
    'cinematic': 'Cinématique',
    'bento':     'Bento',
    'organique': 'Organique',
}

def _create_github_repo(slug: str, github_token: str) -> dict:
    gh_user = _github_username()
    repo_name = f"site-{slug}"
    r = _req.post(
        'https://api.github.com/user/repos',
        headers={'Authorization': f'token {github_token}', 'Accept': 'application/vnd.github.v3+json'},
        json={'name': repo_name, 'private': True, 'auto_init': False, 'description': f'Site Cocktail Média — {slug}'}
    )
    if r.status_code not in (200, 201):
        raise RuntimeError(f"GitHub repo creation failed: {r.text}")
    return {'full_name': f"{gh_user}/{repo_name}", 'url': r.json().get('html_url', '')}

def _push_template_to_repo(template: str, slug: str, config_ts: str, github_token: str, repo_full_name: str):
    template_src = os.path.join(_TEMPLATES_PATH, template)
    with _tempfile.TemporaryDirectory() as tmpdir:
        _shutil.copytree(template_src, tmpdir, dirs_exist_ok=True)
        with open(os.path.join(tmpdir, 'client.config.ts'), 'w') as f:
            f.write(config_ts)
        remote_url = f"https://{github_token}@github.com/{repo_full_name}.git"
        git_env = {**os.environ,
                   'GIT_AUTHOR_NAME': 'Cocktail Média', 'GIT_AUTHOR_EMAIL': 'felix.dumont@cocktailmedia.ca',
                   'GIT_COMMITTER_NAME': 'Cocktail Média', 'GIT_COMMITTER_EMAIL': 'felix.dumont@cocktailmedia.ca'}
        for cmd in [
            ['git', 'init'],
            ['git', 'add', '.'],
            ['git', 'commit', '-m', f'feat: init {slug}'],
            ['git', 'branch', '-M', 'main'],
            ['git', 'remote', 'add', 'origin', remote_url],
            ['git', 'push', '-u', 'origin', 'main'],
        ]:
            _sub.run(cmd, cwd=tmpdir, check=True, capture_output=True, env=git_env)

def _create_sanity_project(display_name: str, sanity_token: str) -> str:
    r = _req.post(
        'https://api.sanity.io/v1/projects',
        headers={'Authorization': f'Bearer {sanity_token}', 'Content-Type': 'application/json'},
        json={'displayName': display_name}
    )
    if r.status_code not in (200, 201):
        raise RuntimeError(f"Sanity create project failed: {r.text}")
    project_id = r.json()['id']
    # La création d'un projet ne provisionne PAS de dataset — sans cet appel, le site
    # généré interroge un dataset "production" inexistant et retombe silencieusement
    # sur les textes placeholder (try/except dans lib/sanity.ts), sans jamais planter.
    # Bug découvert le 2026-07-09 : au moins 2 sites réels (dont un en prod) n'avaient
    # jamais eu de dataset créé depuis la mise en place de ce pipeline.
    ds = _req.put(
        f'https://api.sanity.io/v1/projects/{project_id}/datasets/production',
        headers={'Authorization': f'Bearer {sanity_token}', 'Content-Type': 'application/json'},
        json={'aclMode': 'public'}
    )
    if ds.status_code not in (200, 201):
        raise RuntimeError(f"Sanity create dataset failed: {ds.text}")
    return project_id

def _invite_sanity_member(project_id: str, email: str, sanity_token: str):
    _req.post(
        f'https://api.sanity.io/v1/invitations/project/{project_id}',
        headers={'Authorization': f'Bearer {sanity_token}', 'Content-Type': 'application/json'},
        json={'email': email, 'role': 'editor'}
    )

# Texte de repli JSX Baseline, extrait mécaniquement des 8 directions artistiques
# (templates/vitrine/components/{Accueil,Services,Equipe,Contact}.tsx) — c'est le texte
# affiché sur le site TANT QUE Sanity n'a rien pour ce champ (`content?.champ ?? "..."`).
# Le semer explicitement dans Sanity à la création évite qu'un client voie un champ vide
# dans /mon-site alors que son site affiche déjà du texte (confus sans ça).
VITRINE_BASELINE_DEFAULTS = {
    'editorial': {
        'pageAccueil': {'heroEyebrow': "VOTRE PARTENAIRE DE CONFIANCE", 'heroTitre': "Un service qui tient ses promesses", 'heroSousTitre': "Une équipe locale qui connaît votre réalité, reste présente après le mandat, et traite chaque projet comme s'il était le seul.", 'servicesEyebrow': "NOS SERVICES", 'servicesTitre': "Un accompagnement complet, du début à la fin", 'equipeTitle': "Les gens derrière chaque mandat", 'faqTitle': "QUESTIONS FRÉQUENTES", 'ctaTitre': "Prêt·e à démarrer votre projet ?", 'ctaSousTitre': "Une soumission gratuite, sans engagement."},
        'pageServices': {'heroEyebrow': "ACCUEIL / SERVICES", 'heroTitre': "Ce qu'on fait, et comment on le fait bien", 'heroSousTitre': "Un accompagnement complet, du premier échange à la livraison — sans mauvaise surprise.", 'ctaTitre': "Un projet en tête ?"},
        'pageEquipe': {'heroEyebrow': "ACCUEIL / ÉQUIPE", 'heroTitre': "Les visages derrière chaque mandat", 'heroSousTitre': "Une équipe locale, stable, qui connaît vos dossiers d'un projet à l'autre.", 'ctaTitre': "Envie de nous rencontrer ?", 'missionTexte': "Offrir un accompagnement humain et rigoureux, du premier contact à la livraison de chaque mandat.", 'visionTexte': "Devenir un partenaire de confiance sur lequel on peut compter, année après année.", 'valeursTexte': "L'écoute, la rigueur et le respect du rythme de chacun."},
        'pageContact': {'heroTitre': "Parlons de votre projet", 'heroSousTitre': "Une soumission gratuite, une question, ou simplement envie d'échanger — écrivez-nous."},
    },
    'manifesto': {
        'pageAccueil': {'heroEyebrow': "/ AGENCE DE CRÉATION", 'heroTitre': "On fait\ndu bruit.", 'heroSousTitre': "Pas de demi-mesure. On construit des marques qu'on ne peut pas ignorer.", 'equipeTitle': "Les gens derrière le travail", 'ctaTitre': "On y va ?"},
        'pageServices': {'heroTitre': "Ce qu'on\nsait faire", 'ctaTitre': "On attaque ?"},
        'pageEquipe': {'heroEyebrow': "/ ÉQUIPE", 'heroTitre': "On ne\nrecule pas.", 'ctaTitre': "Rejoins la meute", 'missionTexte': "Livrer vite, livrer bien, sans jamais sacrifier l'audace. On ne fait pas de projets timides.", 'visionTexte': "Devenir la référence pour les marques qui refusent de se fondre dans la masse.", 'valeursTexte': "Franchise. Intensité. Zéro compromis sur la qualité."},
        'pageContact': {'heroTitre': "Écris-\nnous.", 'heroSousTitre': "Un projet, une idée, un défi. On répond vite, on parle vrai."},
    },
    'feutre': {
        'pageAccueil': {'heroEyebrow': "MAISON FONDÉE EN 2016", 'heroTitre': "L'art du détail, au service de votre projet", 'heroSousTitre': "Un accompagnement discret et rigoureux, pensé pour durer bien au-delà du mandat.", 'equipeTitle': "L'équipe derrière la Maison", 'ctaTitre': "Écrivons votre histoire"},
        'pageServices': {'heroEyebrow': "NOS SERVICES", 'heroTitre': "Un métier, décliné avec soin", 'ctaTitre': "Discutons de votre projet"},
        'pageEquipe': {'heroEyebrow': "L'ÉQUIPE", 'heroTitre': "Un savoir-faire porté par des mains attentives", 'ctaTitre': "Faisons connaissance", 'missionTexte': "Offrir un accompagnement sur mesure, où chaque détail est pensé avec soin et discrétion.", 'visionTexte': "Devenir le partenaire de confiance qu'on garde pour la vie, projet après projet.", 'valeursTexte': "L'élégance du geste, la rigueur du métier, le respect du temps de chacun."},
        'pageContact': {'heroTitre': "Prenons le temps de discuter", 'heroSousTitre': "Nous vous répondrons personnellement, sous 24 heures."},
    },
    'ludique': {
        'pageAccueil': {'heroEyebrow': "On adore ce qu'on fait", 'heroTitre': "Créons ensemble quelque chose de génial", 'heroSousTitre': "Une équipe joyeuse et sérieuse à la fois, qui prend vos idées au sérieux, pas nous-mêmes.", 'equipeTitle': "L'équipe qui vous accueille", 'ctaTitre': "On se lance ?"},
        'pageServices': {'heroEyebrow': "Nos petits talents", 'heroTitre': "Tout ce qu'on peut faire pour vous", 'ctaTitre': "On commence par quoi ?"},
        'pageEquipe': {'heroEyebrow': "Notre joyeuse bande", 'heroTitre': "Des gens qu'on a hâte de vous présenter", 'ctaTitre': "On a une place pour vous !", 'missionTexte': "Rendre chaque collaboration légère, joyeuse et sans prise de tête, tout en livrant du solide.", 'visionTexte': "Prouver qu'on peut s'amuser et faire un travail sérieux en même temps.", 'valeursTexte': "La bonne humeur, la curiosité, et une bonne dose d'audace créative."},
        'pageContact': {'heroTitre': "On adore recevoir de vos nouvelles !", 'heroSousTitre': "Écrivez-nous, on répond toujours avec le sourire."},
    },
    'clean': {
        'pageAccueil': {'heroEyebrow': "Nouvelle génération", 'heroTitre': "La solution claire pour votre entreprise", 'heroSousTitre': "Structuré, fiable, sans friction. Tout ce qu’il faut pour avancer vite et bien.", 'equipeTitle': "L'équipe", 'ctaTitre': "Commencez aujourd'hui"},
        'pageServices': {'heroTitre': "Des offres claires, sans surprise", 'heroSousTitre': "Choisissez le niveau d'accompagnement qui vous convient.", 'ctaTitre': "Pas sûr du bon choix ?"},
        'pageEquipe': {'heroEyebrow': "Notre équipe", 'heroTitre': "Les bonnes personnes, au bon endroit", 'heroSousTitre': "Un annuaire clair — vous savez toujours à qui parler.", 'ctaTitre': "Parlez à la bonne personne", 'missionTexte': "Simplifier la collaboration grâce à des processus clairs et une communication directe.", 'visionTexte': "Être le partenaire opérationnel le plus fiable de notre industrie.", 'valeursTexte': "Clarté, efficacité, transparence à chaque étape."},
        'pageContact': {'heroTitre': "Contactez-nous", 'heroSousTitre': "Choisissez le canal qui vous convient. On répond sous 24 h."},
    },
    'cinematic': {
        'pageAccueil': {'heroEyebrow': "STUDIO DE PRODUCTION", 'heroTitre': "Des histoires\nqui marquent", 'heroSousTitre': "Image, film, motion — on donne du relief à votre marque.", 'equipeTitle': "L'équipe", 'ctaTitre': "On tourne ?"},
        'pageServices': {'heroEyebrow': "/ SERVICES", 'heroTitre': "Nos savoir-faire", 'ctaTitre': "On tourne ?"},
        'pageEquipe': {'heroEyebrow': "ACCUEIL / ÉQUIPE", 'heroTitre': "Le collectif.", 'ctaTitre': "Envie de bosser avec nous ?", 'missionTexte': "Créer des images qui marquent, qui restent gravées bien après le générique.", 'visionTexte': "Redéfinir ce qu'une marque peut raconter à l'écran.", 'valeursTexte': "L'exigence visuelle, l'audace narrative, aucun compromis sur le cadre."},
        'pageContact': {'heroTitre': "On se parle ?"},
    },
    'bento': {
        'pageAccueil': {'heroTitre': "Construisez l'avenir, bloc par bloc", 'equipeTitle': "L'équipe", 'ctaTitre': "Prêt à décoller ?"},
        'pageServices': {'heroTitre': "Nos modules, à la carte", 'ctaTitre': "Composez votre offre"},
        'pageEquipe': {'heroEyebrow': "L'ÉQUIPE", 'heroTitre': "Des esprits qui gravitent ensemble", 'ctaTitre': "Rejoindre l'orbite ?", 'missionTexte': "Connecter design et technologie pour créer des expériences qui gravitent autour de vos utilisateurs.", 'visionTexte': "Construire les interfaces qui définiront la prochaine décennie du web.", 'valeursTexte': "Innovation, précision, une curiosité sans limite."},
        'pageContact': {'heroTitre': "Entrons en contact"},
    },
    'organique': {
        'pageAccueil': {'heroEyebrow': "Du champ à la table", 'heroTitre': "Un savoir-faire qui a des racines", 'heroSousTitre': "Des produits honnêtes, une approche patiente, et le goût du travail bien fait.", 'equipeTitle': "L'équipe qui vous accompagne", 'ctaTitre': "Passez nous voir"},
        'pageServices': {'heroEyebrow': "Nos services", 'heroTitre': "Un savoir-faire, plusieurs récoltes", 'ctaTitre': "Cultivons votre projet"},
        'pageEquipe': {'heroEyebrow': "Notre monde", 'heroTitre': "Des artisans, pas des numéros", 'ctaTitre': "Venez cultiver avec nous", 'missionTexte': "Cultiver des projets qui ont du sens, avec la patience et le soin d'un artisan.", 'visionTexte': "Créer un monde où chaque marque pousse à son propre rythme, sans forcer la nature des choses.", 'valeursTexte': "La patience, l'authenticité, le respect du vivant et du temps qui passe."},
        'pageContact': {'heroTitre': "Passez nous voir, ou écrivez-nous"},
    },
}

def _seed_vitrine_baseline_defaults(sanity_project_id, direction):
    """Écrit le texte de repli Baseline de la direction donnée dans Sanity, mais
    UNIQUEMENT pour les champs encore vides (`setIfMissing`, jamais `set`) — ne touche
    jamais un champ déjà personnalisé par un client réel."""
    defaults = VITRINE_BASELINE_DEFAULTS.get(direction, {})
    for doc_type, fields in defaults.items():
        if not fields:
            continue
        try:
            _sanity_mutate(sanity_project_id, [
                {'createIfNotExists': {'_id': doc_type, '_type': doc_type}},
                {'patch': {'id': doc_type, 'setIfMissing': fields}},
            ])
        except Exception as exc:
            print(f"[SEED] échec {doc_type} pour {sanity_project_id}: {exc}", flush=True)


def _push_checklist_content_to_sanity(projet_id: int, sanity_project_id: str):
    """Pousse vers Sanity les réponses de checklist qui déclarent un content_key —
    remplace la ressaisie manuelle par une transmission directe et déterministe
    (chaque item a une destination unique, jamais devinée par correspondance de libellé).
    Réutilise _sanity_mutate (même fonction que l'auto-édition client, /mon-site/contenu).
    Une erreur sur un document/membre n'empêche pas les autres d'être poussés."""
    conn = get_db_connection()
    try:
        items = conn.execute(
            """SELECT ci.content_key, ci.text_value
               FROM checklist_items ci JOIN checklistes ck ON ck.id = ci.id_checklist
               WHERE ck.id_projet = ? AND ci.content_key IS NOT NULL
                     AND ci.text_value IS NOT NULL AND ci.text_value != ''""",
            (projet_id,)
        ).fetchall()
    finally:
        conn.close()

    docs: dict = {}
    membres_raw = None
    for it in items:
        key = it['content_key']
        if key == 'membre[]':
            membres_raw = it['text_value']
        elif '.' in key:
            doc_type, field = key.split('.', 1)
            docs.setdefault(doc_type, {})[field] = it['text_value']

    for doc_type, fields in docs.items():
        try:
            _sanity_mutate(sanity_project_id, [
                {'createIfNotExists': {'_id': doc_type, '_type': doc_type}},
                {'patch': {'id': doc_type, 'set': fields}},
            ])
        except Exception as exc:
            print(f"[SANITY PUSH] échec {doc_type} pour projet {projet_id}: {exc}", flush=True)

    if membres_raw:
        try:
            membres = json.loads(membres_raw)
        except Exception:
            membres = []
        for idx, mem in enumerate(membres if isinstance(membres, list) else []):
            name = (mem.get('name') or '').strip()
            if not name:
                continue
            parts = name.split(' ', 1)
            prenom, nom = parts[0], (parts[1] if len(parts) > 1 else '')
            try:
                _sanity_mutate(sanity_project_id, [
                    {'createOrReplace': {
                        '_id': f'membre-projet{projet_id}-{idx}', '_type': 'membre',
                        'ordre': idx + 1, 'prenom': prenom, 'nom': nom,
                        'titre': (mem.get('title') or '').strip(),
                        'bio': (mem.get('desc') or '').strip(),
                        'visible': True,
                    }},
                ])
            except Exception as exc:
                print(f"[SANITY PUSH] échec membre {idx} pour projet {projet_id}: {exc}", flush=True)

def _get_github_repo_id(repo_full_name: str, github_token: str) -> int:
    r = _req.get(f'https://api.github.com/repos/{repo_full_name}',
                 headers={'Authorization': f'token {github_token}', 'Accept': 'application/vnd.github.v3+json'})
    return r.json()['id']

def _deploy_to_vercel(repo_full_name: str, project_name: str, env_vars: dict, vercel_token: str) -> dict:
    headers = {'Authorization': f'Bearer {vercel_token}', 'Content-Type': 'application/json'}
    team_qs = f'?teamId={_VERCEL_TEAM_ID}' if _VERCEL_TEAM_ID else ''
    r = _req.post(f'https://api.vercel.com/v9/projects{team_qs}', headers=headers, json={
        'name': project_name,
        'gitRepository': {'type': 'github', 'repo': repo_full_name},
        'framework': 'nextjs',
    })
    if r.status_code not in (200, 201):
        raise RuntimeError(f"Vercel create project failed: {r.text}")
    project_id = r.json()['id']
    env_list = [
        {'key': k, 'value': v, 'type': 'encrypted', 'target': ['production', 'preview', 'development']}
        for k, v in env_vars.items() if v
    ]
    if env_list:
        _req.post(f'https://api.vercel.com/v9/projects/{project_id}/env{team_qs}', headers=headers, json=env_list)
    repo_id = _get_github_repo_id(repo_full_name, _GITHUB_TOKEN)
    deploy_r = _req.post(f'https://api.vercel.com/v13/deployments{team_qs}', headers=headers, json={
        'name': project_name,
        'target': 'production',
        'gitSource': {'type': 'github', 'ref': 'main', 'repoId': repo_id},
    })
    deploy_url = ''
    if deploy_r.status_code in (200, 201, 202):
        raw = deploy_r.json().get('url', '')
        deploy_url = f"https://{raw}" if raw and not raw.startswith('http') else raw
    return {'id': project_id, 'url': deploy_url}


@app.route('/api/v1/admin/projet/<int:projet_id>/site-prefill', methods=['GET'])
@admin_required
def api_admin_projet_site_prefill(projet_id):
    conn = get_db_connection()
    projet = conn.execute("SELECT * FROM projets WHERE id=?", (projet_id,)).fetchone()
    if not projet:
        conn.close()
        return jsonify({'error': 'Projet introuvable'}), 404

    checklist = conn.execute("SELECT id FROM checklistes WHERE id_projet=?", (projet_id,)).fetchone()
    items = []
    if checklist:
        items = conn.execute(
            "SELECT nom_item, field_type, text_value FROM checklist_items WHERE id_checklist=? AND text_value IS NOT NULL AND text_value != ''",
            (checklist['id'],)
        ).fetchall()

    client = conn.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
    conn.close()

    # Normaliser les noms d'items pour le lookup
    item_map = {i['nom_item'].lower().strip(): i['text_value'] for i in items}

    def get(*keys):
        for k in keys:
            val = item_map.get(k.lower().strip())
            if val:
                return val
        return ''

    # Parsing adresse : "825 rue X ville Province"
    raw_addr = get('adresse', 'adresse complète')
    address, city, province = raw_addr, '', 'QC'
    if raw_addr:
        # Détection province
        prov_map = {'quebec': 'QC', 'qc': 'QC', 'ontario': 'ON', 'on': 'ON',
                    'colombie-britannique': 'BC', 'bc': 'BC', 'alberta': 'AB', 'ab': 'AB'}
        tokens = raw_addr.rsplit(None, 2)  # split depuis la droite
        if len(tokens) >= 2 and tokens[-1].lower().strip().rstrip('.') in prov_map:
            province = prov_map[tokens[-1].lower().strip().rstrip('.')]
            city = tokens[-2] if len(tokens) >= 2 else ''
            address = ' '.join(tokens[:-2]) if len(tokens) > 2 else raw_addr
        elif len(tokens) >= 3 and ' '.join(tokens[-2:]).lower().rstrip('.') in prov_map:
            province = prov_map[' '.join(tokens[-2:]).lower().rstrip('.')]
            city = tokens[-3] if len(tokens) >= 3 else ''
            address = ' '.join(tokens[:-3]) if len(tokens) > 3 else raw_addr

    # Owner depuis membres équipe
    owner_name, owner_title = '', ''
    try:
        membres_raw = item_map.get('équipe', item_map.get('equipe', ''))
        if membres_raw:
            membres = json.loads(membres_raw)
            if membres and isinstance(membres, list):
                owner_name = membres[0].get('name', '').strip()
                owner_title = membres[0].get('title', '').strip()
    except Exception:
        pass

    # Description : préférer "texte section a propos" sinon description
    description = get('texte section a propos', 'texte section à propos', 'description de l\'entreprise', 'description')

    # Tagline depuis mission
    tagline = get('mission de l\'entreprise', 'mission', 'accroche', 'tagline')

    return jsonify({
        'business_name':  get('nom du business', 'nom de l\'entreprise', 'nom'),
        'email':          get('courriel professionnel', 'courriel', 'email'),
        'client_email':   client['email'] if client else '',
        'phone':          get('telephone', 'téléphone'),
        'address':        address.strip(),
        'city':           city.strip(),
        'province':       province,
        'postal_code':    get('code postal'),
        'description':    description,
        'tagline':        tagline,
        'owner_name':     owner_name,
        'owner_title':    owner_title,
        'instagram':      get('instagram'),
        'facebook':       get('facebook'),
        'linkedin':       get('linkedin'),
    })


@app.route('/api/v1/admin/sites', methods=['GET'])
@admin_required
def api_admin_sites_list():
    conn = get_db_connection()
    sites = conn.execute(
        "SELECT id, business_name, template, slug, status, vercel_url, github_repo, sanity_project_id, created_at FROM sites ORDER BY created_at DESC"
    ).fetchall()
    conn.close()
    return jsonify([dict(s) for s in sites])


@app.route('/api/v1/admin/sites/<int:site_id>', methods=['GET'])
@admin_required
def api_admin_site_get(site_id):
    conn = get_db_connection()
    site = conn.execute("SELECT * FROM sites WHERE id=?", (site_id,)).fetchone()
    conn.close()
    if not site:
        return jsonify({'error': 'Site introuvable'}), 404
    return jsonify(dict(site))


@app.route('/api/v1/admin/sites/<int:site_id>/status', methods=['GET'])
@admin_required
def api_admin_site_status(site_id):
    conn = get_db_connection()
    site = conn.execute(
        "SELECT id, status, github_repo, sanity_project_id, vercel_project_id, vercel_url, error_message FROM sites WHERE id=?",
        (site_id,)
    ).fetchone()
    conn.close()
    if not site:
        return jsonify({'error': 'Site introuvable'}), 404
    return jsonify(dict(site))


TEST_CLIENT_EMAIL = 'felixrbk@gmail.com'

_SITE_CLONE_COLUMNS = [
    'template', 'slug', 'business_name', 'owner_name', 'owner_title', 'tagline', 'description',
    'address', 'city', 'province', 'postal_code', 'phone', 'email', 'acuity_url',
    'instagram', 'facebook', 'linkedin', 'hero_style',
    'seo_meta_title', 'seo_meta_description', 'seo_keywords',
    'seo_og_image', 'seo_twitter_handle', 'seo_logo_url', 'seo_business_type', 'seo_price_range',
    'resend_api_key', 'site_url', 'github_repo', 'sanity_project_id', 'vercel_project_id', 'vercel_url',
    'status', 'error_message', 'theme_font_pair', 'theme_accent_color', 'theme_bg_color',
    'style_variant', 'direction',
]

def _ensure_test_shadow_site(conn, source_site_id):
    """Clone un site vers le compte test TEST_CLIENT_EMAIL (même sanity_project_id/template) pour
       que ce compte puisse toujours prévisualiser/éditer n'importe quel site via /mon-site — sans
       toucher au client_email réel du site d'origine. Ne fait rien si le site source appartient
       déjà à ce compte test, ou si une copie test existe déjà pour lui."""
    row = conn.execute(
        f"SELECT {','.join(_SITE_CLONE_COLUMNS)}, client_email FROM sites WHERE id=?", (source_site_id,)
    ).fetchone()
    if not row or (row['client_email'] or '').lower() == TEST_CLIENT_EMAIL.lower():
        return
    existing = conn.execute(
        "SELECT id FROM sites WHERE LOWER(client_email)=LOWER(?) AND slug=? AND id != ?",
        (TEST_CLIENT_EMAIL, row['slug'], source_site_id)
    ).fetchone()
    if existing:
        return
    values = [row[c] for c in _SITE_CLONE_COLUMNS] + [TEST_CLIENT_EMAIL]
    placeholders = ','.join(['?'] * (len(_SITE_CLONE_COLUMNS) + 1))
    conn.execute(
        f"INSERT INTO sites ({','.join(_SITE_CLONE_COLUMNS)}, client_email) VALUES ({placeholders})",
        values
    )

@app.route('/api/v1/admin/sites/create', methods=['POST'])
@admin_required
def api_admin_sites_create():
    data = request.get_json() or {}
    required = ['template', 'business_name', 'owner_name', 'email', 'client_email']
    missing = [f for f in required if not str(data.get(f, '')).strip()]
    if missing:
        return jsonify({'error': f"Champs manquants: {', '.join(missing)}"}), 400
    template = data['template']
    if template not in ('reservation', 'vitrine'):
        return jsonify({'error': 'Template invalide'}), 400

    direction = (data.get('direction') or '').strip() or DEFAULT_DIRECTION
    if template == 'vitrine' and direction not in DIRECTIONS:
        return jsonify({'error': 'Direction artistique invalide'}), 400

    slug = _slugify_site(data['business_name'])

    if template == 'reservation':
        owner_title_db  = _json.dumps({'fr': data.get('owner_title_fr', ''), 'en': data.get('owner_title_en', '')})
        tagline_db      = _json.dumps({'fr': data.get('tagline_fr', ''), 'en': data.get('tagline_en', '')})
        description_db  = _json.dumps({'fr': data.get('description_fr', ''), 'en': data.get('description_en', '')})
        seo_title_db    = _json.dumps({'fr': data.get('seo_meta_title_fr', ''), 'en': data.get('seo_meta_title_en', '')})
        seo_desc_db     = _json.dumps({'fr': data.get('seo_meta_description_fr', ''), 'en': data.get('seo_meta_description_en', '')})
        seo_kw_db       = _json.dumps({'fr': data.get('seo_keywords_fr', ''), 'en': data.get('seo_keywords_en', '')})
    else:
        owner_title_db = data.get('owner_title', '')
        tagline_db     = data.get('tagline', '')
        description_db = data.get('description', '')
        seo_title_db   = data.get('seo_meta_title', '')
        seo_desc_db    = data.get('seo_meta_description', '')
        seo_kw_db      = data.get('seo_keywords', '')

    conn = get_db_connection()
    try:
        _id_projet = data.get('id_projet') or None
        cur = conn.execute("""
            INSERT INTO sites (template, slug, business_name, owner_name, owner_title, tagline, description,
                address, city, province, postal_code, phone, email, acuity_url,
                instagram, facebook, linkedin, hero_style, direction,
                seo_meta_title, seo_meta_description, seo_keywords,
                seo_og_image, seo_twitter_handle, seo_logo_url, seo_business_type, seo_price_range,
                client_email, resend_api_key, site_url, id_projet, status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'creating')
        """, (
            template, slug, data['business_name'], data['owner_name'],
            owner_title_db, tagline_db, description_db,
            data.get('address'), data.get('city'), data.get('province', 'QC'), data.get('postal_code'),
            data.get('phone'), data['email'], data.get('acuity_url'),
            data.get('instagram'), data.get('facebook'), data.get('linkedin'),
            data.get('hero_style', 'A' if template == 'reservation' else 'luxe'), direction,
            seo_title_db, seo_desc_db, seo_kw_db,
            data.get('seo_og_image'), data.get('seo_twitter_handle'), data.get('seo_logo_url'),
            data.get('seo_business_type'), data.get('seo_price_range', '$$'),
            data['client_email'], data.get('resend_api_key', ''), data.get('site_url', ''), _id_projet,
        ))
        site_id = cur.lastrowid
        conn.commit()
    finally:
        conn.close()

    def _build():
        try:
            config_ts  = _generate_config_ts(template, data)
            repo_info  = _create_github_repo(slug, _GITHUB_TOKEN)
            repo_name  = repo_info['full_name']
            _push_template_to_repo(template, slug, config_ts, _GITHUB_TOKEN, repo_name)
            sanity_id  = _create_sanity_project(data['business_name'], _SANITY_TOKEN)
            _invite_sanity_member(sanity_id, data['client_email'], _SANITY_TOKEN)
            if template == 'vitrine':
                _seed_vitrine_baseline_defaults(sanity_id, direction)
                if _id_projet:
                    _push_checklist_content_to_sanity(_id_projet, sanity_id)
            conn2 = get_db_connection()
            conn2.execute(
                "UPDATE sites SET github_repo=?, sanity_project_id=?, status='active', updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (repo_name, sanity_id, site_id)
            )
            _ensure_test_shadow_site(conn2, site_id)
            conn2.commit()
            conn2.close()
        except Exception as exc:
            print(f"[SITES] build error site {site_id}: {exc}")
            conn2 = get_db_connection()
            conn2.execute("UPDATE sites SET status='error', error_message=? WHERE id=?", (str(exc), site_id))
            conn2.commit()
            conn2.close()

    _threading.Thread(target=_build, daemon=True).start()
    return jsonify({'id': site_id, 'slug': slug, 'status': 'creating'}), 202


@app.route('/api/v1/admin/sites/<int:site_id>/deploy', methods=['POST'])
@admin_required
def api_admin_site_deploy(site_id):
    conn = get_db_connection()
    site = conn.execute("SELECT * FROM sites WHERE id=?", (site_id,)).fetchone()
    conn.close()
    if not site:
        return jsonify({'error': 'Site introuvable'}), 404
    if not site['github_repo']:
        return jsonify({'error': 'Repo GitHub non encore créé'}), 400

    body = request.get_json() or {}
    site_url   = body.get('site_url', '') or site['site_url'] or ''
    resend_key = body.get('resend_api_key', '') or site['resend_api_key'] or ''
    site_url_env = f"https://{site_url}" if site_url and not re.match(r'^https?://', site_url) else site_url

    env_vars = {
        'NEXT_PUBLIC_SANITY_PROJECT_ID': site['sanity_project_id'] or '',
        'NEXT_PUBLIC_SANITY_DATASET': 'production',
        'NEXT_PUBLIC_SITE_URL': site_url_env,
        'RESEND_API_KEY': resend_key,
        'SANITY_STUDIO_PROJECT_ID': site['sanity_project_id'] or '',
        'SANITY_STUDIO_DATASET': 'production',
    }
    if site['template'] == 'reservation' and site['acuity_url']:
        env_vars['NEXT_PUBLIC_ACUITY_URL'] = site['acuity_url']

    project_name = f"site-{site['slug']}"
    try:
        vercel = _deploy_to_vercel(site['github_repo'], project_name, env_vars, _VERCEL_TOKEN)
        conn2 = get_db_connection()
        conn2.execute(
            "UPDATE sites SET vercel_project_id=?, vercel_url=?, site_url=?, resend_api_key=?, status='deployed', updated_at=CURRENT_TIMESTAMP WHERE id=?",
            (vercel['id'], vercel['url'], site_url, resend_key, site_id)
        )
        conn2.commit()
        conn2.close()

        # Site déployé et lié à un projet : le site étant préparé, les travaux peuvent
        # démarrer automatiquement (facture, todos, courriel) — sauf si le projet a déjà
        # dépassé cette étape (évite de redéclencher au redéploiement d'un site existant).
        if site['id_projet']:
            try:
                conn3 = get_db_connection()
                projet = conn3.execute("SELECT * FROM projets WHERE id=?", (site['id_projet'],)).fetchone()
                if projet and projet['statut'] in ('En attente de rendez-vous', 'Documents à donner', 'Documents reçus'):
                    client = conn3.execute("SELECT * FROM clients WHERE id=?", (projet['id_client'],)).fetchone()
                    _do_start_travaux(conn3, projet, client)
                conn3.close()
            except Exception as e:
                print(f"[SITES] Auto-start travaux après déploiement échoué: {e}")

        return jsonify({'vercel_id': vercel['id'], 'url': vercel['url']})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


_ASSET_SLOTS = {
    'logo-icone':             'public/logo-icone',
    'logo-texte':             'public/logo-texte',
    'logo-complet':           'public/logo-complet',
    'favicon':                'public/favicon',
    'hero':                   'public/hero_section_picture',
    'approche':               'public/approche-soin',
}

# Slot spécial « banque photo » : dépose le fichier dans public/services/
# en CONSERVANT son nom d'origine (aucun renommage, contrairement aux slots ci-dessus).
_ASSET_BANK_SLOT = 'banque-photo'
_ASSET_BANK_DIR  = 'public/services'

def _github_push_file(repo: str, path: str, content_bytes: bytes, message: str, github_token: str):
    """Pousse un fichier via la Git Data API (supporte les binaires lourds)."""
    import base64 as _b64
    headers = {'Authorization': f'token {github_token}', 'Accept': 'application/vnd.github.v3+json'}
    base = f'https://api.github.com/repos/{repo}'

    # 1. Créer un blob
    blob_r = _req.post(f'{base}/git/blobs', headers=headers, json={
        'content': _b64.b64encode(content_bytes).decode(),
        'encoding': 'base64',
    })
    if blob_r.status_code not in (200, 201):
        raise RuntimeError(f"blob: {blob_r.text}")
    blob_sha = blob_r.json()['sha']

    # 2. Récupérer le commit HEAD de main
    ref_r = _req.get(f'{base}/git/ref/heads/main', headers=headers)
    if ref_r.status_code != 200:
        raise RuntimeError(f"ref: {ref_r.text}")
    head_sha = ref_r.json()['object']['sha']

    # 3. Récupérer le tree du commit HEAD
    commit_r = _req.get(f'{base}/git/commits/{head_sha}', headers=headers)
    if commit_r.status_code != 200:
        raise RuntimeError(f"commit: {commit_r.text}")
    base_tree_sha = commit_r.json()['tree']['sha']

    # 4. Créer un nouveau tree avec le fichier
    tree_r = _req.post(f'{base}/git/trees', headers=headers, json={
        'base_tree': base_tree_sha,
        'tree': [{'path': path, 'mode': '100644', 'type': 'blob', 'sha': blob_sha}],
    })
    if tree_r.status_code not in (200, 201):
        raise RuntimeError(f"tree: {tree_r.text}")
    new_tree_sha = tree_r.json()['sha']

    # 5. Créer un commit
    new_commit_r = _req.post(f'{base}/git/commits', headers=headers, json={
        'message': message,
        'tree': new_tree_sha,
        'parents': [head_sha],
        'author': {'name': 'Cocktail Média', 'email': 'felix.dumont@cocktailmedia.ca'},
    })
    if new_commit_r.status_code not in (200, 201):
        raise RuntimeError(f"new commit: {new_commit_r.text}")
    new_commit_sha = new_commit_r.json()['sha']

    # 6. Mettre à jour la ref main
    update_r = _req.patch(f'{base}/git/refs/heads/main', headers=headers, json={'sha': new_commit_sha})
    if update_r.status_code not in (200, 201):
        raise RuntimeError(f"update ref: {update_r.text}")


@app.route('/api/v1/admin/sites/<int:site_id>/assets', methods=['POST'])
@admin_required
def api_admin_site_assets(site_id):
    import traceback as _tb
    try:
        print(f"[ASSETS] Requête reçue site_id={site_id} content_length={request.content_length}", flush=True)
        conn = get_db_connection()
        site = conn.execute("SELECT github_repo FROM sites WHERE id=?", (site_id,)).fetchone()
        conn.close()
        if not site or not site['github_repo']:
            return jsonify({'error': 'Repo GitHub non disponible'}), 400

        print("[ASSETS] lecture form", flush=True)
        slot = request.form.get('slot')
        print(f"[ASSETS] slot={slot}", flush=True)
        if slot != _ASSET_BANK_SLOT and slot not in _ASSET_SLOTS:
            return jsonify({'error': f"Slot invalide. Options: {', '.join(list(_ASSET_SLOTS) + [_ASSET_BANK_SLOT])}"}), 400

        print("[ASSETS] lecture fichier", flush=True)
        file = request.files.get('file')
        if not file or not file.filename:
            return jsonify({'error': 'Aucun fichier fourni'}), 400
        print(f"[ASSETS] fichier={file.filename}", flush=True)

        if slot == _ASSET_BANK_SLOT:
            # Banque photo : on conserve EXACTEMENT le nom d'origine (espaces et accents inclus).
            # On ne garde que le nom de base pour empêcher toute traversée de chemin.
            original = os.path.basename(file.filename.replace('\\', '/'))
            if not original or original.startswith('.') or '..' in original:
                return jsonify({'error': 'Nom de fichier invalide'}), 400
            gh_path    = f"{_ASSET_BANK_DIR}/{original}"
            commit_msg = f'assets: ajout banque photo {original}'
        else:
            ext        = os.path.splitext(file.filename)[1].lower() or '.bin'
            gh_path    = f"{_ASSET_SLOTS[slot]}{ext}"
            commit_msg = f'assets: upload {slot}'
        repo     = site['github_repo']

        print(f"[ASSETS] → {gh_path}", flush=True)
        content_bytes = file.read()
        print(f"[ASSETS] {len(content_bytes)} bytes lus, push GitHub…", flush=True)

        _github_push_file(
            repo=repo,
            path=gh_path,
            content_bytes=content_bytes,
            message=commit_msg,
            github_token=_GITHUB_TOKEN,
        )
        print("[ASSETS] succès", flush=True)
        return jsonify({'success': True, 'path': gh_path})
    except BaseException as exc:
        print(f"[ASSETS ERROR] {type(exc).__name__}: {exc}\n{_tb.format_exc()}", flush=True)
        return jsonify({'error': str(exc)}), 500


@app.route('/api/v1/admin/sites/<int:site_id>/commit', methods=['POST'])
@admin_required
def api_admin_site_commit(site_id):
    conn = get_db_connection()
    site = conn.execute("SELECT github_repo FROM sites WHERE id=?", (site_id,)).fetchone()
    conn.close()
    if not site or not site['github_repo']:
        return jsonify({'error': 'Repo GitHub non disponible'}), 400
    try:
        from datetime import datetime as _dt
        ts = _dt.utcnow().strftime('%Y-%m-%d %H:%M UTC')
        _github_push_file(
            repo=site['github_repo'],
            path='.deploy',
            content_bytes=ts.encode(),
            message=f'deploy: trigger redeploy {ts}',
            github_token=_GITHUB_TOKEN,
        )
        return jsonify({'success': True, 'ts': ts})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.route('/api/v1/admin/sites/<int:site_id>', methods=['DELETE'])
@admin_required
def api_admin_site_delete(site_id):
    conn = get_db_connection()
    site = conn.execute("SELECT id FROM sites WHERE id=?", (site_id,)).fetchone()
    if not site:
        conn.close()
        return jsonify({'error': 'Site introuvable'}), 404
    conn.execute("DELETE FROM sites WHERE id=?", (site_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ───────────────────────────────────────────────────────────
# API v1 — Soumissions
# ───────────────────────────────────────────────────────────

@app.route('/api/v1/admin/soumissions', methods=['GET'])
@admin_required
def api_admin_list_soumissions():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT s.id, s.titre, s.statut, s.date_expiration, s.created_at,
               c.nom_complet AS nom_client, c.email AS email_client,
               so.nom AS option_acceptee_nom
        FROM soumissions s
        LEFT JOIN clients c ON c.id = s.id_client
        LEFT JOIN soumission_options so ON so.id = s.option_acceptee_id
        ORDER BY s.created_at DESC
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/v1/admin/soumission/creer', methods=['POST'])
@admin_required
def api_admin_creer_soumission():
    import json as _json
    data = request.get_json(force=True) or {}
    id_client = data.get('id_client')
    titre = (data.get('titre') or '').strip()
    message_intro = (data.get('message_intro') or '').strip() or None
    date_expiration = data.get('date_expiration') or None
    options = data.get('options') or []

    if not id_client or not titre:
        return jsonify({'error': 'id_client et titre requis'}), 400

    conn = get_db_connection()
    client = conn.execute("SELECT id, nom_complet, email, is_email_confirmed FROM clients WHERE id=?", (id_client,)).fetchone()
    if not client:
        conn.close()
        return jsonify({'error': 'Client introuvable'}), 404

    conn.execute("""
        INSERT INTO soumissions (id_client, titre, message_intro, statut, date_expiration)
        VALUES (?, ?, ?, 'envoyee', ?)
    """, (id_client, titre, message_intro, date_expiration))
    soumission_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    for idx, opt in enumerate(options):
        conn.execute("""
            INSERT INTO soumission_options
                (id_soumission, nom, description, prix_setup, prix_mensuel, prix_horaire,
                 delai_livraison, conditions_paiement, inclus_json, couts_tiers_json,
                 couts_supplementaires_json, scenarios_json, est_recommande, badge_texte, ordre,
                 features_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            soumission_id,
            (opt.get('nom') or '').strip(),
            (opt.get('description') or '').strip() or None,
            float(opt.get('prix_setup') or 0),
            float(opt.get('prix_mensuel') or 0),
            float(opt.get('prix_horaire') or 0),
            (opt.get('delai_livraison') or '').strip() or None,
            (opt.get('conditions_paiement') or '').strip() or None,
            _json.dumps(opt.get('inclus') or [], ensure_ascii=False),
            _json.dumps(opt.get('couts_tiers') or [], ensure_ascii=False),
            _json.dumps(opt.get('couts_supplementaires') or [], ensure_ascii=False),
            _json.dumps(opt.get('scenarios') or [], ensure_ascii=False),
            1 if opt.get('est_recommande') else 0,
            (opt.get('badge_texte') or '').strip() or None,
            idx,
            _json.dumps(opt.get('features') or {}, ensure_ascii=False),
        ))
    conn.commit()

    # Génération PDF soumission
    import pathlib as _pl
    import soumission_service as _soum_svc
    soum_pdf_path = None
    try:
        upload_root = os.getenv("UPLOAD_ROOT", "/data/uploads")
        soum_dir = os.path.join(upload_root, "soumissions", f"client_{id_client}")
        _pl.Path(soum_dir).mkdir(parents=True, exist_ok=True)
        soum_pdf_path = os.path.join(soum_dir, f"SOUM-{soumission_id}.pdf")
        opts_rows = conn.execute(
            "SELECT * FROM soumission_options WHERE id_soumission=? ORDER BY ordre, id",
            (soumission_id,)
        ).fetchall()
        opts_dicts = []
        for o in opts_rows:
            d = dict(o)
            for k in ('inclus_json', 'couts_tiers_json', 'couts_supplementaires_json', 'scenarios_json'):
                try:
                    d[k] = _json.loads(d[k]) if d[k] else []
                except Exception:
                    d[k] = []
            try:
                d['features_json'] = _json.loads(d.get('features_json') or '{}') or {}
            except Exception:
                d['features_json'] = {}
            opts_dicts.append(d)
        soum_dict = {
            'id': soumission_id,
            'titre': titre,
            'message_intro': message_intro,
            'date_expiration': date_expiration,
        }
        _soum_svc.generer_pdf_soumission(soum_dict, opts_dicts, dict(client), soum_pdf_path)
        conn.execute("UPDATE soumissions SET pdf_path=? WHERE id=?", (soum_pdf_path, soumission_id))
        conn.commit()
    except Exception as e:
        print(f"[PDF] Erreur génération soumission: {e}")
        soum_pdf_path = None

    settings = get_notification_settings()
    lien = f"{PORTAIL_URL}/soumission/{soumission_id}"
    from email_templates import email_soumission_disponible
    html = email_soumission_disponible(client['nom_complet'], titre, lien)
    send_email_client(
        dict(client),
        f"Votre soumission est disponible — {titre}",
        f"Bonjour {client['nom_complet']}, votre soumission \"{titre}\" est disponible dans votre portail : {lien}",
        html=html,
    )
    conn.close()
    return jsonify({'id': soumission_id}), 201


@app.route('/api/v1/admin/soumission/<int:soumission_id>', methods=['GET'])
@admin_required
def api_admin_get_soumission(soumission_id):
    import json as _json
    conn = get_db_connection()
    s = conn.execute("""
        SELECT s.*, c.nom_complet AS nom_client, c.email AS email_client,
               c.telephone AS telephone_client
        FROM soumissions s
        LEFT JOIN clients c ON c.id = s.id_client
        WHERE s.id = ?
    """, (soumission_id,)).fetchone()
    if not s:
        conn.close()
        return jsonify({'error': 'Soumission introuvable'}), 404
    opts = conn.execute(
        "SELECT * FROM soumission_options WHERE id_soumission=? ORDER BY ordre, id",
        (soumission_id,)
    ).fetchall()
    conn.close()

    def _parse(row):
        d = dict(row)
        for k in ('inclus_json', 'couts_tiers_json', 'couts_supplementaires_json', 'scenarios_json'):
            try:
                d[k] = _json.loads(d[k]) if d[k] else []
            except Exception:
                d[k] = []
        try:
            d['features_json'] = _json.loads(d.get('features_json') or '{}') or {}
        except Exception:
            d['features_json'] = {}
        return d

    return jsonify({
        **dict(s),
        'options': [_parse(o) for o in opts],
    })


@app.route('/api/v1/admin/soumission/<int:soumission_id>', methods=['DELETE'])
@admin_required
def api_admin_delete_soumission(soumission_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM soumissions WHERE id=?", (soumission_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/admin/soumission/<int:soumission_id>/renvoyer', methods=['POST'])
@admin_required
def api_admin_renvoyer_soumission(soumission_id):
    conn = get_db_connection()
    row = conn.execute(
        "SELECT s.id, s.titre, s.id_client, c.nom_complet, c.email, c.is_email_confirmed "
        "FROM soumissions s JOIN clients c ON c.id = s.id_client WHERE s.id=?",
        (soumission_id,)
    ).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Soumission introuvable'}), 404
    lien = f"{PORTAIL_URL}/soumission/{soumission_id}"
    from email_templates import email_soumission_disponible
    html = email_soumission_disponible(row['nom_complet'], row['titre'], lien)
    send_email_client(
        dict(row),
        f"Votre soumission est disponible — {row['titre']}",
        f"Bonjour {row['nom_complet']}, votre soumission \"{row['titre']}\" est disponible dans votre portail : {lien}",
        html=html
    )
    return jsonify({'success': True})


@app.route('/api/v1/soumission/<int:soumission_id>', methods=['GET'])
@login_required
def api_client_get_soumission(soumission_id):
    import json as _json
    from datetime import datetime as _dt
    user_id = session.get('user_id')
    conn = get_db_connection()
    s = conn.execute(
        "SELECT s.*, c.nom_complet as nom_client FROM soumissions s JOIN clients c ON c.id = s.id_client WHERE s.id=? AND s.id_client=?",
        (soumission_id, user_id)
    ).fetchone()
    if not s:
        conn.close()
        return jsonify({'error': 'Soumission introuvable'}), 404

    statut = s['statut']
    if statut == 'envoyee' and s['date_expiration']:
        try:
            exp = _dt.strptime(s['date_expiration'], '%Y-%m-%d').date()
            if exp < _dt.utcnow().date():
                conn.execute("UPDATE soumissions SET statut='expiree', updated_at=CURRENT_TIMESTAMP WHERE id=?", (soumission_id,))
                conn.commit()
                statut = 'expiree'
        except Exception:
            pass

    opts = conn.execute(
        "SELECT * FROM soumission_options WHERE id_soumission=? ORDER BY ordre, id",
        (soumission_id,)
    ).fetchall()
    conn.close()

    def _parse(row):
        d = dict(row)
        for k in ('inclus_json', 'couts_tiers_json', 'couts_supplementaires_json', 'scenarios_json'):
            try:
                d[k] = _json.loads(d[k]) if d[k] else []
            except Exception:
                d[k] = []
        try:
            d['features_json'] = _json.loads(d.get('features_json') or '{}') or {}
        except Exception:
            d['features_json'] = {}
        return d

    result = dict(s)
    result['statut'] = statut
    result['options'] = [_parse(o) for o in opts]
    return jsonify(result)


@app.route('/api/v1/soumission/<int:soumission_id>/accepter', methods=['POST'])
@login_required
def api_client_accepter_soumission(soumission_id):
    from datetime import datetime as _dt
    user_id = session.get('user_id')
    data = request.get_json(force=True) or {}
    id_option = data.get('id_option')
    extras_selectionnes = data.get('extras_selectionnes') or []
    if not id_option:
        return jsonify({'error': 'id_option requis'}), 400

    conn = get_db_connection()
    s = conn.execute(
        "SELECT s.*, c.nom_complet, c.email FROM soumissions s JOIN clients c ON c.id=s.id_client WHERE s.id=? AND s.id_client=?",
        (soumission_id, user_id)
    ).fetchone()
    if not s:
        conn.close()
        return jsonify({'error': 'Soumission introuvable'}), 404
    if s['statut'] != 'envoyee':
        conn.close()
        return jsonify({'error': f"Soumission non modifiable (statut: {s['statut']})"}), 409

    opt = conn.execute(
        "SELECT * FROM soumission_options WHERE id=? AND id_soumission=?",
        (id_option, soumission_id)
    ).fetchone()
    if not opt:
        conn.close()
        return jsonify({'error': "Option introuvable"}), 404

    now_str = _dt.now().strftime('%Y-%m-%d %H:%M:%S')
    conn.execute("""
        UPDATE soumissions
        SET statut='acceptee', option_acceptee_id=?, updated_at=CURRENT_TIMESTAMP
        WHERE id=?
    """, (id_option, soumission_id))
    conn.commit()

    settings = get_notification_settings()
    admin_emails = settings.get('admin_emails') or []
    lien_admin = f"{PORTAIL_URL}/admin/soumissions/{soumission_id}"
    now_affiche = _dt.now().strftime('%d %B %Y a %H:%M')

    from email_templates import email_soumission_acceptee_admin
    html = email_soumission_acceptee_admin(
        nom_client=s['nom_complet'],
        email_client=s['email'],
        titre_soumission=s['titre'],
        nom_option=opt['nom'],
        prix_setup=float(opt['prix_setup'] or 0),
        prix_mensuel=float(opt['prix_mensuel'] or 0),
        date_acceptation=now_affiche,
        lien_admin=lien_admin,
        extras_selectionnes=extras_selectionnes or [],
    )
    sujet = f"[CocktailOS] Soumission acceptee - {s['nom_complet']} - {opt['nom']}"
    extras_plain = ''
    if extras_selectionnes:
        lignes = '\n'.join(f"  - {e.get('situation','')}: {e.get('cout','')}" for e in extras_selectionnes)
        extras_plain = f"\nExtras sélectionnés par le client :\n{lignes}"
    plain = (
        f"Client : {s['nom_complet']} ({s['email']})\n"
        f"Soumission : {s['titre']}\n"
        f"Option choisie : {opt['nom']}\n"
        f"Prix setup : {opt['prix_setup']:.2f} $\n"
        f"Prix mensuel : {opt['prix_mensuel']:.2f} $\n"
        f"Date : {now_affiche}\n"
        f"Lien : {lien_admin}"
        f"{extras_plain}"
    )
    if admin_emails:
        send_email(admin_emails, sujet, plain, html=html)
    push_admin_notif(
        conn,
        titre=f"Soumission acceptée — {s['nom_complet']}",
        message=f"{s['titre']} — option « {opt['nom']} »",
        type='info',
        lien=f"/admin/soumissions/{soumission_id}",
    )
    conn.commit()

    conn.close()
    return jsonify({'success': True, 'statut': 'acceptee'})


@app.route('/api/v1/soumission/<int:soumission_id>/pdf', methods=['GET'])
@login_required
def api_client_download_soumission_pdf(soumission_id):
    user_id = session.get('user_id')
    conn = get_db_connection()
    s = conn.execute(
        "SELECT s.pdf_path, s.titre, c.nom_complet FROM soumissions s JOIN clients c ON c.id = s.id_client WHERE s.id=? AND s.id_client=?",
        (soumission_id, user_id)
    ).fetchone()
    conn.close()
    if not s:
        return jsonify({'error': 'Soumission introuvable'}), 404
    pdf_path = s['pdf_path']
    if not pdf_path or not os.path.exists(pdf_path):
        return jsonify({'error': 'PDF non disponible'}), 404
    nom_client = (s['nom_complet'] or s['titre'] or f"SOUM-{soumission_id}").strip()
    safe_name = "".join(c for c in nom_client if c.isalnum() or c in " _-").strip()
    download_name = f"{safe_name} — Soumission Cocktail Média.pdf"
    return send_file(pdf_path, as_attachment=True, download_name=download_name)


@app.route('/api/v1/client/soumissions', methods=['GET'])
@login_required
def api_client_list_soumissions():
    from datetime import datetime as _dt
    user_id = session.get('user_id')
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT s.id, s.titre, s.statut, s.date_expiration, s.created_at,
               so.nom AS option_acceptee_nom
        FROM soumissions s
        LEFT JOIN soumission_options so ON so.id = s.option_acceptee_id
        WHERE s.id_client = ?
        ORDER BY s.created_at DESC
    """, (user_id,)).fetchall()
    result = []
    for r in rows:
        row = dict(r)
        if row['statut'] == 'envoyee' and row['date_expiration']:
            try:
                exp = _dt.strptime(row['date_expiration'], '%Y-%m-%d').date()
                if exp < _dt.utcnow().date():
                    conn.execute("UPDATE soumissions SET statut='expiree', updated_at=CURRENT_TIMESTAMP WHERE id=?", (r['id'],))
                    row['statut'] = 'expiree'
            except Exception:
                pass
        result.append(row)
    if any(r['statut'] == 'expiree' for r in result):
        conn.commit()
    conn.close()
    return jsonify(result)


# ─── SOUMISSION TEMPLATES ────────────────────────────────────────────────────

@app.route('/api/v1/admin/soumissions/templates', methods=['GET'])
@admin_required
def api_admin_list_templates():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT t.id, t.nom, t.description, t.titre_template, t.est_actif, t.created_at,
               COUNT(o.id) AS nb_options
        FROM soumission_templates t
        LEFT JOIN soumission_template_options o ON o.id_template = t.id
        GROUP BY t.id
        ORDER BY t.nom
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/v1/admin/soumissions/templates/<int:template_id>', methods=['GET'])
@admin_required
def api_admin_get_template(template_id):
    conn = get_db_connection()
    t = conn.execute("SELECT * FROM soumission_templates WHERE id = ?", (template_id,)).fetchone()
    if not t:
        conn.close()
        return jsonify({'error': 'Template introuvable'}), 404
    options = conn.execute(
        "SELECT * FROM soumission_template_options WHERE id_template = ? ORDER BY ordre",
        (template_id,)
    ).fetchall()
    conn.close()
    data = dict(t)
    data['options'] = [dict(o) for o in options]
    return jsonify(data)


@app.route('/api/v1/admin/soumissions/templates', methods=['POST'])
@admin_required
def api_admin_creer_template():
    body = request.get_json(force=True) or {}
    nom = (body.get('nom') or '').strip()
    if not nom:
        return jsonify({'error': 'Le champ nom est requis'}), 400
    conn = get_db_connection()
    cur = conn.execute("""
        INSERT INTO soumission_templates
            (nom, description, message_intro_template, titre_template, est_actif)
        VALUES (?, ?, ?, ?, ?)
    """, (
        nom,
        body.get('description'),
        body.get('message_intro_template'),
        body.get('titre_template') or 'Soumission - {nom_entreprise}',
        1 if body.get('est_actif', True) else 0,
    ))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'message': 'Template cree'}), 201


@app.route('/api/v1/admin/soumissions/templates/<int:template_id>', methods=['PUT'])
@admin_required
def api_admin_update_template(template_id):
    body = request.get_json(force=True) or {}
    conn = get_db_connection()
    t = conn.execute("SELECT id FROM soumission_templates WHERE id = ?", (template_id,)).fetchone()
    if not t:
        conn.close()
        return jsonify({'error': 'Template introuvable'}), 404
    fields = []
    values = []
    for col in ('nom', 'description', 'message_intro_template', 'titre_template'):
        if col in body:
            fields.append(f"{col} = ?")
            values.append(body[col])
    if 'est_actif' in body:
        fields.append("est_actif = ?")
        values.append(1 if body['est_actif'] else 0)
    if not fields:
        conn.close()
        return jsonify({'error': 'Aucun champ a mettre a jour'}), 400
    fields.append("updated_at = CURRENT_TIMESTAMP")
    values.append(template_id)
    conn.execute(f"UPDATE soumission_templates SET {', '.join(fields)} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return jsonify({'message': 'Template mis a jour'})


@app.route('/api/v1/admin/soumissions/templates/<int:template_id>', methods=['DELETE'])
@admin_required
def api_admin_delete_template(template_id):
    conn = get_db_connection()
    t = conn.execute("SELECT id FROM soumission_templates WHERE id = ?", (template_id,)).fetchone()
    if not t:
        conn.close()
        return jsonify({'error': 'Template introuvable'}), 404
    conn.execute("DELETE FROM soumission_templates WHERE id = ?", (template_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Template supprime'})


@app.route('/api/v1/admin/soumissions/templates/<int:template_id>/options', methods=['POST'])
@admin_required
def api_admin_creer_template_option(template_id):
    conn = get_db_connection()
    t = conn.execute("SELECT id FROM soumission_templates WHERE id = ?", (template_id,)).fetchone()
    if not t:
        conn.close()
        return jsonify({'error': 'Template introuvable'}), 404
    body = request.get_json(force=True) or {}
    nom = (body.get('nom') or '').strip()
    if not nom:
        conn.close()
        return jsonify({'error': 'Le champ nom est requis'}), 400
    import json as _json
    cur = conn.execute("""
        INSERT INTO soumission_template_options
            (id_template, nom, description, prix_setup, prix_mensuel, prix_horaire,
             delai_livraison, conditions_paiement, badge_texte, est_recommande, ordre,
             features_json, inclus_json, couts_tiers_json, couts_supplementaires_json,
             scenarios_json, rachat_disponible, prix_rachat, inclus_rachat_json)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        template_id,
        nom,
        body.get('description'),
        float(body.get('prix_setup') or 0),
        float(body.get('prix_mensuel') or 0),
        float(body.get('prix_horaire') or 0),
        body.get('delai_livraison'),
        body.get('conditions_paiement'),
        body.get('badge_texte'),
        1 if body.get('est_recommande') else 0,
        int(body.get('ordre') or 0),
        body.get('features_json') or '{}',
        body.get('inclus_json') or '[]',
        body.get('couts_tiers_json') or '[]',
        body.get('couts_supplementaires_json') or '[]',
        body.get('scenarios_json') or '[]',
        1 if body.get('rachat_disponible') else 0,
        float(body.get('prix_rachat') or 0),
        body.get('inclus_rachat_json') or '[]',
    ))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({'id': new_id, 'message': 'Option creee'}), 201


@app.route('/api/v1/admin/soumissions/templates/<int:template_id>/options/<int:option_id>', methods=['PUT'])
@admin_required
def api_admin_update_template_option(template_id, option_id):
    conn = get_db_connection()
    o = conn.execute(
        "SELECT id FROM soumission_template_options WHERE id = ? AND id_template = ?",
        (option_id, template_id)
    ).fetchone()
    if not o:
        conn.close()
        return jsonify({'error': 'Option introuvable'}), 404
    body = request.get_json(force=True) or {}
    fields = []
    values = []
    scalar_cols = ('nom', 'description', 'delai_livraison', 'conditions_paiement',
                   'badge_texte', 'features_json', 'inclus_json', 'couts_tiers_json',
                   'couts_supplementaires_json', 'scenarios_json', 'inclus_rachat_json')
    for col in scalar_cols:
        if col in body:
            fields.append(f"{col} = ?")
            values.append(body[col])
    for col in ('prix_setup', 'prix_mensuel', 'prix_horaire', 'prix_rachat'):
        if col in body:
            fields.append(f"{col} = ?")
            values.append(float(body[col]))
    for col in ('ordre',):
        if col in body:
            fields.append(f"{col} = ?")
            values.append(int(body[col]))
    for col in ('est_recommande', 'rachat_disponible'):
        if col in body:
            fields.append(f"{col} = ?")
            values.append(1 if body[col] else 0)
    if not fields:
        conn.close()
        return jsonify({'error': 'Aucun champ a mettre a jour'}), 400
    values.append(option_id)
    conn.execute(f"UPDATE soumission_template_options SET {', '.join(fields)} WHERE id = ?", values)
    conn.commit()
    conn.close()
    return jsonify({'message': 'Option mise a jour'})


@app.route('/api/v1/admin/soumissions/templates/<int:template_id>/options/<int:option_id>', methods=['DELETE'])
@admin_required
def api_admin_delete_template_option(template_id, option_id):
    conn = get_db_connection()
    o = conn.execute(
        "SELECT id FROM soumission_template_options WHERE id = ? AND id_template = ?",
        (option_id, template_id)
    ).fetchone()
    if not o:
        conn.close()
        return jsonify({'error': 'Option introuvable'}), 404
    conn.execute("DELETE FROM soumission_template_options WHERE id = ?", (option_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Option supprimee'})


@app.route('/api/v1/admin/soumissions/templates/<int:template_id>/appliquer/<int:id_client>', methods=['GET'])
@admin_required
def api_admin_appliquer_template(template_id, id_client):
    conn = get_db_connection()
    t = conn.execute("SELECT * FROM soumission_templates WHERE id = ?", (template_id,)).fetchone()
    if not t:
        conn.close()
        return jsonify({'error': 'Template introuvable'}), 404
    client = conn.execute(
        "SELECT id, nom_complet, nom_entreprise FROM clients WHERE id = ?", (id_client,)
    ).fetchone()
    if not client:
        conn.close()
        return jsonify({'error': 'Client introuvable'}), 404
    options = conn.execute(
        "SELECT * FROM soumission_template_options WHERE id_template = ? ORDER BY ordre",
        (template_id,)
    ).fetchall()
    conn.close()
    nom_complet = client['nom_complet'] or ''
    nom_entreprise = client['nom_entreprise'] or nom_complet
    titre = t['titre_template'].replace('{nom_entreprise}', nom_entreprise).replace('{nom_complet}', nom_complet)
    message_intro = (t['message_intro_template'] or '').replace('{nom_entreprise}', nom_entreprise).replace('{nom_complet}', nom_complet)
    return jsonify({
        'titre': titre,
        'message_intro': message_intro,
        'options': [dict(o) for o in options],
    })


# ───────────────────────────────────────────────────────────
# Comptabilité — Paramètres de facturation de l'abonné
# ───────────────────────────────────────────────────────────
@app.route('/api/v1/admin/parametres-facturation', methods=['GET'])
@admin_required
def api_admin_get_parametres_facturation():
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM parametres_facturation WHERE id = 1").fetchone()
    conn.close()
    if not row:
        return jsonify({
            'charge_taxes': False, 'neq': '', 'numero_tps': '', 'numero_tvq': '',
            'nom_entreprise': '', 'couleur_marque': '#c0321a'
        })
    return jsonify({
        'charge_taxes': bool(row['charge_taxes']),
        'neq': row['neq'] or '',
        'numero_tps': row['numero_tps'] or '',
        'numero_tvq': row['numero_tvq'] or '',
        'nom_entreprise': row['nom_entreprise'] or '',
        'couleur_marque': row['couleur_marque'] or '#c0321a',
    })

@app.route('/api/v1/admin/parametres-facturation', methods=['PUT'])
@admin_required
def api_admin_update_parametres_facturation():
    data = request.get_json(silent=True) or {}
    charge_taxes = 1 if data.get('charge_taxes') else 0
    neq = (data.get('neq') or '').strip()
    numero_tps = (data.get('numero_tps') or '').strip()
    numero_tvq = (data.get('numero_tvq') or '').strip()
    nom_entreprise = (data.get('nom_entreprise') or '').strip()
    couleur_marque = (data.get('couleur_marque') or '#c0321a').strip()

    conn = get_db_connection()
    cur = conn.execute("""
        UPDATE parametres_facturation
        SET charge_taxes = ?, neq = ?, numero_tps = ?, numero_tvq = ?, nom_entreprise = ?, couleur_marque = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = 1
    """, (charge_taxes, neq, numero_tps, numero_tvq, nom_entreprise, couleur_marque))
    if cur.rowcount == 0:
        conn.execute("""
            INSERT INTO parametres_facturation (id, charge_taxes, neq, numero_tps, numero_tvq, nom_entreprise, couleur_marque)
            VALUES (1, ?, ?, ?, ?, ?, ?)
        """, (charge_taxes, neq, numero_tps, numero_tvq, nom_entreprise, couleur_marque))
    conn.commit()
    conn.close()
    return jsonify({
        'success': True,
        'charge_taxes': bool(charge_taxes),
        'neq': neq,
        'numero_tps': numero_tps,
        'numero_tvq': numero_tvq,
        'nom_entreprise': nom_entreprise,
        'couleur_marque': couleur_marque,
    })


# Catégories de dépenses + lignes fiscales (T2125 / TP-80) : voir LIGNES_FISCALES
# et CATEGORIES_DEPENSES définis en haut du fichier (avant init_db).

@app.route('/api/v1/admin/depenses', methods=['POST'])
@admin_required
def api_admin_create_depense():
    data = request.get_json(silent=True) or {}
    date_transaction = (data.get('date_transaction') or '').strip()
    description = (data.get('description') or '').strip()
    categorie = (data.get('categorie') or '').strip()

    # Validation montant
    try:
        montant_total = round(float(data.get('montant_total') or 0), 2)
    except (ValueError, TypeError):
        return jsonify({'error': 'Montant invalide'}), 400

    if not date_transaction:
        return jsonify({'error': 'La date est requise'}), 400
    if not description:
        return jsonify({'error': 'La description est requise'}), 400
    if montant_total <= 0:
        return jsonify({'error': 'Le montant doit être supérieur à 0'}), 400
    if categorie and categorie not in CATEGORIES_DEPENSES:
        return jsonify({'error': 'Catégorie invalide'}), 400
    if not categorie:
        categorie = 'Autre'

    lignes = LIGNES_FISCALES.get(categorie, LIGNES_FISCALES['Autre'])
    ligne_t2125 = lignes['t2125']
    ligne_tp80 = lignes['tp80']
    piece_jointe = (data.get('piece_jointe') or '').strip() or None

    conn = get_db_connection()
    cur = conn.execute("""
        INSERT INTO transactions
            (type, date_transaction, description, categorie,
             montant_avant_taxes, montant_tps, montant_tvq, montant_total, source,
             ligne_t2125, ligne_tp80, piece_jointe)
        VALUES ('depense', ?, ?, ?, 0, 0, 0, ?, 'manuel', ?, ?, ?)
    """, (date_transaction, description, categorie, montant_total, ligne_t2125, ligne_tp80, piece_jointe))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()

    return jsonify({
        'success': True,
        'id': new_id,
        'type': 'depense',
        'date_transaction': date_transaction,
        'description': description,
        'categorie': categorie,
        'montant_total': montant_total,
        'ligne_t2125': ligne_t2125,
        'ligne_tp80': ligne_tp80,
    })

@app.route('/api/v1/admin/depenses', methods=['GET'])
@admin_required
def api_admin_list_depenses():
    annee = request.args.get('annee', '').strip()
    conn = get_db_connection()
    if annee:
        rows = conn.execute("""
            SELECT id, type, date_transaction, description, categorie,
                   montant_avant_taxes, montant_tps, montant_tvq, montant_total, source, created_at,
                   ligne_t2125, ligne_tp80
            FROM transactions
            WHERE type = 'depense' AND strftime('%Y', date_transaction) = ?
            ORDER BY date_transaction DESC, id DESC
        """, (annee,)).fetchall()
    else:
        rows = conn.execute("""
            SELECT id, type, date_transaction, description, categorie,
                   montant_avant_taxes, montant_tps, montant_tvq, montant_total, source, created_at,
                   ligne_t2125, ligne_tp80
            FROM transactions
            WHERE type = 'depense'
            ORDER BY date_transaction DESC, id DESC
        """).fetchall()
    conn.close()
    return jsonify([{
        'id': r['id'],
        'date_transaction': r['date_transaction'],
        'description': r['description'],
        'categorie': r['categorie'],
        'montant_total': r['montant_total'],
        'source': r['source'],
        'ligne_t2125': r['ligne_t2125'],
        'ligne_tp80': r['ligne_tp80'],
    } for r in rows])


# ───────────────────────────────────────────────────────────
# Comptabilité — Revenus (grand livre unifié : factures + manuel + Square/Shopify)
# ───────────────────────────────────────────────────────────
def _row_revenu(r):
    return {
        'id': r['id'],
        'date_transaction': r['date_transaction'],
        'description': r['description'],
        'categorie': r['categorie'],
        'montant_avant_taxes': r['montant_avant_taxes'],
        'montant_tps': r['montant_tps'],
        'montant_tvq': r['montant_tvq'],
        'montant_total': r['montant_total'],
        'source': r['source'],
        'source_ref': r['source_ref'],
        'id_facture': r['id_facture'],
        'ligne_t2125': r['ligne_t2125'],
        'ligne_tp80': r['ligne_tp80'],
    }

@app.route('/api/v1/admin/revenus', methods=['GET'])
@admin_required
def api_admin_list_revenus():
    annee = request.args.get('annee', '').strip()
    conn = get_db_connection()
    if annee:
        rows = conn.execute("""
            SELECT id, date_transaction, description, categorie,
                   montant_avant_taxes, montant_tps, montant_tvq, montant_total,
                   source, source_ref, id_facture, ligne_t2125, ligne_tp80
            FROM transactions
            WHERE type = 'revenu' AND strftime('%Y', date_transaction) = ?
            ORDER BY date_transaction DESC, id DESC
        """, (annee,)).fetchall()
        attente = conn.execute("""
            SELECT COALESCE(SUM(f.total), 0) t, COUNT(*) n FROM factures f
            JOIN clients c ON c.id = f.id_client
            WHERE f.statut IN ('envoyee', 'ouverte') AND strftime('%Y', f.date_emission) = ?
              AND c.is_test_client = 0
        """, (annee,)).fetchone()
    else:
        rows = conn.execute("""
            SELECT id, date_transaction, description, categorie,
                   montant_avant_taxes, montant_tps, montant_tvq, montant_total,
                   source, source_ref, id_facture, ligne_t2125, ligne_tp80
            FROM transactions
            WHERE type = 'revenu'
            ORDER BY date_transaction DESC, id DESC
        """).fetchall()
        attente = conn.execute("""
            SELECT COALESCE(SUM(f.total), 0) t, COUNT(*) n FROM factures f
            JOIN clients c ON c.id = f.id_client
            WHERE f.statut IN ('envoyee', 'ouverte') AND c.is_test_client = 0
        """).fetchone()
    conn.close()
    items = [_row_revenu(r) for r in rows]
    return jsonify({
        'items': items,
        'total_encaisse': round(sum(i['montant_total'] or 0 for i in items), 2),
        'en_attente_total': round(attente['t'] or 0, 2),
        'en_attente_count': attente['n'] or 0,
    })

@app.route('/api/v1/admin/revenus', methods=['POST'])
@admin_required
def api_admin_create_revenu():
    data = request.get_json(silent=True) or {}
    date_transaction = (data.get('date_transaction') or '').strip()
    description = (data.get('description') or '').strip()
    categorie = (data.get('categorie') or '').strip()
    try:
        montant_total = round(float(data.get('montant_total') or 0), 2)
    except (ValueError, TypeError):
        return jsonify({'error': 'Montant invalide'}), 400
    if not date_transaction:
        return jsonify({'error': 'La date est requise'}), 400
    if not description:
        return jsonify({'error': 'La description est requise'}), 400
    if montant_total <= 0:
        return jsonify({'error': 'Le montant doit être supérieur à 0'}), 400
    if categorie and categorie not in CATEGORIES_REVENUS:
        return jsonify({'error': 'Catégorie invalide'}), 400
    if not categorie:
        categorie = 'Ventes et honoraires professionnels'

    lg = LIGNES_FISCALES_REVENUS[categorie]
    piece_jointe = (data.get('piece_jointe') or '').strip() or None
    conn = get_db_connection()
    cur = conn.execute("""
        INSERT INTO transactions
            (type, date_transaction, description, categorie,
             montant_avant_taxes, montant_tps, montant_tvq, montant_total,
             source, ligne_t2125, ligne_tp80, piece_jointe)
        VALUES ('revenu', ?, ?, ?, 0, 0, 0, ?, 'manuel', ?, ?, ?)
    """, (date_transaction, description, categorie, montant_total, lg['t2125'], lg['tp80'], piece_jointe))
    conn.commit()
    row = conn.execute("""
        SELECT id, date_transaction, description, categorie,
               montant_avant_taxes, montant_tps, montant_tvq, montant_total,
               source, source_ref, id_facture, ligne_t2125, ligne_tp80
        FROM transactions WHERE id = ?
    """, (cur.lastrowid,)).fetchone()
    conn.close()
    return jsonify({'success': True, **_row_revenu(row)})

@app.route('/api/v1/admin/revenus/import', methods=['POST'])
@admin_required
def api_admin_import_revenus():
    data = request.get_json(silent=True) or {}
    lignes_in = data.get('lignes')
    if not isinstance(lignes_in, list) or not lignes_in:
        return jsonify({'error': 'Aucune ligne à importer'}), 400
    inserted, erreurs = 0, []
    conn = get_db_connection()
    for i, l in enumerate(lignes_in):
        try:
            d = (str(l.get('date_transaction') or '')).strip()
            desc = (str(l.get('description') or '')).strip()
            cat = (str(l.get('categorie') or '')).strip() or 'Ventes et honoraires professionnels'
            montant = round(float(l.get('montant_total') or 0), 2)
            if not d or not desc or montant <= 0:
                erreurs.append(f"Ligne {i + 1} : date, description et montant requis")
                continue
            if cat not in CATEGORIES_REVENUS:
                cat = 'Ventes et honoraires professionnels'
            lg = LIGNES_FISCALES_REVENUS[cat]
            conn.execute("""
                INSERT INTO transactions
                    (type, date_transaction, description, categorie,
                     montant_avant_taxes, montant_tps, montant_tvq, montant_total,
                     source, ligne_t2125, ligne_tp80)
                VALUES ('revenu', ?, ?, ?, 0, 0, 0, ?, 'manuel', ?, ?)
            """, (d, desc, cat, montant, lg['t2125'], lg['tp80']))
            inserted += 1
        except Exception as e:
            erreurs.append(f"Ligne {i + 1} : {e}")
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'inserted': inserted, 'erreurs': erreurs})

@app.route('/api/v1/admin/revenus/<int:revenu_id>', methods=['DELETE'])
@admin_required
def api_admin_delete_revenu(revenu_id):
    conn = get_db_connection()
    row = conn.execute("SELECT source FROM transactions WHERE id = ? AND type = 'revenu'", (revenu_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Revenu introuvable'}), 404
    if row['source'] != 'manuel':
        conn.close()
        return jsonify({'error': "Ce revenu provient d'une facture — gérez-le depuis la facture."}), 400
    conn.execute("DELETE FROM transactions WHERE id = ? AND type = 'revenu' AND source = 'manuel'", (revenu_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ───────────────────────────────────────────────────────────
# Comptabilité — Scan de reçu par photo (dépenses + revenus)
# ───────────────────────────────────────────────────────────
def _scanner_recu(categories, type_libelle):
    import base64
    f = request.files.get('image')
    if not f:
        return jsonify({'error': 'Aucune image reçue'}), 400
    raw = f.read()
    if not raw:
        return jsonify({'error': 'Image vide'}), 400
    if len(raw) > 18 * 1024 * 1024:
        return jsonify({'error': "Fichier trop lourd (max ~18 Mo)"}), 400
    mime = (f.mimetype or 'image/jpeg').lower()
    fields = analyser_recu(base64.b64encode(raw).decode('ascii'), mime, categories, type_libelle)
    if 'error' in fields:
        return jsonify(fields), 502
    try:
        fields['piece_jointe'] = sauver_piece_jointe(raw, f.filename, mime)
    except Exception:
        fields['piece_jointe'] = None  # l'extraction reste utilisable même si la sauvegarde échoue
    return jsonify({'success': True, **fields})

@app.route('/api/v1/admin/depenses/scan', methods=['POST'])
@admin_required
def api_admin_scan_depense():
    return _scanner_recu(CATEGORIES_DEPENSES, "dépense")

@app.route('/api/v1/admin/revenus/scan', methods=['POST'])
@admin_required
def api_admin_scan_revenu():
    return _scanner_recu(CATEGORIES_REVENUS, "rentrée d'argent (revenu)")


# ───────────────────────────────────────────────────────────
# App de capture (/capture) — jumelage d'appareil (jeton persistant par appareil,
# rattaché au compte de l'abonné, indépendant de la session CRM).
# ───────────────────────────────────────────────────────────
def _device_cookie_kwargs():
    """Attributs communs des cookies de jeton d'appareil (PWA Tâches/Capture) —
    httpOnly pour ne plus exposer le jeton à un script JS (cf. audit sécurité 19/07)."""
    return dict(httponly=True, secure=True, samesite='Lax', max_age=400 * 24 * 3600, path='/')


def _capture_user():
    """Résout l'appareil via son jeton (cookie httpOnly en priorité, puis header
    X-Capture-Token / form / JSON pour compatibilité avec les appareils déjà jumelés
    avant le passage au cookie). Retourne {user_id, organisation_id, token} ou None.
    Marque last_used_at."""
    import hashlib
    token = request.cookies.get('cos_capture_token')
    if not token:
        token = request.headers.get('X-Capture-Token')
    if not token and request.form:
        token = request.form.get('token')
    if not token:
        token = (request.get_json(silent=True) or {}).get('token')
    if not token:
        return None
    th = hashlib.sha256(token.encode()).hexdigest()
    conn = get_db_connection()
    row = conn.execute(
        "SELECT id, user_id, organisation_id FROM capture_devices WHERE token_hash = ? AND revoked = 0",
        (th,)
    ).fetchone()
    if row:
        conn.execute("UPDATE capture_devices SET last_used_at = ? WHERE id = ?",
                     (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), row['id']))
        conn.commit()
    conn.close()
    return {'user_id': row['user_id'], 'organisation_id': row['organisation_id'], 'token': token} if row else None

@app.route('/api/v1/capture/pair', methods=['POST'])
@admin_required
def api_capture_pair():
    """Jumelle l'appareil courant au compte connecté. Émet un jeton (une seule fois)."""
    import secrets, hashlib
    token = secrets.token_urlsafe(32)
    th = hashlib.sha256(token.encode()).hexdigest()
    label = ((request.get_json(silent=True) or {}).get('label') or 'Appareil').strip()[:60]
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO capture_devices (user_id, organisation_id, token_hash, label) VALUES (?, ?, ?, ?)",
        (session.get('user_id'), session.get('organisation_id'), th, label)
    )
    conn.commit()
    conn.close()
    resp = jsonify({'success': True, 'token': token})
    resp.set_cookie('cos_capture_token', token, **_device_cookie_kwargs())
    return resp

@app.route('/api/v1/capture/login', methods=['POST'])
@limiter.limit("5 per minute")
def api_capture_login():
    """Connexion PROPRE à l'app de capture : vérifie les identifiants et jumelle
    l'appareil (émet un jeton). N'ouvre PAS de session CRM — le jeton ne donne
    accès qu'au scan + ajout de transactions, jamais au portail complet."""
    import secrets, hashlib
    data = request.get_json(silent=True) or {}
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''
    if email and _redis_login_locked(email):
        return jsonify({'error': 'Trop de tentatives. Réessayez dans 15 min.'}), 429
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    ok = False
    if user and user['auth_provider'] == 'password' and user['mot_de_passe_hash']:
        try:
            ok = bcrypt.check_password_hash(user['mot_de_passe_hash'], password)
        except Exception:
            ok = False
    # Pour l'instant (mono-locataire) : seul le compte admin (l'abonné) peut jumeler.
    if not ok or not int(user['is_admin'] or 0):
        conn.close()
        _redis_login_fail_inc(email)
        return jsonify({'error': 'Identifiants invalides.'}), 401
    if not int(user['is_email_confirmed'] or 0):
        conn.close()
        return jsonify({'error': 'Courriel non confirmé.'}), 403
    _redis_login_reset(email)
    token = secrets.token_urlsafe(32)
    th = hashlib.sha256(token.encode()).hexdigest()
    conn.execute(
        "INSERT INTO capture_devices (user_id, organisation_id, token_hash, label) VALUES (?, ?, ?, ?)",
        (user['id'], None, th, 'Appareil')
    )
    conn.commit()
    conn.close()
    resp = jsonify({'success': True, 'token': token})
    resp.set_cookie('cos_capture_token', token, **_device_cookie_kwargs())
    return resp

@app.route('/api/v1/capture/verify', methods=['POST'])
def api_capture_verify():
    u = _capture_user()
    if not u:
        return jsonify({'error': 'Appareil non lié'}), 401
    conn = get_db_connection()
    row = conn.execute("SELECT nom_complet, email FROM clients WHERE id = ?", (u['user_id'],)).fetchone()
    conn.close()
    compte = (row['nom_complet'] or row['email']) if row else None
    resp = jsonify({'success': True, 'compte': compte, 'email': row['email'] if row else None})
    # Ré-émet le cookie httpOnly même si l'appareil s'est authentifié via l'ancien
    # header/localStorage — bascule les appareils déjà jumelés sans les forcer à se reconnecter.
    resp.set_cookie('cos_capture_token', u['token'], **_device_cookie_kwargs())
    return resp

@app.route('/api/v1/capture/logout', methods=['POST'])
def api_capture_logout():
    """Déconnexion de l'appareil : révoque le jeton (il devient inutilisable)."""
    import hashlib
    token = request.cookies.get('cos_capture_token') or request.headers.get('X-Capture-Token') or (request.get_json(silent=True) or {}).get('token')
    if token:
        th = hashlib.sha256(token.encode()).hexdigest()
        conn = get_db_connection()
        conn.execute("UPDATE capture_devices SET revoked = 1 WHERE token_hash = ?", (th,))
        conn.commit()
        conn.close()
    resp = jsonify({'success': True})
    resp.set_cookie('cos_capture_token', '', max_age=0, path='/')
    return resp


def _task_user():
    """Retourne {'user_id': ..., 'token': ...} si le jeton d'appareil de la PWA Tâches
    (/taches) est valide, sinon None. Même principe que _capture_user (cookie httpOnly
    en priorité, header/JSON en repli pour les appareils déjà jumelés)."""
    import hashlib
    token = request.cookies.get('cos_task_token')
    if not token:
        token = request.headers.get('X-Task-Token')
    if not token:
        token = (request.get_json(silent=True) or {}).get('token')
    if not token:
        return None
    th = hashlib.sha256(token.encode()).hexdigest()
    conn = get_db_connection()
    row = conn.execute(
        "SELECT id, user_id FROM task_devices WHERE token_hash = ? AND revoked = 0",
        (th,)
    ).fetchone()
    if row:
        conn.execute("UPDATE task_devices SET last_used_at = ? WHERE id = ?",
                     (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), row['id']))
        conn.commit()
    conn.close()
    return {'user_id': row['user_id'], 'token': token} if row else None


@app.route('/api/v1/taches/login', methods=['POST'])
@limiter.limit("5 per minute")
def api_taches_login():
    """Connexion PROPRE à la PWA Tâches : vérifie les identifiants et jumelle l'appareil
    (émet un jeton). N'ouvre PAS de session CRM. Chaque membre de l'équipe (compte admin
    avec un rôle — même critère que /api/v1/admin/team) jumelle son propre appareil."""
    import secrets, hashlib
    data = request.get_json(silent=True) or {}
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''
    if email and _redis_login_locked(email):
        return jsonify({'error': 'Trop de tentatives. Réessayez dans 15 min.'}), 429
    conn = get_db_connection()
    user = conn.execute("SELECT * FROM clients WHERE email = ?", (email,)).fetchone()
    ok = False
    if user and user['auth_provider'] == 'password' and user['mot_de_passe_hash']:
        try:
            ok = bcrypt.check_password_hash(user['mot_de_passe_hash'], password)
        except Exception:
            ok = False
    if not ok or not int(user['is_admin'] or 0) or not user['role']:
        conn.close()
        _redis_login_fail_inc(email)
        return jsonify({'error': 'Identifiants invalides.'}), 401
    if not int(user['is_email_confirmed'] or 0):
        conn.close()
        return jsonify({'error': 'Courriel non confirmé.'}), 403
    _redis_login_reset(email)
    token = secrets.token_urlsafe(32)
    th = hashlib.sha256(token.encode()).hexdigest()
    conn.execute(
        "INSERT INTO task_devices (user_id, token_hash, label) VALUES (?, ?, ?)",
        (user['id'], th, 'Appareil')
    )
    conn.commit()
    conn.close()
    resp = jsonify({'success': True, 'token': token})
    resp.set_cookie('cos_task_token', token, **_device_cookie_kwargs())
    return resp


@app.route('/api/v1/taches/verify', methods=['POST'])
def api_taches_verify():
    u = _task_user()
    if not u:
        return jsonify({'error': 'Appareil non lié'}), 401
    conn = get_db_connection()
    row = conn.execute("SELECT nom_complet, email FROM clients WHERE id = ?", (u['user_id'],)).fetchone()
    conn.close()
    compte = (row['nom_complet'] or row['email']) if row else None
    resp = jsonify({'success': True, 'compte': compte, 'email': row['email'] if row else None})
    resp.set_cookie('cos_task_token', u['token'], **_device_cookie_kwargs())
    return resp


@app.route('/api/v1/taches/logout', methods=['POST'])
def api_taches_logout():
    """Déconnexion de l'appareil : révoque le jeton (il devient inutilisable)."""
    import hashlib
    token = request.cookies.get('cos_task_token') or request.headers.get('X-Task-Token') or (request.get_json(silent=True) or {}).get('token')
    if token:
        th = hashlib.sha256(token.encode()).hexdigest()
        conn = get_db_connection()
        conn.execute("UPDATE task_devices SET revoked = 1 WHERE token_hash = ?", (th,))
        conn.commit()
        conn.close()
    resp = jsonify({'success': True})
    resp.set_cookie('cos_task_token', '', max_age=0, path='/')
    return resp


@app.route('/api/v1/taches/todos', methods=['GET'])
def api_taches_todos_list():
    u = _task_user()
    if not u:
        return jsonify({'error': 'Appareil non lié'}), 401
    return jsonify(_todos_query(u['user_id'], 'mine'))


@app.route('/api/v1/taches/todos', methods=['POST'])
def api_taches_todos_create():
    u = _task_user()
    if not u:
        return jsonify({'error': 'Appareil non lié'}), 401
    body, status = _create_todo(u['user_id'], request.get_json() or {})
    return jsonify(body), status


@app.route('/api/v1/taches/todos/<int:todo_id>/toggle', methods=['POST'])
def api_taches_todos_toggle(todo_id):
    if not _task_user():
        return jsonify({'error': 'Appareil non lié'}), 401
    body, status = _toggle_todo(todo_id)
    return jsonify(body), status


@app.route('/api/v1/taches/todos/<int:todo_id>', methods=['PATCH'])
def api_taches_todos_update(todo_id):
    if not _task_user():
        return jsonify({'error': 'Appareil non lié'}), 401
    body, status = _update_todo(todo_id, request.get_json() or {})
    return jsonify(body), status


@app.route('/api/v1/taches/todos/<int:todo_id>', methods=['DELETE'])
def api_taches_todos_delete(todo_id):
    if not _task_user():
        return jsonify({'error': 'Appareil non lié'}), 401
    body, status = _delete_todo(todo_id)
    return jsonify(body), status


@app.route('/api/v1/taches/todos/<int:todo_id>/planifier', methods=['POST'])
def api_taches_todos_planifier(todo_id):
    if not _task_user():
        return jsonify({'error': 'Appareil non lié'}), 401
    body, status = _planifier_todo(todo_id, request.get_json() or {})
    return jsonify(body), status


@app.route('/api/v1/taches/todos/<int:todo_id>/deplanifier', methods=['POST'])
def api_taches_todos_deplanifier(todo_id):
    if not _task_user():
        return jsonify({'error': 'Appareil non lié'}), 401
    body, status = _deplanifier_todo(todo_id)
    return jsonify(body), status


@app.route('/api/v1/taches/team', methods=['GET'])
def api_taches_team():
    """Équipe assignable — même critère que /api/v1/admin/team (comptes admin avec un rôle)."""
    if not _task_user():
        return jsonify({'error': 'Appareil non lié'}), 401
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT id, nom_complet FROM clients WHERE is_admin = 1 AND role IS NOT NULL ORDER BY nom_complet"
    ).fetchall()
    conn.close()
    return jsonify([{'id': r['id'], 'nom_complet': r['nom_complet']} for r in rows])


@app.route('/api/v1/taches/clients', methods=['GET'])
def api_taches_clients():
    """Liste allégée des clients — pour le sélecteur d'assignation de la PWA Tâches."""
    if not _task_user():
        return jsonify({'error': 'Appareil non lié'}), 401
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT id, nom_complet FROM clients WHERE is_admin = 0 ORDER BY nom_complet"
    ).fetchall()
    conn.close()
    return jsonify([{'id': r['id'], 'nom_complet': r['nom_complet']} for r in rows])


@app.route('/api/v1/taches/projets', methods=['GET'])
def api_taches_projets():
    """Liste allégée des projets actifs — pour le sélecteur d'assignation de la PWA Tâches."""
    if not _task_user():
        return jsonify({'error': 'Appareil non lié'}), 401
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT p.id, p.nom_projet, c.nom_complet AS client_nom
        FROM projets p
        LEFT JOIN clients c ON c.id = p.id_client
        WHERE COALESCE(p.is_archived, 0) = 0
        ORDER BY p.created_at DESC
    """).fetchall()
    conn.close()
    return jsonify([{'id': r['id'], 'nom_projet': r['nom_projet'], 'client_nom': r['client_nom']} for r in rows])


@app.route('/api/v1/taches/push/vapid-public-key', methods=['GET'])
def api_taches_push_vapid_key():
    if not _task_user():
        return jsonify({'error': 'Appareil non lié'}), 401
    return jsonify({'key': VAPID_PUBLIC_KEY})


@app.route('/api/v1/taches/push/subscribe', methods=['POST'])
def api_taches_push_subscribe():
    """Abonnement Web Push depuis la PWA Tâches — même table que le portail desktop
    (push_subscriptions), rattaché au courriel du compte jumelé pour que l'assignation
    d'une tâche puisse cibler cette personne précisément."""
    u = _task_user()
    if not u:
        return jsonify({'error': 'Appareil non lié'}), 401
    data = request.get_json() or {}
    endpoint = data.get('endpoint')
    keys = data.get('keys') or {}
    p256dh, auth = keys.get('p256dh'), keys.get('auth')
    if not (endpoint and p256dh and auth):
        return jsonify({'error': 'Abonnement invalide'}), 400
    conn = get_db_connection()
    user = conn.execute("SELECT email FROM clients WHERE id=?", (u['user_id'],)).fetchone()
    conn.execute(
        "INSERT INTO push_subscriptions (email, endpoint, p256dh, auth) VALUES (?, ?, ?, ?) "
        "ON CONFLICT(endpoint) DO UPDATE SET email=excluded.email, p256dh=excluded.p256dh, auth=excluded.auth",
        (user['email'] if user else None, endpoint, p256dh, auth)
    )
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/taches/push/unsubscribe', methods=['POST'])
def api_taches_push_unsubscribe():
    if not _task_user():
        return jsonify({'error': 'Appareil non lié'}), 401
    data = request.get_json() or {}
    endpoint = data.get('endpoint')
    if endpoint:
        conn = get_db_connection()
        conn.execute("DELETE FROM push_subscriptions WHERE endpoint=?", (endpoint,))
        conn.commit()
        conn.close()
    return jsonify({'success': True})


@app.route('/api/v1/taches/marketing', methods=['GET'])
def api_taches_marketing():
    """Posts marketing en attente de dépôt de visuels (todo_felix_done=0), pour la
    personne au rôle 'production' — convergence Phase 4 : ces posts apparaissaient déjà
    dans /admin/marketing et un widget dashboard, jamais dans la PWA Tâches. On exclut
    les posts déjà liés à un item roadmap (linked_roadmap_todo_id) : ceux-là ont déjà
    une tâche todos_perso jumelle qui apparaît dans la liste principale — les inclure
    ici les ferait apparaître en double. todo_marie_done n'est pas une case à cocher
    personnelle (c'est un indicateur « Félix a notifié Marie » posé par un envoi groupé
    de courriel côté /admin/marketing), donc rien à afficher ici pour le rôle gestion."""
    import json as _json
    u = _task_user()
    if not u:
        return jsonify({'error': 'Appareil non lié'}), 401
    conn = get_db_connection()
    user = conn.execute("SELECT role FROM clients WHERE id=?", (u['user_id'],)).fetchone()
    if not user or user['role'] != 'production':
        conn.close()
        return jsonify([])
    posts = conn.execute("""
        SELECT id, titre, date_publication, plateformes, statut FROM marketing_posts
        WHERE todo_felix_done = 0 AND linked_roadmap_todo_id IS NULL
        ORDER BY date_publication ASC
    """).fetchall()
    conn.close()
    return jsonify([{
        'id': p['id'], 'titre': p['titre'], 'date_publication': p['date_publication'],
        'plateformes': _json.loads(p['plateformes']) if p['plateformes'] else [],
        'statut': p['statut'],
    } for p in posts])


@app.route('/api/v1/taches/marketing/<int:post_id>/toggle', methods=['POST'])
def api_taches_marketing_toggle(post_id):
    u = _task_user()
    if not u:
        return jsonify({'error': 'Appareil non lié'}), 401
    conn = get_db_connection()
    user = conn.execute("SELECT role FROM clients WHERE id=?", (u['user_id'],)).fetchone()
    if not user or user['role'] != 'production':
        conn.close()
        return jsonify({'error': 'Réservé au rôle production'}), 403
    post = conn.execute("SELECT todo_felix_done FROM marketing_posts WHERE id=?", (post_id,)).fetchone()
    if not post:
        conn.close()
        return jsonify({'error': 'Post introuvable'}), 404
    new_val = 0 if int(post['todo_felix_done'] or 0) else 1
    conn.execute("UPDATE marketing_posts SET todo_felix_done=? WHERE id=?", (new_val, post_id))
    conn.commit()
    conn.close()
    return jsonify({'done': bool(new_val)})


@app.route('/api/v1/capture/scan', methods=['POST'])
def api_capture_scan():
    if not _capture_user():
        return jsonify({'error': 'Appareil non lié'}), 401
    est_revenu = (request.form.get('type') == 'revenu')
    cats = CATEGORIES_REVENUS if est_revenu else CATEGORIES_DEPENSES
    libelle = "rentrée d'argent (revenu)" if est_revenu else "dépense"
    return _scanner_recu(cats, libelle)

@app.route('/api/v1/capture/transaction', methods=['POST'])
def api_capture_transaction():
    u = _capture_user()
    if not u:
        return jsonify({'error': 'Appareil non lié'}), 401
    data = request.get_json(silent=True) or {}
    type_ = 'revenu' if data.get('type') == 'revenu' else 'depense'
    date_transaction = (data.get('date_transaction') or '').strip()
    description = (data.get('description') or '').strip()
    categorie = (data.get('categorie') or '').strip()
    piece_jointe = (data.get('piece_jointe') or '').strip() or None
    try:
        montant_total = round(float(data.get('montant_total') or 0), 2)
    except (ValueError, TypeError):
        return jsonify({'error': 'Montant invalide'}), 400
    if not date_transaction or not description or montant_total <= 0:
        return jsonify({'error': 'Date, description et montant requis'}), 400
    if type_ == 'revenu':
        if categorie not in CATEGORIES_REVENUS:
            categorie = 'Ventes et honoraires professionnels'
        lg = LIGNES_FISCALES_REVENUS[categorie]
    else:
        if categorie not in CATEGORIES_DEPENSES:
            categorie = 'Autre'
        lg = LIGNES_FISCALES[categorie]
    conn = get_db_connection()
    conn.execute("""
        INSERT INTO transactions
            (organisation_id, type, date_transaction, description, categorie,
             montant_avant_taxes, montant_tps, montant_tvq, montant_total,
             source, ligne_t2125, ligne_tp80, piece_jointe)
        VALUES (?, ?, ?, ?, ?, 0, 0, 0, ?, 'manuel', ?, ?, ?)
    """, (u['organisation_id'], type_, date_transaction, description, categorie,
          montant_total, lg['t2125'], lg['tp80'], piece_jointe))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# ───────────────────────────────────────────────────────────
# Boîte de réception comptable — factures « à valider »
# (réceptacle de l'ingestion Gmail + éventuellement scans en brouillon)
# ───────────────────────────────────────────────────────────
def deposer_facture_a_valider(fields, source, source_ref=None, expediteur=None, organisation_id=None):
    """Dépose une facture détectée dans la file « à valider ». GARDE-FOUS :
    - jamais un revenu (dépenses seulement — les revenus viennent factures + Square) ;
    - écarte les documents ÉMIS par nous (sens != 'recu') ;
    - anti-doublon via (source, source_ref).
    Retourne 'ajoute' | 'ignore_sens' | 'doublon' | 'erreur'."""
    sens = (fields.get('sens') or 'recu').lower()
    if sens not in ('recu', 'reçu'):
        return 'ignore_sens'  # facture émise par nous → jamais dans les dépenses
    try:
        conn = get_db_connection()
        # anti-doublon explicite (en plus de l'index) pour un retour propre
        if source_ref:
            dup = conn.execute(
                "SELECT 1 FROM factures_a_valider WHERE source = ? AND source_ref = ?",
                (source, source_ref)
            ).fetchone()
            if dup:
                conn.close()
                return 'doublon'
        cat = (fields.get('categorie') or '').strip()
        if cat not in CATEGORIES_DEPENSES:
            cat = 'Autre'
        conn.execute("""
            INSERT INTO factures_a_valider
                (organisation_id, source, source_ref, sens, expediteur, date_transaction,
                 fournisseur, description, categorie, montant_avant_taxes, montant_tps,
                 montant_tvq, montant_total, piece_jointe, confiance)
            VALUES (?, ?, ?, 'recu', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (organisation_id, source, source_ref, expediteur,
              (fields.get('date_transaction') or '').strip(),
              (fields.get('fournisseur') or '').strip(),
              (fields.get('description') or '').strip(), cat,
              float(fields.get('montant_avant_taxes') or 0),
              float(fields.get('montant_tps') or 0),
              float(fields.get('montant_tvq') or 0),
              float(fields.get('montant_total') or 0),
              fields.get('piece_jointe'), fields.get('confiance')))
        conn.commit()
        conn.close()
        return 'ajoute'
    except Exception as e:
        print(f"[A-VALIDER] dépôt échoué: {e}")
        return 'erreur'

@app.route('/api/v1/admin/factures-a-valider', methods=['GET'])
@admin_required
def api_admin_list_a_valider():
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT id, source, source_ref, sens, expediteur, date_transaction, fournisseur,
               description, categorie, montant_avant_taxes, montant_tps, montant_tvq,
               montant_total, piece_jointe, confiance, note, created_at
        FROM factures_a_valider
        WHERE statut = 'en_attente'
        ORDER BY created_at DESC, id DESC
    """).fetchall()
    conn.close()
    return jsonify({'items': [dict(r) for r in rows], 'count': len(rows)})

@app.route('/api/v1/admin/factures-a-valider/count', methods=['GET'])
@admin_required
def api_admin_count_a_valider():
    conn = get_db_connection()
    n = conn.execute("SELECT COUNT(*) FROM factures_a_valider WHERE statut = 'en_attente'").fetchone()[0]
    conn.close()
    return jsonify({'count': n})

@app.route('/api/v1/admin/factures-a-valider/<int:item_id>/approuver', methods=['POST'])
@admin_required
def api_admin_approuver_a_valider(item_id):
    data = request.get_json(silent=True) or {}
    conn = get_db_connection()
    row = conn.execute("SELECT * FROM factures_a_valider WHERE id = ? AND statut = 'en_attente'", (item_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Élément introuvable ou déjà traité'}), 404
    # Champs éventuellement corrigés par l'admin avant approbation
    date_transaction = (data.get('date_transaction') or row['date_transaction'] or '').strip()
    description = (data.get('description') or row['description'] or '').strip()
    categorie = (data.get('categorie') or row['categorie'] or 'Autre').strip()
    try:
        montant_total = round(float(data.get('montant_total') if data.get('montant_total') is not None else row['montant_total'] or 0), 2)
    except (ValueError, TypeError):
        conn.close()
        return jsonify({'error': 'Montant invalide'}), 400
    if not date_transaction or not description or montant_total <= 0:
        conn.close()
        return jsonify({'error': 'Date, description et montant requis'}), 400
    if categorie not in CATEGORIES_DEPENSES:
        categorie = 'Autre'
    lg = LIGNES_FISCALES[categorie]
    # GARDE-FOU : toujours une dépense, jamais un revenu
    conn.execute("""
        INSERT INTO transactions
            (organisation_id, type, date_transaction, description, categorie,
             montant_avant_taxes, montant_tps, montant_tvq, montant_total,
             source, ligne_t2125, ligne_tp80, piece_jointe)
        VALUES (?, 'depense', ?, ?, ?, ?, ?, ?, ?, 'manuel', ?, ?, ?)
    """, (row['organisation_id'], date_transaction, description, categorie,
          row['montant_avant_taxes'] or 0, row['montant_tps'] or 0, row['montant_tvq'] or 0,
          montant_total, lg['t2125'], lg['tp80'], row['piece_jointe']))
    conn.execute("UPDATE factures_a_valider SET statut = 'approuve', traite_at = ? WHERE id = ?",
                 (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), item_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/factures-a-valider/<int:item_id>/rejeter', methods=['POST'])
@admin_required
def api_admin_rejeter_a_valider(item_id):
    conn = get_db_connection()
    cur = conn.execute("UPDATE factures_a_valider SET statut = 'rejete', traite_at = ? WHERE id = ? AND statut = 'en_attente'",
                       (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), item_id))
    conn.commit()
    n = cur.rowcount
    conn.close()
    if not n:
        return jsonify({'error': 'Élément introuvable ou déjà traité'}), 404
    return jsonify({'success': True})

# ───────────────────────────────────────────────────────────
# Intégration Gmail — ingestion automatique des factures fournisseurs.
# OAuth par abonné (scope gmail.readonly), jeton rangé dans `integrations`
# (prêt multi-locataire). Le job dépose dans la file « à valider ».
# ───────────────────────────────────────────────────────────
GMAIL_LABEL_FACTURES = os.getenv('GMAIL_LABEL_FACTURES', 'Factures')
GMAIL_OAUTH_SCOPE = 'https://www.googleapis.com/auth/gmail.readonly'

def _gmail_redirect_uri():
    return f"{PORTAIL_URL}/api/v1/admin/integrations/gmail/callback"

@app.route('/api/v1/admin/integrations/gmail/status', methods=['GET'])
@admin_required
def api_gmail_status():
    configured = bool(os.getenv('GOOGLE_CLIENT_ID') and os.getenv('GOOGLE_CLIENT_SECRET'))
    conn = get_db_connection()
    row = conn.execute(
        "SELECT merchant_id, last_sync_at FROM integrations "
        "WHERE provider = 'gmail' AND statut = 'actif' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    conn.close()
    return jsonify({
        'configured': configured,
        'connected': bool(row),
        'compte': row['merchant_id'] if row else None,
        'last_sync_at': row['last_sync_at'] if row else None,
        'label': GMAIL_LABEL_FACTURES,
    })

@app.route('/api/v1/admin/integrations/gmail/connect', methods=['GET'])
@admin_required
def api_gmail_connect():
    from urllib.parse import urlencode
    if not (os.getenv('GOOGLE_CLIENT_ID') and os.getenv('GOOGLE_CLIENT_SECRET')):
        return jsonify({'error': 'OAuth Google non configuré (identifiants manquants).'}), 400
    params = {
        'client_id': os.getenv('GOOGLE_CLIENT_ID'),
        'redirect_uri': _gmail_redirect_uri(),
        'response_type': 'code',
        'scope': GMAIL_OAUTH_SCOPE,
        'access_type': 'offline',
        'prompt': 'consent',
        'include_granted_scopes': 'true',
        'state': str(session.get('user_id') or ''),
    }
    return redirect('https://accounts.google.com/o/oauth2/v2/auth?' + urlencode(params))

@app.route('/api/v1/admin/integrations/gmail/callback', methods=['GET'])
@admin_required
def api_gmail_callback():
    import requests
    code = request.args.get('code')
    if not code:
        return redirect(f"{PORTAIL_URL}/admin/comptabilite/a-valider?gmail=refus")
    try:
        tok = requests.post('https://oauth2.googleapis.com/token', data={
            'code': code,
            'client_id': os.getenv('GOOGLE_CLIENT_ID'),
            'client_secret': os.getenv('GOOGLE_CLIENT_SECRET'),
            'redirect_uri': _gmail_redirect_uri(),
            'grant_type': 'authorization_code',
        }, timeout=30).json()
        refresh = tok.get('refresh_token')
        if not refresh:
            return redirect(f"{PORTAIL_URL}/admin/comptabilite/a-valider?gmail=sansjeton")
        # Adresse de la boîte connectée (pour affichage + exclusion des propres envois)
        email_compte = ''
        try:
            import gmail_service
            email_compte = gmail_service.profil_email(gmail_service.build_gmail(refresh))
        except Exception:
            pass
        conn = get_db_connection()
        # une seule intégration Gmail active par organisation (ici org NULL = compte actuel)
        conn.execute("UPDATE integrations SET statut = 'revoque' WHERE provider = 'gmail' AND COALESCE(organisation_id,0) = 0")
        conn.execute(
            "INSERT INTO integrations (organisation_id, provider, merchant_id, refresh_token, scopes, statut) "
            "VALUES (?, 'gmail', ?, ?, ?, 'actif')",
            (session.get('organisation_id'), email_compte, refresh, GMAIL_OAUTH_SCOPE)
        )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[GMAIL] callback échoué: {e}")
        return redirect(f"{PORTAIL_URL}/admin/comptabilite/a-valider?gmail=erreur")
    return redirect(f"{PORTAIL_URL}/admin/comptabilite/a-valider?gmail=ok")

@app.route('/api/v1/admin/integrations/gmail/disconnect', methods=['POST'])
@admin_required
def api_gmail_disconnect():
    conn = get_db_connection()
    conn.execute("UPDATE integrations SET statut = 'revoque' WHERE provider = 'gmail' AND statut = 'actif'")
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/v1/admin/integrations/gmail/sync', methods=['POST'])
@admin_required
def api_gmail_sync_now():
    res = sync_gmail_factures(app)
    if isinstance(res, dict) and res.get('error'):
        return jsonify(res), 400
    return jsonify({'success': True, **(res or {})})

def sync_gmail_factures(app_ref=None):
    """Lit les courriels étiquetés « Factures », extrait les pièces PDF/image,
    les analyse et dépose les factures REÇUES dans la file « à valider ».
    Garde-fous : INBOX seulement, jamais nos propres envois, anti-doublon, sens='recu'."""
    import base64 as _b64, time
    from email.utils import parseaddr
    try:
        import gmail_service
    except Exception as e:
        return {'error': f"Module Gmail indisponible : {e}"}
    conn = get_db_connection()
    integ = conn.execute(
        "SELECT * FROM integrations WHERE provider = 'gmail' AND statut = 'actif' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if not integ:
        conn.close()
        return {'error': 'Gmail non connecté.'}
    conn.close()

    try:
        service = gmail_service.build_gmail(integ['refresh_token'])
        mon_email = (gmail_service.profil_email(service) or integ['merchant_id'] or '').lower()
        mon_domaine = mon_email.split('@')[-1] if '@' in mon_email else ''
        label_id = gmail_service.trouver_label_id(service, GMAIL_LABEL_FACTURES)
        if not label_id:
            return {'error': f"Libellé Gmail « {GMAIL_LABEL_FACTURES} » introuvable."}
        # Fenêtre : depuis le dernier sync (moins une marge), sinon 30 jours
        apres = None
        if integ['last_sync_at']:
            try:
                from datetime import datetime as _dt
                apres = int(_dt.strptime(integ['last_sync_at'], '%Y-%m-%d %H:%M:%S').timestamp()) - 86400
            except Exception:
                apres = None
        if apres is None:
            apres = int(time.time()) - 30 * 86400
        msg_ids = gmail_service.lister_messages(service, label_id, apres_epoch=apres, maximum=25)
    except Exception as e:
        print(f"[GMAIL] connexion/liste échouée: {e}")
        return {'error': "Connexion Gmail échouée (jeton expiré ?)."}

    proprietaire = integ['merchant_id'] or mon_email or 'le titulaire du compte'
    ajoutes = ignores = doublons = erreurs = 0
    for mid in msg_ids:
        try:
            m = gmail_service.lire_message(service, mid)
            _, exp_email = parseaddr(m.get('expediteur', ''))
            exp_email = (exp_email or '').lower()
            # Garde-fou : ignorer nos propres envois (même adresse ou même domaine)
            if exp_email and (exp_email == mon_email or (mon_domaine and exp_email.endswith('@' + mon_domaine))):
                ignores += 1
                continue
            if not m.get('pieces'):
                continue
            for i, piece in enumerate(m['pieces']):
                fields = analyser_recu(piece['data'], piece['mime'], CATEGORIES_DEPENSES,
                                       "dépense (facture fournisseur)", sens_contexte=proprietaire)
                if 'error' in fields:
                    erreurs += 1
                    continue
                # pièce justificative sur Drive
                try:
                    fields['piece_jointe'] = sauver_piece_jointe(
                        _b64.b64decode(piece['data']), piece['filename'], piece['mime'])
                except Exception:
                    fields['piece_jointe'] = None
                res = deposer_facture_a_valider(
                    fields, source='gmail', source_ref=f"{mid}:{i}",
                    expediteur=m.get('expediteur'), organisation_id=integ['organisation_id'])
                if res == 'ajoute':
                    ajoutes += 1
                elif res == 'doublon':
                    doublons += 1
                elif res == 'ignore_sens':
                    ignores += 1
                else:
                    erreurs += 1
        except Exception as e:
            print(f"[GMAIL] message {mid} échoué: {e}")
            erreurs += 1

    conn = get_db_connection()
    conn.execute("UPDATE integrations SET last_sync_at = ? WHERE id = ?",
                 (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), integ['id']))
    conn.commit()
    conn.close()
    print(f"[GMAIL] sync: +{ajoutes} à valider, {doublons} doublons, {ignores} ignorés, {erreurs} erreurs")
    return {'ajoutes': ajoutes, 'doublons': doublons, 'ignores': ignores, 'erreurs': erreurs}

def _calculer_bilan(annee, mois):
    if mois:
        like_periode = f"{annee}-{mois.zfill(2)}"
        where_date = "strftime('%Y-%m', date_transaction) = ?"
        param_date = like_periode
        periode_label = f"{mois.zfill(2)}/{annee}"
    else:
        where_date = "strftime('%Y', date_transaction) = ?"
        param_date = annee
        periode_label = annee

    conn = get_db_connection()

    def _somme(type_tx):
        row = conn.execute(f"""
            SELECT
                COALESCE(SUM(montant_avant_taxes),0) AS avant_taxes,
                COALESCE(SUM(montant_tps),0)         AS tps,
                COALESCE(SUM(montant_tvq),0)         AS tvq,
                COALESCE(SUM(montant_total),0)       AS total,
                COUNT(*)                             AS nb
            FROM transactions
            WHERE type = ? AND {where_date}
        """, (type_tx, param_date)).fetchone()
        return {
            'avant_taxes': round(row['avant_taxes'], 2),
            'tps': round(row['tps'], 2),
            'tvq': round(row['tvq'], 2),
            'total': round(row['total'], 2),
            'nb': row['nb'],
        }

    def _par_categorie(type_tx):
        rows = conn.execute(f"""
            SELECT
                COALESCE(categorie,'Autre') AS categorie,
                ligne_t2125, ligne_tp80,
                COALESCE(SUM(montant_total),0) AS total,
                COUNT(*) AS nb
            FROM transactions
            WHERE type = ? AND {where_date}
            GROUP BY categorie
            ORDER BY total DESC
        """, (type_tx, param_date)).fetchall()
        return [{
            'categorie': r['categorie'],
            'ligne_t2125': r['ligne_t2125'] or '',
            'ligne_tp80': r['ligne_tp80'] or '',
            'total': round(r['total'], 2),
            'nb': r['nb'],
        } for r in rows]

    revenus = _somme('revenu')
    depenses = _somme('depense')
    revenus_par_categorie = _par_categorie('revenu')
    depenses_par_categorie = _par_categorie('depense')

    parametres_row = conn.execute("SELECT * FROM parametres_facturation WHERE id = 1").fetchone()
    conn.close()

    profit_net = round(revenus['total'] - depenses['total'], 2)
    tps_a_remettre = round(revenus['tps'] - depenses['tps'], 2)
    tvq_a_remettre = round(revenus['tvq'] - depenses['tvq'], 2)

    return {
        'periode': periode_label,
        'annee': annee,
        'mois': mois or None,
        'revenus': revenus,
        'depenses': depenses,
        'profit_net': profit_net,
        'revenus_par_categorie': revenus_par_categorie,
        'depenses_par_categorie': depenses_par_categorie,
        'taxes': {
            'tps_percue': revenus['tps'],
            'tvq_percue': revenus['tvq'],
            'tps_payee': depenses['tps'],
            'tvq_payee': depenses['tvq'],
            'tps_a_remettre': tps_a_remettre,
            'tvq_a_remettre': tvq_a_remettre,
        },
        'nom_entreprise': (parametres_row['nom_entreprise'] if parametres_row and parametres_row['nom_entreprise'] else 'Mon entreprise'),
        'couleur_marque': (parametres_row['couleur_marque'] if parametres_row and parametres_row['couleur_marque'] else '#c0321a'),
    }

# ───────────────────────────────────────────────────────────
# Comptabilité — Bilan (agrégation lecture seule)
# ───────────────────────────────────────────────────────────
@app.route('/api/v1/admin/bilan', methods=['GET'])
@admin_required
def api_admin_bilan():
    annee = (request.args.get('annee') or '').strip()
    mois = (request.args.get('mois') or '').strip()
    if not annee:
        return jsonify({'error': "L'année est requise"}), 400
    return jsonify(_calculer_bilan(annee, mois))

import csv
import io as _io_export

def _nom_fichier_propre(txt):
    txt = (txt or '').strip()
    for c in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        txt = txt.replace(c, '')
    return txt or 'Rapport'

@app.route('/api/v1/admin/bilan/export', methods=['GET'])
@admin_required
def api_admin_bilan_export():
    annee = (request.args.get('annee') or '').strip()
    mois = (request.args.get('mois') or '').strip()
    fmt = (request.args.get('format') or 'csv').strip().lower()
    rapport = (request.args.get('rapport') or 'bilan').strip().lower()  # bilan | revenus | depenses

    if not annee:
        return jsonify({'error': "L'année est requise"}), 400
    if fmt not in ('csv', 'xlsx', 'pdf'):
        return jsonify({'error': 'Format invalide'}), 400

    data = _calculer_bilan(annee, mois)
    entreprise = _nom_fichier_propre(data['nom_entreprise'])
    periode_label = data['periode']

    noms_rapport = {
        'bilan': 'Bilan',
        'revenus': 'Revenus',
        'depenses': 'Dépenses',
    }
    nom_rapport = noms_rapport.get(rapport, 'Bilan')
    if mois:
        nom_periode = f"{nom_rapport} - {periode_label.replace('/', '-')}"
    else:
        nom_periode = f"{nom_rapport} - année {annee}" if annee != str(__import__('datetime').date.today().year) else f"{nom_rapport} - année en cours {annee}"

    nom_fichier_base = _nom_fichier_propre(f"{nom_periode} - {entreprise}")

    if rapport == 'revenus':
        lignes = data['revenus_par_categorie']
        total = data['revenus']['total']
    elif rapport == 'depenses':
        lignes = data['depenses_par_categorie']
        total = data['depenses']['total']
    else:
        lignes = None
        total = None

    # ── CSV ──
    if fmt == 'csv':
        buf = _io_export.StringIO()
        buf.write('﻿')  # BOM pour Excel FR
        writer = csv.writer(buf, delimiter=';')
        if rapport == 'bilan':
            writer.writerow(['Rapport', nom_rapport])
            writer.writerow(['Période', periode_label])
            writer.writerow([])
            writer.writerow(['Revenus totaux', f"{data['revenus']['total']:.2f}"])
            writer.writerow(['Dépenses totales', f"{data['depenses']['total']:.2f}"])
            writer.writerow(['Profit net', f"{data['profit_net']:.2f}"])
            writer.writerow([])
            writer.writerow(['TPS perçue', f"{data['taxes']['tps_percue']:.2f}"])
            writer.writerow(['TVQ perçue', f"{data['taxes']['tvq_percue']:.2f}"])
            writer.writerow(['TPS payée', f"{data['taxes']['tps_payee']:.2f}"])
            writer.writerow(['TVQ payée', f"{data['taxes']['tvq_payee']:.2f}"])
            writer.writerow(['TPS à remettre (estimation)', f"{data['taxes']['tps_a_remettre']:.2f}"])
            writer.writerow(['TVQ à remettre (estimation)', f"{data['taxes']['tvq_a_remettre']:.2f}"])
        else:
            writer.writerow(['Catégorie', 'Ligne T2125', 'Ligne TP-80', 'Montant', 'Nb transactions'])
            for l in lignes:
                writer.writerow([l['categorie'], l['ligne_t2125'], l['ligne_tp80'], f"{l['total']:.2f}", l['nb']])
            writer.writerow([])
            writer.writerow(['Total', '', '', f"{total:.2f}", ''])
        contenu = buf.getvalue().encode('utf-8')
        return Response(contenu, mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{nom_fichier_base}.csv"'})

    # ── XLSX ──
    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = nom_rapport[:31]
        bold = Font(bold=True)

        if rapport == 'bilan':
            ws.append([nom_rapport, periode_label])
            ws['A1'].font = Font(bold=True, size=14)
            ws.append([])
            ws.append(['Revenus totaux', data['revenus']['total']])
            ws.append(['Dépenses totales', data['depenses']['total']])
            ws.append(['Profit net', data['profit_net']])
            ws.append([])
            ws.append(['TPS perçue', data['taxes']['tps_percue']])
            ws.append(['TVQ perçue', data['taxes']['tvq_percue']])
            ws.append(['TPS payée', data['taxes']['tps_payee']])
            ws.append(['TVQ payée', data['taxes']['tvq_payee']])
            ws.append(['TPS à remettre (estimation)', data['taxes']['tps_a_remettre']])
            ws.append(['TVQ à remettre (estimation)', data['taxes']['tvq_a_remettre']])
            for row in ws.iter_rows(min_row=3, max_col=2):
                row[0].font = bold
                row[1].number_format = '#,##0.00 $'
            ws.column_dimensions['A'].width = 32
            ws.column_dimensions['B'].width = 18
        else:
            headers = ['Catégorie', 'Ligne T2125', 'Ligne TP-80', 'Montant', 'Nb transactions']
            ws.append(headers)
            for c in ws[1]:
                c.font = bold
            for l in lignes:
                ws.append([l['categorie'], l['ligne_t2125'], l['ligne_tp80'], l['total'], l['nb']])
            for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
                row[0].number_format = '#,##0.00 $'
            ws.append([])
            r = ws.max_row + 1
            ws.cell(row=r, column=1, value='Total').font = bold
            ws.cell(row=r, column=4, value=total).font = bold
            ws.cell(row=r, column=4).number_format = '#,##0.00 $'
            widths = [30, 14, 14, 14, 16]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = w

        buf = _io_export.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(buf.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{nom_fichier_base}.xlsx"'})

    # ── PDF ──
    if fmt == 'pdf':
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        buf = _io_export.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6*inch, bottomMargin=0.6*inch)
        styles = getSampleStyleSheet()
        accent = colors.HexColor(data['couleur_marque'] or '#c0321a')

        titre_style = ParagraphStyle('titre', parent=styles['Title'], textColor=accent, fontSize=20)
        sous_style = ParagraphStyle('sous', parent=styles['Normal'], textColor=colors.grey, fontSize=10)

        elements = []
        elements.append(Paragraph(data['nom_entreprise'], titre_style))
        elements.append(Paragraph(f"{nom_rapport} — {periode_label}", sous_style))
        elements.append(Spacer(1, 0.3*inch))

        if rapport == 'bilan':
            table_data = [
                ['Revenus totaux', f"{data['revenus']['total']:.2f} $"],
                ['Dépenses totales', f"{data['depenses']['total']:.2f} $"],
                ['Profit net', f"{data['profit_net']:.2f} $"],
            ]
            t = Table(table_data, colWidths=[3.5*inch, 2*inch])
            t.setStyle(TableStyle([
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                ('FONTNAME', (0,2), (-1,2), 'Helvetica-Bold'),
                ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 0.3*inch))
            elements.append(Paragraph('Taxes (estimation)', ParagraphStyle('h2', parent=styles['Heading2'], textColor=accent)))
            taxes_data = [
                ['TPS perçue', f"{data['taxes']['tps_percue']:.2f} $"],
                ['TVQ perçue', f"{data['taxes']['tvq_percue']:.2f} $"],
                ['TPS payée', f"{data['taxes']['tps_payee']:.2f} $"],
                ['TVQ payée', f"{data['taxes']['tvq_payee']:.2f} $"],
                ['TPS à remettre', f"{data['taxes']['tps_a_remettre']:.2f} $"],
                ['TVQ à remettre', f"{data['taxes']['tvq_a_remettre']:.2f} $"],
            ]
            t2 = Table(taxes_data, colWidths=[3.5*inch, 2*inch])
            t2.setStyle(TableStyle([
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('TOPPADDING', (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            elements.append(t2)
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph("Ces montants sont une estimation. Confirmez avec votre comptable avant de remettre.", sous_style))
        else:
            headers = ['Catégorie', 'Ligne T2125', 'Ligne TP-80', 'Montant']
            rows = [[l['categorie'], l['ligne_t2125'], l['ligne_tp80'], f"{l['total']:.2f} $"] for l in lignes]
            rows.append(['Total', '', '', f"{total:.2f} $"])
            t = Table([headers] + rows, colWidths=[2.3*inch, 1.2*inch, 1.2*inch, 1.3*inch])
            t.setStyle(TableStyle([
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
                ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('LINEBELOW', (0,0), (-1,0), 1, accent),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ]))
            elements.append(t)

        elements.append(Spacer(1, 0.4*inch))
        elements.append(Paragraph(f"Généré via CocktailOS", ParagraphStyle('foot', parent=styles['Normal'], textColor=colors.grey, fontSize=8)))

        doc.build(elements)
        buf.seek(0)
        return Response(buf.read(), mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename="{nom_fichier_base}.pdf"'})

# ───────────────────────────────────────────────────────────
# Mon site — édition du contenu Sanity par le client (self-service)
# ───────────────────────────────────────────────────────────
_SANITY_API_VERSION = '2021-06-07'
_SANITY_DATASET = 'production'

# Champs texte/image autorisés en écriture par page, PAR GABARIT — les deux gabarits
# réutilisent les mêmes noms de type de document (pageAccueil, siteSettings, etc.) mais
# avec des champs différents, donc un seul dict partagé casse silencieusement l'un des
# deux gabarits dès que l'autre est corrigé (vécu : le fix vitrine a effacé pageAPropos
# de sante, cassant l'auto-édition d'un client réel). Miroir des schémas Sanity de
# Site_web-cocktailmedia/templates/{vitrine,sante}/studio/schemas/*.ts.
MON_SITE_PAGE_FIELDS = {
    'vitrine': {
        'pageAccueil':  {'heroEyebrow', 'heroTitre', 'heroSousTitre', 'servicesEyebrow', 'servicesTitre',
                         'equipeTitle', 'faqTitle', 'ctaTitre', 'ctaSousTitre'},
        'pageServices': {'heroEyebrow', 'heroTitre', 'heroSousTitre', 'ctaTitre'},
        'pageEquipe':   {'heroEyebrow', 'heroTitre', 'heroSousTitre', 'missionTexte', 'visionTexte', 'valeursTexte', 'ctaTitre'},
        'pageContact':  {'heroTitre', 'heroSousTitre'},
        'siteSettings': {'adresse', 'telephone', 'courriel', 'instagram', 'facebook', 'linkedin'},
    },
    'sante': {
        'pageAccueil':  {'heroEyebrow', 'heroTitre', 'heroTitre2', 'heroSousTitre',
                         'traitementEyebrow', 'traitementTitre', 'traitementTexte1', 'traitementTexte2', 'traitementTexte3',
                         'servicesEyebrow', 'servicesTitre',
                         'approcheTitre', 'approcheTexte1', 'approcheTexte2', 'approcheImage',
                         'faqEyebrow', 'faqTitre', 'faqTexte'},
        'pageAPropos':  {'heroTitre', 'heroSousTitre', 'bioImage', 'bioTexte1', 'bioTexte2', 'missionTexte',
                         'formationTitre', 'formationTexte1'},
        'pageContact':  {'heroTitre', 'heroSousTitre'},
        'siteSettings': {'adresse', 'telephone', 'cellulaire', 'courriel', 'instagram', 'facebook', 'linkedin', 'assurances'},
    },
}

# Contenu répétable (collections de documents) éditable par le client, par gabarit.
MON_SITE_COLLECTION_FIELDS = {
    'vitrine': {
        'service': {'numero', 'eyebrow', 'titre', 'description', 'image', 'inclus', 'variations', 'featured'},
        'membre':  {'prenom', 'nom', 'titre', 'photo', 'bio', 'histoire', 'visible'},
        'faq':     {'question', 'reponse'},
    },
    'sante': {
        'service':     {'numero', 'eyebrow', 'titre', 'description', 'image', 'inclus', 'variations', 'featured', 'requiresRx'},
        'membre':      {'prenom', 'nom', 'titre', 'photo', 'bio', 'ordreProf', 'langues', 'visible'},
        'faq':         {'question', 'reponse', 'visible'},
        'temoignage':  {'nom', 'texte', 'service', 'note', 'visible'},
        'statistique': {'stat', 'texte', 'source'},
        'reference':   {'texte'},
    },
}

def _get_client_sites():
    """Retourne la liste de tous les sites (id, business_name, slug, template,
    sanity_project_id) liés au compte client connecté, du plus récent au plus ancien —
    un même client peut avoir plusieurs sites (ex: agence de test, ou client avec
    plusieurs mandats)."""
    conn = get_db_connection()
    try:
        client = conn.execute("SELECT email FROM clients WHERE id = ?", (session['user_id'],)).fetchone()
        if not client or not client['email']:
            return []
        rows = conn.execute(
            "SELECT id, business_name, slug, template, sanity_project_id FROM sites "
            "WHERE LOWER(client_email) = LOWER(?) AND sanity_project_id IS NOT NULL AND sanity_project_id != '' "
            "ORDER BY created_at DESC",
            (client['email'],)
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()

def _get_client_site(site_id=None):
    """Retourne (sanity_project_id, template) d'un site précis du compte client connecté
    (si site_id fourni et lui appartient), sinon le site le plus récent — (None, None)
    si aucun site trouvé. site_id optionnel garde la compatibilité pour les clients qui
    n'ont qu'un seul site (comportement historique inchangé)."""
    sites = _get_client_sites()
    if not sites:
        return None, None
    if site_id:
        for s in sites:
            if str(s['id']) == str(site_id):
                return s['sanity_project_id'], s['template']
        return None, None
    return sites[0]['sanity_project_id'], sites[0]['template']

@app.route('/api/v1/mon-site/sites', methods=['GET'])
@login_required
def api_mon_site_sites():
    sites = _get_client_sites()
    return jsonify([
        {'id': s['id'], 'business_name': s['business_name'], 'slug': s['slug'], 'template': s['template']}
        for s in sites
    ])

@app.route('/api/v1/mon-site/info', methods=['GET'])
@login_required
def api_mon_site_info():
    project_id, template = _get_client_site(request.args.get('site_id'))
    if not project_id:
        return jsonify({'error': 'Aucun site associé à ce compte'}), 404
    return jsonify({
        'template': template,
        'sanity_project_id': project_id,
        'pages': sorted(MON_SITE_PAGE_FIELDS.get(template, {}).keys()),
        'collections': sorted(MON_SITE_COLLECTION_FIELDS.get(template, {}).keys()),
    })

def _sanity_query(project_id, groq):
    r = _req.get(
        f'https://{project_id}.api.sanity.io/v{_SANITY_API_VERSION}/data/query/{_SANITY_DATASET}',
        params={'query': groq},
        headers={'Authorization': f'Bearer {_SANITY_TOKEN}'},
    )
    if r.status_code != 200:
        raise RuntimeError(f"Sanity query failed: {r.text}")
    return r.json().get('result')

def _sanity_mutate(project_id, mutations):
    r = _req.post(
        f'https://{project_id}.api.sanity.io/v{_SANITY_API_VERSION}/data/mutate/{_SANITY_DATASET}',
        headers={'Authorization': f'Bearer {_SANITY_TOKEN}', 'Content-Type': 'application/json'},
        json={'mutations': mutations},
    )
    if r.status_code not in (200, 201):
        raise RuntimeError(f"Sanity mutate failed: {r.text}")
    return r.json()

@app.route('/api/v1/mon-site/contenu/<page>', methods=['GET', 'PUT'])
@login_required
def api_mon_site_contenu(page):
    project_id, template = _get_client_site(request.args.get('site_id'))
    if not project_id:
        return jsonify({'error': 'Aucun site associé à ce compte'}), 404

    page_fields = MON_SITE_PAGE_FIELDS.get(template, {})
    if page not in page_fields:
        return jsonify({'error': 'Page inconnue'}), 400

    if request.method == 'GET':
        try:
            doc = _sanity_query(project_id, f'*[_type == "{page}"][0]')
        except RuntimeError:
            return jsonify({'error': 'Erreur de communication avec Sanity'}), 502
        return jsonify(doc or {})

    # PUT — seulement les champs reconnus du schéma de ce gabarit, jamais _id/_type/_rev
    data = request.get_json(silent=True) or {}
    fields = {k: v for k, v in data.items() if k in page_fields[page]}
    if not fields:
        return jsonify({'error': 'Aucun champ valide à mettre à jour'}), 400

    try:
        existing_id = _sanity_query(project_id, f'*[_type == "{page}"][0]._id')
        doc_id = existing_id or page  # singleton : id de repli = nom du type si le document n'existe pas encore
        _sanity_mutate(project_id, [
            {'createIfNotExists': {'_id': doc_id, '_type': page}},
            {'patch': {'id': doc_id, 'set': fields}},
        ])
    except RuntimeError:
        return jsonify({'error': 'Erreur lors de la sauvegarde sur Sanity'}), 502

    return jsonify({'success': True})

def _sanity_upload_image(project_id, content_bytes, content_type):
    """Upload un asset image vers Sanity, retourne la référence à stocker dans un
    champ de type `image` (ex: {photo: <retour de cette fonction>})."""
    r = _req.post(
        f'https://{project_id}.api.sanity.io/v{_SANITY_API_VERSION}/assets/images/{_SANITY_DATASET}',
        headers={'Authorization': f'Bearer {_SANITY_TOKEN}', 'Content-Type': content_type or 'application/octet-stream'},
        data=content_bytes,
    )
    if r.status_code not in (200, 201):
        raise RuntimeError(f"Sanity asset upload failed: {r.text}")
    asset_id = r.json()['document']['_id']
    return {'_type': 'image', 'asset': {'_type': 'reference', '_ref': asset_id}}

@app.route('/api/v1/mon-site/image', methods=['POST'])
@login_required
def api_mon_site_image():
    project_id, template = _get_client_site(request.args.get('site_id'))
    if not project_id:
        return jsonify({'error': 'Aucun site associé à ce compte'}), 404

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'Aucun fichier fourni'}), 400

    try:
        image_ref = _sanity_upload_image(project_id, file.read(), file.mimetype)
    except RuntimeError:
        return jsonify({'error': "Erreur lors de l'envoi de l'image à Sanity"}), 502

    return jsonify(image_ref)

def _verify_doc_type(project_id, doc_id, expected_type):
    """Confirme que doc_id est bien du type attendu avant patch/delete — défense en
    profondeur en plus du scoping par projet (chaque client n'a accès qu'à son propre
    dataset Sanity, donc le risque réel est une faute de frappe côté client, pas une
    fuite inter-tenant)."""
    actual = _sanity_query(project_id, f'*[_id == "{doc_id}"][0]._type')
    return actual == expected_type

@app.route('/api/v1/mon-site/collection/<doc_type>', methods=['GET', 'POST'])
@login_required
def api_mon_site_collection(doc_type):
    project_id, template = _get_client_site(request.args.get('site_id'))
    if not project_id:
        return jsonify({'error': 'Aucun site associé à ce compte'}), 404

    collection_fields = MON_SITE_COLLECTION_FIELDS.get(template, {})
    if doc_type not in collection_fields:
        return jsonify({'error': 'Collection inconnue'}), 400

    if request.method == 'GET':
        try:
            docs = _sanity_query(project_id, f'*[_type == "{doc_type}"] | order(ordre asc)')
        except RuntimeError:
            return jsonify({'error': 'Erreur de communication avec Sanity'}), 502
        return jsonify(docs or [])

    # POST — créer un nouvel item, ordre = max existant + 1
    data = request.get_json(silent=True) or {}
    fields = {k: v for k, v in data.items() if k in collection_fields[doc_type]}
    try:
        max_ordre = _sanity_query(project_id, f'*[_type == "{doc_type}"] | order(ordre desc)[0].ordre') or 0
        new_doc = {'_type': doc_type, 'ordre': max_ordre + 1, **fields}
        if 'visible' in collection_fields[doc_type] and 'visible' not in fields:
            new_doc['visible'] = True
        result = _sanity_mutate(project_id, [{'create': new_doc}])
        new_id = result['results'][0]['id']
    except RuntimeError:
        return jsonify({'error': 'Erreur lors de la création'}), 502

    return jsonify({'success': True, 'id': new_id}), 201

@app.route('/api/v1/mon-site/collection/<doc_type>/reorder', methods=['PUT'])
@login_required
def api_mon_site_collection_reorder(doc_type):
    project_id, template = _get_client_site(request.args.get('site_id'))
    if not project_id:
        return jsonify({'error': 'Aucun site associé à ce compte'}), 404
    if doc_type not in MON_SITE_COLLECTION_FIELDS.get(template, {}):
        return jsonify({'error': 'Collection inconnue'}), 400

    order = (request.get_json(silent=True) or {}).get('order') or []
    if not isinstance(order, list) or not order:
        return jsonify({'error': 'Liste order requise'}), 400

    try:
        mutations = [
            {'patch': {'id': doc_id, 'set': {'ordre': i + 1}}}
            for i, doc_id in enumerate(order)
        ]
        _sanity_mutate(project_id, mutations)
    except RuntimeError:
        return jsonify({'error': 'Erreur lors du réordonnancement'}), 502

    return jsonify({'success': True})

@app.route('/api/v1/mon-site/collection/<doc_type>/<doc_id>', methods=['PUT', 'DELETE'])
@login_required
def api_mon_site_collection_item(doc_type, doc_id):
    project_id, template = _get_client_site(request.args.get('site_id'))
    if not project_id:
        return jsonify({'error': 'Aucun site associé à ce compte'}), 404

    collection_fields = MON_SITE_COLLECTION_FIELDS.get(template, {})
    if doc_type not in collection_fields:
        return jsonify({'error': 'Collection inconnue'}), 400

    try:
        if not _verify_doc_type(project_id, doc_id, doc_type):
            return jsonify({'error': 'Item introuvable'}), 404
    except RuntimeError:
        return jsonify({'error': 'Erreur de communication avec Sanity'}), 502

    if request.method == 'DELETE':
        try:
            _sanity_mutate(project_id, [{'delete': {'id': doc_id}}])
        except RuntimeError:
            return jsonify({'error': 'Erreur lors de la suppression'}), 502
        return jsonify({'success': True})

    # PUT — patch des champs reconnus seulement
    data = request.get_json(silent=True) or {}
    fields = {k: v for k, v in data.items() if k in collection_fields[doc_type]}
    if not fields:
        return jsonify({'error': 'Aucun champ valide à mettre à jour'}), 400

    try:
        _sanity_mutate(project_id, [{'patch': {'id': doc_id, 'set': fields}}])
    except RuntimeError:
        return jsonify({'error': 'Erreur lors de la sauvegarde'}), 502

    return jsonify({'success': True})

# Exempter toutes les routes /api/v1/ du CSRF — endpoints JSON appelés par Next.js,
# protégés par la politique CORS (origins restreintes) et non par formulaires HTML.
for _rule in app.url_map.iter_rules():
    if _rule.rule.startswith('/api/v1/'):
        _view = app.view_functions.get(_rule.endpoint)
        if _view:
            csrf.exempt(_view)

if __name__ == '__main__':
    app.run(debug=True)
